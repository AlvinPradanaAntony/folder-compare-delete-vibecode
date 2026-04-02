import csv
import hashlib
import os
import queue
import threading
import tkinter as tk
from dataclasses import dataclass, field
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Dict, Iterable, List, Optional, Tuple

APP_TITLE = "Folder Compare & Delete"
CHUNK_SIZE = 1024 * 1024  # 1 MB


@dataclass
class FileRecord:
    path: Path
    base_folder: Path
    base_label: str
    relative_path: str
    size: int
    sha256: Optional[str] = None


@dataclass
class MatchResult:
    target_path: Path
    target_relative_path: str
    size: int
    match_type: str
    exact_matches: List[FileRecord] = field(default_factory=list)
    same_name_different_content: List[FileRecord] = field(default_factory=list)
    missing_from_folders: List[str] = field(default_factory=list)
    only_in_target: bool = False

    @property
    def exact_folder_labels(self) -> str:
        labels = sorted({item.base_label for item in self.exact_matches})
        return ", ".join(labels) if labels else "-"

    @property
    def exact_paths_text(self) -> str:
        if not self.exact_matches:
            return "-"
        lines = [f"[{item.base_label}] {item.path}" for item in self.exact_matches]
        return " | ".join(lines)

    @property
    def diff_paths_text(self) -> str:
        if not self.same_name_different_content:
            return "-"
        lines = [f"[{item.base_label}] {item.path}" for item in self.same_name_different_content]
        return " | ".join(lines)

    @property
    def status_text(self) -> str:
        if self.exact_matches:
            return "Duplikat ditemukan"
        if self.same_name_different_content:
            return "Nama sama, isi berbeda"
        return "Hanya ada di Folder A"

    @property
    def tree_tag(self) -> str:
        if self.exact_matches:
            return "exact_match"
        if self.same_name_different_content:
            return "different_content"
        return "only_target"


class FolderCompareDeleteApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1450x840")
        self.root.minsize(1180, 720)

        self.target_folder_var = tk.StringVar()
        self.compare_mode_var = tk.StringVar(value="name_size")
        self.include_subfolders_var = tk.BooleanVar(value=True)
        self.delete_mode_var = tk.StringVar(value="recycle_bin")
        self.show_only_matches_var = tk.BooleanVar(value=False)
        self.status_var = tk.StringVar(value="Siap.")
        self.progress_var = tk.DoubleVar(value=0.0)

        self.compare_folder_vars: List[tk.StringVar] = []
        self.result_rows: List[MatchResult] = []
        self.scan_thread: Optional[threading.Thread] = None
        self.ui_queue: queue.Queue = queue.Queue()
        self.openpyxl_available = False
        self._tree_item_to_result: Dict[str, MatchResult] = {}

        self._build_ui()
        self._poll_queue()
        self._check_excel_support()

    def _build_ui(self) -> None:
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill="both", expand=True)

        title = ttk.Label(
            main,
            text="Bandingkan folder dan hapus file duplikat dari Folder A",
            font=("Segoe UI", 14, "bold"),
        )
        title.pack(anchor="w", pady=(0, 12))

        config = ttk.LabelFrame(main, text="Konfigurasi", padding=12)
        config.pack(fill="x")
        config.columnconfigure(1, weight=1)

        self._folder_row(config, "Folder A (target hapus):", self.target_folder_var, 0)

        compare_container = ttk.LabelFrame(config, text="Folder pembanding", padding=8)
        compare_container.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(8, 0))
        compare_container.columnconfigure(0, weight=1)
        self.compare_list_frame = ttk.Frame(compare_container)
        self.compare_list_frame.pack(fill="x", expand=True)

        compare_buttons = ttk.Frame(compare_container)
        compare_buttons.pack(fill="x", pady=(8, 0))
        ttk.Button(compare_buttons, text="+ Tambah Folder Pembanding", command=self.add_compare_folder_row).pack(side="left")
        ttk.Button(compare_buttons, text="- Hapus Folder Terakhir", command=self.remove_compare_folder_row).pack(side="left", padx=(8, 0))

        self.add_compare_folder_row()
        self.add_compare_folder_row()

        options = ttk.Frame(config)
        options.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(12, 0))
        options.columnconfigure(0, weight=1)
        options.columnconfigure(1, weight=1)

        match_frame = ttk.LabelFrame(options, text="Metode pencocokan", padding=8)
        match_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))

        ttk.Radiobutton(
            match_frame,
            text="Nama file + ukuran (cepat)",
            variable=self.compare_mode_var,
            value="name_size",
        ).pack(anchor="w")
        ttk.Radiobutton(
            match_frame,
            text="Hash SHA-256 + ukuran (akurat)",
            variable=self.compare_mode_var,
            value="hash",
        ).pack(anchor="w")

        other_frame = ttk.LabelFrame(options, text="Opsi lain", padding=8)
        other_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

        ttk.Checkbutton(
            other_frame,
            text="Sertakan subfolder",
            variable=self.include_subfolders_var,
        ).pack(anchor="w")
        ttk.Checkbutton(
            other_frame,
            text="Tampilkan hanya file yang punya kecocokan / perbedaan",
            variable=self.show_only_matches_var,
        ).pack(anchor="w")

        ttk.Radiobutton(
            other_frame,
            text="Hapus ke Recycle Bin (disarankan)",
            variable=self.delete_mode_var,
            value="recycle_bin",
        ).pack(anchor="w")
        ttk.Radiobutton(
            other_frame,
            text="Hapus permanen",
            variable=self.delete_mode_var,
            value="permanent",
        ).pack(anchor="w")

        buttons = ttk.Frame(main)
        buttons.pack(fill="x", pady=12)

        self.scan_button = ttk.Button(buttons, text="Scan & Bandingkan", command=self.start_scan)
        self.scan_button.pack(side="left")

        self.export_csv_button = ttk.Button(
            buttons,
            text="Simpan ke CSV",
            command=self.export_csv,
            state="disabled",
        )
        self.export_csv_button.pack(side="left", padx=(8, 0))

        self.export_excel_button = ttk.Button(
            buttons,
            text="Simpan ke Excel",
            command=self.export_excel,
            state="disabled",
        )
        self.export_excel_button.pack(side="left", padx=(8, 0))

        self.delete_button = ttk.Button(
            buttons,
            text="Hapus File Terpilih",
            command=self.delete_selected,
            state="disabled",
        )
        self.delete_button.pack(side="left", padx=(8, 0))

        self.delete_all_button = ttk.Button(
            buttons,
            text="Hapus Semua Duplikat Hijau",
            command=self.delete_all_results,
            state="disabled",
        )
        self.delete_all_button.pack(side="left", padx=(8, 0))

        self.clear_button = ttk.Button(buttons, text="Bersihkan Hasil", command=self.clear_results)
        self.clear_button.pack(side="left", padx=(8, 0))

        progress_frame = ttk.Frame(main)
        progress_frame.pack(fill="x", pady=(0, 10))
        self.progress = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100, mode="determinate")
        self.progress.pack(fill="x", side="left", expand=True)
        self.progress_label = ttk.Label(progress_frame, text="0%")
        self.progress_label.pack(side="left", padx=(8, 0))

        result_frame = ttk.LabelFrame(
            main,
            text="Hasil perbandingan visual (hijau = cocok/aman dihapus, merah = berbeda/tidak ada)",
            padding=8,
        )
        result_frame.pack(fill="both", expand=True)

        columns = (
            "status",
            "target",
            "target_rel",
            "size",
            "found_in",
            "exact_paths",
            "diff_paths",
            "missing_in",
            "match_type",
        )
        self.tree = ttk.Treeview(result_frame, columns=columns, show="headings", selectmode="extended")
        self.tree.heading("status", text="Status")
        self.tree.heading("target", text="Path File di Folder A")
        self.tree.heading("target_rel", text="Relative Path")
        self.tree.heading("size", text="Ukuran")
        self.tree.heading("found_in", text="Ditemukan di Folder")
        self.tree.heading("exact_paths", text="Path File yang Cocok")
        self.tree.heading("diff_paths", text="Path Nama Sama Isi Beda")
        self.tree.heading("missing_in", text="Tidak Ada di Folder")
        self.tree.heading("match_type", text="Pencocokan")

        self.tree.column("status", width=170, anchor="center")
        self.tree.column("target", width=320, anchor="w")
        self.tree.column("target_rel", width=220, anchor="w")
        self.tree.column("size", width=100, anchor="center")
        self.tree.column("found_in", width=150, anchor="center")
        self.tree.column("exact_paths", width=340, anchor="w")
        self.tree.column("diff_paths", width=320, anchor="w")
        self.tree.column("missing_in", width=180, anchor="center")
        self.tree.column("match_type", width=120, anchor="center")

        self.tree.tag_configure("exact_match", background="#d9f2d9")
        self.tree.tag_configure("different_content", background="#ffd9d9")
        self.tree.tag_configure("only_target", background="#ffe7cc")

        xscroll = ttk.Scrollbar(result_frame, orient="horizontal", command=self.tree.xview)
        yscroll = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(xscrollcommand=xscroll.set, yscrollcommand=yscroll.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)

        legend = ttk.Label(
            main,
            text="Hijau: file di Folder A punya pasangan sama di folder lain | Merah: nama sama tetapi isi/ukuran berbeda | Oranye: hanya ada di Folder A",
        )
        legend.pack(anchor="w", pady=(8, 0))

        status_bar = ttk.Label(main, textvariable=self.status_var, relief="sunken", anchor="w")
        status_bar.pack(fill="x", pady=(10, 0))

    def _check_excel_support(self) -> None:
        try:
            import openpyxl  # noqa: F401
            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False
            self.export_excel_button.config(state="disabled")

    def add_compare_folder_row(self) -> None:
        row_index = len(self.compare_folder_vars)
        variable = tk.StringVar()
        self.compare_folder_vars.append(variable)

        row = ttk.Frame(self.compare_list_frame)
        row.pack(fill="x", pady=3)
        setattr(variable, "_row_widget", row)

        ttk.Label(row, text=f"Folder pembanding {row_index + 1}:", width=22).pack(side="left")
        ttk.Entry(row, textvariable=variable).pack(side="left", fill="x", expand=True, padx=8)
        ttk.Button(row, text="Pilih...", command=lambda v=variable: self.pick_folder(v)).pack(side="left")

    def remove_compare_folder_row(self) -> None:
        if len(self.compare_folder_vars) <= 1:
            messagebox.showinfo(APP_TITLE, "Minimal harus ada satu folder pembanding.")
            return
        variable = self.compare_folder_vars.pop()
        row = getattr(variable, "_row_widget", None)
        if row is not None:
            row.destroy()

    def _folder_row(self, parent: ttk.Widget, label: str, variable: tk.StringVar, row: int) -> None:
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=4)
        entry = ttk.Entry(parent, textvariable=variable)
        entry.grid(row=row, column=1, sticky="ew", padx=8, pady=4)
        ttk.Button(parent, text="Pilih...", command=lambda v=variable: self.pick_folder(v)).grid(
            row=row, column=2, sticky="ew", pady=4
        )
        parent.columnconfigure(1, weight=1)

    def pick_folder(self, variable: tk.StringVar) -> None:
        folder = filedialog.askdirectory()
        if folder:
            variable.set(folder)

    def start_scan(self) -> None:
        if self.scan_thread and self.scan_thread.is_alive():
            messagebox.showinfo(APP_TITLE, "Proses scan masih berjalan.")
            return

        target_folder = self.target_folder_var.get().strip()
        compare_folders = []
        for index, var in enumerate(self.compare_folder_vars, start=1):
            folder_value = var.get().strip()
            if folder_value:
                compare_folders.append(folder_value)

        if not target_folder:
            messagebox.showwarning(APP_TITLE, "Folder A wajib diisi.")
            return
        if not compare_folders:
            messagebox.showwarning(APP_TITLE, "Minimal isi satu folder pembanding.")
            return
        if any(Path(target_folder) == Path(c) for c in compare_folders):
            messagebox.showwarning(APP_TITLE, "Folder A tidak boleh sama dengan folder pembanding.")
            return

        self.scan_button.config(state="disabled")
        self.delete_button.config(state="disabled")
        self.delete_all_button.config(state="disabled")
        self.export_csv_button.config(state="disabled")
        self.export_excel_button.config(state="disabled")
        self.clear_results(reset_status=False)
        self._set_progress(0, "Memulai scan...")

        self.scan_thread = threading.Thread(
            target=self._scan_worker,
            args=(target_folder, compare_folders, self.compare_mode_var.get(), self.include_subfolders_var.get()),
            daemon=True,
        )
        self.scan_thread.start()

    def _scan_worker(self, target_folder: str, compare_folders: List[str], mode: str, include_subfolders: bool) -> None:
        try:
            self._queue_progress(5, "Mengumpulkan file Folder A...")
            target_files = self._collect_files(Path(target_folder), include_subfolders, base_label="A")

            compare_groups: List[Tuple[str, Path, List[FileRecord]]] = []
            total_compare = len(compare_folders)
            for idx, folder in enumerate(compare_folders, start=1):
                label = f"F{idx}"
                percent = 5 + (idx / max(total_compare, 1)) * 25
                self._queue_progress(percent, f"Mengumpulkan file pembanding {label}...")
                compare_records = self._collect_files(Path(folder), include_subfolders, base_label=label)
                compare_groups.append((label, Path(folder), compare_records))

            self._queue_progress(35, "Menyusun indeks pembanding...")
            results = self._build_comparison_results(target_files, compare_groups, mode)
            self.ui_queue.put(("scan_done", results))
        except Exception as exc:
            self.ui_queue.put(("scan_error", str(exc)))

    def _build_comparison_results(
        self,
        target_files: List[FileRecord],
        compare_groups: List[Tuple[str, Path, List[FileRecord]]],
        mode: str,
    ) -> List[MatchResult]:
        by_name_size: Dict[Tuple[str, int], List[FileRecord]] = {}
        by_relative_size: Dict[Tuple[str, int], List[FileRecord]] = {}
        by_relative_name: Dict[str, List[FileRecord]] = {}
        by_hash: Dict[Tuple[int, str], List[FileRecord]] = {}
        all_compare_labels = [label for label, _, _ in compare_groups]

        flat_compare_records: List[FileRecord] = []
        for label, _, records in compare_groups:
            for record in records:
                flat_compare_records.append(record)
                by_name_size.setdefault((record.path.name.lower(), record.size), []).append(record)
                by_relative_size.setdefault((record.relative_path.lower(), record.size), []).append(record)
                by_relative_name.setdefault(record.relative_path.lower(), []).append(record)

        if mode == "hash":
            total_hash_items = len(flat_compare_records) + len(target_files)
            processed = 0
            for record in flat_compare_records:
                record.sha256 = self._hash_file(record.path)
                by_hash.setdefault((record.size, record.sha256), []).append(record)
                processed += 1
                self._queue_progress(35 + (processed / max(total_hash_items, 1)) * 45, f"Hash file pembanding {processed}/{total_hash_items}...")

        results: List[MatchResult] = []
        total_targets = len(target_files)
        base_progress = 80 if mode == "hash" else 40
        progress_span = 18

        for idx, target in enumerate(target_files, start=1):
            exact_matches: List[FileRecord] = []
            same_name_different_content: List[FileRecord] = []

            if mode == "name_size":
                exact_matches.extend(by_name_size.get((target.path.name.lower(), target.size), []))
                exact_matches.extend(
                    item
                    for item in by_relative_size.get((target.relative_path.lower(), target.size), [])
                    if item not in exact_matches
                )
                same_name_different_content = [
                    item for item in by_relative_name.get(target.relative_path.lower(), []) if item.size != target.size
                ]
                match_type = "nama+ukuran"
            else:
                target.sha256 = self._hash_file(target.path)
                exact_matches = by_hash.get((target.size, target.sha256), [])
                same_name_different_content = [
                    item
                    for item in by_relative_name.get(target.relative_path.lower(), [])
                    if item.sha256 != target.sha256 or item.size != target.size
                ]
                match_type = "hash+ukuran"
                self._queue_progress(
                    35 + ((len(flat_compare_records) + idx) / max(len(flat_compare_records) + len(target_files), 1)) * 45,
                    f"Hash file target {idx}/{total_targets}...",
                )

            exact_labels = {item.base_label for item in exact_matches}
            diff_labels = {item.base_label for item in same_name_different_content}
            involved_labels = exact_labels.union(diff_labels)
            missing = [label for label in all_compare_labels if label not in involved_labels]

            result = MatchResult(
                target_path=target.path,
                target_relative_path=target.relative_path,
                size=target.size,
                match_type=match_type,
                exact_matches=sorted(exact_matches, key=lambda x: (x.base_label, str(x.path))),
                same_name_different_content=sorted(same_name_different_content, key=lambda x: (x.base_label, str(x.path))),
                missing_from_folders=missing,
                only_in_target=not exact_matches and not same_name_different_content,
            )
            results.append(result)
            self._queue_progress(base_progress + (idx / max(total_targets, 1)) * progress_span, f"Membandingkan file {idx}/{total_targets}...")

        self._queue_progress(99, "Menyiapkan tampilan hasil...")
        return results

    def _collect_files(self, folder: Path, include_subfolders: bool, base_label: str) -> List[FileRecord]:
        if not folder.exists() or not folder.is_dir():
            raise FileNotFoundError(f"Folder tidak valid: {folder}")

        records: List[FileRecord] = []
        iterator: Iterable[Path] = folder.rglob("*") if include_subfolders else folder.glob("*")
        for path in iterator:
            if path.is_file():
                try:
                    size = path.stat().st_size
                except OSError:
                    continue
                records.append(
                    FileRecord(
                        path=path,
                        base_folder=folder,
                        base_label=base_label,
                        relative_path=str(path.relative_to(folder)),
                        size=size,
                    )
                )
        return records

    def _hash_file(self, path: Path) -> str:
        sha = hashlib.sha256()
        with path.open("rb") as f:
            while True:
                chunk = f.read(CHUNK_SIZE)
                if not chunk:
                    break
                sha.update(chunk)
        return sha.hexdigest()

    def _format_size(self, size: int) -> str:
        value = float(size)
        units = ["B", "KB", "MB", "GB", "TB"]
        for unit in units:
            if value < 1024 or unit == units[-1]:
                return f"{value:.2f} {unit}" if unit != "B" else f"{int(value)} B"
            value /= 1024
        return f"{size} B"

    def _set_progress(self, value: float, text: str) -> None:
        self.progress_var.set(max(0.0, min(100.0, value)))
        self.progress_label.config(text=f"{int(self.progress_var.get())}%")
        self.status_var.set(text)
        self.root.update_idletasks()

    def _queue_progress(self, value: float, text: str) -> None:
        self.ui_queue.put(("progress", (value, text)))

    def _poll_queue(self) -> None:
        try:
            while True:
                message = self.ui_queue.get_nowait()
                self._handle_queue_message(message)
        except queue.Empty:
            pass
        self.root.after(100, self._poll_queue)

    def _handle_queue_message(self, message: Tuple[str, object]) -> None:
        kind, payload = message
        if kind == "progress":
            value, text = payload  # type: ignore[misc]
            self._set_progress(float(value), str(text))

        elif kind == "scan_done":
            self.scan_button.config(state="normal")
            self.result_rows = payload  # type: ignore[assignment]
            self._populate_tree()
            self._set_progress(100, "Scan selesai.")

            duplicate_count = sum(1 for row in self.result_rows if row.exact_matches)
            diff_count = sum(1 for row in self.result_rows if row.same_name_different_content)
            only_count = sum(1 for row in self.result_rows if row.only_in_target)
            self.status_var.set(
                f"Scan selesai. Hijau: {duplicate_count}, Merah: {diff_count}, Oranye: {only_count}."
            )

            self.export_csv_button.config(state="normal")
            if self.openpyxl_available:
                self.export_excel_button.config(state="normal")
            if duplicate_count > 0:
                self.delete_button.config(state="normal")
                self.delete_all_button.config(state="normal")
            else:
                messagebox.showinfo(APP_TITLE, "Tidak ada file duplikat hijau untuk dihapus. Cek hasil merah/oranye untuk analisis.")

        elif kind == "scan_error":
            self.scan_button.config(state="normal")
            self._set_progress(0, "Terjadi kesalahan saat scan.")
            messagebox.showerror(APP_TITLE, str(payload))

    def _populate_tree(self) -> None:
        self._tree_item_to_result.clear()
        for item in self.tree.get_children():
            self.tree.delete(item)

        filtered_rows = self.result_rows
        if self.show_only_matches_var.get():
            filtered_rows = [row for row in self.result_rows if row.exact_matches or row.same_name_different_content]

        for row in filtered_rows:
            item_id = self.tree.insert(
                "",
                "end",
                values=(
                    row.status_text,
                    str(row.target_path),
                    row.target_relative_path,
                    self._format_size(row.size),
                    row.exact_folder_labels,
                    row.exact_paths_text,
                    row.diff_paths_text,
                    ", ".join(row.missing_from_folders) if row.missing_from_folders else "-",
                    row.match_type,
                ),
                tags=(row.tree_tag,),
            )
            self._tree_item_to_result[item_id] = row

    def clear_results(self, reset_status: bool = True) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.result_rows = []
        self._tree_item_to_result.clear()
        self.delete_button.config(state="disabled")
        self.delete_all_button.config(state="disabled")
        self.export_csv_button.config(state="disabled")
        self.export_excel_button.config(state="disabled")
        self.progress_var.set(0.0)
        self.progress_label.config(text="0%")
        if reset_status:
            self.status_var.set("Hasil dibersihkan.")

    def delete_selected(self) -> None:
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo(APP_TITLE, "Pilih minimal satu file dari hasil scan.")
            return

        deletable = []
        tree_items = []
        skipped = 0
        for item in selected:
            result = self._tree_item_to_result.get(item)
            if result and result.exact_matches:
                deletable.append(result.target_path)
                tree_items.append(item)
            else:
                skipped += 1

        if not deletable:
            messagebox.showinfo(APP_TITLE, "Pilihan Anda tidak mengandung file hijau yang aman dihapus.")
            return
        if skipped:
            messagebox.showinfo(APP_TITLE, f"{skipped} item merah/oranye diabaikan karena tidak aman untuk dihapus otomatis.")

        self._confirm_and_delete(deletable, tree_items)

    def delete_all_results(self) -> None:
        tree_items: List[str] = []
        paths: List[Path] = []
        for item, result in self._tree_item_to_result.items():
            if result.exact_matches:
                tree_items.append(item)
                paths.append(result.target_path)

        if not paths:
            messagebox.showinfo(APP_TITLE, "Belum ada hasil hijau untuk dihapus.")
            return

        self._confirm_and_delete(paths, tree_items)

    def _confirm_and_delete(self, paths: List[Path], tree_items: List[str]) -> None:
        preview = "\n".join(str(p) for p in paths[:10])
        extra = "" if len(paths) <= 10 else f"\n... dan {len(paths) - 10} file lainnya"
        mode_text = "Recycle Bin" if self.delete_mode_var.get() == "recycle_bin" else "hapus permanen"

        confirmed = messagebox.askyesno(
            APP_TITLE,
            f"Yakin ingin memproses {len(paths)} file HIJAU dari Folder A?\n\n"
            f"Mode: {mode_text}\n\n"
            f"Contoh file:\n{preview}{extra}",
            icon="warning",
        )
        if not confirmed:
            return

        deleted_count = 0
        errors: List[str] = []
        use_recycle_bin = self.delete_mode_var.get() == "recycle_bin"

        for path, item in zip(paths, tree_items):
            try:
                if use_recycle_bin:
                    self._send_to_recycle_bin(path)
                else:
                    path.unlink()
                if item in self.tree.get_children():
                    self.tree.delete(item)
                deleted_count += 1
            except Exception as exc:
                errors.append(f"{path}: {exc}")

        self.result_rows = [row for row in self.result_rows if row.target_path not in set(paths)]
        self._tree_item_to_result = {
            item_id: result
            for item_id, result in self._tree_item_to_result.items()
            if item_id in self.tree.get_children()
        }

        remaining = len(self.tree.get_children())
        self.status_var.set(f"Selesai menghapus {deleted_count} file. Sisa hasil terlihat: {remaining}.")

        if remaining == 0:
            self.delete_button.config(state="disabled")
            self.delete_all_button.config(state="disabled")

        if errors:
            messagebox.showwarning(
                APP_TITLE,
                f"Sebagian file gagal diproses.\n\n{os.linesep.join(errors[:10])}",
            )
        else:
            messagebox.showinfo(APP_TITLE, f"Berhasil memproses {deleted_count} file.")

    def export_csv(self) -> None:
        if not self.result_rows:
            messagebox.showinfo(APP_TITLE, "Belum ada hasil scan untuk disimpan.")
            return

        save_path = filedialog.asksaveasfilename(
            title="Simpan hasil ke CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
        )
        if not save_path:
            return

        with open(save_path, "w", newline="", encoding="utf-8-sig") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(
                [
                    "status",
                    "target_path",
                    "target_relative_path",
                    "size_bytes",
                    "size_display",
                    "found_in_folders",
                    "exact_match_paths",
                    "different_content_paths",
                    "missing_in_folders",
                    "match_type",
                ]
            )
            for row in self.result_rows:
                writer.writerow(
                    [
                        row.status_text,
                        str(row.target_path),
                        row.target_relative_path,
                        row.size,
                        self._format_size(row.size),
                        row.exact_folder_labels,
                        row.exact_paths_text,
                        row.diff_paths_text,
                        ", ".join(row.missing_from_folders) if row.missing_from_folders else "-",
                        row.match_type,
                    ]
                )
        messagebox.showinfo(APP_TITLE, f"Berhasil menyimpan CSV:\n{save_path}")

    def export_excel(self) -> None:
        if not self.result_rows:
            messagebox.showinfo(APP_TITLE, "Belum ada hasil scan untuk disimpan.")
            return
        if not self.openpyxl_available:
            messagebox.showwarning(APP_TITLE, "Modul openpyxl belum terpasang. Jalankan: pip install openpyxl")
            return

        save_path = filedialog.asksaveasfilename(
            title="Simpan hasil ke Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not save_path:
            return

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Hasil Scan"
        headers = [
            "Status",
            "Path File di Folder A",
            "Relative Path",
            "Ukuran (bytes)",
            "Ukuran",
            "Ditemukan di Folder",
            "Path File yang Cocok",
            "Path Nama Sama Isi Beda",
            "Tidak Ada di Folder",
            "Pencocokan",
        ]
        ws.append(headers)

        fill_green = PatternFill("solid", fgColor="D9F2D9")
        fill_red = PatternFill("solid", fgColor="FFD9D9")
        fill_orange = PatternFill("solid", fgColor="FFE7CC")

        for row in self.result_rows:
            ws.append(
                [
                    row.status_text,
                    str(row.target_path),
                    row.target_relative_path,
                    row.size,
                    self._format_size(row.size),
                    row.exact_folder_labels,
                    row.exact_paths_text,
                    row.diff_paths_text,
                    ", ".join(row.missing_from_folders) if row.missing_from_folders else "-",
                    row.match_type,
                ]
            )
            excel_row = ws.max_row
            fill = fill_green if row.tree_tag == "exact_match" else fill_red if row.tree_tag == "different_content" else fill_orange
            for col in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=col).fill = fill

        for column_letter, width in {
            "A": 22,
            "B": 50,
            "C": 30,
            "D": 16,
            "E": 14,
            "F": 18,
            "G": 70,
            "H": 70,
            "I": 22,
            "J": 14,
        }.items():
            ws.column_dimensions[column_letter].width = width

        wb.save(save_path)
        messagebox.showinfo(APP_TITLE, f"Berhasil menyimpan Excel:\n{save_path}")

    def _send_to_recycle_bin(self, path: Path) -> None:
        try:
            from send2trash import send2trash  # type: ignore
        except ImportError as exc:
            raise RuntimeError(
                "Paket 'send2trash' belum terpasang. Jalankan: pip install send2trash"
            ) from exc
        send2trash(str(path))


if __name__ == "__main__":
    root = tk.Tk()
    try:
        from ctypes import windll

        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    app = FolderCompareDeleteApp(root)
    root.mainloop()
