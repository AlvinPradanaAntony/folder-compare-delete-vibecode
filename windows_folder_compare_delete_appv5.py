from __future__ import annotations

import csv
import hashlib
import os
import queue
import sys
import threading
import traceback
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

try:
    from PySide6.QtCore import QAbstractAnimation, QAbstractTableModel, QEasingCurve, QItemSelectionModel, QModelIndex, QPoint, QPropertyAnimation, QParallelAnimationGroup, QRect, QSortFilterProxyModel, QTimer, Qt, Signal
    from PySide6.QtGui import QBrush, QColor, QDragEnterEvent, QDropEvent, QFocusEvent, QWheelEvent
    from PySide6.QtWidgets import (
        QAbstractItemView,
        QApplication,
        QButtonGroup,
        QDialog,
        QFileDialog,
        QFrame,
        QGraphicsBlurEffect,
        QGridLayout,
        QHBoxLayout,
        QHeaderView,
        QLabel,
        QLineEdit,
        QMainWindow,
        QMessageBox,
        QPlainTextEdit,
        QProgressBar,
        QPushButton,
        QRadioButton,
        QCheckBox,
        QScrollArea,
        QSplitter,
        QStackedLayout,
        QTableView,
        QVBoxLayout,
        QWidget,
    )
except ImportError as exc:
    raise SystemExit("PySide6 belum terpasang. Jalankan: pip install PySide6") from exc


APP_TITLE = "Folder Compare & Delete"
CHUNK_SIZE = 1024 * 1024  # 1 MB
BG_COLOR = "#f4f7fb"
SURFACE = "#ffffff"
SURFACE_ALT = "#f7f9fc"
PRIMARY = "#1f5eff"
PRIMARY_DARK = "#1948c7"
TEXT = "#152033"
MUTED = "#62708a"
GREEN = "#27c281"
RED = "#eb5d70"
ORANGE = "#f4a23a"
CYAN = "#20b7d6"
BORDER = "#dbe3f3"
SUCCESS_ROW = "#eafaf1"
DANGER_ROW = "#fff0f3"
WARNING_ROW = "#fff7eb"


@dataclass
class FileRecord:
    path: Path
    base_folder: Path
    base_label: str
    relative_path: str
    size: int
    sha256: Optional[str] = None


@dataclass
class MatchResult:
    target_path: Path
    target_relative_path: str
    size: int
    match_type: str
    exact_matches: List[FileRecord] = field(default_factory=list)
    same_name_different_content: List[FileRecord] = field(default_factory=list)
    missing_from_folders: List[str] = field(default_factory=list)
    only_in_target: bool = False

    @property
    def exact_folder_labels(self) -> str:
        labels = sorted({item.base_label for item in self.exact_matches})
        return ", ".join(labels) if labels else "-"

    @property
    def exact_paths_text(self) -> str:
        if not self.exact_matches:
            return "-"
        return " | ".join(f"[{item.base_label}] {item.path}" for item in self.exact_matches)

    @property
    def diff_paths_text(self) -> str:
        if not self.same_name_different_content:
            return "-"
        return " | ".join(f"[{item.base_label}] {item.path}" for item in self.same_name_different_content)

    @property
    def status_text(self) -> str:
        if self.exact_matches:
            return "Duplikat ditemukan"
        if self.same_name_different_content:
            return "Nama sama, isi berbeda"
        return "Hanya ada di Folder A"

    @property
    def tree_tag(self) -> str:
        if self.exact_matches:
            return "exact_match"
        if self.same_name_different_content:
            return "different_content"
        return "only_target"


class ResponsiveTableWidget(QTableView):
    focusReleased = Signal()

    def wheelEvent(self, event: QWheelEvent) -> None:
        if event.modifiers() & Qt.ShiftModifier:
            delta = event.angleDelta().y() or event.angleDelta().x()
            if delta:
                scrollbar = self.horizontalScrollBar()
                step = scrollbar.singleStep() or 24
                direction = -1 if delta > 0 else 1
                scrollbar.setValue(scrollbar.value() + (direction * step * 3))
                event.accept()
                return
        super().wheelEvent(event)

    def focusOutEvent(self, event: QFocusEvent) -> None:
        if self.selectionModel() is not None:
            self.selectionModel().setCurrentIndex(QModelIndex(), QItemSelectionModel.SelectionFlag.NoUpdate)
        self.focusReleased.emit()
        super().focusOutEvent(event)


class MatchResultTableModel(QAbstractTableModel):
    HEADERS = [
        "Status",
        "Path File di Folder A",
        "Relative Path",
        "Ukuran",
        "Ditemukan di",
        "Path Cocok",
        "Path Beda",
        "Tidak Ada di",
        "Mode",
    ]

    def __init__(self) -> None:
        super().__init__()
        self._rows: List[MatchResult] = []

    def set_rows(self, rows: List[MatchResult]) -> None:
        self.beginResetModel()
        self._rows = list(rows)
        self.endResetModel()

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:
        if parent.isValid():
            return 0
        return len(self._rows)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:
        if parent.isValid():
            return 0
        return len(self.HEADERS)

    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.DisplayRole) -> Any:
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal and 0 <= section < len(self.HEADERS):
            return self.HEADERS[section]
        return section + 1 if orientation == Qt.Vertical else None

    def data(self, index: QModelIndex, role: int = Qt.DisplayRole) -> Any:
        if not index.isValid() or not (0 <= index.row() < len(self._rows)):
            return None

        row = self._rows[index.row()]
        display_values = [
            row.status_text,
            str(row.target_path),
            row.target_relative_path,
            self._format_size(row.size),
            row.exact_folder_labels,
            row.exact_paths_text,
            row.diff_paths_text,
            ", ".join(row.missing_from_folders) if row.missing_from_folders else "-",
            row.match_type,
        ]
        sort_values = [
            row.status_text.lower(),
            str(row.target_path).lower(),
            row.target_relative_path.lower(),
            row.size,
            row.exact_folder_labels.lower(),
            row.exact_paths_text.lower(),
            row.diff_paths_text.lower(),
            (", ".join(row.missing_from_folders) if row.missing_from_folders else "-").lower(),
            row.match_type.lower(),
        ]

        if role == Qt.DisplayRole:
            return display_values[index.column()]
        if role == Qt.ToolTipRole:
            return display_values[index.column()]
        if role == Qt.TextAlignmentRole:
            return Qt.AlignVCenter | (Qt.AlignCenter if index.column() in {0, 3, 4, 7, 8} else Qt.AlignLeft)
        if role == Qt.BackgroundRole:
            return QBrush(self._result_background(row.tree_tag))
        if role == Qt.ForegroundRole:
            return QBrush(self._result_foreground(row.tree_tag))
        if role == Qt.UserRole:
            return row
        if role == Qt.UserRole + 1:
            return sort_values[index.column()]
        return None

    def result_at(self, row_index: int) -> Optional[MatchResult]:
        if 0 <= row_index < len(self._rows):
            return self._rows[row_index]
        return None

    @staticmethod
    def _format_size(size: int) -> str:
        value = float(size)
        units = ["B", "KB", "MB", "GB", "TB"]
        for unit in units:
            if value < 1024 or unit == units[-1]:
                return f"{value:.2f} {unit}" if unit != "B" else f"{int(value)} B"
            value /= 1024
        return f"{size} B"

    @staticmethod
    def _result_background(tree_tag: str) -> QColor:
        if tree_tag == "exact_match":
            return QColor(SUCCESS_ROW)
        if tree_tag == "different_content":
            return QColor(DANGER_ROW)
        return QColor(WARNING_ROW)

    @staticmethod
    def _result_foreground(tree_tag: str) -> QColor:
        if tree_tag == "exact_match":
            return QColor("#1f6f4a")
        if tree_tag == "different_content":
            return QColor("#9f3143")
        return QColor("#8a5a19")


class MatchResultFilterProxyModel(QSortFilterProxyModel):
    def __init__(self) -> None:
        super().__init__()
        self.search_text = ""
        self.status_filter = "all"
        self.matches_only = False
        self.setSortRole(Qt.UserRole + 1)
        self.setDynamicSortFilter(True)

    def set_search_text(self, text: str) -> None:
        self.beginFilterChange()
        self.search_text = text.strip().lower()
        self.endFilterChange(QSortFilterProxyModel.Direction.Rows)

    def set_status_filter(self, status_filter: str) -> None:
        self.beginFilterChange()
        self.status_filter = status_filter
        self.endFilterChange(QSortFilterProxyModel.Direction.Rows)

    def set_matches_only(self, enabled: bool) -> None:
        self.beginFilterChange()
        self.matches_only = enabled
        self.endFilterChange(QSortFilterProxyModel.Direction.Rows)

    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:
        source_model = self.sourceModel()
        if source_model is None:
            return True

        source_index = source_model.index(source_row, 0, source_parent)
        result = source_index.data(Qt.UserRole)
        if not isinstance(result, MatchResult):
            return True

        if self.matches_only and not (result.exact_matches or result.same_name_different_content):
            return False
        if self.status_filter != "all" and result.tree_tag != self.status_filter:
            return False
        if not self.search_text:
            return True

        haystacks = [
            result.status_text,
            str(result.target_path),
            result.target_relative_path,
            MatchResultTableModel._format_size(result.size),
            result.exact_folder_labels,
            result.exact_paths_text,
            result.diff_paths_text,
            ", ".join(result.missing_from_folders) if result.missing_from_folders else "-",
            result.match_type,
        ]
        searchable_text = " ".join(haystacks).lower()
        return self.search_text in searchable_text


class FolderPathLineEdit(QLineEdit):
    folderDropped = Signal(str)

    def __init__(self, placeholder: str = "", parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setAcceptDrops(True)
        if placeholder:
            self.setPlaceholderText(placeholder)

    def dragEnterEvent(self, event: QDragEnterEvent) -> None:
        if self._extract_path_from_event(event) is not None:
            event.acceptProposedAction()
            return
        super().dragEnterEvent(event)

    def dropEvent(self, event: QDropEvent) -> None:
        dropped_path = self._extract_path_from_event(event)
        if dropped_path is not None:
            self.setText(dropped_path)
            self.folderDropped.emit(dropped_path)
            event.acceptProposedAction()
            return
        super().dropEvent(event)

    def _extract_path_from_event(self, event: QDragEnterEvent | QDropEvent) -> Optional[str]:
        mime_data = event.mimeData()

        if mime_data.hasUrls():
            for url in mime_data.urls():
                local_path = url.toLocalFile()
                if not local_path:
                    continue
                path = Path(local_path)
                if path.is_dir():
                    return os.path.normpath(str(path))
                if path.exists():
                    return os.path.normpath(str(path.parent))

        if mime_data.hasText():
            raw_text = mime_data.text().strip().strip('"')
            if raw_text:
                path = Path(raw_text)
                if path.is_dir():
                    return os.path.normpath(str(path))
                if path.exists():
                    return os.path.normpath(str(path.parent))

        return None


class ErrorOverlayDialog(QDialog):
    def __init__(self, parent: Optional[QWidget], title: str, summary: str, details: str) -> None:
        super().__init__(parent)
        self._blur_target = parent.centralWidget() if isinstance(parent, QMainWindow) else parent
        self._owns_blur_effect = False
        self._details_text = details.strip()
        self._is_finalizing = False
        self._close_result = QDialog.Rejected
        self._open_animation_started = False
        self._blur_effect: Optional[QGraphicsBlurEffect] = None

        self.setModal(True)
        self.setObjectName("ErrorOverlayDialog")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setMinimumSize(720, 480)
        self.setWindowOpacity(0.0)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        overlay = QFrame()
        self.overlay = overlay
        overlay.setObjectName("ErrorOverlayBackdrop")
        overlay_layout = QVBoxLayout(overlay)
        overlay_layout.setContentsMargins(0, 0, 0, 0)
        overlay_layout.addStretch(1)

        card = QFrame()
        self.card = card
        card.setObjectName("ErrorOverlayCard")
        card.setMinimumWidth(640)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 24, 28, 24)
        card_layout.setSpacing(16)

        header_row = QHBoxLayout()
        header_row.setSpacing(14)

        icon_badge = QLabel("!")
        icon_badge.setObjectName("ErrorOverlayIcon")
        icon_badge.setAlignment(Qt.AlignCenter)
        icon_badge.setFixedSize(48, 48)

        title_wrap = QVBoxLayout()
        title_wrap.setSpacing(4)

        title_label = QLabel(title)
        title_label.setObjectName("ErrorOverlayTitle")
        title_label.setWordWrap(True)

        summary_label = QLabel(summary or "Terjadi kesalahan yang tidak terduga pada aplikasi.")
        summary_label.setObjectName("ErrorOverlaySummary")
        summary_label.setWordWrap(True)
        summary_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        title_wrap.addWidget(title_label)
        title_wrap.addWidget(summary_label)
        header_row.addWidget(icon_badge, 0, Qt.AlignTop)
        header_row.addLayout(title_wrap, 1)

        detail_label = QLabel("Detail teknis")
        detail_label.setObjectName("ErrorOverlaySectionTitle")

        self.details_box = QPlainTextEdit()
        self.details_box.setObjectName("ErrorOverlayDetails")
        self.details_box.setReadOnly(True)
        self.details_box.setPlainText(self._details_text or summary)
        self.details_box.setMinimumHeight(220)

        button_row = QHBoxLayout()
        button_row.setSpacing(10)
        button_row.addStretch(1)

        copy_button = QPushButton("Salin Detail")
        copy_button.setObjectName("ErrorOverlaySecondaryButton")
        copy_button.clicked.connect(self._copy_details)

        close_button = QPushButton("Tutup")
        close_button.setObjectName("ErrorOverlayPrimaryButton")
        close_button.clicked.connect(lambda: self.done(QDialog.Accepted))
        close_button.setAutoDefault(True)
        close_button.setDefault(True)

        button_row.addWidget(copy_button)
        button_row.addWidget(close_button)

        card_layout.addLayout(header_row)
        card_layout.addWidget(detail_label)
        card_layout.addWidget(self.details_box)
        card_layout.addLayout(button_row)

        overlay_layout.addWidget(card, 0, Qt.AlignCenter)
        overlay_layout.addStretch(1)
        outer.addWidget(overlay)

        self.setStyleSheet(
            f"""
            QDialog#ErrorOverlayDialog {{
                background: transparent;
            }}
            QFrame#ErrorOverlayBackdrop {{
                background: rgba(11, 18, 32, 168);
                border-radius: 0px;
            }}
            QFrame#ErrorOverlayCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #f5f8ff
                );
                border: 1px solid #d7e1f4;
                border-radius: 24px;
            }}
            QLabel#ErrorOverlayIcon {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #f56d86,
                    stop: 1 #e64d67
                );
                color: white;
                border-radius: 24px;
                font: 700 18pt "Segoe UI";
            }}
            QLabel#ErrorOverlayTitle {{
                color: {TEXT};
                font: 700 15pt "Segoe UI";
            }}
            QLabel#ErrorOverlaySummary {{
                color: {MUTED};
                font: 10pt "Segoe UI";
                line-height: 1.4em;
            }}
            QLabel#ErrorOverlaySectionTitle {{
                color: {TEXT};
                font: 700 10pt "Segoe UI";
            }}
            QPlainTextEdit#ErrorOverlayDetails {{
                background: #f7faff;
                color: #25324a;
                border: 1px solid #d7e1f4;
                border-radius: 16px;
                padding: 12px;
                font: 9.5pt "Consolas";
                selection-background-color: #cfe0ff;
            }}
            QPushButton#ErrorOverlayPrimaryButton {{
                min-width: 120px;
                padding: 10px 18px;
                border: none;
                border-radius: 14px;
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 {PRIMARY},
                    stop: 1 #4d84ff
                );
                color: white;
                font: 700 10pt "Segoe UI";
            }}
            QPushButton#ErrorOverlayPrimaryButton:hover {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 {PRIMARY_DARK},
                    stop: 1 #386cf0
                );
            }}
            QPushButton#ErrorOverlaySecondaryButton {{
                min-width: 120px;
                padding: 10px 18px;
                border-radius: 14px;
                border: 1px solid #d3def2;
                background: white;
                color: {TEXT};
                font: 600 10pt "Segoe UI";
            }}
            QPushButton#ErrorOverlaySecondaryButton:hover {{
                background: #f4f8ff;
            }}
            """
        )

    def showEvent(self, event) -> None:
        parent = self.parentWidget()
        if parent is not None:
            self.resize(parent.size())
            self.move(parent.mapToGlobal(QPoint(0, 0)))
        if self._blur_target is not None and self._blur_target.graphicsEffect() is None:
            blur = QGraphicsBlurEffect(self._blur_target)
            blur.setBlurRadius(8.0)
            self._blur_target.setGraphicsEffect(blur)
            self._owns_blur_effect = True
            self._blur_effect = blur
        super().showEvent(event)
        QTimer.singleShot(0, self._start_open_animation)

    def done(self, result: int) -> None:
        if self._is_finalizing:
            self._clear_blur()
            super().done(result)
            return
        self._close_result = result
        self._start_close_animation()

    def hideEvent(self, event) -> None:
        self._clear_blur()
        super().hideEvent(event)

    def closeEvent(self, event) -> None:
        if not self._is_finalizing:
            event.ignore()
            self.done(QDialog.Rejected)
            return
        self._clear_blur()
        super().closeEvent(event)

    def _clear_blur(self) -> None:
        if self._owns_blur_effect and self._blur_target is not None:
            self._blur_target.setGraphicsEffect(None)
            self._owns_blur_effect = False
        self._blur_effect = None

    @staticmethod
    def _scaled_rect(rect: QRect, scale: float) -> QRect:
        width = max(1, int(rect.width() * scale))
        height = max(1, int(rect.height() * scale))
        center = rect.center()
        scaled = QRect(0, 0, width, height)
        scaled.moveCenter(center)
        return scaled

    def _start_open_animation(self) -> None:
        if self._open_animation_started:
            return
        end_rect = self.card.geometry()
        if not end_rect.isValid():
            return

        self._open_animation_started = True
        start_rect = self._scaled_rect(end_rect, 0.985)
        self.card.setGeometry(start_rect)
        self.setWindowOpacity(0.0)

        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(170)
        opacity_animation.setStartValue(0.0)
        opacity_animation.setEndValue(1.0)
        opacity_animation.setEasingCurve(QEasingCurve.OutCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(180)
        geometry_animation.setStartValue(start_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.OutCubic)

        animation.start()
        self._open_animation = animation

    def _start_close_animation(self) -> None:
        if self._is_finalizing:
            return

        current_rect = self.card.geometry()
        if not current_rect.isValid():
            self._finalize_close()
            return

        end_rect = self._scaled_rect(current_rect, 0.99)
        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(140)
        opacity_animation.setStartValue(self.windowOpacity())
        opacity_animation.setEndValue(0.0)
        opacity_animation.setEasingCurve(QEasingCurve.InCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(140)
        geometry_animation.setStartValue(current_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.InCubic)

        animation.finished.connect(self._finalize_close)
        animation.start()
        self._close_animation = animation

    def _finalize_close(self) -> None:
        if self._is_finalizing:
            return
        self._is_finalizing = True
        try:
            self._clear_blur()
            super().done(self._close_result)
        finally:
            self._is_finalizing = False

    def _copy_details(self) -> None:
        QApplication.clipboard().setText(self._details_text or self.details_box.toPlainText())


class ConfirmOverlayDialog(QDialog):
    confirmRequested = Signal()

    def __init__(self, parent: Optional[QWidget], title: str, summary: str, details: str) -> None:
        super().__init__(parent)
        self._blur_target = parent.centralWidget() if isinstance(parent, QMainWindow) else parent
        self._owns_blur_effect = False
        self._details_text = details.strip()
        self._is_finalizing = False
        self._close_result = QDialog.Rejected
        self._open_animation_started = False
        self._blur_effect: Optional[QGraphicsBlurEffect] = None
        self._processing = False
        self._success_mode = False
        self._confirm_button_text = "Ya, Proses Hapus"
        self._spinner_frames = ["|", "/", "-", "\\"]
        self._spinner_index = 0

        self.setModal(True)
        self.setObjectName("ConfirmOverlayDialog")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setMinimumSize(720, 500)
        self.setWindowOpacity(0.0)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        overlay = QFrame()
        self.overlay = overlay
        overlay.setObjectName("ConfirmOverlayBackdrop")
        overlay_layout = QVBoxLayout(overlay)
        overlay_layout.setContentsMargins(0, 0, 0, 0)
        overlay_layout.addStretch(1)

        card = QFrame()
        self.card = card
        card.setObjectName("ConfirmOverlayCard")
        card.setMinimumWidth(660)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 24, 28, 24)
        card_layout.setSpacing(16)

        header_row = QHBoxLayout()
        header_row.setSpacing(14)

        icon_badge = QLabel("!")
        self.icon_badge = icon_badge
        icon_badge.setObjectName("ConfirmOverlayIcon")
        icon_badge.setProperty("successMode", False)
        icon_badge.setAlignment(Qt.AlignCenter)
        icon_badge.setFixedSize(48, 48)

        title_wrap = QVBoxLayout()
        title_wrap.setSpacing(4)

        title_label = QLabel(title)
        self.title_label = title_label
        title_label.setObjectName("ConfirmOverlayTitle")
        title_label.setWordWrap(True)

        summary_label = QLabel(summary)
        self.summary_label = summary_label
        summary_label.setObjectName("ConfirmOverlaySummary")
        summary_label.setWordWrap(True)
        summary_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        title_wrap.addWidget(title_label)
        title_wrap.addWidget(summary_label)
        header_row.addWidget(icon_badge, 0, Qt.AlignTop)
        header_row.addLayout(title_wrap, 1)

        detail_label = QLabel("Pratinjau file")
        self.detail_label = detail_label
        detail_label.setObjectName("ConfirmOverlaySectionTitle")

        self.details_box = QPlainTextEdit()
        self.details_box.setObjectName("ConfirmOverlayDetails")
        self.details_box.setProperty("successMode", False)
        self.details_box.setReadOnly(True)
        self.details_box.setPlainText(self._details_text)
        self.details_box.setMinimumHeight(240)

        footnote = QLabel("Pastikan Anda sudah meninjau file hijau sebelum melanjutkan proses hapus.")
        self.footnote_label = footnote
        footnote.setObjectName("ConfirmOverlayFootnote")
        footnote.setProperty("successMode", False)
        footnote.setWordWrap(True)

        button_row = QHBoxLayout()
        button_row.setSpacing(10)
        button_row.addStretch(1)

        cancel_button = QPushButton("Batal")
        self.cancel_button = cancel_button
        cancel_button.setObjectName("ConfirmOverlaySecondaryButton")
        cancel_button.clicked.connect(lambda: self.done(QDialog.Rejected))

        confirm_button = QPushButton(self._confirm_button_text)
        self.confirm_button = confirm_button
        confirm_button.setObjectName("ConfirmOverlayPrimaryButton")
        confirm_button.setProperty("successMode", False)
        confirm_button.clicked.connect(self._on_confirm_clicked)
        confirm_button.setAutoDefault(True)
        confirm_button.setDefault(True)

        button_row.addWidget(cancel_button)
        button_row.addWidget(confirm_button)

        card_layout.addLayout(header_row)
        card_layout.addWidget(detail_label)
        card_layout.addWidget(self.details_box)
        card_layout.addWidget(footnote)
        card_layout.addLayout(button_row)

        overlay_layout.addWidget(card, 0, Qt.AlignCenter)
        overlay_layout.addStretch(1)
        outer.addWidget(overlay)

        self.spinner_timer = QTimer(self)
        self.spinner_timer.setInterval(120)
        self.spinner_timer.timeout.connect(self._advance_spinner)

        self.setStyleSheet(
            f"""
            QDialog#ConfirmOverlayDialog {{
                background: transparent;
            }}
            QFrame#ConfirmOverlayBackdrop {{
                background: rgba(11, 18, 32, 168);
                border-radius: 0px;
            }}
            QFrame#ConfirmOverlayCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #fef8f3
                );
                border: 1px solid #f0d7c3;
                border-radius: 24px;
            }}
            QFrame#ConfirmOverlayCard[successMode="true"] {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #f5fff9
                );
                border: 1px solid #cce9db;
            }}
            QLabel#ConfirmOverlayIcon {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #f29a3b,
                    stop: 1 #e07a16
                );
                color: white;
                border-radius: 24px;
                font: 700 18pt "Segoe UI";
            }}
            QLabel#ConfirmOverlayIcon[successMode="true"] {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #16a56f,
                    stop: 1 #27c281
                );
            }}
            QLabel#ConfirmOverlayTitle {{
                color: {TEXT};
                font: 700 15pt "Segoe UI";
            }}
            QLabel#ConfirmOverlaySummary {{
                color: {MUTED};
                font: 10pt "Segoe UI";
                line-height: 1.4em;
            }}
            QLabel#ConfirmOverlaySectionTitle {{
                color: {TEXT};
                font: 700 10pt "Segoe UI";
            }}
            QLabel#ConfirmOverlayFootnote {{
                color: #8a5a19;
                font: 9.3pt "Segoe UI";
            }}
            QLabel#ConfirmOverlayFootnote[successMode="true"] {{
                color: #1f7a58;
            }}
            QPlainTextEdit#ConfirmOverlayDetails {{
                background: #fffaf5;
                color: #25324a;
                border: 1px solid #f0d7c3;
                border-radius: 16px;
                padding: 12px;
                font: 9.5pt "Consolas";
                selection-background-color: #ffe2c0;
            }}
            QPlainTextEdit#ConfirmOverlayDetails[successMode="true"] {{
                background: #f6fff9;
                border: 1px solid #cce9db;
                selection-background-color: #cfeedd;
            }}
            QPushButton#ConfirmOverlayPrimaryButton {{
                min-width: 150px;
                padding: 10px 18px;
                border: none;
                border-radius: 14px;
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ed7b24,
                    stop: 1 #f29a3b
                );
                color: white;
                font: 700 10pt "Segoe UI";
            }}
            QPushButton#ConfirmOverlayPrimaryButton:hover {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #db6912,
                    stop: 1 #eb8c28
                );
            }}
            QPushButton#ConfirmOverlayPrimaryButton[successMode="true"] {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #169766,
                    stop: 1 #27c281
                );
            }}
            QPushButton#ConfirmOverlayPrimaryButton[successMode="true"]:hover {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #118557,
                    stop: 1 #1fb172
                );
            }}
            QPushButton#ConfirmOverlayPrimaryButton:disabled {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ed7b24,
                    stop: 1 #f29a3b
                );
                color: rgba(255, 255, 255, 0.96);
                border: none;
            }}
            QPushButton#ConfirmOverlayPrimaryButton[successMode="true"]:disabled {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #169766,
                    stop: 1 #27c281
                );
            }}
            QPushButton#ConfirmOverlaySecondaryButton {{
                min-width: 120px;
                padding: 10px 18px;
                border-radius: 14px;
                border: 1px solid #e8d8c9;
                background: white;
                color: {TEXT};
                font: 600 10pt "Segoe UI";
            }}
            QPushButton#ConfirmOverlaySecondaryButton:hover {{
                background: #fff8f1;
            }}
            QPushButton#ConfirmOverlaySecondaryButton:disabled {{
                background: #fff7ef;
                border: 1px solid #ecdccf;
                color: #b18a66;
            }}
            """
        )
        self.card.setProperty("successMode", False)

    def _refresh_theme_state(self) -> None:
        success_mode = self._success_mode
        themed_widgets = [
            self.card,
            self.icon_badge,
            self.details_box,
            self.footnote_label,
            self.confirm_button,
        ]
        for widget in themed_widgets:
            widget.setProperty("successMode", success_mode)
            widget.style().unpolish(widget)
            widget.style().polish(widget)
        self.style().unpolish(self)
        self.style().polish(self)
        self.overlay.update()
        self.card.update()

    def showEvent(self, event) -> None:
        parent = self.parentWidget()
        if parent is not None:
            self.resize(parent.size())
            self.move(parent.mapToGlobal(QPoint(0, 0)))
        if self._blur_target is not None and self._blur_target.graphicsEffect() is None:
            blur = QGraphicsBlurEffect(self._blur_target)
            blur.setBlurRadius(8.0)
            self._blur_target.setGraphicsEffect(blur)
            self._owns_blur_effect = True
            self._blur_effect = blur
        super().showEvent(event)
        QTimer.singleShot(0, self._start_open_animation)

    def done(self, result: int) -> None:
        if self._processing:
            return
        if self._is_finalizing:
            self._clear_blur()
            super().done(result)
            return
        self._close_result = result
        self._start_close_animation()

    def hideEvent(self, event) -> None:
        self._clear_blur()
        super().hideEvent(event)

    def closeEvent(self, event) -> None:
        if self._processing:
            event.ignore()
            return
        if not self._is_finalizing:
            event.ignore()
            self.done(QDialog.Rejected)
            return
        self._clear_blur()
        super().closeEvent(event)

    def _clear_blur(self) -> None:
        if self._owns_blur_effect and self._blur_target is not None:
            self._blur_target.setGraphicsEffect(None)
            self._owns_blur_effect = False
        self._blur_effect = None

    @staticmethod
    def _scaled_rect(rect: QRect, scale: float) -> QRect:
        width = max(1, int(rect.width() * scale))
        height = max(1, int(rect.height() * scale))
        center = rect.center()
        scaled = QRect(0, 0, width, height)
        scaled.moveCenter(center)
        return scaled

    def _start_open_animation(self) -> None:
        if self._open_animation_started:
            return
        end_rect = self.card.geometry()
        if not end_rect.isValid():
            return

        self._open_animation_started = True
        start_rect = self._scaled_rect(end_rect, 0.985)
        self.card.setGeometry(start_rect)
        self.setWindowOpacity(0.0)

        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(170)
        opacity_animation.setStartValue(0.0)
        opacity_animation.setEndValue(1.0)
        opacity_animation.setEasingCurve(QEasingCurve.OutCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(180)
        geometry_animation.setStartValue(start_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.OutCubic)

        animation.start()
        self._open_animation = animation

    def _start_close_animation(self) -> None:
        if self._is_finalizing:
            return

        current_rect = self.card.geometry()
        if not current_rect.isValid():
            self._finalize_close()
            return

        end_rect = self._scaled_rect(current_rect, 0.99)
        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(140)
        opacity_animation.setStartValue(self.windowOpacity())
        opacity_animation.setEndValue(0.0)
        opacity_animation.setEasingCurve(QEasingCurve.InCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(140)
        geometry_animation.setStartValue(current_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.InCubic)

        animation.finished.connect(self._finalize_close)
        animation.start()
        self._close_animation = animation

    def _finalize_close(self) -> None:
        if self._is_finalizing:
            return
        self._is_finalizing = True
        try:
            self._clear_blur()
            super().done(self._close_result)
        finally:
            self._is_finalizing = False

    def _on_confirm_clicked(self) -> None:
        if self._processing:
            return
        if self._success_mode:
            self.done(QDialog.Accepted)
            return
        self.confirmRequested.emit()

    def set_processing(self, processing: bool, summary: Optional[str] = None) -> None:
        self._processing = processing
        if processing:
            self._success_mode = False
            self._refresh_theme_state()
        if summary:
            self.summary_label.setText(summary)

        self.cancel_button.setEnabled(not processing)
        self.confirm_button.setEnabled(not processing)

        if processing:
            self.footnote_label.setText("Penghapusan sedang berjalan. Mohon tunggu sampai proses selesai.")
            self._spinner_index = 0
            self.confirm_button.setText(f"Memproses {self._spinner_frames[self._spinner_index]}")
            self.spinner_timer.start()
        else:
            self.spinner_timer.stop()
            self.confirm_button.setText(self._confirm_button_text)
            self.footnote_label.setText("Pastikan Anda sudah meninjau file hijau sebelum melanjutkan proses hapus.")
            self.cancel_button.show()

    def flush_visual_state(self) -> None:
        self.raise_()
        self.activateWindow()
        self.overlay.update()
        self.card.update()
        self.repaint()
        QApplication.processEvents()

    def finish_processing(self) -> None:
        self.set_processing(False)
        self.done(QDialog.Accepted)

    def force_close(self, result: int = QDialog.Rejected) -> None:
        self.spinner_timer.stop()
        self._processing = False
        self.done(result)

    def _advance_spinner(self) -> None:
        self._spinner_index = (self._spinner_index + 1) % len(self._spinner_frames)
        self.confirm_button.setText(f"Memproses {self._spinner_frames[self._spinner_index]}")

    def show_success_state(self, summary: str, details: str = "") -> None:
        self.spinner_timer.stop()
        self._processing = False
        self._success_mode = True
        self._refresh_theme_state()

        self.title_label.setText("Penghapusan Berhasil")
        self.summary_label.setText(summary)
        self.icon_badge.setText("OK")
        self.detail_label.setText("Ringkasan hasil")
        self.detail_label.setVisible(bool(details))
        self.details_box.setPlainText(details)
        self.details_box.setVisible(bool(details))
        self.footnote_label.setText("Proses selesai. Anda dapat menutup dialog ini.")

        self.cancel_button.hide()
        self.confirm_button.setEnabled(True)
        self.confirm_button.setText("Tutup")


class SuccessOverlayDialog(QDialog):
    def __init__(self, parent: Optional[QWidget], title: str, summary: str, details: str = "") -> None:
        super().__init__(parent)
        self._blur_target = parent.centralWidget() if isinstance(parent, QMainWindow) else parent
        self._owns_blur_effect = False
        self._details_text = details.strip()
        self._is_finalizing = False
        self._close_result = QDialog.Accepted
        self._open_animation_started = False
        self._blur_effect: Optional[QGraphicsBlurEffect] = None

        self.setModal(True)
        self.setObjectName("SuccessOverlayDialog")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setMinimumSize(680, 420)
        self.setWindowOpacity(0.0)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        overlay = QFrame()
        self.overlay = overlay
        overlay.setObjectName("SuccessOverlayBackdrop")
        overlay_layout = QVBoxLayout(overlay)
        overlay_layout.setContentsMargins(0, 0, 0, 0)
        overlay_layout.addStretch(1)

        card = QFrame()
        self.card = card
        card.setObjectName("SuccessOverlayCard")
        card.setMinimumWidth(560)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 24, 28, 24)
        card_layout.setSpacing(16)

        header_row = QHBoxLayout()
        header_row.setSpacing(14)

        icon_badge = QLabel("OK")
        icon_badge.setObjectName("SuccessOverlayIcon")
        icon_badge.setAlignment(Qt.AlignCenter)
        icon_badge.setFixedSize(56, 56)

        title_wrap = QVBoxLayout()
        title_wrap.setSpacing(4)

        title_label = QLabel(title)
        title_label.setObjectName("SuccessOverlayTitle")
        title_label.setWordWrap(True)

        summary_label = QLabel(summary)
        summary_label.setObjectName("SuccessOverlaySummary")
        summary_label.setWordWrap(True)
        summary_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        title_wrap.addWidget(title_label)
        title_wrap.addWidget(summary_label)
        header_row.addWidget(icon_badge, 0, Qt.AlignTop)
        header_row.addLayout(title_wrap, 1)

        detail_label = QLabel("Ringkasan hasil")
        detail_label.setObjectName("SuccessOverlaySectionTitle")
        detail_label.setVisible(bool(self._details_text))

        self.details_box = QPlainTextEdit()
        self.details_box.setObjectName("SuccessOverlayDetails")
        self.details_box.setReadOnly(True)
        self.details_box.setPlainText(self._details_text)
        self.details_box.setMinimumHeight(140)
        self.details_box.setVisible(bool(self._details_text))

        button_row = QHBoxLayout()
        button_row.setSpacing(10)
        button_row.addStretch(1)

        close_button = QPushButton("Tutup")
        close_button.setObjectName("SuccessOverlayPrimaryButton")
        close_button.clicked.connect(lambda: self.done(QDialog.Accepted))
        close_button.setAutoDefault(True)
        close_button.setDefault(True)

        button_row.addWidget(close_button)

        card_layout.addLayout(header_row)
        card_layout.addWidget(detail_label)
        card_layout.addWidget(self.details_box)
        card_layout.addLayout(button_row)

        overlay_layout.addWidget(card, 0, Qt.AlignCenter)
        overlay_layout.addStretch(1)
        outer.addWidget(overlay)

        self.setStyleSheet(
            f"""
            QDialog#SuccessOverlayDialog {{
                background: transparent;
            }}
            QFrame#SuccessOverlayBackdrop {{
                background: rgba(11, 18, 32, 168);
                border-radius: 0px;
            }}
            QFrame#SuccessOverlayCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #f5fff9
                );
                border: 1px solid #cce9db;
                border-radius: 24px;
            }}
            QLabel#SuccessOverlayIcon {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #16a56f,
                    stop: 1 #27c281
                );
                color: white;
                border-radius: 28px;
                font: 800 11pt "Segoe UI";
            }}
            QLabel#SuccessOverlayTitle {{
                color: {TEXT};
                font: 700 15pt "Segoe UI";
            }}
            QLabel#SuccessOverlaySummary {{
                color: {MUTED};
                font: 10pt "Segoe UI";
                line-height: 1.4em;
            }}
            QLabel#SuccessOverlaySectionTitle {{
                color: {TEXT};
                font: 700 10pt "Segoe UI";
            }}
            QPlainTextEdit#SuccessOverlayDetails {{
                background: #f6fff9;
                color: #25324a;
                border: 1px solid #cce9db;
                border-radius: 16px;
                padding: 12px;
                font: 9.5pt "Consolas";
                selection-background-color: #cfeedd;
            }}
            QPushButton#SuccessOverlayPrimaryButton {{
                min-width: 120px;
                padding: 10px 18px;
                border: none;
                border-radius: 14px;
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #169766,
                    stop: 1 #27c281
                );
                color: white;
                font: 700 10pt "Segoe UI";
            }}
            QPushButton#SuccessOverlayPrimaryButton:hover {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #118557,
                    stop: 1 #1fb172
                );
            }}
            """
        )

    def showEvent(self, event) -> None:
        parent = self.parentWidget()
        if parent is not None:
            self.resize(parent.size())
            self.move(parent.mapToGlobal(QPoint(0, 0)))
        if self._blur_target is not None and self._blur_target.graphicsEffect() is None:
            blur = QGraphicsBlurEffect(self._blur_target)
            blur.setBlurRadius(8.0)
            self._blur_target.setGraphicsEffect(blur)
            self._owns_blur_effect = True
            self._blur_effect = blur
        super().showEvent(event)
        QTimer.singleShot(0, self._start_open_animation)

    def done(self, result: int) -> None:
        if self._is_finalizing:
            self._clear_blur()
            super().done(result)
            return
        self._close_result = result
        self._start_close_animation()

    def hideEvent(self, event) -> None:
        self._clear_blur()
        super().hideEvent(event)

    def closeEvent(self, event) -> None:
        if not self._is_finalizing:
            event.ignore()
            self.done(QDialog.Accepted)
            return
        self._clear_blur()
        super().closeEvent(event)

    def _clear_blur(self) -> None:
        if self._owns_blur_effect and self._blur_target is not None:
            self._blur_target.setGraphicsEffect(None)
            self._owns_blur_effect = False
        self._blur_effect = None

    @staticmethod
    def _scaled_rect(rect: QRect, scale: float) -> QRect:
        width = max(1, int(rect.width() * scale))
        height = max(1, int(rect.height() * scale))
        center = rect.center()
        scaled = QRect(0, 0, width, height)
        scaled.moveCenter(center)
        return scaled

    def _start_open_animation(self) -> None:
        if self._open_animation_started:
            return
        end_rect = self.card.geometry()
        if not end_rect.isValid():
            return

        self._open_animation_started = True
        start_rect = self._scaled_rect(end_rect, 0.985)
        self.card.setGeometry(start_rect)
        self.setWindowOpacity(0.0)

        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(170)
        opacity_animation.setStartValue(0.0)
        opacity_animation.setEndValue(1.0)
        opacity_animation.setEasingCurve(QEasingCurve.OutCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(180)
        geometry_animation.setStartValue(start_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.OutCubic)

        animation.start()
        self._open_animation = animation

    def _start_close_animation(self) -> None:
        if self._is_finalizing:
            return

        current_rect = self.card.geometry()
        if not current_rect.isValid():
            self._finalize_close()
            return

        end_rect = self._scaled_rect(current_rect, 0.99)
        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(140)
        opacity_animation.setStartValue(self.windowOpacity())
        opacity_animation.setEndValue(0.0)
        opacity_animation.setEasingCurve(QEasingCurve.InCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(140)
        geometry_animation.setStartValue(current_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.InCubic)

        animation.finished.connect(self._finalize_close)
        animation.start()
        self._close_animation = animation

    def _finalize_close(self) -> None:
        if self._is_finalizing:
            return
        self._is_finalizing = True
        try:
            self._clear_blur()
            super().done(self._close_result)
        finally:
            self._is_finalizing = False


class ProcessingOverlayDialog(QDialog):
    def __init__(self, parent: Optional[QWidget], title: str, summary: str) -> None:
        super().__init__(parent)
        self._blur_target = parent.centralWidget() if isinstance(parent, QMainWindow) else parent
        self._owns_blur_effect = False

        self.setModal(True)
        self.setObjectName("ProcessingOverlayDialog")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setMinimumSize(620, 320)
        self.setWindowOpacity(0.0)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        overlay = QFrame()
        self.overlay = overlay
        overlay.setObjectName("ProcessingOverlayBackdrop")
        overlay_layout = QVBoxLayout(overlay)
        overlay_layout.setContentsMargins(0, 0, 0, 0)
        overlay_layout.addStretch(1)

        card = QFrame()
        self.card = card
        card.setObjectName("ProcessingOverlayCard")
        card.setMinimumWidth(500)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 24, 28, 24)
        card_layout.setSpacing(16)

        title_label = QLabel(title)
        title_label.setObjectName("ProcessingOverlayTitle")
        title_label.setAlignment(Qt.AlignCenter)

        summary_label = QLabel(summary)
        summary_label.setObjectName("ProcessingOverlaySummary")
        summary_label.setAlignment(Qt.AlignCenter)
        summary_label.setWordWrap(True)

        progress = QProgressBar()
        progress.setObjectName("ProcessingOverlayProgress")
        progress.setRange(0, 0)
        progress.setTextVisible(False)

        footnote = QLabel("Mohon tunggu, proses ini bisa memakan waktu tergantung jumlah file.")
        footnote.setObjectName("ProcessingOverlayFootnote")
        footnote.setAlignment(Qt.AlignCenter)
        footnote.setWordWrap(True)

        card_layout.addWidget(title_label)
        card_layout.addWidget(summary_label)
        card_layout.addWidget(progress)
        card_layout.addWidget(footnote)

        overlay_layout.addWidget(card, 0, Qt.AlignCenter)
        overlay_layout.addStretch(1)
        outer.addWidget(overlay)

        self.setStyleSheet(
            f"""
            QDialog#ProcessingOverlayDialog {{
                background: transparent;
            }}
            QFrame#ProcessingOverlayBackdrop {{
                background: rgba(11, 18, 32, 168);
                border-radius: 0px;
            }}
            QFrame#ProcessingOverlayCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #f7fbff
                );
                border: 1px solid #d7e1f4;
                border-radius: 24px;
            }}
            QLabel#ProcessingOverlayTitle {{
                color: {TEXT};
                font: 700 15pt "Segoe UI";
            }}
            QLabel#ProcessingOverlaySummary {{
                color: {MUTED};
                font: 10pt "Segoe UI";
            }}
            QLabel#ProcessingOverlayFootnote {{
                color: {MUTED};
                font: 9.2pt "Segoe UI";
            }}
            QProgressBar#ProcessingOverlayProgress {{
                background: #edf2fb;
                border: none;
                border-radius: 8px;
                min-height: 12px;
                max-height: 12px;
            }}
            QProgressBar#ProcessingOverlayProgress::chunk {{
                background: {PRIMARY};
                border-radius: 8px;
            }}
            """
        )

    def showEvent(self, event) -> None:
        parent = self.parentWidget()
        if parent is not None:
            self.resize(parent.size())
            self.move(parent.mapToGlobal(QPoint(0, 0)))
        if self._blur_target is not None and self._blur_target.graphicsEffect() is None:
            blur = QGraphicsBlurEffect(self._blur_target)
            blur.setBlurRadius(8.0)
            self._blur_target.setGraphicsEffect(blur)
            self._owns_blur_effect = True
        super().showEvent(event)
        QTimer.singleShot(0, self._start_open_animation)

    def closeEvent(self, event) -> None:
        self._clear_blur()
        super().closeEvent(event)

    def hideEvent(self, event) -> None:
        self._clear_blur()
        super().hideEvent(event)

    def _clear_blur(self) -> None:
        if self._owns_blur_effect and self._blur_target is not None:
            self._blur_target.setGraphicsEffect(None)
            self._owns_blur_effect = False

    @staticmethod
    def _scaled_rect(rect: QRect, scale: float) -> QRect:
        width = max(1, int(rect.width() * scale))
        height = max(1, int(rect.height() * scale))
        center = rect.center()
        scaled = QRect(0, 0, width, height)
        scaled.moveCenter(center)
        return scaled

    def _start_open_animation(self) -> None:
        end_rect = self.card.geometry()
        if not end_rect.isValid():
            return

        start_rect = self._scaled_rect(end_rect, 0.985)
        self.card.setGeometry(start_rect)
        self.setWindowOpacity(0.0)

        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(140)
        opacity_animation.setStartValue(0.0)
        opacity_animation.setEndValue(1.0)
        opacity_animation.setEasingCurve(QEasingCurve.OutCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(150)
        geometry_animation.setStartValue(start_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.OutCubic)

        animation.start()
        self._open_animation = animation


class SafeApplication(QApplication):
    errorRaised = Signal(str, str, str)

    def __init__(self, argv: List[str]) -> None:
        super().__init__(argv)
        self._error_presenter = None
        self._presenting_error = False
        self.errorRaised.connect(self._dispatch_error)

    def set_error_presenter(self, presenter) -> None:
        self._error_presenter = presenter

    def notify(self, receiver, event) -> bool:
        try:
            return super().notify(receiver, event)
        except Exception as exc:
            self.report_error(
                "Terjadi kesalahan pada antarmuka aplikasi",
                str(exc) or exc.__class__.__name__,
                "".join(traceback.format_exception(type(exc), exc, exc.__traceback__)),
            )
            return False

    def report_error(self, title: str, summary: str, details: str) -> None:
        if self._presenting_error:
            sys.stderr.write(f"{title}\n{summary}\n{details}\n")
            return
        self.errorRaised.emit(title, summary, details)

    def _dispatch_error(self, title: str, summary: str, details: str) -> None:
        if self._presenting_error:
            return

        self._presenting_error = True
        try:
            if callable(self._error_presenter):
                self._error_presenter(title, summary, details)
            else:
                QMessageBox.critical(None, title, summary)
        finally:
            self._presenting_error = False


class FolderCompareDeleteApp(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(1560, 940)
        self.setMinimumSize(1280, 780)

        self.result_rows: List[MatchResult] = []
        self.compare_folder_rows: List[Dict[str, Any]] = []
        self.scan_thread: Optional[threading.Thread] = None
        self.delete_thread: Optional[threading.Thread] = None
        self.ui_queue: queue.Queue[Tuple[str, object]] = queue.Queue()
        self.openpyxl_available = False
        self.current_selected_result: Optional[MatchResult] = None
        self._last_queued_progress = -1.0
        self._last_queued_progress_text = ""
        self._pending_progress: Optional[Tuple[float, str]] = None
        self._pending_scan_results: Optional[List[MatchResult]] = None
        self._awaiting_scan_finalize = False
        self._pending_delete_result: Optional[Dict[str, Any]] = None
        self.delete_confirm_dialog: Optional[ConfirmOverlayDialog] = None
        self.delete_processing_dialog: Optional[ProcessingOverlayDialog] = None
        self._progress_lock = threading.Lock()

        self.stat_labels: Dict[str, QLabel] = {}
        self.detail_labels: Dict[str, QLabel] = {}

        self._build_ui()
        self._apply_styles()
        self.progress_animation = QPropertyAnimation(self.progress_bar, b"value", self)
        self.progress_animation.setEasingCurve(QEasingCurve.OutCubic)
        self.progress_animation.finished.connect(self._on_progress_animation_finished)
        app_instance = QApplication.instance()
        if isinstance(app_instance, SafeApplication):
            app_instance.set_error_presenter(self.show_error_dialog)
        self._check_excel_support()
        self._reset_detail_panel()
        self._refresh_stats()
        self._update_table_empty_state()
        self._set_progress(0, "Siap untuk scan folder.")

        self.filter_timer = QTimer(self)
        self.filter_timer.setSingleShot(True)
        self.filter_timer.setInterval(180)
        self.filter_timer.timeout.connect(self._apply_debounced_filter)

        self.queue_timer = QTimer(self)
        self.queue_timer.timeout.connect(self._poll_queue)
        self.queue_timer.start(33)

        self.show_only_matches_checkbox.toggled.connect(self._on_filter_changed)
        self.add_compare_folder_row()
        self.add_compare_folder_row()

    def _build_ui(self) -> None:
        central = QWidget()
        self.setCentralWidget(central)

        outer = QVBoxLayout(central)
        outer.setContentsMargins(20, 20, 20, 20)
        outer.setSpacing(16)

        outer.addWidget(self._build_header())
        outer.addLayout(self._build_stat_cards())

        self.main_splitter = QSplitter(Qt.Horizontal)
        self.main_splitter.setObjectName("MainSplitter")
        self.main_splitter.setChildrenCollapsible(False)
        self.main_splitter.setHandleWidth(10)
        self.main_splitter.addWidget(self._build_left_panel())
        self.main_splitter.addWidget(self._build_right_panel())
        self.main_splitter.setSizes([420, 1080])
        outer.addWidget(self.main_splitter, 1)

        footer = QFrame()
        footer.setObjectName("FooterBar")
        footer_layout = QHBoxLayout(footer)
        footer_layout.setContentsMargins(16, 10, 16, 10)
        footer_layout.setSpacing(12)

        self.status_label = QLabel("Siap untuk scan folder.")
        self.status_label.setObjectName("FooterLabel")
        self.status_label.setWordWrap(True)
        footer_layout.addWidget(self.status_label)

        outer.addWidget(footer)

    def _build_header(self) -> QWidget:
        header = QFrame()
        header.setObjectName("HeaderCard")
        layout = QHBoxLayout(header)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(24)

        left_wrap = QHBoxLayout()
        left_wrap.setSpacing(18)

        logo_card = QFrame()
        logo_card.setObjectName("HeaderLogoCard")
        logo_card.setFixedSize(84, 84)
        logo_layout = QVBoxLayout(logo_card)
        logo_layout.setContentsMargins(0, 0, 0, 0)
        logo_layout.setSpacing(0)

        logo_mark = QLabel("FC")
        logo_mark.setObjectName("HeaderLogoText")
        logo_mark.setAlignment(Qt.AlignCenter)
        logo_layout.addWidget(logo_mark)

        text_group = QVBoxLayout()
        text_group.setSpacing(6)

        eyebrow = QLabel("FILE OPERATIONS SUITE")
        eyebrow.setObjectName("HeaderEyebrow")

        title = QLabel("Folder Compare Delete")
        title.setObjectName("HeaderTitle")

        subtitle = QLabel(
            "Bandingkan isi beberapa folder, identifikasi duplikat dan perbedaan secara visual, lalu proses file dari Folder A dengan lebih aman."
        )
        subtitle.setObjectName("HeaderSubtitle")
        subtitle.setWordWrap(True)

        meta_row = QHBoxLayout()
        meta_row.setSpacing(8)
        meta_row.addWidget(self._create_header_chip("Multi-folder compare"))
        meta_row.addWidget(self._create_header_chip("Visual diff review"))
        meta_row.addWidget(self._create_header_chip("Safe delete flow"))
        meta_row.addStretch(1)

        text_group.addWidget(eyebrow)
        text_group.addWidget(title)
        text_group.addWidget(subtitle)
        text_group.addLayout(meta_row)

        left_wrap.addWidget(logo_card, 0, Qt.AlignTop)
        left_wrap.addLayout(text_group, 1)
        layout.addLayout(left_wrap, 1)

        hero = QFrame()
        hero.setObjectName("HeroInfo")
        hero_layout = QVBoxLayout(hero)
        hero_layout.setContentsMargins(18, 16, 18, 16)
        hero_layout.setSpacing(4)

        tip_title = QLabel("Workflow")
        tip_title.setObjectName("HeroInfoTitle")
        tip_body = QLabel(
            "1. Pilih Folder A dan folder pembanding.\n"
            "2. Scan via nama+ukuran atau hash+ukuran.\n"
            "3. Tinjau hasil hijau sebelum hapus."
        )
        tip_body.setObjectName("HeroInfoBody")
        tip_body.setWordWrap(True)

        hero_layout.addWidget(tip_title)
        hero_layout.addWidget(tip_body)
        layout.addWidget(hero, 0)

        return header

    def _create_header_chip(self, text: str) -> QWidget:
        chip = QFrame()
        chip.setObjectName("HeaderChip")
        chip_layout = QHBoxLayout(chip)
        chip_layout.setContentsMargins(10, 6, 10, 6)
        chip_layout.setSpacing(6)

        dot = QFrame()
        dot.setObjectName("HeaderChipDot")
        dot.setFixedSize(8, 8)

        label = QLabel(text)
        label.setObjectName("HeaderChipLabel")

        chip_layout.addWidget(dot, 0, Qt.AlignVCenter)
        chip_layout.addWidget(label, 0, Qt.AlignVCenter)
        return chip

    def _build_stat_cards(self) -> QHBoxLayout:
        layout = QHBoxLayout()
        layout.setSpacing(12)

        cards = [
            ("total", "Total Hasil", CYAN, "Semua file target yang dianalisis"),
            ("exact", "Duplikat Hijau", GREEN, "Aman dipertimbangkan dihapus"),
            ("diff", "Perbedaan Merah", RED, "Nama sama tetapi isi berbeda"),
            ("only", "Hanya di A", ORANGE, "Tidak ditemukan di folder lain"),
        ]
        for key, title, accent, subtitle in cards:
            layout.addWidget(self._create_stat_card(title, accent, subtitle, key))

        return layout

    def _create_stat_card(self, title: str, accent: str, subtitle: str, key: str) -> QWidget:
        card = QFrame()
        card.setObjectName("StatCard")
        background_map = {
            CYAN: (("#1d63f2", "#3b82ff"), "#5d9bff"),
            GREEN: (("#0d9b6b", "#24c78b"), "#5ee1af"),
            RED: (("#d94b6a", "#ff6f8b"), "#ff96ab"),
            ORANGE: (("#d88916", "#ffb23d"), "#ffc96f"),
        }
        (gradient_start, gradient_end), card_border = background_map.get(accent, ((SURFACE, SURFACE_ALT), BORDER))
        card.setStyleSheet(
            f"""
            QFrame#StatCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 {gradient_start},
                    stop: 1 {gradient_end}
                );
                border: 1px solid {card_border};
                border-radius: 20px;
            }}
            QLabel#StatTitle {{
                color: rgba(255, 255, 255, 0.88);
                font: 700 10pt "Segoe UI";
                background: transparent;
            }}
            QLabel#StatValue {{
                color: #ffffff;
                font: 700 25px "Segoe UI";
                background: transparent;
            }}
            QLabel#StatSubtitle {{
                color: rgba(255, 255, 255, 0.9);
                font: 9pt "Segoe UI";
                background: transparent;
            }}
            """
        )
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(18, 16, 18, 16)
        card_layout.setSpacing(6)

        body = QWidget()
        body_layout = QVBoxLayout(body)
        body_layout.setContentsMargins(0, 0, 0, 0)
        body_layout.setSpacing(6)

        title_label = QLabel(title)
        title_label.setObjectName("StatTitle")
        value_label = QLabel("0")
        value_label.setObjectName("StatValue")
        subtitle_label = QLabel(subtitle)
        subtitle_label.setObjectName("StatSubtitle")
        subtitle_label.setWordWrap(True)

        self.stat_labels[key] = value_label

        body_layout.addWidget(title_label)
        body_layout.addWidget(value_label)
        body_layout.addWidget(subtitle_label)

        card_layout.addWidget(body)
        return card

    def _build_left_panel(self) -> QWidget:
        container = QFrame()
        container.setObjectName("SurfaceCard")
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setObjectName("PanelScroll")
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.viewport().setObjectName("PanelViewport")

        body = QWidget()
        body.setObjectName("PanelScrollBody")
        layout = QVBoxLayout(body)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(14)

        heading = QLabel("Konfigurasi Scan")
        heading.setObjectName("SectionTitle")
        description = QLabel("Atur folder target, folder pembanding, mode pencocokan, dan aksi hasil scan.")
        description.setObjectName("SectionSubtitle")
        description.setWordWrap(True)
        layout.addWidget(heading)
        layout.addWidget(description)

        layout.addWidget(self._build_folder_input("Folder A (target hapus)", is_target=True))
        layout.addWidget(self._build_compare_group())
        layout.addWidget(self._build_option_group())
        layout.addWidget(self._build_action_group())
        layout.addStretch(1)

        scroll.setWidget(body)
        container_layout.addWidget(scroll)
        return container

    def _build_folder_input(self, label_text: str, is_target: bool = False) -> QWidget:
        card = QFrame()
        card.setObjectName("SubCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        label = QLabel(label_text)
        label.setObjectName("FieldLabel")

        row = QHBoxLayout()
        row.setSpacing(10)

        line_edit = FolderPathLineEdit("Pilih atau drop folder...")
        line_edit.setClearButtonEnabled(True)
        browse_button = QPushButton("Pilih Folder")
        browse_button.setObjectName("GhostButton")
        browse_button.clicked.connect(lambda: self.pick_folder(line_edit))
        line_edit.folderDropped.connect(lambda path: self.status_label.setText(f"Folder didrop: {path}"))

        status_label = self._create_path_status_label()
        self._bind_path_field(line_edit, status_label)

        row.addWidget(line_edit, 1)
        row.addWidget(browse_button, 0)

        layout.addWidget(label)
        layout.addLayout(row)
        layout.addWidget(status_label)

        if is_target:
            self.target_folder_edit = line_edit
            self.target_folder_status_label = status_label

        return card

    def _build_compare_group(self) -> QWidget:
        card = QFrame()
        card.setObjectName("SubCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        title = QLabel("Folder Pembanding")
        title.setObjectName("FieldLabel")
        description = QLabel("Tambahkan satu atau lebih folder untuk dibandingkan terhadap Folder A.")
        description.setObjectName("MutedText")
        description.setWordWrap(True)

        self.compare_list_widget = QWidget()
        self.compare_list_layout = QVBoxLayout(self.compare_list_widget)
        self.compare_list_layout.setContentsMargins(0, 0, 0, 0)
        self.compare_list_layout.setSpacing(8)

        buttons = QHBoxLayout()
        buttons.setSpacing(8)

        add_button = QPushButton("+ Tambah Folder")
        add_button.setObjectName("GhostButton")
        add_button.clicked.connect(self.add_compare_folder_row)

        remove_button = QPushButton("- Hapus Terakhir")
        remove_button.setObjectName("OutlineButton")
        remove_button.clicked.connect(self.remove_compare_folder_row)

        buttons.addWidget(add_button)
        buttons.addWidget(remove_button)
        buttons.addStretch(1)

        layout.addWidget(title)
        layout.addWidget(description)
        layout.addWidget(self.compare_list_widget)
        layout.addLayout(buttons)
        return card

    def _build_option_group(self) -> QWidget:
        wrapper = QWidget()
        layout = QVBoxLayout(wrapper)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        method_card = QFrame()
        method_card.setObjectName("SubCard")
        method_layout = QVBoxLayout(method_card)
        method_layout.setContentsMargins(14, 14, 14, 14)
        method_layout.setSpacing(8)

        method_title = QLabel("Metode Pencocokan")
        method_title.setObjectName("FieldLabel")

        self.compare_mode_group = QButtonGroup(self)
        self.compare_mode_name_size = QRadioButton("Nama file + ukuran (cepat)")
        self.compare_mode_hash = QRadioButton("Hash SHA-256 + ukuran (akurat)")
        self.compare_mode_name_size.setChecked(True)
        self.compare_mode_group.addButton(self.compare_mode_name_size)
        self.compare_mode_group.addButton(self.compare_mode_hash)

        method_layout.addWidget(method_title)
        method_layout.addWidget(self.compare_mode_name_size)
        method_layout.addWidget(self.compare_mode_hash)

        extra_card = QFrame()
        extra_card.setObjectName("SubCard")
        extra_layout = QVBoxLayout(extra_card)
        extra_layout.setContentsMargins(14, 14, 14, 14)
        extra_layout.setSpacing(8)

        extra_title = QLabel("Opsi Tampilan dan Hapus")
        extra_title.setObjectName("FieldLabel")

        self.include_subfolders_checkbox = QCheckBox("Sertakan subfolder")
        self.include_subfolders_checkbox.setChecked(True)
        self.show_only_matches_checkbox = QCheckBox("Tampilkan hanya file dengan kecocokan / perbedaan")
        self.delete_mode_recycle = QRadioButton("Hapus ke Recycle Bin")
        self.delete_mode_permanent = QRadioButton("Hapus permanen")
        self.delete_mode_recycle.setChecked(True)

        self.delete_mode_group = QButtonGroup(self)
        self.delete_mode_group.addButton(self.delete_mode_recycle)
        self.delete_mode_group.addButton(self.delete_mode_permanent)

        extra_layout.addWidget(extra_title)
        extra_layout.addWidget(self.include_subfolders_checkbox)
        extra_layout.addWidget(self.show_only_matches_checkbox)
        extra_layout.addSpacing(6)
        extra_layout.addWidget(self.delete_mode_recycle)
        extra_layout.addWidget(self.delete_mode_permanent)

        layout.addWidget(method_card)
        layout.addWidget(extra_card)
        return wrapper

    def _build_action_group(self) -> QWidget:
        card = QFrame()
        card.setObjectName("SubCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        self.scan_button = QPushButton("Scan dan Bandingkan")
        self.scan_button.setObjectName("PrimaryButton")
        self.scan_button.clicked.connect(self.start_scan)

        first_row = QHBoxLayout()
        first_row.setSpacing(8)
        self.export_csv_button = QPushButton("CSV")
        self.export_csv_button.setObjectName("GhostButton")
        self.export_csv_button.clicked.connect(self.export_csv)

        self.export_excel_button = QPushButton("Excel")
        self.export_excel_button.setObjectName("GhostButton")
        self.export_excel_button.clicked.connect(self.export_excel)

        self.clear_button = QPushButton("Reset")
        self.clear_button.setObjectName("OutlineButton")
        self.clear_button.clicked.connect(self.clear_results)

        first_row.addWidget(self.export_csv_button)
        first_row.addWidget(self.export_excel_button)
        first_row.addWidget(self.clear_button)

        second_row = QHBoxLayout()
        second_row.setSpacing(8)

        self.delete_button = QPushButton("Hapus Terpilih")
        self.delete_button.setObjectName("DangerButton")
        self.delete_button.clicked.connect(self.delete_selected)

        self.delete_all_button = QPushButton("Hapus Semua Hijau")
        self.delete_all_button.setObjectName("DangerButton")
        self.delete_all_button.clicked.connect(self.delete_all_results)

        second_row.addWidget(self.delete_button)
        second_row.addWidget(self.delete_all_button)

        layout.addWidget(self.scan_button)
        layout.addLayout(first_row)
        layout.addLayout(second_row)

        self.export_csv_button.setEnabled(False)
        self.export_excel_button.setEnabled(False)
        self.delete_button.setEnabled(False)
        self.delete_all_button.setEnabled(False)
        return card

    def _build_right_panel(self) -> QWidget:
        container = QFrame()
        container.setObjectName("SurfaceCard")
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setObjectName("PanelScroll")
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.viewport().setObjectName("PanelViewport")

        body = QWidget()
        body.setObjectName("PanelScrollBody")
        body_layout = QVBoxLayout(body)
        body_layout.setContentsMargins(18, 18, 18, 18)
        body_layout.setSpacing(12)

        heading = QLabel("Visual Diff Result")
        heading.setObjectName("SectionTitle")

        body_layout.addWidget(heading)
        body_layout.addWidget(self._build_navigation_card())
        self.progress_card = self._build_progress_card()
        self.progress_card.setVisible(False)
        body_layout.addWidget(self.progress_card)
        body_layout.addWidget(self._build_results_table())
        self.detail_panel = self._build_detail_panel()
        self.detail_panel.setVisible(False)
        body_layout.addWidget(self.detail_panel)
        body_layout.addStretch(1)

        scroll.setWidget(body)
        layout.addWidget(scroll)
        return container

    def _build_progress_card(self) -> QWidget:
        card = QFrame()
        card.setObjectName("SubCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        top = QHBoxLayout()
        top.setSpacing(8)

        label = QLabel("Progress Scan")
        label.setObjectName("FieldLabel")
        self.progress_badge = QLabel("0%")
        self.progress_badge.setObjectName("ProgressBadge")
        self.progress_badge.setAlignment(Qt.AlignCenter)

        top.addWidget(label)
        top.addStretch(1)
        top.addWidget(self.progress_badge)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 1000)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(False)

        layout.addLayout(top)
        layout.addWidget(self.progress_bar)
        return card

    def _build_navigation_card(self) -> QWidget:
        card = QFrame()
        card.setObjectName("SubCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        top_row = QHBoxLayout()
        top_row.setSpacing(10)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Cari status, path, relative path, atau folder...")
        self.search_input.setClearButtonEnabled(True)
        self.search_input.textChanged.connect(self._schedule_search_filter)

        self.reset_filter_button = QPushButton("Reset Filter")
        self.reset_filter_button.setObjectName("OutlineButton")
        self.reset_filter_button.clicked.connect(self._reset_quick_filters)

        top_row.addWidget(self.search_input, 1)
        top_row.addWidget(self.reset_filter_button)

        chip_row = QHBoxLayout()
        chip_row.setSpacing(8)

        self.quick_filter_group = QButtonGroup(self)
        self.quick_filter_group.setExclusive(True)
        self.quick_filter_buttons: Dict[str, QPushButton] = {}

        chip_specs = [
            ("all", "Semua", CYAN),
            ("exact_match", "Duplikat", GREEN),
            ("different_content", "Berbeda", RED),
            ("only_target", "Hanya di A", ORANGE),
        ]

        for key, label, color in chip_specs:
            button = QPushButton(label)
            button.setCheckable(True)
            button.setObjectName("FilterChip")
            button.setProperty("chipColor", color)
            button.clicked.connect(self._on_filter_changed)
            self.quick_filter_group.addButton(button)
            self.quick_filter_buttons[key] = button
            chip_row.addWidget(button)

        self.quick_filter_buttons["all"].setChecked(True)
        chip_row.addStretch(1)

        layout.addLayout(top_row)
        layout.addLayout(chip_row)
        return card

    def _build_results_table(self) -> QWidget:
        card = QFrame()
        card.setObjectName("TableCard")
        card.setMinimumHeight(360)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(12, 12, 12, 12)

        self.table_stack_host = QWidget()
        self.table_stack = QStackedLayout(self.table_stack_host)
        self.table_stack.setContentsMargins(0, 0, 0, 0)
        self.table_stack.setStackingMode(QStackedLayout.StackOne)

        self.results_table = ResponsiveTableWidget()
        self.table_model = MatchResultTableModel()
        self.table_proxy = MatchResultFilterProxyModel()
        self.table_proxy.setSourceModel(self.table_model)
        self.results_table.setModel(self.table_proxy)
        self.results_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.results_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.results_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.results_table.setAlternatingRowColors(False)
        self.results_table.setWordWrap(False)
        self.results_table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.results_table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.results_table.setTextElideMode(Qt.ElideMiddle)
        self.results_table.setShowGrid(True)
        self.results_table.verticalHeader().setVisible(False)
        self.results_table.verticalHeader().setDefaultSectionSize(34)
        self.results_table.setMinimumHeight(320)
        self.results_table.setSortingEnabled(True)

        header = self.results_table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setMinimumSectionSize(90)
        header.setSectionsMovable(False)
        header.setSortIndicatorShown(True)
        for column in range(self.table_model.columnCount()):
            header.setSectionResizeMode(column, QHeaderView.Interactive)

        self._apply_default_table_widths()
        self.results_table.sortByColumn(0, Qt.AscendingOrder)

        selection_model = self.results_table.selectionModel()
        selection_model.selectionChanged.connect(self._on_table_selection_changed)
        selection_model.currentRowChanged.connect(self._on_current_row_changed)
        self.results_table.focusReleased.connect(self._on_table_focus_released)

        self.table_empty_state = self._build_table_empty_state()

        self.table_stack.addWidget(self.results_table)
        self.table_stack.addWidget(self.table_empty_state)
        layout.addWidget(self.table_stack_host)
        return card

    def _build_table_empty_state(self) -> QWidget:
        empty = QFrame()
        empty.setObjectName("TableEmptyState")
        layout = QVBoxLayout(empty)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(10)
        layout.setAlignment(Qt.AlignCenter)

        icon = QLabel("[]")
        icon.setObjectName("EmptyStateIcon")
        icon.setAlignment(Qt.AlignCenter)

        self.empty_state_title = QLabel("Belum ada hasil scan")
        self.empty_state_title.setObjectName("EmptyStateTitle")
        self.empty_state_title.setAlignment(Qt.AlignCenter)

        self.empty_state_description = QLabel(
            "Pilih folder target dan folder pembanding, lalu jalankan Scan dan Bandingkan untuk melihat hasil."
        )
        self.empty_state_description.setObjectName("EmptyStateDescription")
        self.empty_state_description.setAlignment(Qt.AlignCenter)
        self.empty_state_description.setWordWrap(True)

        layout.addWidget(icon)
        layout.addWidget(self.empty_state_title)
        layout.addWidget(self.empty_state_description)
        return empty

    def _build_detail_panel(self) -> QWidget:
        card = QFrame()
        card.setObjectName("SubCard")
        card.setMinimumHeight(320)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(12)

        top = QHBoxLayout()
        top.setSpacing(8)
        title = QLabel("Detail File Terpilih")
        title.setObjectName("FieldLabel")
        copy_button = QPushButton("Copy Path")
        copy_button.setObjectName("GhostButton")
        copy_button.clicked.connect(self.copy_selected_path)

        top.addWidget(title)
        top.addStretch(1)
        top.addWidget(copy_button)

        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(12)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)
        grid.setRowStretch(4, 1)

        fields = [
            ("status", "Status"),
            ("target", "Path Target"),
            ("relative", "Relative Path"),
            ("size", "Ukuran"),
            ("found", "Ditemukan di"),
            ("missing", "Tidak Ada di"),
            ("mode", "Mode"),
            ("exact", "Path Cocok"),
            ("diff", "Path Beda"),
        ]
        positions = {
            "status": (0, 0),
            "target": (1, 0),
            "relative": (2, 0),
            "size": (3, 0),
            "found": (4, 0),
            "missing": (0, 1),
            "mode": (1, 1),
            "exact": (2, 1),
            "diff": (3, 1),
        }

        for key, title_text in fields:
            if key not in positions:
                continue
            row, column = positions[key]
            grid.addWidget(self._create_detail_card(key, title_text), row, column)

        layout.addLayout(top)
        layout.addLayout(grid)
        return card

    def _create_detail_card(self, key: str, title: str) -> QWidget:
        card = QFrame()
        card.setObjectName("DetailCard")
        card.setMinimumHeight(78)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(6)

        title_label = QLabel(title)
        title_label.setObjectName("DetailTitle")
        value_label = QLabel("-")
        value_label.setObjectName("DetailValue")
        value_label.setWordWrap(True)
        value_label.setMinimumHeight(22)
        value_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        layout.addWidget(title_label)
        layout.addWidget(value_label)

        self.detail_labels[key] = value_label
        return card

    def _create_path_status_label(self) -> QLabel:
        label = QLabel("Belum ada folder dipilih.")
        label.setObjectName("PathStatus")
        label.setWordWrap(True)
        return label

    def _bind_path_field(self, line_edit: QLineEdit, status_label: QLabel) -> None:
        line_edit.textChanged.connect(
            lambda _text, edit=line_edit, status=status_label: self._update_path_field_state(edit, status)
        )
        self._update_path_field_state(line_edit, status_label)

    def _normalize_folder_path(self, raw_path: str) -> str:
        cleaned = raw_path.strip().strip('"')
        if not cleaned:
            return ""
        return os.path.normpath(str(Path(cleaned).expanduser()))

    def _update_path_field_state(self, line_edit: QLineEdit, status_label: QLabel) -> None:
        raw_value = line_edit.text()
        normalized = self._normalize_folder_path(raw_value)

        if not normalized:
            line_edit.setProperty("pathState", "empty")
            status_label.setProperty("pathState", "empty")
            status_label.setText("Belum ada folder dipilih.")
        elif Path(normalized).is_dir():
            line_edit.setProperty("pathState", "valid")
            status_label.setProperty("pathState", "valid")
            status_label.setText(f"Folder valid: {normalized}")
        else:
            line_edit.setProperty("pathState", "invalid")
            status_label.setProperty("pathState", "invalid")
            status_label.setText(f"Folder belum valid / tidak ditemukan: {normalized}")

        line_edit.setToolTip(normalized or "Belum ada folder dipilih")
        status_label.setToolTip(status_label.text())
        self._refresh_widget_style(line_edit)
        self._refresh_widget_style(status_label)

    def _refresh_widget_style(self, widget: QWidget) -> None:
        style = widget.style()
        style.unpolish(widget)
        style.polish(widget)
        widget.update()

    def _stylesheet_url(self, path: Path) -> str:
        return path.resolve().as_posix()

    def _apply_styles(self) -> None:
        checkbox_check_url = self._stylesheet_url(Path(__file__).with_name("checkbox_check.svg"))
        sort_up_url = self._stylesheet_url(Path(__file__).with_name("sort_up.svg"))
        sort_down_url = self._stylesheet_url(Path(__file__).with_name("sort_down.svg"))
        self.setStyleSheet(
            f"""
            QMainWindow {{
                background: {BG_COLOR};
                color: {TEXT};
            }}
            QLabel {{
                color: {TEXT};
            }}
            QFrame#HeaderCard {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #173fbb, stop:0.55 #2b61df, stop:1 #3f86ff);
                border-radius: 24px;
            }}
            QFrame#HeaderLogoCard {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 rgba(255,255,255,0.18), stop:1 rgba(255,255,255,0.08));
                border: 1px solid rgba(255, 255, 255, 0.18);
                border-radius: 22px;
            }}
            QLabel#HeaderLogoText {{
                color: white;
                font: 800 24pt "Segoe UI";
                letter-spacing: 1px;
            }}
            QLabel#HeaderEyebrow {{
                color: rgba(236, 243, 255, 0.76);
                font: 700 8.5pt "Segoe UI";
                letter-spacing: 1.6px;
            }}
            QLabel#HeaderTitle {{
                color: white;
                font: 700 24pt "Segoe UI Semibold";
                letter-spacing: 0.3px;
            }}
            QLabel#HeaderSubtitle {{
                color: rgba(242, 246, 255, 0.92);
                font: 10.5pt "Segoe UI";
                line-height: 1.35;
            }}
            QFrame#HeroInfo {{
                background: rgba(255, 255, 255, 0.12);
                border: 1px solid rgba(255, 255, 255, 0.18);
                border-radius: 18px;
                min-width: 290px;
            }}
            QLabel#HeroInfoTitle {{
                color: white;
                font: 700 11.5pt "Segoe UI Semibold";
            }}
            QLabel#HeroInfoBody {{
                color: rgba(241, 246, 255, 0.95);
                font: 9.5pt "Segoe UI";
            }}
            QFrame#HeaderChip {{
                background: rgba(255, 255, 255, 0.12);
                border: 1px solid rgba(255, 255, 255, 0.16);
                border-radius: 12px;
            }}
            QFrame#HeaderChipDot {{
                background: #b9d4ff;
                border-radius: 4px;
            }}
            QLabel#HeaderChipLabel {{
                color: rgba(248, 250, 255, 0.92);
                font: 700 8.8pt "Segoe UI";
            }}
            QFrame#SurfaceCard, QFrame#StatCard, QFrame#SubCard, QFrame#TableCard {{
                background: {SURFACE};
                border: 1px solid {BORDER};
                border-radius: 20px;
            }}
            QWidget#PanelScrollBody {{
                background: #f8fbff;
                border-radius: 20px;
            }}
            QWidget#PanelViewport {{
                background: #f8fbff;
                border-radius: 20px;
            }}
            QScrollArea#PanelScroll {{
                background: #f8fbff;
                border: none;
                border-radius: 20px;
            }}
            QSplitter#MainSplitter::handle {{
                background: transparent;
            }}
            QSplitter#MainSplitter::handle:hover {{
                background: #dde6f6;
                border-radius: 4px;
            }}
            QLabel#SectionTitle {{
                font: 700 14pt "Segoe UI";
                color: {TEXT};
            }}
            QLabel#SectionSubtitle {{
                color: {MUTED};
                font: 9.5pt "Segoe UI";
            }}
            QLabel#FieldLabel {{
                font: 700 10pt "Segoe UI";
                color: {TEXT};
            }}
            QLabel#MutedText {{
                font: 9pt "Segoe UI";
                color: {MUTED};
            }}
            QLabel#StatTitle {{
                font: 700 10pt "Segoe UI";
                color: {MUTED};
            }}
            QLabel#StatValue {{
                font: 700 25px "Segoe UI";
                color: {TEXT};
            }}
            QLabel#StatSubtitle {{
                font: 9pt "Segoe UI";
                color: {MUTED};
            }}
            QLineEdit {{
                background: white;
                color: {TEXT};
                border: 1px solid {BORDER};
                border-radius: 12px;
                padding: 10px 12px;
                selection-background-color: #cfe0ff;
                min-height: 20px;
            }}
            QLineEdit[pathState="valid"] {{
                border: 1px solid #a8e2c4;
                background: #fbfffd;
            }}
            QLineEdit[pathState="invalid"] {{
                border: 1px solid #ffc3ce;
                background: #fffafb;
            }}
            QLineEdit:focus {{
                border: 1px solid {PRIMARY};
            }}
            QLineEdit::placeholder {{
                color: #8d98ab;
            }}
            QPushButton {{
                border-radius: 12px;
                padding: 10px 14px;
                font: 600 9.5pt "Segoe UI";
            }}
            QPushButton#PrimaryButton {{
                background: {PRIMARY};
                color: white;
                border: none;
            }}
            QPushButton#PrimaryButton:hover {{
                background: {PRIMARY_DARK};
            }}
            QPushButton#GhostButton {{
                background: #edf3ff;
                color: {TEXT};
                border: 1px solid #dce6fb;
            }}
            QPushButton#GhostButton:hover {{
                background: #e2ecff;
            }}
            QPushButton#OutlineButton {{
                background: white;
                color: {TEXT};
                border: 1px solid {BORDER};
            }}
            QPushButton#OutlineButton:hover {{
                background: {SURFACE_ALT};
            }}
            QPushButton#FilterChip {{
                background: #f6f9ff;
                color: {TEXT};
                border: 1px solid #d8e2f4;
                border-radius: 16px;
                padding: 8px 14px;
                font: 700 9pt "Segoe UI";
            }}
            QPushButton#FilterChip:hover {{
                background: #eef4ff;
                border: 1px solid #c6d6f4;
            }}
            QPushButton#FilterChip:checked {{
                background: #e9f1ff;
                color: {PRIMARY};
                border: 1px solid #aecaef;
            }}
            QPushButton#DangerButton {{
                background: #fff0f3;
                color: #9d263a;
                border: 1px solid #ffd2da;
            }}
            QPushButton#DangerButton:hover {{
                background: #ffe5eb;
            }}
            QPushButton:disabled {{
                background: #f0f3f8;
                color: #99a4b8;
                border: 1px solid #e2e7f0;
            }}
            QRadioButton, QCheckBox {{
                spacing: 8px;
                color: {TEXT};
                font: 9.5pt "Segoe UI";
            }}
            QRadioButton::indicator, QCheckBox::indicator {{
                width: 18px;
                height: 18px;
                background: #f7faff;
                border: 1px solid #c7d5ec;
            }}
            QCheckBox::indicator {{
                border-radius: 5px;
            }}
            QRadioButton::indicator {{
                border-radius: 9px;
            }}
            QCheckBox::indicator:hover, QRadioButton::indicator:hover {{
                border: 1px solid #9fb6dc;
                background: #eef4ff;
            }}
            QCheckBox::indicator:checked, QRadioButton::indicator:checked {{
                background: #2d64f1;
                border: 1px solid #2d64f1;
            }}
            QCheckBox::indicator:unchecked, QRadioButton::indicator:unchecked {{
                background: #f9fbff;
                border: 1px solid #c7d5ec;
            }}
            QCheckBox::indicator:disabled, QRadioButton::indicator:disabled {{
                background: #eef2f8;
                border: 1px solid #d9e1ee;
            }}
            QCheckBox::indicator:checked {{
                image: url({checkbox_check_url});
            }}
            QRadioButton::indicator:checked {{
                background: qradialgradient(
                    cx: 0.5, cy: 0.5, radius: 0.7,
                    fx: 0.5, fy: 0.5,
                    stop: 0 #2d64f1,
                    stop: 0.45 #2d64f1,
                    stop: 0.46 #f7faff,
                    stop: 1 #f7faff
                );
                border: 1px solid #2d64f1;
            }}
            QProgressBar {{
                background: #edf2fb;
                border: none;
                border-radius: 6px;
                min-height: 8px;
                max-height: 8px;
            }}
            QProgressBar::chunk {{
                background: {PRIMARY};
                border-radius: 6px;
            }}
            QLabel#ProgressBadge {{
                background: #edf3ff;
                color: {PRIMARY};
                border-radius: 12px;
                font: 700 9pt "Segoe UI";
                padding: 4px 10px;
                min-width: 48px;
            }}
            QFrame#LegendChip {{
                background: #f7f9fc;
                border: 1px solid {BORDER};
                border-radius: 12px;
            }}
            QLabel#LegendLabel {{
                font: 700 9pt "Segoe UI";
                color: {TEXT};
            }}
            QTableView {{
                background: white;
                border: 1px solid {BORDER};
                border-radius: 16px;
                gridline-color: #edf1f7;
                color: {TEXT};
                selection-background-color: transparent;
                selection-color: {TEXT};
                font: 9pt "Segoe UI";
                outline: 0;
            }}
            QTableView::item {{
                padding: 4px 6px;
            }}
            QTableView::item:selected:active {{
                background: #dce8ff;
                color: {TEXT};
            }}
            QTableView::item:selected:!active {{
                background: transparent;
                color: {TEXT};
            }}
            QHeaderView::section {{
                background: #f0f4fb;
                color: {TEXT};
                padding: 10px 18px 10px 8px;
                border: none;
                border-right: 1px solid #e5ebf5;
                border-bottom: 1px solid #e5ebf5;
                font: 700 9pt "Segoe UI";
            }}
            QHeaderView::up-arrow {{
                image: url({sort_up_url});
                width: 12px;
                height: 12px;
            }}
            QHeaderView::down-arrow {{
                image: url({sort_down_url});
                width: 12px;
                height: 12px;
            }}
            QFrame#DetailCard {{
                background: #f8fafc;
                border: 1px solid {BORDER};
                border-radius: 14px;
            }}
            QFrame#TableEmptyState {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #fbfdff, stop:1 #f4f8ff);
                border: 1px dashed #cdd9ee;
                border-radius: 16px;
            }}
            QLabel#EmptyStateIcon {{
                color: #8eabdd;
                font: 700 22pt "Consolas";
                letter-spacing: 2px;
            }}
            QLabel#EmptyStateTitle {{
                color: {TEXT};
                font: 700 13pt "Segoe UI Semibold";
            }}
            QLabel#EmptyStateDescription {{
                color: {MUTED};
                font: 9.5pt "Segoe UI";
            }}
            QLabel#DetailTitle {{
                color: {MUTED};
                font: 700 9pt "Segoe UI";
            }}
            QLabel#DetailValue {{
                color: {TEXT};
                font: 9pt "Segoe UI";
            }}
            QLabel#PathStatus {{
                font: 8.8pt "Segoe UI";
                color: {MUTED};
            }}
            QLabel#PathStatus[pathState="valid"] {{
                color: #1c8f5b;
            }}
            QLabel#PathStatus[pathState="invalid"] {{
                color: #b24154;
            }}
            QFrame#FooterBar {{
                background: #182235;
                border-radius: 16px;
            }}
            QLabel#FooterLabel {{
                color: #edf2ff;
                font: 9.5pt "Segoe UI";
            }}
            QScrollArea {{
                background: transparent;
                border: none;
            }}
            QScrollArea#PanelScroll {{
                background: #f8fbff;
                border: none;
            }}
            QAbstractScrollArea {{
                background: #f8fbff;
                border: none;
            }}
            QScrollBar:vertical {{
                background: transparent;
                width: 12px;
                margin: 4px 0 4px 0;
            }}
            QScrollBar::handle:vertical {{
                background: #c5d2e8;
                min-height: 36px;
                border-radius: 6px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: #aebdd9;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
                background: transparent;
            }}
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                background: transparent;
            }}
            QScrollBar:horizontal {{
                background: transparent;
                height: 12px;
                margin: 0 4px 0 4px;
            }}
            QScrollBar::handle:horizontal {{
                background: #c5d2e8;
                min-width: 36px;
                border-radius: 6px;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: #aebdd9;
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
                background: transparent;
            }}
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{
                background: transparent;
            }}
            QMessageBox {{
                background: #ffffff;
                padding: 12px 14px 10px 14px;
            }}
            QMessageBox QWidget {{
                background: #ffffff;
                color: {TEXT};
            }}
            QMessageBox QLabel {{
                background: transparent;
                color: {TEXT};
                font: 9.5pt "Segoe UI";
                padding: 0px;
            }}
            QMessageBox QPushButton {{
                min-width: 92px;
                padding: 8px 16px;
                border-radius: 12px;
                border: 1px solid #d8e2f4;
                background: #edf3ff;
                color: {TEXT};
                font: 600 9.5pt "Segoe UI";
            }}
            QMessageBox QPushButton:hover {{
                background: #e2ecff;
                border: 1px solid #c6d6f4;
            }}
            """
        )
        self.status_label.setObjectName("FooterLabel")

    def _check_excel_support(self) -> None:
        try:
            import openpyxl  # noqa: F401
            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False
        self.export_excel_button.setEnabled(False)

    def add_compare_folder_row(self) -> None:
        row_index = len(self.compare_folder_rows) + 1

        row_widget = QFrame()
        row_layout = QVBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 0, 0, 0)
        row_layout.setSpacing(6)

        top_row = QHBoxLayout()
        top_row.setContentsMargins(0, 0, 0, 0)
        top_row.setSpacing(8)

        label = QLabel(f"Folder {row_index}")
        label.setObjectName("FieldLabel")
        label.setMinimumWidth(72)

        line_edit = FolderPathLineEdit("Pilih atau drop folder pembanding...")
        line_edit.setClearButtonEnabled(True)

        browse_button = QPushButton("Browse")
        browse_button.setObjectName("OutlineButton")
        browse_button.clicked.connect(lambda: self.pick_folder(line_edit))
        line_edit.folderDropped.connect(lambda path: self.status_label.setText(f"Folder didrop: {path}"))

        status_label = self._create_path_status_label()
        self._bind_path_field(line_edit, status_label)

        top_row.addWidget(label)
        top_row.addWidget(line_edit, 1)
        top_row.addWidget(browse_button)

        row_layout.addLayout(top_row)
        row_layout.addWidget(status_label)

        self.compare_folder_rows.append(
            {"widget": row_widget, "edit": line_edit, "label": label, "status": status_label}
        )
        self.compare_list_layout.addWidget(row_widget)

    def remove_compare_folder_row(self) -> None:
        if len(self.compare_folder_rows) <= 1:
            QMessageBox.information(self, APP_TITLE, "Minimal harus ada satu folder pembanding.")
            return

        last = self.compare_folder_rows.pop()
        widget = last["widget"]
        widget.setParent(None)
        widget.deleteLater()

    def pick_folder(self, line_edit: QLineEdit) -> None:
        start_dir = line_edit.text().strip() or str(Path.home())
        folder = QFileDialog.getExistingDirectory(self, "Pilih Folder", start_dir)
        if folder:
            normalized = self._normalize_folder_path(folder)
            line_edit.setText(normalized)
            line_edit.setCursorPosition(0)
            line_edit.setFocus()
            self.status_label.setText(f"Folder dipilih: {normalized}")

    def start_scan(self) -> None:
        if self.scan_thread and self.scan_thread.is_alive():
            QMessageBox.information(self, APP_TITLE, "Proses scan masih berjalan.")
            return

        target_folder = self._normalize_folder_path(self.target_folder_edit.text())
        compare_folders = [
            self._normalize_folder_path(row["edit"].text())
            for row in self.compare_folder_rows
            if self._normalize_folder_path(row["edit"].text())
        ]

        self.target_folder_edit.setText(target_folder)
        for row in self.compare_folder_rows:
            normalized = self._normalize_folder_path(row["edit"].text())
            if normalized != row["edit"].text():
                row["edit"].setText(normalized)

        if not target_folder:
            QMessageBox.warning(self, APP_TITLE, "Folder A wajib diisi.")
            return
        if not compare_folders:
            QMessageBox.warning(self, APP_TITLE, "Minimal isi satu folder pembanding.")
            return
        if any(Path(target_folder) == Path(folder) for folder in compare_folders):
            QMessageBox.warning(self, APP_TITLE, "Folder A tidak boleh sama dengan folder pembanding.")
            return

        self.clear_results(reset_status=False)
        self._set_action_state(scanning=True)
        self.progress_card.setVisible(True)
        self._last_queued_progress = -1.0
        self._last_queued_progress_text = ""
        with self._progress_lock:
            self._pending_progress = None
        self._set_progress(0, "Memulai scan folder...")

        self.scan_thread = threading.Thread(
            target=self._scan_worker,
            args=(
                target_folder,
                compare_folders,
                self._current_compare_mode(),
                self.include_subfolders_checkbox.isChecked(),
            ),
            daemon=True,
        )
        self.scan_thread.start()

    def _current_compare_mode(self) -> str:
        return "hash" if self.compare_mode_hash.isChecked() else "name_size"

    def _current_delete_mode(self) -> str:
        return "permanent" if self.delete_mode_permanent.isChecked() else "recycle_bin"

    def _set_action_state(self, scanning: bool) -> None:
        self.scan_button.setEnabled(not scanning)
        self.export_csv_button.setEnabled(False)
        self.export_excel_button.setEnabled(False)
        self.delete_button.setEnabled(False)
        self.delete_all_button.setEnabled(False)

    def _set_delete_processing_state(self, deleting: bool) -> None:
        self.scan_button.setEnabled(not deleting)
        self.export_csv_button.setEnabled(not deleting and bool(self.result_rows))
        self.export_excel_button.setEnabled(not deleting and bool(self.result_rows) and self.openpyxl_available)
        self.delete_button.setEnabled(not deleting and any(row.exact_matches for row in self.result_rows))
        self.delete_all_button.setEnabled(not deleting and any(row.exact_matches for row in self.result_rows))

    def _scan_worker(self, target_folder: str, compare_folders: List[str], mode: str, include_subfolders: bool) -> None:
        try:
            self._queue_progress(2.0, "Menyiapkan proses scan...")

            compare_paths = [Path(folder) for folder in compare_folders]
            folder_plan = [("A", Path(target_folder), "Folder A")] + [
                (f"F{idx}", folder_path, f"folder pembanding F{idx}")
                for idx, folder_path in enumerate(compare_paths, start=1)
            ]
            file_counts: Dict[str, int] = {}
            total_folders = len(folder_plan)

            for idx, (label, folder_path, folder_name) in enumerate(folder_plan, start=1):
                self._queue_progress(
                    self._progress_in_range(2.0, 10.0, idx - 1, total_folders),
                    f"Menghitung file di {folder_name}...",
                )
                file_counts[label] = self._count_files(folder_path, include_subfolders)
                self._queue_progress(
                    self._progress_in_range(2.0, 10.0, idx, total_folders),
                    f"Perkiraan file {folder_name}: {file_counts[label]} item.",
                )

            total_collect_files = file_counts.get("A", 0) + sum(file_counts.get(f"F{idx}", 0) for idx in range(1, len(compare_folders) + 1))
            collected_so_far = 0
            collect_stride = self._progress_stride(total_collect_files, 220)

            def emit_collect_progress(processed: int, folder_name: str) -> None:
                if total_collect_files <= 0:
                    return
                if processed not in {1, total_collect_files} and processed % collect_stride != 0:
                    return
                self._queue_progress(
                    self._progress_in_range(10.0, 45.0, processed, total_collect_files),
                    f"Mengambil metadata file dari {folder_name} ({processed}/{total_collect_files})...",
                )

            target_files = self._collect_files(
                Path(target_folder),
                include_subfolders,
                base_label="A",
                progress_callback=lambda folder_done: emit_collect_progress(folder_done, "Folder A"),
            )
            collected_so_far += len(target_files)

            compare_groups: List[Tuple[str, Path, List[FileRecord]]] = []
            for idx, folder_path in enumerate(compare_paths, start=1):
                label = f"F{idx}"
                folder_name = f"folder pembanding {label}"
                base_processed = collected_so_far
                compare_records = self._collect_files(
                    folder_path,
                    include_subfolders,
                    base_label=label,
                    progress_callback=lambda folder_done, base=base_processed, name=folder_name: emit_collect_progress(
                        base + folder_done,
                        name,
                    ),
                )
                collected_so_far += len(compare_records)
                compare_groups.append((label, folder_path, compare_records))

            self._queue_progress(45.0, "Menyusun indeks data pembanding...")
            results = self._build_comparison_results(target_files, compare_groups, mode)
            self.ui_queue.put(("scan_done", results))
        except Exception as exc:
            self.ui_queue.put(
                (
                    "scan_error",
                    {
                        "title": "Terjadi kesalahan saat scan dan bandingkan",
                        "summary": str(exc) or exc.__class__.__name__,
                        "details": "".join(traceback.format_exception(type(exc), exc, exc.__traceback__)),
                    },
                )
            )

    def _build_comparison_results(
        self,
        target_files: List[FileRecord],
        compare_groups: List[Tuple[str, Path, List[FileRecord]]],
        mode: str,
    ) -> List[MatchResult]:
        by_name_size: Dict[Tuple[str, int], List[FileRecord]] = {}
        by_relative_size: Dict[Tuple[str, int], List[FileRecord]] = {}
        by_relative_name: Dict[str, List[FileRecord]] = {}
        by_hash: Dict[Tuple[int, str], List[FileRecord]] = {}
        all_compare_labels = [label for label, _, _ in compare_groups]

        flat_compare_records: List[FileRecord] = []
        total_compare_records = sum(len(records) for _, _, records in compare_groups)
        indexed_records = 0
        index_stride = self._progress_stride(total_compare_records, 160)
        for _, _, records in compare_groups:
            for record in records:
                flat_compare_records.append(record)
                by_name_size.setdefault((record.path.name.lower(), record.size), []).append(record)
                by_relative_size.setdefault((record.relative_path.lower(), record.size), []).append(record)
                by_relative_name.setdefault(record.relative_path.lower(), []).append(record)
                indexed_records += 1
                if total_compare_records > 0 and (
                    indexed_records in {1, total_compare_records} or indexed_records % index_stride == 0
                ):
                    self._queue_progress(
                        self._progress_in_range(45.0, 55.0, indexed_records, total_compare_records),
                        f"Menyusun indeks data pembanding ({indexed_records}/{total_compare_records})...",
                    )
        if total_compare_records == 0:
            self._queue_progress(55.0, "Tidak ada file pada folder pembanding yang perlu diindeks.")

        if mode == "hash":
            total_hash_items = len(flat_compare_records) + len(target_files)
            processed = 0
            hash_stride = self._progress_stride(total_hash_items, 180)
            for record in flat_compare_records:
                record.sha256 = self._hash_file(record.path)
                by_hash.setdefault((record.size, record.sha256), []).append(record)
                processed += 1
                if processed in {1, total_hash_items} or processed % hash_stride == 0:
                    self._queue_progress(
                        self._progress_in_range(55.0, 82.0, processed, total_hash_items),
                        f"Hash file pembanding {processed}/{total_hash_items}...",
                    )

        results: List[MatchResult] = []
        total_targets = len(target_files)
        compare_progress_start = 82.0 if mode == "hash" else 55.0
        compare_progress_end = 98.5
        compare_stride = self._progress_stride(total_targets, 220)

        for idx, target in enumerate(target_files, start=1):
            exact_matches: List[FileRecord] = []
            same_name_different_content: List[FileRecord] = []

            if mode == "name_size":
                exact_matches.extend(by_name_size.get((target.path.name.lower(), target.size), []))
                exact_matches.extend(
                    item
                    for item in by_relative_size.get((target.relative_path.lower(), target.size), [])
                    if item not in exact_matches
                )
                same_name_different_content = [
                    item for item in by_relative_name.get(target.relative_path.lower(), []) if item.size != target.size
                ]
                match_type = "nama+ukuran"
            else:
                target.sha256 = self._hash_file(target.path)
                exact_matches = by_hash.get((target.size, target.sha256), [])
                same_name_different_content = [
                    item
                    for item in by_relative_name.get(target.relative_path.lower(), [])
                    if item.sha256 != target.sha256 or item.size != target.size
                ]
                match_type = "hash+ukuran"
                total_hash_steps = len(flat_compare_records) + len(target_files)
                processed_hash_steps = len(flat_compare_records) + idx
                if processed_hash_steps in {1, total_hash_steps} or processed_hash_steps % max(1, self._progress_stride(total_hash_steps, 180)) == 0:
                    self._queue_progress(
                        self._progress_in_range(55.0, 82.0, processed_hash_steps, total_hash_steps),
                        f"Hash file target {idx}/{total_targets}...",
                    )

            exact_labels = {item.base_label for item in exact_matches}
            diff_labels = {item.base_label for item in same_name_different_content}
            involved_labels = exact_labels.union(diff_labels)
            missing = [label for label in all_compare_labels if label not in involved_labels]

            results.append(
                MatchResult(
                    target_path=target.path,
                    target_relative_path=target.relative_path,
                    size=target.size,
                    match_type=match_type,
                    exact_matches=sorted(exact_matches, key=lambda x: (x.base_label, str(x.path))),
                    same_name_different_content=sorted(
                        same_name_different_content,
                        key=lambda x: (x.base_label, str(x.path)),
                    ),
                    missing_from_folders=missing,
                    only_in_target=not exact_matches and not same_name_different_content,
                )
            )
            if idx in {1, total_targets} or idx % compare_stride == 0:
                self._queue_progress(
                    self._progress_in_range(compare_progress_start, compare_progress_end, idx, total_targets),
                    f"Membandingkan file {idx}/{total_targets}...",
                )

        if total_targets == 0:
            self._queue_progress(compare_progress_end, "Tidak ada file di Folder A yang perlu dibandingkan.")

        self._queue_progress(99.5, "Menyiapkan tampilan visual...")
        return results

    def _count_files(self, folder: Path, include_subfolders: bool) -> int:
        if not folder.exists() or not folder.is_dir():
            raise FileNotFoundError(f"Folder tidak valid: {folder}")

        total = 0
        iterator: Iterable[Path] = folder.rglob("*") if include_subfolders else folder.glob("*")
        for path in iterator:
            if path.is_file():
                total += 1
        return total

    @staticmethod
    def _progress_in_range(start: float, end: float, completed: int, total: int) -> float:
        if total <= 0:
            return end
        fraction = max(0.0, min(1.0, completed / total))
        return start + ((end - start) * fraction)

    @staticmethod
    def _progress_stride(total: int, target_updates: int = 180) -> int:
        if total <= 0:
            return 1
        return max(1, total // target_updates)

    def _collect_files(
        self,
        folder: Path,
        include_subfolders: bool,
        base_label: str,
        progress_callback: Optional[Callable[[int], None]] = None,
    ) -> List[FileRecord]:
        if not folder.exists() or not folder.is_dir():
            raise FileNotFoundError(f"Folder tidak valid: {folder}")

        records: List[FileRecord] = []
        iterator: Iterable[Path] = folder.rglob("*") if include_subfolders else folder.glob("*")
        for path in iterator:
            if path.is_file():
                try:
                    size = path.stat().st_size
                except OSError:
                    continue
                records.append(
                    FileRecord(
                        path=path,
                        base_folder=folder,
                        base_label=base_label,
                        relative_path=str(path.relative_to(folder)),
                        size=size,
                    )
                )
                if progress_callback is not None:
                    progress_callback(len(records))
        return records

    def _hash_file(self, path: Path) -> str:
        sha = hashlib.sha256()
        with path.open("rb") as file_handle:
            while True:
                chunk = file_handle.read(CHUNK_SIZE)
                if not chunk:
                    break
                sha.update(chunk)
        return sha.hexdigest()

    def _format_size(self, size: int) -> str:
        value = float(size)
        units = ["B", "KB", "MB", "GB", "TB"]
        for unit in units:
            if value < 1024 or unit == units[-1]:
                return f"{value:.2f} {unit}" if unit != "B" else f"{int(value)} B"
            value /= 1024
        return f"{size} B"

    @staticmethod
    def _format_progress_value(value: float) -> str:
        rounded = round(value, 1)
        if abs(rounded - round(rounded)) < 0.05:
            return f"{int(round(rounded))}%"
        return f"{rounded:.1f}%"

    def _set_progress(self, value: float, text: str) -> None:
        safe_value = max(0.0, min(100.0, value))
        target_value = int(round(safe_value * 10))
        current_value = self.progress_bar.value()

        self.progress_badge.setText(self._format_progress_value(safe_value))
        self.status_label.setText(text)

        if target_value == current_value:
            return

        if target_value <= current_value or target_value == 0:
            self.progress_animation.stop()
            self.progress_bar.setValue(target_value)
            return

        duration = max(140, min(520, 120 + abs(target_value - current_value)))
        self.progress_animation.stop()
        self.progress_animation.setDuration(duration)
        self.progress_animation.setStartValue(current_value)
        self.progress_animation.setEndValue(target_value)
        self.progress_animation.start()

    def _queue_progress(self, value: float, text: str) -> None:
        safe_value = max(0.0, min(100.0, value))
        rounded_value = round(safe_value, 1)

        if rounded_value not in {0.0, 100.0} and rounded_value < self._last_queued_progress:
            rounded_value = self._last_queued_progress
        if rounded_value == self._last_queued_progress and text == self._last_queued_progress_text:
            return

        self._last_queued_progress = rounded_value
        self._last_queued_progress_text = text
        with self._progress_lock:
            self._pending_progress = (rounded_value, text)

    def _poll_queue(self) -> None:
        pending_progress: Optional[Tuple[float, str]] = None
        with self._progress_lock:
            if self._pending_progress is not None:
                pending_progress = self._pending_progress
                self._pending_progress = None

        if pending_progress is not None:
            value, text = pending_progress
            self._set_progress(float(value), str(text))

        processed_messages = 0
        try:
            while processed_messages < 12:
                message = self.ui_queue.get_nowait()
                self._handle_queue_message(message)
                processed_messages += 1
        except queue.Empty:
            return

    def _handle_queue_message(self, message: Tuple[str, object]) -> None:
        kind, payload = message
        if kind == "progress":
            value, text = payload  # type: ignore[misc]
            self._set_progress(float(value), str(text))
            return

        if kind == "scan_done":
            self.scan_button.setEnabled(True)
            with self._progress_lock:
                self._pending_progress = None
            self._pending_scan_results = payload  # type: ignore[assignment]
            self._awaiting_scan_finalize = True
            self._set_progress(100, "Menyiapkan tampilan hasil scan...")
            self._maybe_finalize_scan_after_progress()
            return

        if kind == "scan_error":
            self.scan_button.setEnabled(True)
            self._pending_scan_results = None
            self._awaiting_scan_finalize = False
            with self._progress_lock:
                self._pending_progress = None
            self.progress_animation.stop()
            self._set_progress(0, "Terjadi kesalahan saat scan.")
            self.progress_card.setVisible(False)
            if isinstance(payload, dict):
                self.show_error_dialog(
                    str(payload.get("title", APP_TITLE)),
                    str(payload.get("summary", "Terjadi kesalahan saat scan.")),
                    str(payload.get("details", payload.get("summary", ""))),
                )
            else:
                self.show_error_dialog(
                    "Terjadi kesalahan saat scan dan bandingkan",
                    str(payload),
                    str(payload),
                )
            return

        if kind == "delete_done":
            self._pending_delete_result = payload if isinstance(payload, dict) else {}
            QTimer.singleShot(0, self._finalize_delete_results)
            return

    def _on_progress_animation_finished(self) -> None:
        self._maybe_finalize_scan_after_progress()

    def _maybe_finalize_scan_after_progress(self) -> None:
        if not self._awaiting_scan_finalize or self._pending_scan_results is None:
            return
        if self.progress_animation.state() == QAbstractAnimation.Running:
            return
        if self.progress_bar.value() < self.progress_bar.maximum():
            return

        self._awaiting_scan_finalize = False
        QTimer.singleShot(0, self._finalize_scan_results)

    def _finalize_scan_results(self) -> None:
        if self._pending_scan_results is None:
            return

        self.result_rows = self._pending_scan_results
        self._pending_scan_results = None

        self._apply_default_table_widths()
        self._populate_table(recompute_widths=False)
        self._refresh_stats()
        self.progress_animation.stop()
        self.progress_bar.setValue(self.progress_bar.maximum())
        self.progress_badge.setText("100%")
        self.status_label.setText("Scan selesai. Data siap dianalisis.")
        self.export_csv_button.setEnabled(True)
        if self.openpyxl_available:
            self.export_excel_button.setEnabled(True)

        duplicate_count = sum(1 for row in self.result_rows if row.exact_matches)
        if duplicate_count > 0:
            self.delete_button.setEnabled(True)
            self.delete_all_button.setEnabled(True)
        else:
            QMessageBox.information(
                self,
                APP_TITLE,
                "Tidak ada file duplikat hijau untuk dihapus. Silakan cek hasil merah/oranye.",
            )

        QTimer.singleShot(40, self._finalize_table_layout_after_scan)
        self.progress_card.setVisible(False)

    def _finalize_table_layout_after_scan(self) -> None:
        if self.result_rows and self.table_proxy.rowCount() > 0:
            self._apply_responsive_column_widths()

    def _finalize_delete_results(self) -> None:
        payload_map = self._pending_delete_result or {}
        self._pending_delete_result = None

        deleted_paths = [Path(path) for path in payload_map.get("deleted_paths", [])]
        deleted_count = int(payload_map.get("deleted_count", len(deleted_paths)))
        errors = [str(item) for item in payload_map.get("errors", [])]
        active_delete_dialog = self.delete_confirm_dialog

        if active_delete_dialog is not None and not errors:
            active_delete_dialog.set_processing(
                True,
                f"Penghapusan selesai. Menyegarkan hasil visual untuk {deleted_count} file...",
            )
            active_delete_dialog.flush_visual_state()

        path_set = set(deleted_paths)
        self.result_rows = [row for row in self.result_rows if row.target_path not in path_set]
        self._apply_default_table_widths()
        self._populate_table(recompute_widths=False)
        self._refresh_stats()
        self.current_selected_result = None
        self._reset_detail_panel()

        remaining = self.table_proxy.rowCount()
        self.status_label.setText(f"Selesai menghapus {deleted_count} file. Sisa hasil terlihat: {remaining}.")
        self._set_delete_processing_state(False)

        if self.result_rows and self.table_proxy.rowCount() > 0:
            QTimer.singleShot(40, self._finalize_table_layout_after_delete)

        if errors:
            if active_delete_dialog is not None:
                active_delete_dialog.force_close(QDialog.Rejected)
                self.delete_confirm_dialog = None
            error_preview = os.linesep.join(errors[:10])
            QTimer.singleShot(
                180,
                lambda preview=error_preview, total_errors=len(errors): self.show_error_dialog(
                    "Sebagian file gagal diproses",
                    f"{total_errors} file gagal diproses saat penghapusan.",
                    preview,
                ),
            )
        else:
            success_summary = f"Berhasil memproses {deleted_count} file dari Folder A."
            success_details = f"Jumlah file diproses: {deleted_count}\nSisa hasil terlihat: {remaining}"
            if active_delete_dialog is not None:
                active_delete_dialog.show_success_state(success_summary, success_details)
                active_delete_dialog.flush_visual_state()
            else:
                self.show_success_dialog(
                    "Penghapusan Berhasil",
                    success_summary,
                    success_details,
                )

    def _finalize_table_layout_after_delete(self) -> None:
        if self.result_rows and self.table_proxy.rowCount() > 0:
            self._apply_responsive_column_widths()

    def show_error_dialog(self, title: str, summary: str, details: str) -> None:
        self.status_label.setText(summary or "Terjadi kesalahan pada aplikasi.")
        try:
            dialog = ErrorOverlayDialog(self, title, summary, details)
            dialog.exec()
        except Exception:
            QMessageBox.critical(self, title or APP_TITLE, summary or "Terjadi kesalahan pada aplikasi.")

    def show_success_dialog(self, title: str, summary: str, details: str = "") -> None:
        self.status_label.setText(summary or "Aksi berhasil diproses.")
        try:
            dialog = SuccessOverlayDialog(self, title, summary, details)
            dialog.exec()
        except Exception:
            QMessageBox.information(self, title or APP_TITLE, summary or "Aksi berhasil diproses.")

    def _refresh_stats(self) -> None:
        total = len(self.result_rows)
        exact = sum(1 for row in self.result_rows if row.exact_matches)
        diff = sum(1 for row in self.result_rows if row.same_name_different_content)
        only = sum(1 for row in self.result_rows if row.only_in_target)

        self.stat_labels["total"].setText(str(total))
        self.stat_labels["exact"].setText(str(exact))
        self.stat_labels["diff"].setText(str(diff))
        self.stat_labels["only"].setText(str(only))

    def _apply_default_table_widths(self) -> None:
        default_widths = [170, 240, 180, 110, 120, 260, 260, 140, 110]
        for column, width in enumerate(default_widths):
            self.results_table.setColumnWidth(column, width)

    def _apply_responsive_column_widths(self) -> None:
        min_widths = [140, 200, 170, 110, 120, 260, 260, 140, 110]
        max_widths = [240, 420, 320, 140, 180, 520, 520, 220, 150]
        flexible_columns = [1, 2, 5, 6]

        self.results_table.resizeColumnsToContents()

        computed_widths: List[int] = []
        for column in range(self.table_model.columnCount()):
            content_width = self.results_table.columnWidth(column) + 18
            target_width = max(min_widths[column], min(content_width, max_widths[column]))
            computed_widths.append(target_width)

        viewport_width = self.results_table.viewport().width()
        used_width = sum(computed_widths)

        if viewport_width > used_width and flexible_columns:
            extra = viewport_width - used_width
            per_column = max(extra // len(flexible_columns), 0)
            for column in flexible_columns:
                computed_widths[column] = min(computed_widths[column] + per_column, max_widths[column])

        for column, width in enumerate(computed_widths):
            self.results_table.setColumnWidth(column, width)

    def _update_table_empty_state(self) -> None:
        if not self.result_rows:
            self.empty_state_title.setText("Belum ada hasil scan")
            self.empty_state_description.setText(
                "Pilih folder target dan folder pembanding, lalu jalankan Scan dan Bandingkan untuk melihat hasil."
            )
            self.table_stack.setCurrentWidget(self.table_empty_state)
            return

        if self.table_proxy.rowCount() == 0:
            self.empty_state_title.setText("Tidak ada hasil yang cocok")
            self.empty_state_description.setText(
                "Coba ubah kata kunci pencarian, quick filter status, atau opsi filter tampilan untuk menampilkan data."
            )
            self.table_stack.setCurrentWidget(self.table_empty_state)
            return

        self.table_stack.setCurrentWidget(self.results_table)

    def _populate_table(self, recompute_widths: bool = True, refresh_source: bool = True) -> None:
        previously_selected_target = self.current_selected_result.target_path if self.current_selected_result else None

        self.results_table.setUpdatesEnabled(False)
        if refresh_source:
            self.table_model.set_rows(self.result_rows)
        self.table_proxy.set_matches_only(self.show_only_matches_checkbox.isChecked())
        self.table_proxy.set_status_filter(self._current_quick_filter())
        self.table_proxy.set_search_text(self.search_input.text() if hasattr(self, "search_input") else "")

        if self.results_table.selectionModel() is not None:
            self.results_table.clearSelection()
            self.results_table.selectionModel().setCurrentIndex(QModelIndex(), QItemSelectionModel.SelectionFlag.NoUpdate)

        if recompute_widths and self.table_proxy.rowCount() > 0:
            self._apply_responsive_column_widths()
        elif recompute_widths:
            self._apply_default_table_widths()

        self._update_table_empty_state()
        self.results_table.setUpdatesEnabled(True)

        selected_row_index = self._find_table_row_by_target_path(previously_selected_target) if previously_selected_target else None
        if selected_row_index is not None:
            proxy_index = self.table_proxy.index(selected_row_index, 0)
            self.results_table.selectRow(selected_row_index)
            self.results_table.setCurrentIndex(proxy_index)
            self.results_table.scrollTo(proxy_index)
            self._update_detail_from_row(selected_row_index)
        else:
            self.current_selected_result = None
            self._reset_detail_panel()

    def _schedule_search_filter(self) -> None:
        self.filter_timer.start()

    def _apply_debounced_filter(self) -> None:
        self._populate_table(recompute_widths=False, refresh_source=False)
        if self.result_rows:
            self.status_label.setText("Hasil pencarian diperbarui.")

    def _on_filter_changed(self, *args) -> None:
        self.filter_timer.stop()
        self._populate_table(recompute_widths=False, refresh_source=False)
        if self.result_rows:
            self.status_label.setText("Filter tampilan diperbarui.")

    def _current_quick_filter(self) -> str:
        for key, button in getattr(self, "quick_filter_buttons", {}).items():
            if button.isChecked():
                return key
        return "all"

    def _reset_quick_filters(self) -> None:
        self.filter_timer.stop()
        if hasattr(self, "search_input"):
            self.search_input.blockSignals(True)
            self.search_input.clear()
            self.search_input.blockSignals(False)
        if hasattr(self, "quick_filter_buttons"):
            self.quick_filter_buttons["all"].setChecked(True)
        self._on_filter_changed()

    def _on_table_selection_changed(self, *args) -> None:
        selected_rows = self.results_table.selectionModel().selectedRows()
        if not selected_rows:
            self.current_selected_result = None
            self._reset_detail_panel()
            return
        self._update_detail_from_row(selected_rows[0].row())

    def _on_current_row_changed(self, current: QModelIndex, previous: QModelIndex) -> None:
        if current.isValid():
            self._update_detail_from_row(current.row())

    def _on_table_focus_released(self) -> None:
        self.current_selected_result = None
        self._reset_detail_panel()

    def _result_for_table_row(self, row_index: int) -> Optional[MatchResult]:
        if row_index < 0:
            return None
        proxy_index = self.table_proxy.index(row_index, 0)
        if not proxy_index.isValid():
            return None
        result = proxy_index.data(Qt.UserRole)
        return result if isinstance(result, MatchResult) else None

    def _find_table_row_by_target_path(self, target_path: Path) -> Optional[int]:
        for row_index in range(self.table_proxy.rowCount()):
            result = self._result_for_table_row(row_index)
            if result is not None and result.target_path == target_path:
                return row_index
        return None

    def _update_detail_from_row(self, row_index: int) -> None:
        result = self._result_for_table_row(row_index)
        if result is None:
            self.current_selected_result = None
            self._reset_detail_panel()
            return

        self.current_selected_result = result
        self.detail_panel.setVisible(True)
        self.detail_labels["status"].setText(result.status_text)
        self.detail_labels["target"].setText(str(result.target_path))
        self.detail_labels["relative"].setText(result.target_relative_path)
        self.detail_labels["size"].setText(f"{self._format_size(result.size)} ({result.size} bytes)")
        self.detail_labels["found"].setText(result.exact_folder_labels)
        self.detail_labels["missing"].setText(", ".join(result.missing_from_folders) if result.missing_from_folders else "-")
        self.detail_labels["mode"].setText(result.match_type)
        self.detail_labels["exact"].setText(result.exact_paths_text)
        self.detail_labels["diff"].setText(result.diff_paths_text)

    def _reset_detail_panel(self) -> None:
        defaults = {
            "status": "Pilih satu item untuk melihat detail.",
            "target": "-",
            "relative": "-",
            "size": "-",
            "found": "-",
            "missing": "-",
            "mode": "-",
            "exact": "-",
            "diff": "-",
        }
        for key, value in defaults.items():
            self.detail_labels[key].setText(value)
        if hasattr(self, "detail_panel"):
            self.detail_panel.setVisible(False)

    def copy_selected_path(self) -> None:
        if self.current_selected_result is None:
            QMessageBox.information(self, APP_TITLE, "Pilih satu item terlebih dahulu.")
            return

        QApplication.clipboard().setText(str(self.current_selected_result.target_path))
        self.status_label.setText("Path file target berhasil disalin ke clipboard.")

    def clear_results(self, reset_status: bool = True) -> None:
        if self.delete_confirm_dialog is not None:
            self.delete_confirm_dialog.force_close(QDialog.Rejected)
            self.delete_confirm_dialog = None
        if self.delete_processing_dialog is not None:
            self.delete_processing_dialog.close()
            self.delete_processing_dialog = None
        self.result_rows = []
        self._pending_scan_results = None
        self._awaiting_scan_finalize = False
        self.current_selected_result = None
        self.table_model.set_rows([])
        if self.results_table.selectionModel() is not None:
            self.results_table.clearSelection()
            self.results_table.selectionModel().setCurrentIndex(QModelIndex(), QItemSelectionModel.SelectionFlag.NoUpdate)

        self.export_csv_button.setEnabled(False)
        self.export_excel_button.setEnabled(False)
        self.delete_button.setEnabled(False)
        self.delete_all_button.setEnabled(False)

        for label in self.stat_labels.values():
            label.setText("0")

        self.progress_animation.stop()
        self.progress_bar.setValue(0)
        self.progress_badge.setText("0%")
        self.progress_card.setVisible(False)
        self._last_queued_progress = -1.0
        self._last_queued_progress_text = ""
        with self._progress_lock:
            self._pending_progress = None
        self._update_table_empty_state()
        self._reset_detail_panel()

        if reset_status:
            self.status_label.setText("Hasil dibersihkan. Silakan pilih folder baru.")

    def delete_selected(self) -> None:
        selected_rows = sorted({index.row() for index in self.results_table.selectionModel().selectedRows()})
        if not selected_rows:
            QMessageBox.information(self, APP_TITLE, "Pilih minimal satu file dari hasil scan.")
            return

        deletable: List[Path] = []
        skipped = 0
        for row_index in selected_rows:
            result = self._result_for_table_row(row_index)
            if result is None:
                continue
            if result.exact_matches:
                deletable.append(result.target_path)
            else:
                skipped += 1

        if not deletable:
            QMessageBox.information(self, APP_TITLE, "Pilihan Anda tidak mengandung file hijau yang aman dihapus.")
            return
        if skipped:
            QMessageBox.information(
                self,
                APP_TITLE,
                f"{skipped} item merah/oranye diabaikan karena tidak aman dihapus otomatis.",
            )

        self._confirm_and_delete(deletable)

    def delete_all_results(self) -> None:
        paths = [row.target_path for row in self.result_rows if row.exact_matches]
        if not paths:
            QMessageBox.information(self, APP_TITLE, "Belum ada hasil hijau untuk dihapus.")
            return
        self._confirm_and_delete(paths)

    def _confirm_and_delete(self, paths: List[Path]) -> None:
        if self.delete_thread and self.delete_thread.is_alive():
            QMessageBox.information(self, APP_TITLE, "Proses penghapusan masih berjalan.")
            return

        preview = "\n".join(str(path) for path in paths[:10])
        extra = "" if len(paths) <= 10 else f"\n... dan {len(paths) - 10} file lainnya"
        mode_text = "Recycle Bin" if self._current_delete_mode() == "recycle_bin" else "hapus permanen"
        summary = (
            f"Anda akan memproses {len(paths)} file hijau dari Folder A.\n"
            f"Mode penghapusan: {mode_text}."
        )
        details = f"{preview}{extra}"

        confirm_dialog = ConfirmOverlayDialog(
            self,
            "Konfirmasi Penghapusan File",
            summary,
            details,
        )
        self.delete_confirm_dialog = confirm_dialog
        confirm_dialog.confirmRequested.connect(
            lambda delete_paths=list(paths): self._start_delete_from_confirm_dialog(delete_paths)
        )

        result = confirm_dialog.exec()
        if self.delete_confirm_dialog is confirm_dialog:
            self.delete_confirm_dialog = None
        confirm_dialog.deleteLater()

        if result != QDialog.Accepted:
            return

    def _start_delete_from_confirm_dialog(self, paths: List[Path]) -> None:
        if self.delete_thread and self.delete_thread.is_alive():
            return

        dialog = self.delete_confirm_dialog
        if dialog is None:
            return

        self.status_label.setText("Memproses penghapusan file terpilih...")
        dialog.set_processing(True, f"Sedang memproses {len(paths)} file dari Folder A. Mohon tunggu...")
        self._set_delete_processing_state(True)
        dialog.flush_visual_state()

        use_recycle_bin = self._current_delete_mode() == "recycle_bin"
        QTimer.singleShot(
            0,
            lambda delete_paths=list(paths), recycle_mode=use_recycle_bin: self._launch_delete_worker(
                delete_paths, recycle_mode
            ),
        )

    def _launch_delete_worker(self, paths: List[Path], use_recycle_bin: bool) -> None:
        if self.delete_thread and self.delete_thread.is_alive():
            return

        self.delete_thread = threading.Thread(
            target=self._delete_worker,
            args=(list(paths), use_recycle_bin),
            daemon=True,
        )
        self.delete_thread.start()

    def _delete_worker(self, paths: List[Path], use_recycle_bin: bool) -> None:
        deleted_count = 0
        deleted_paths: List[Path] = []
        errors: List[str] = []

        for path in paths:
            try:
                if use_recycle_bin:
                    self._send_to_recycle_bin(path)
                else:
                    path.unlink()
                deleted_count += 1
                deleted_paths.append(path)
            except Exception as exc:
                errors.append(f"{path}: {exc}")

        self.ui_queue.put(
            (
                "delete_done",
                {
                    "deleted_count": deleted_count,
                    "deleted_paths": [str(path) for path in deleted_paths],
                    "errors": errors,
                },
            )
        )

    def export_csv(self) -> None:
        if not self.result_rows:
            QMessageBox.information(self, APP_TITLE, "Belum ada hasil scan untuk disimpan.")
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Simpan hasil ke CSV",
            str(Path.home() / "hasil_scan.csv"),
            "CSV files (*.csv)",
        )
        if not save_path:
            return

        with open(save_path, "w", newline="", encoding="utf-8-sig") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(
                [
                    "status",
                    "target_path",
                    "target_relative_path",
                    "size_bytes",
                    "size_display",
                    "found_in_folders",
                    "exact_match_paths",
                    "different_content_paths",
                    "missing_in_folders",
                    "match_type",
                ]
            )
            for row in self.result_rows:
                writer.writerow(
                    [
                        row.status_text,
                        str(row.target_path),
                        row.target_relative_path,
                        row.size,
                        self._format_size(row.size),
                        row.exact_folder_labels,
                        row.exact_paths_text,
                        row.diff_paths_text,
                        ", ".join(row.missing_from_folders) if row.missing_from_folders else "-",
                        row.match_type,
                    ]
                )

        QMessageBox.information(self, APP_TITLE, f"Berhasil menyimpan CSV:\n{save_path}")

    def export_excel(self) -> None:
        if not self.result_rows:
            QMessageBox.information(self, APP_TITLE, "Belum ada hasil scan untuk disimpan.")
            return
        if not self.openpyxl_available:
            QMessageBox.warning(self, APP_TITLE, "Modul openpyxl belum terpasang. Jalankan: pip install openpyxl")
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Simpan hasil ke Excel",
            str(Path.home() / "hasil_scan.xlsx"),
            "Excel files (*.xlsx)",
        )
        if not save_path:
            return

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Hasil Scan"
        headers = [
            "Status",
            "Path File di Folder A",
            "Relative Path",
            "Ukuran (bytes)",
            "Ukuran",
            "Ditemukan di Folder",
            "Path File yang Cocok",
            "Path Nama Sama Isi Beda",
            "Tidak Ada di Folder",
            "Pencocokan",
        ]
        worksheet.append(headers)

        fill_green = PatternFill("solid", fgColor="E8FFF4")
        fill_red = PatternFill("solid", fgColor="FFF0F2")
        fill_orange = PatternFill("solid", fgColor="FFF6E8")

        for row in self.result_rows:
            worksheet.append(
                [
                    row.status_text,
                    str(row.target_path),
                    row.target_relative_path,
                    row.size,
                    self._format_size(row.size),
                    row.exact_folder_labels,
                    row.exact_paths_text,
                    row.diff_paths_text,
                    ", ".join(row.missing_from_folders) if row.missing_from_folders else "-",
                    row.match_type,
                ]
            )
            excel_row = worksheet.max_row
            fill = fill_green if row.tree_tag == "exact_match" else fill_red if row.tree_tag == "different_content" else fill_orange
            for column in range(1, len(headers) + 1):
                worksheet.cell(row=excel_row, column=column).fill = fill

        for column_letter, width in {"A": 22, "B": 50, "C": 30, "D": 16, "E": 14, "F": 18, "G": 70, "H": 70, "I": 22, "J": 14}.items():
            worksheet.column_dimensions[column_letter].width = width

        workbook.save(save_path)
        QMessageBox.information(self, APP_TITLE, f"Berhasil menyimpan Excel:\n{save_path}")

    def _send_to_recycle_bin(self, path: Path) -> None:
        try:
            from send2trash import send2trash  # type: ignore
        except ImportError as exc:
            raise RuntimeError("Paket 'send2trash' belum terpasang. Jalankan: pip install send2trash") from exc
        send2trash(str(path))


def _forward_exception_to_app(title: str, summary: str, details: str) -> None:
    app = QApplication.instance()
    if isinstance(app, SafeApplication):
        app.report_error(title, summary, details)
    else:
        QMessageBox.critical(None, title or APP_TITLE, summary or "Terjadi kesalahan pada aplikasi.")


def _handle_uncaught_exception(exc_type, exc_value, exc_traceback) -> None:
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return

    _forward_exception_to_app(
        "Terjadi kesalahan tak terduga",
        str(exc_value) or exc_type.__name__,
        "".join(traceback.format_exception(exc_type, exc_value, exc_traceback)),
    )


def _handle_thread_exception(args: threading.ExceptHookArgs) -> None:
    _forward_exception_to_app(
        "Terjadi kesalahan pada proses latar belakang",
        str(args.exc_value) or args.exc_type.__name__,
        "".join(traceback.format_exception(args.exc_type, args.exc_value, args.exc_traceback)),
    )


def main() -> int:
    sys.excepthook = _handle_uncaught_exception
    threading.excepthook = _handle_thread_exception

    app = SafeApplication(sys.argv)
    app.setStyle("Fusion")

    window = FolderCompareDeleteApp()
    window.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
