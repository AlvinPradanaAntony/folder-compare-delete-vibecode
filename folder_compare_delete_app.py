from __future__ import annotations

import csv
import hashlib
import os
import queue
import shutil
import subprocess
import sys
import tempfile
import threading
import traceback
import json
from components.pyqtspinner import WaitingSpinner
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple
from uuid import uuid4

try:
    from PySide6.QtCore import QAbstractAnimation, QAbstractTableModel, QEasingCurve, QItemSelectionModel, QModelIndex, QPoint, QPropertyAnimation, QParallelAnimationGroup, QRect, QSize, QSortFilterProxyModel, QTimer, Qt, Signal
    from PySide6.QtGui import QAction, QBrush, QColor, QFont, QDragEnterEvent, QDropEvent, QFocusEvent, QIcon, QWheelEvent
    from PySide6.QtWidgets import (
        QAbstractItemView,
        QApplication,
        QButtonGroup,
        QDialog,
        QFileDialog,
        QFrame,
        QGraphicsBlurEffect,
        QGridLayout,
        QHBoxLayout,
        QHeaderView,
        QLabel,
        QLineEdit,
        QMainWindow,
        QMenu,
        QMessageBox,
        QPlainTextEdit,
        QProgressBar,
        QPushButton,
        QRadioButton,
        QCheckBox,
        QScrollArea,
        QSplitter,
        QStackedLayout,
        QStackedWidget,
        QTableView,
        QTableWidget,
        QTableWidgetItem,
        QVBoxLayout,
        QWidget,
    )
except ImportError as exc:
    raise SystemExit("PySide6 belum terpasang. Jalankan: pip install PySide6") from exc


APP_TITLE = "Folder Compare & Delete"
APP_VERSION = "2.4.3"
APP_DEVELOPER = "Tonzdev"
CHUNK_SIZE = 1024 * 1024  # 1 MB
BG_COLOR = "#f4f7fb"
SURFACE = "#ffffff"
SURFACE_ALT = "#f7f9fc"
PRIMARY = "#1f5eff"
PRIMARY_DARK = "#1948c7"
TEXT = "#152033"
MUTED = "#62708a"
GREEN = "#27c281"
RED = "#eb5d70"
ORANGE = "#f4a23a"
CYAN = "#20b7d6"
BORDER = "#dbe3f3"
SUCCESS_ROW = "#eafaf1"
DANGER_ROW = "#fff0f3"
WARNING_ROW = "#fff7eb"


@dataclass
class FileRecord:
    path: Path
    base_folder: Path
    base_label: str
    relative_path: str
    size: int
    sha256: Optional[str] = None


@dataclass
class MatchResult:
    target_path: Path
    target_relative_path: str
    size: int
    match_type: str
    exact_matches: List[FileRecord] = field(default_factory=list)
    same_name_different_content: List[FileRecord] = field(default_factory=list)
    missing_from_folders: List[str] = field(default_factory=list)
    only_in_target: bool = False
    temp_synced_labels: List[str] = field(default_factory=list)
    temp_synced_paths: List[str] = field(default_factory=list)

    @property
    def exact_folder_labels(self) -> str:
        labels = sorted({item.base_label for item in self.exact_matches})
        if self.temp_synced_labels:
            labels.extend(f"{lbl} (Disalin)" for lbl in self.temp_synced_labels)
        return ", ".join(labels) if labels else "-"

    @property
    def exact_paths_text(self) -> str:
        parts = [f"[{item.base_label}] {item.path}" for item in self.exact_matches]
        if self.temp_synced_paths:
            parts.extend(f"(Baru) {p}" for p in self.temp_synced_paths)
        return " | ".join(parts) if parts else "-"

    @property
    def diff_paths_text(self) -> str:
        if not self.same_name_different_content:
            return "-"
        return " | ".join(f"[{item.base_label}] {item.path}" for item in self.same_name_different_content)

    @property
    def missing_display_text(self) -> str:
        if not self.missing_from_folders:
            return "-"
        display = []
        for lbl in self.missing_from_folders:
            if lbl in self.temp_synced_labels:
                display.append(f"{lbl} (Disalin)")
            else:
                display.append(lbl)
        return ", ".join(display)

    @property
    def status_text(self) -> str:
        if self.exact_matches:
            return "Duplikat ditemukan"
        if self.same_name_different_content:
            return "Nama sama, isi berbeda"
        return "Hanya ada di Folder A"

    @property
    def tree_tag(self) -> str:
        if self.exact_matches:
            return "exact_match"
        if self.same_name_different_content:
            return "different_content"
        return "only_target"


@dataclass
class HistoryEntry:
    timestamp: str
    action: str
    file_name: str
    status: str
    detail: str
    tone: str = "info"


@dataclass
class UndoAction:
    label: str
    detail: str
    operations: List[Dict[str, str]] = field(default_factory=list)
    action_dir: str = ""
    file_name_hint: str = "-"


@dataclass
class TrashEntry:
    entry_id: str
    original_path: str
    trash_path: str
    deleted_at: str
    size: int


class ResponsiveTableWidget(QTableView):
    focusReleased = Signal()

    def wheelEvent(self, event: QWheelEvent) -> None:
        if event.modifiers() & Qt.ShiftModifier:
            delta = event.angleDelta().y() or event.angleDelta().x()
            if delta:
                scrollbar = self.horizontalScrollBar()
                step = scrollbar.singleStep() or 24
                direction = -1 if delta > 0 else 1
                scrollbar.setValue(scrollbar.value() + (direction * step * 3))
                event.accept()
                return
        super().wheelEvent(event)

    def focusOutEvent(self, event: QFocusEvent) -> None:
        if self.selectionModel() is not None:
            self.selectionModel().setCurrentIndex(QModelIndex(), QItemSelectionModel.SelectionFlag.NoUpdate)
        self.focusReleased.emit()
        super().focusOutEvent(event)


class MatchResultTableModel(QAbstractTableModel):
    HEADERS = [
        "Status",
        "Path File di Folder A",
        "Relative Path",
        "Ukuran",
        "Ditemukan di",
        "Path Cocok",
        "Path Beda",
        "Tidak Ada di",
        "Mode",
    ]

    def __init__(self) -> None:
        super().__init__()
        self._rows: List[MatchResult] = []

    def set_rows(self, rows: List[MatchResult]) -> None:
        self.beginResetModel()
        self._rows = list(rows)
        self.endResetModel()

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:
        if parent.isValid():
            return 0
        return len(self._rows)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:
        if parent.isValid():
            return 0
        return len(self.HEADERS)

    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.DisplayRole) -> Any:
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal and 0 <= section < len(self.HEADERS):
            return self.HEADERS[section]
        return section + 1 if orientation == Qt.Vertical else None

    def data(self, index: QModelIndex, role: int = Qt.DisplayRole) -> Any:
        if not index.isValid() or not (0 <= index.row() < len(self._rows)):
            return None

        row = self._rows[index.row()]
        display_values = [
            row.status_text,
            str(row.target_path),
            row.target_relative_path,
            self._format_size(row.size),
            row.exact_folder_labels,
            row.exact_paths_text,
            row.diff_paths_text,
            row.missing_display_text,
            row.match_type,
        ]
        sort_values = [
            row.status_text.lower(),
            str(row.target_path).lower(),
            row.target_relative_path.lower(),
            row.size,
            row.exact_folder_labels.lower(),
            row.exact_paths_text.lower(),
            row.diff_paths_text.lower(),
            row.missing_display_text.lower(),
            row.match_type.lower(),
        ]

        if role == Qt.DisplayRole:
            return display_values[index.column()]
        if role == Qt.ToolTipRole:
            return display_values[index.column()]
        if role == Qt.TextAlignmentRole:
            return Qt.AlignVCenter | (Qt.AlignCenter if index.column() in {0, 3, 4, 7, 8} else Qt.AlignLeft)
        if role == Qt.BackgroundRole:
            return QBrush(self._result_background(row.tree_tag))
        if role == Qt.ForegroundRole:
            if row.temp_synced_labels:
                if index.column() in {4, 5}:
                    return QBrush(QColor(PRIMARY))
                elif index.column() == 7:
                    return QBrush(QColor(RED))
            return QBrush(self._result_foreground(row.tree_tag))
        if role == Qt.FontRole:
            if row.temp_synced_labels and index.column() == 7:
                font = QFont()
                font.setBold(True)
                return font
        if role == Qt.UserRole:
            return row
        if role == Qt.UserRole + 1:
            return sort_values[index.column()]
        return None

    def result_at(self, row_index: int) -> Optional[MatchResult]:
        if 0 <= row_index < len(self._rows):
            return self._rows[row_index]
        return None

    @staticmethod
    def _format_size(size: int) -> str:
        value = float(size)
        units = ["B", "KB", "MB", "GB", "TB"]
        for unit in units:
            if value < 1024 or unit == units[-1]:
                return f"{value:.2f} {unit}" if unit != "B" else f"{int(value)} B"
            value /= 1024
        return f"{size} B"

    @staticmethod
    def _result_background(tree_tag: str) -> QColor:
        if tree_tag == "exact_match":
            return QColor(SUCCESS_ROW)
        if tree_tag == "different_content":
            return QColor(DANGER_ROW)
        return QColor(WARNING_ROW)

    @staticmethod
    def _result_foreground(tree_tag: str) -> QColor:
        if tree_tag == "exact_match":
            return QColor("#1f6f4a")
        if tree_tag == "different_content":
            return QColor("#9f3143")
        return QColor("#8a5a19")


class MatchResultFilterProxyModel(QSortFilterProxyModel):
    def __init__(self) -> None:
        super().__init__()
        self.search_text = ""
        self.status_filter = "all"
        self.matches_only = False
        self.setSortRole(Qt.UserRole + 1)
        self.setDynamicSortFilter(True)

    def set_search_text(self, text: str) -> None:
        self.beginFilterChange()
        self.search_text = text.strip().lower()
        self.endFilterChange(QSortFilterProxyModel.Direction.Rows)

    def set_status_filter(self, status_filter: str) -> None:
        self.beginFilterChange()
        self.status_filter = status_filter
        self.endFilterChange(QSortFilterProxyModel.Direction.Rows)

    def set_matches_only(self, enabled: bool) -> None:
        self.beginFilterChange()
        self.matches_only = enabled
        self.endFilterChange(QSortFilterProxyModel.Direction.Rows)

    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:
        source_model = self.sourceModel()
        if source_model is None:
            return True

        source_index = source_model.index(source_row, 0, source_parent)
        result = source_index.data(Qt.UserRole)
        if not isinstance(result, MatchResult):
            return True

        if self.matches_only and not (result.exact_matches or result.same_name_different_content):
            return False
        if self.status_filter != "all" and result.tree_tag != self.status_filter:
            return False
        if not self.search_text:
            return True

        haystacks = [
            result.status_text,
            str(result.target_path),
            result.target_relative_path,
            MatchResultTableModel._format_size(result.size),
            result.exact_folder_labels,
            result.exact_paths_text,
            result.diff_paths_text,
            ", ".join(result.missing_from_folders) if result.missing_from_folders else "-",
            result.match_type,
        ]
        searchable_text = " ".join(haystacks).lower()
        return self.search_text in searchable_text


class HistoryTableModel(QAbstractTableModel):
    HEADERS = ["Waktu", "Aksi", "Nama File", "Status", "Detail"]

    def __init__(self) -> None:
        super().__init__()
        self._rows: List[HistoryEntry] = []

    def set_rows(self, rows: List[HistoryEntry]) -> None:
        self.beginResetModel()
        self._rows = list(rows)
        self.endResetModel()

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:
        if parent.isValid():
            return 0
        return len(self._rows)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:
        if parent.isValid():
            return 0
        return len(self.HEADERS)

    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.DisplayRole) -> Any:
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal and 0 <= section < len(self.HEADERS):
            return self.HEADERS[section]
        return section + 1 if orientation == Qt.Vertical else None

    def data(self, index: QModelIndex, role: int = Qt.DisplayRole) -> Any:
        if not index.isValid() or not (0 <= index.row() < len(self._rows)):
            return None

        entry = self._rows[index.row()]
        display_values = [entry.timestamp, entry.action, entry.file_name, entry.status, entry.detail]

        if role == Qt.DisplayRole:
            return display_values[index.column()]
        if role == Qt.ToolTipRole:
            return display_values[index.column()]
        if role == Qt.TextAlignmentRole:
            return Qt.AlignVCenter | (Qt.AlignCenter if index.column() in {0, 3} else Qt.AlignLeft)
        if role == Qt.BackgroundRole:
            return QBrush(self._tone_background(entry.tone))
        if role == Qt.ForegroundRole:
            return QBrush(self._tone_foreground(entry.tone))
        return None

    @staticmethod
    def _tone_background(tone: str) -> QColor:
        palette = {
            "success": QColor("#eafaf1"),
            "warning": QColor("#fff7eb"),
            "error": QColor("#fff0f3"),
            "info": QColor("#eef4ff"),
        }
        return palette.get(tone, QColor("#eef4ff"))

    @staticmethod
    def _tone_foreground(tone: str) -> QColor:
        palette = {
            "success": QColor("#1f6f4a"),
            "warning": QColor("#8a5a19"),
            "error": QColor("#9f3143"),
            "info": QColor("#2d5bca"),
        }
        return palette.get(tone, QColor("#2d5bca"))


class FolderPathLineEdit(QLineEdit):
    folderDropped = Signal(str)

    def __init__(self, placeholder: str = "", parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setAcceptDrops(True)
        if placeholder:
            self.setPlaceholderText(placeholder)

    def dragEnterEvent(self, event: QDragEnterEvent) -> None:
        if self._extract_path_from_event(event) is not None:
            event.acceptProposedAction()
            return
        super().dragEnterEvent(event)

    def dropEvent(self, event: QDropEvent) -> None:
        dropped_path = self._extract_path_from_event(event)
        if dropped_path is not None:
            self.setText(dropped_path)
            self.folderDropped.emit(dropped_path)
            event.acceptProposedAction()
            return
        super().dropEvent(event)

    def _extract_path_from_event(self, event: QDragEnterEvent | QDropEvent) -> Optional[str]:
        mime_data = event.mimeData()

        if mime_data.hasUrls():
            for url in mime_data.urls():
                local_path = url.toLocalFile()
                if not local_path:
                    continue
                path = Path(local_path)
                if path.is_dir():
                    return os.path.normpath(str(path))
                if path.exists():
                    return os.path.normpath(str(path.parent))

        if mime_data.hasText():
            raw_text = mime_data.text().strip().strip('"')
            if raw_text:
                path = Path(raw_text)
                if path.is_dir():
                    return os.path.normpath(str(path))
                if path.exists():
                    return os.path.normpath(str(path.parent))

        return None


class ErrorOverlayDialog(QDialog):
    def __init__(self, parent: Optional[QWidget], title: str, summary: str, details: str) -> None:
        super().__init__(parent)
        self._blur_target = parent.centralWidget() if isinstance(parent, QMainWindow) else parent
        self._owns_blur_effect = False
        self._details_text = details.strip()
        self._is_finalizing = False
        self._close_result = QDialog.Rejected
        self._open_animation_started = False
        self._blur_effect: Optional[QGraphicsBlurEffect] = None

        self.setModal(True)
        self.setObjectName("ErrorOverlayDialog")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setMinimumSize(720, 480)
        self.setWindowOpacity(0.0)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        overlay = QFrame()
        self.overlay = overlay
        overlay.setObjectName("ErrorOverlayBackdrop")
        overlay_layout = QVBoxLayout(overlay)
        overlay_layout.setContentsMargins(0, 0, 0, 0)
        overlay_layout.addStretch(1)

        card = QFrame()
        self.card = card
        card.setObjectName("ErrorOverlayCard")
        card.setMinimumWidth(640)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 24, 28, 24)
        card_layout.setSpacing(16)

        header_row = QHBoxLayout()
        header_row.setSpacing(14)

        icon_badge = QLabel("!")
        icon_badge.setObjectName("ErrorOverlayIcon")
        icon_badge.setAlignment(Qt.AlignCenter)
        icon_badge.setFixedSize(48, 48)

        title_wrap = QVBoxLayout()
        title_wrap.setSpacing(4)

        title_label = QLabel(title)
        title_label.setObjectName("ErrorOverlayTitle")
        title_label.setWordWrap(True)

        summary_label = QLabel(summary or "Terjadi kesalahan yang tidak terduga pada aplikasi.")
        summary_label.setObjectName("ErrorOverlaySummary")
        summary_label.setWordWrap(True)
        summary_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        title_wrap.addWidget(title_label)
        title_wrap.addWidget(summary_label)
        header_row.addWidget(icon_badge, 0, Qt.AlignTop)
        header_row.addLayout(title_wrap, 1)

        detail_label = QLabel("Detail teknis")
        detail_label.setObjectName("ErrorOverlaySectionTitle")

        self.details_box = QPlainTextEdit()
        self.details_box.setObjectName("ErrorOverlayDetails")
        self.details_box.setReadOnly(True)
        self.details_box.setPlainText(self._details_text or summary)
        self.details_box.setMinimumHeight(220)

        button_row = QHBoxLayout()
        button_row.setSpacing(10)
        button_row.addStretch(1)

        copy_button = QPushButton("Salin Detail")
        copy_button.setObjectName("ErrorOverlaySecondaryButton")
        copy_button.clicked.connect(self._copy_details)

        close_button = QPushButton("Tutup")
        close_button.setObjectName("ErrorOverlayPrimaryButton")
        close_button.clicked.connect(lambda: self.done(QDialog.Accepted))
        close_button.setAutoDefault(True)
        close_button.setDefault(True)

        button_row.addWidget(copy_button)
        button_row.addWidget(close_button)

        card_layout.addLayout(header_row)
        card_layout.addWidget(detail_label)
        card_layout.addWidget(self.details_box)
        card_layout.addLayout(button_row)

        overlay_layout.addWidget(card, 0, Qt.AlignCenter)
        overlay_layout.addStretch(1)
        outer.addWidget(overlay)

        self.setStyleSheet(
            f"""
            QDialog#ErrorOverlayDialog {{
                background: transparent;
            }}
            QFrame#ErrorOverlayBackdrop {{
                background: rgba(11, 18, 32, 168);
                border-radius: 0px;
            }}
            QFrame#ErrorOverlayCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #f5f8ff
                );
                border: 1px solid #d7e1f4;
                border-radius: 24px;
            }}
            QLabel#ErrorOverlayIcon {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #f56d86,
                    stop: 1 #e64d67
                );
                color: white;
                border-radius: 24px;
                font: 700 18pt "Segoe UI";
            }}
            QLabel#ErrorOverlayTitle {{
                color: {TEXT};
                font: 700 15pt "Segoe UI";
            }}
            QLabel#ErrorOverlaySummary {{
                color: {MUTED};
                font: 10pt "Segoe UI";
                line-height: 1.4em;
            }}
            QLabel#ErrorOverlaySectionTitle {{
                color: {TEXT};
                font: 700 10pt "Segoe UI";
            }}
            QPlainTextEdit#ErrorOverlayDetails {{
                background: #f7faff;
                color: #25324a;
                border: 1px solid #d7e1f4;
                border-radius: 16px;
                padding: 12px;
                font: 9.5pt "Consolas";
                selection-background-color: #cfe0ff;
            }}
            QPushButton#ErrorOverlayPrimaryButton {{
                min-width: 120px;
                padding: 10px 18px;
                border: none;
                border-radius: 14px;
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 {PRIMARY},
                    stop: 1 #4d84ff
                );
                color: white;
                font: 700 10pt "Segoe UI";
            }}
            QPushButton#ErrorOverlayPrimaryButton:hover {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 {PRIMARY_DARK},
                    stop: 1 #386cf0
                );
            }}
            QPushButton#ErrorOverlaySecondaryButton {{
                min-width: 120px;
                padding: 10px 18px;
                border-radius: 14px;
                border: 1px solid #d3def2;
                background: white;
                color: {TEXT};
                font: 600 10pt "Segoe UI";
            }}
            QPushButton#ErrorOverlaySecondaryButton:hover {{
                background: #f4f8ff;
            }}
            """
        )

    def showEvent(self, event) -> None:
        parent = self.parentWidget()
        if parent is not None:
            self.resize(parent.size())
            self.move(parent.mapToGlobal(QPoint(0, 0)))
        if self._blur_target is not None and self._blur_target.graphicsEffect() is None:
            blur = QGraphicsBlurEffect(self._blur_target)
            blur.setBlurRadius(8.0)
            self._blur_target.setGraphicsEffect(blur)
            self._owns_blur_effect = True
            self._blur_effect = blur
        super().showEvent(event)
        QTimer.singleShot(0, self._start_open_animation)

    def done(self, result: int) -> None:
        if self._is_finalizing:
            self._clear_blur()
            super().done(result)
            return
        self._close_result = result
        self._start_close_animation()

    def hideEvent(self, event) -> None:
        self._clear_blur()
        super().hideEvent(event)

    def closeEvent(self, event) -> None:
        if not self._is_finalizing:
            event.ignore()
            self.done(QDialog.Rejected)
            return
        self._clear_blur()
        super().closeEvent(event)

    def _clear_blur(self) -> None:
        if self._owns_blur_effect and self._blur_target is not None:
            self._blur_target.setGraphicsEffect(None)
            self._owns_blur_effect = False
        self._blur_effect = None

    @staticmethod
    def _scaled_rect(rect: QRect, scale: float) -> QRect:
        width = max(1, int(rect.width() * scale))
        height = max(1, int(rect.height() * scale))
        center = rect.center()
        scaled = QRect(0, 0, width, height)
        scaled.moveCenter(center)
        return scaled

    def _start_open_animation(self) -> None:
        if self._open_animation_started:
            return
        end_rect = self.card.geometry()
        if not end_rect.isValid():
            return

        self._open_animation_started = True
        start_rect = self._scaled_rect(end_rect, 0.985)
        self.card.setGeometry(start_rect)
        self.setWindowOpacity(0.0)

        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(170)
        opacity_animation.setStartValue(0.0)
        opacity_animation.setEndValue(1.0)
        opacity_animation.setEasingCurve(QEasingCurve.OutCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(180)
        geometry_animation.setStartValue(start_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.OutCubic)

        animation.start()
        self._open_animation = animation

    def _start_close_animation(self) -> None:
        if self._is_finalizing:
            return

        current_rect = self.card.geometry()
        if not current_rect.isValid():
            self._finalize_close()
            return

        end_rect = self._scaled_rect(current_rect, 0.99)
        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(140)
        opacity_animation.setStartValue(self.windowOpacity())
        opacity_animation.setEndValue(0.0)
        opacity_animation.setEasingCurve(QEasingCurve.InCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(140)
        geometry_animation.setStartValue(current_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.InCubic)

        animation.finished.connect(self._finalize_close)
        animation.start()
        self._close_animation = animation

    def _finalize_close(self) -> None:
        if self._is_finalizing:
            return
        self._is_finalizing = True
        try:
            self._clear_blur()
            super().done(self._close_result)
        finally:
            self._is_finalizing = False

    def _copy_details(self) -> None:
        QApplication.clipboard().setText(self._details_text or self.details_box.toPlainText())


class ConfirmOverlayDialog(QDialog):
    confirmRequested = Signal()

    def __init__(
        self,
        parent: Optional[QWidget],
        title: str,
        summary: str,
        details: str,
        *,
        detail_title: str = "Pratinjau file",
        confirm_button_text: str = "Ya, Proses Hapus",
        confirm_footnote: str = "Pastikan Anda sudah meninjau file hijau sebelum melanjutkan proses hapus.",
        processing_footnote: str = "Penghapusan sedang berjalan. Mohon tunggu sampai proses selesai.",
        processing_button_text: str = "Memproses",
        success_title: str = "Penghapusan Berhasil",
        success_detail_title: str = "Ringkasan hasil",
        success_footnote: str = "Proses selesai. Anda dapat menutup dialog ini.",
    ) -> None:
        super().__init__(parent)
        self._blur_target = parent.centralWidget() if isinstance(parent, QMainWindow) else parent
        self._owns_blur_effect = False
        self._details_text = details.strip()
        self._is_finalizing = False
        self._close_result = QDialog.Rejected
        self._open_animation_started = False
        self._blur_effect: Optional[QGraphicsBlurEffect] = None
        self._processing = False
        self._success_mode = False
        self._confirm_button_text = confirm_button_text
        self._detail_title_text = detail_title
        self._confirm_footnote_text = confirm_footnote
        self._processing_footnote_text = processing_footnote
        self._processing_button_text = processing_button_text
        self._success_title_text = success_title
        self._success_detail_title_text = success_detail_title
        self._success_footnote_text = success_footnote

        self.setModal(True)
        self.setObjectName("ConfirmOverlayDialog")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setMinimumSize(720, 500)
        self.setWindowOpacity(0.0)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        overlay = QFrame()
        self.overlay = overlay
        overlay.setObjectName("ConfirmOverlayBackdrop")
        overlay_layout = QVBoxLayout(overlay)
        overlay_layout.setContentsMargins(0, 0, 0, 0)
        overlay_layout.addStretch(1)

        card = QFrame()
        self.card = card
        card.setObjectName("ConfirmOverlayCard")
        card.setMinimumWidth(660)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 24, 28, 24)
        card_layout.setSpacing(16)

        header_row = QHBoxLayout()
        header_row.setSpacing(14)

        icon_badge = QLabel("!")
        self.icon_badge = icon_badge
        icon_badge.setObjectName("ConfirmOverlayIcon")
        icon_badge.setProperty("successMode", False)
        icon_badge.setAlignment(Qt.AlignCenter)
        icon_badge.setFixedSize(48, 48)

        title_wrap = QVBoxLayout()
        title_wrap.setSpacing(4)

        title_label = QLabel(title)
        self.title_label = title_label
        title_label.setObjectName("ConfirmOverlayTitle")
        title_label.setWordWrap(True)

        summary_label = QLabel(summary)
        self.summary_label = summary_label
        summary_label.setObjectName("ConfirmOverlaySummary")
        summary_label.setWordWrap(True)
        summary_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        title_wrap.addWidget(title_label)
        title_wrap.addWidget(summary_label)
        header_row.addWidget(icon_badge, 0, Qt.AlignTop)
        header_row.addLayout(title_wrap, 1)

        detail_label = QLabel(self._detail_title_text)
        self.detail_label = detail_label
        detail_label.setObjectName("ConfirmOverlaySectionTitle")

        self.details_box = QPlainTextEdit()
        self.details_box.setObjectName("ConfirmOverlayDetails")
        self.details_box.setProperty("successMode", False)
        self.details_box.setReadOnly(True)
        self.details_box.setPlainText(self._details_text)
        self.details_box.setMinimumHeight(240)

        footnote = QLabel(self._confirm_footnote_text)
        self.footnote_label = footnote
        footnote.setObjectName("ConfirmOverlayFootnote")
        footnote.setProperty("successMode", False)
        footnote.setWordWrap(True)

        button_row = QHBoxLayout()
        button_row.setSpacing(10)
        button_row.addStretch(1)

        cancel_button = QPushButton("Batal")
        self.cancel_button = cancel_button
        cancel_button.setObjectName("ConfirmOverlaySecondaryButton")
        cancel_button.clicked.connect(lambda: self.done(QDialog.Rejected))

        confirm_button = QPushButton(self._confirm_button_text)
        self.confirm_button = confirm_button
        confirm_button.setObjectName("ConfirmOverlayPrimaryButton")
        confirm_button.setProperty("successMode", False)
        confirm_button.clicked.connect(self._on_confirm_clicked)
        confirm_button.setAutoDefault(True)
        confirm_button.setDefault(True)

        button_row.addWidget(cancel_button)
        button_row.addWidget(confirm_button)

        self.custom_spinner = WaitingSpinner(confirm_button, False, False)
        self.custom_spinner.roundness = 70.0
        self.custom_spinner.minimum_trail_opacity = 15.0
        self.custom_spinner.trail_fade_percentage = 70.0
        self.custom_spinner.number_of_lines = 12
        self.custom_spinner.line_length = 5
        self.custom_spinner.line_width = 2
        self.custom_spinner.inner_radius = 5
        self.custom_spinner.revolutions_per_second = 1
        self.custom_spinner.color = QColor("white")
        
        btn_layout = QHBoxLayout(confirm_button)
        btn_layout.setContentsMargins(0, 0, 16, 0)
        btn_layout.addWidget(self.custom_spinner, 0, Qt.AlignRight | Qt.AlignVCenter)

        card_layout.addLayout(header_row)
        card_layout.addWidget(detail_label)
        card_layout.addWidget(self.details_box)
        card_layout.addWidget(footnote)
        card_layout.addLayout(button_row)

        overlay_layout.addWidget(card, 0, Qt.AlignCenter)
        overlay_layout.addStretch(1)
        outer.addWidget(overlay)

        self.setStyleSheet(
            f"""
            QDialog#ConfirmOverlayDialog {{
                background: transparent;
            }}
            QFrame#ConfirmOverlayBackdrop {{
                background: rgba(11, 18, 32, 168);
                border-radius: 0px;
            }}
            QFrame#ConfirmOverlayCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #fef8f3
                );
                border: 1px solid #f0d7c3;
                border-radius: 24px;
            }}
            QFrame#ConfirmOverlayCard[successMode="true"] {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #f5fff9
                );
                border: 1px solid #cce9db;
            }}
            QLabel#ConfirmOverlayIcon {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #f29a3b,
                    stop: 1 #e07a16
                );
                color: white;
                border-radius: 24px;
                font: 700 18pt "Segoe UI";
            }}
            QLabel#ConfirmOverlayIcon[successMode="true"] {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #16a56f,
                    stop: 1 #27c281
                );
            }}
            QLabel#ConfirmOverlayTitle {{
                color: {TEXT};
                font: 700 15pt "Segoe UI";
            }}
            QLabel#ConfirmOverlaySummary {{
                color: {MUTED};
                font: 10pt "Segoe UI";
                line-height: 1.4em;
            }}
            QLabel#ConfirmOverlaySectionTitle {{
                color: {TEXT};
                font: 700 10pt "Segoe UI";
            }}
            QLabel#ConfirmOverlayFootnote {{
                color: #8a5a19;
                font: 9.3pt "Segoe UI";
            }}
            QLabel#ConfirmOverlayFootnote[successMode="true"] {{
                color: #1f7a58;
            }}
            QPlainTextEdit#ConfirmOverlayDetails {{
                background: #fffaf5;
                color: #25324a;
                border: 1px solid #f0d7c3;
                border-radius: 16px;
                padding: 12px;
                font: 9.5pt "Consolas";
                selection-background-color: #ffe2c0;
            }}
            QPlainTextEdit#ConfirmOverlayDetails[successMode="true"] {{
                background: #f6fff9;
                border: 1px solid #cce9db;
                selection-background-color: #cfeedd;
            }}
            QPushButton#ConfirmOverlayPrimaryButton {{
                min-width: 150px;
                padding: 10px 46px 10px 46px;
                border: none;
                border-radius: 14px;
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ed7b24,
                    stop: 1 #f29a3b
                );
                color: white;
                font: 700 10pt "Segoe UI";
            }}
            QPushButton#ConfirmOverlayPrimaryButton:hover {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #db6912,
                    stop: 1 #eb8c28
                );
            }}
            QPushButton#ConfirmOverlayPrimaryButton[successMode="true"] {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #169766,
                    stop: 1 #27c281
                );
            }}
            QPushButton#ConfirmOverlayPrimaryButton[successMode="true"]:hover {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #118557,
                    stop: 1 #1fb172
                );
            }}
            QPushButton#ConfirmOverlayPrimaryButton:disabled {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ed7b24,
                    stop: 1 #f29a3b
                );
                color: rgba(255, 255, 255, 0.96);
                border: none;
            }}
            QPushButton#ConfirmOverlayPrimaryButton[successMode="true"]:disabled {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #169766,
                    stop: 1 #27c281
                );
            }}
            QPushButton#ConfirmOverlaySecondaryButton {{
                min-width: 120px;
                padding: 10px 18px;
                border-radius: 14px;
                border: 1px solid #e8d8c9;
                background: white;
                color: {TEXT};
                font: 600 10pt "Segoe UI";
            }}
            QPushButton#ConfirmOverlaySecondaryButton:hover {{
                background: #fff8f1;
            }}
            QPushButton#ConfirmOverlaySecondaryButton:disabled {{
                background: #fff7ef;
                border: 1px solid #ecdccf;
                color: #b18a66;
            }}
            """
        )
        self.card.setProperty("successMode", False)

    def _refresh_theme_state(self) -> None:
        success_mode = self._success_mode
        themed_widgets = [
            self.card,
            self.icon_badge,
            self.details_box,
            self.footnote_label,
            self.confirm_button,
        ]
        for widget in themed_widgets:
            widget.setProperty("successMode", success_mode)
            widget.style().unpolish(widget)
            widget.style().polish(widget)
        self.style().unpolish(self)
        self.style().polish(self)
        self.overlay.update()
        self.card.update()

    def showEvent(self, event) -> None:
        parent = self.parentWidget()
        if parent is not None:
            self.resize(parent.size())
            self.move(parent.mapToGlobal(QPoint(0, 0)))
        if self._blur_target is not None and self._blur_target.graphicsEffect() is None:
            blur = QGraphicsBlurEffect(self._blur_target)
            blur.setBlurRadius(8.0)
            self._blur_target.setGraphicsEffect(blur)
            self._owns_blur_effect = True
            self._blur_effect = blur
        super().showEvent(event)
        QTimer.singleShot(0, self._start_open_animation)

    def done(self, result: int) -> None:
        if self._processing:
            return
        if self._is_finalizing:
            self._clear_blur()
            super().done(result)
            return
        self._close_result = result
        self._start_close_animation()

    def hideEvent(self, event) -> None:
        self._clear_blur()
        super().hideEvent(event)

    def closeEvent(self, event) -> None:
        if self._processing:
            event.ignore()
            return
        if not self._is_finalizing:
            event.ignore()
            self.done(QDialog.Rejected)
            return
        self._clear_blur()
        super().closeEvent(event)

    def _clear_blur(self) -> None:
        if self._owns_blur_effect and self._blur_target is not None:
            self._blur_target.setGraphicsEffect(None)
            self._owns_blur_effect = False
        self._blur_effect = None

    @staticmethod
    def _scaled_rect(rect: QRect, scale: float) -> QRect:
        width = max(1, int(rect.width() * scale))
        height = max(1, int(rect.height() * scale))
        center = rect.center()
        scaled = QRect(0, 0, width, height)
        scaled.moveCenter(center)
        return scaled

    def _start_open_animation(self) -> None:
        if self._open_animation_started:
            return
        end_rect = self.card.geometry()
        if not end_rect.isValid():
            return

        self._open_animation_started = True
        start_rect = self._scaled_rect(end_rect, 0.985)
        self.card.setGeometry(start_rect)
        self.setWindowOpacity(0.0)

        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(170)
        opacity_animation.setStartValue(0.0)
        opacity_animation.setEndValue(1.0)
        opacity_animation.setEasingCurve(QEasingCurve.OutCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(180)
        geometry_animation.setStartValue(start_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.OutCubic)

        animation.start()
        self._open_animation = animation

    def _start_close_animation(self) -> None:
        if self._is_finalizing:
            return

        current_rect = self.card.geometry()
        if not current_rect.isValid():
            self._finalize_close()
            return

        end_rect = self._scaled_rect(current_rect, 0.99)
        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(140)
        opacity_animation.setStartValue(self.windowOpacity())
        opacity_animation.setEndValue(0.0)
        opacity_animation.setEasingCurve(QEasingCurve.InCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(140)
        geometry_animation.setStartValue(current_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.InCubic)

        animation.finished.connect(self._finalize_close)
        animation.start()
        self._close_animation = animation

    def _finalize_close(self) -> None:
        if self._is_finalizing:
            return
        self._is_finalizing = True
        try:
            self._clear_blur()
            super().done(self._close_result)
        finally:
            self._is_finalizing = False

    def _on_confirm_clicked(self) -> None:
        if self._processing:
            return
        if self._success_mode:
            self.done(QDialog.Accepted)
            return
        self.confirmRequested.emit()

    def set_processing(self, processing: bool, summary: Optional[str] = None) -> None:
        self._processing = processing
        if processing:
            self._success_mode = False
            self._refresh_theme_state()
        if summary:
            self.summary_label.setText(summary)

        self.cancel_button.setEnabled(not processing)
        self.confirm_button.setEnabled(not processing)

        if processing:
            self.footnote_label.setText(self._processing_footnote_text)
            self.confirm_button.setText(self._processing_button_text)
            self.custom_spinner.start()
        else:
            self.custom_spinner.stop()
            self.confirm_button.setText(self._confirm_button_text)
            self.footnote_label.setText(self._confirm_footnote_text)
            self.cancel_button.show()

    def flush_visual_state(self) -> None:
        self.raise_()
        self.activateWindow()
        self.overlay.update()
        self.card.update()
        self.repaint()
        QApplication.processEvents()

    def finish_processing(self) -> None:
        self.set_processing(False)
        self.done(QDialog.Accepted)

    def force_close(self, result: int = QDialog.Rejected) -> None:
        self.custom_spinner.stop()
        self._processing = False
        self.done(result)

    def show_success_state(self, summary: str, details: str = "") -> None:
        self.custom_spinner.stop()
        self._processing = False
        self._success_mode = True
        self._refresh_theme_state()

        self.title_label.setText(self._success_title_text)
        self.summary_label.setText(summary)
        self.icon_badge.setText("OK")
        self.detail_label.setText(self._success_detail_title_text)
        self.detail_label.setVisible(bool(details))
        self.details_box.setPlainText(details)
        self.details_box.setVisible(bool(details))
        self.footnote_label.setText(self._success_footnote_text)

        self.cancel_button.hide()
        self.confirm_button.setEnabled(True)
        self.confirm_button.setText("Tutup")


class SuccessOverlayDialog(QDialog):
    def __init__(self, parent: Optional[QWidget], title: str, summary: str, details: str = "") -> None:
        super().__init__(parent)
        self._blur_target = parent.centralWidget() if isinstance(parent, QMainWindow) else parent
        self._owns_blur_effect = False
        self._details_text = details.strip()
        self._is_finalizing = False
        self._close_result = QDialog.Accepted
        self._open_animation_started = False
        self._blur_effect: Optional[QGraphicsBlurEffect] = None

        self.setModal(True)
        self.setObjectName("SuccessOverlayDialog")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setMinimumSize(680, 420)
        self.setWindowOpacity(0.0)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        overlay = QFrame()
        self.overlay = overlay
        overlay.setObjectName("SuccessOverlayBackdrop")
        overlay_layout = QVBoxLayout(overlay)
        overlay_layout.setContentsMargins(0, 0, 0, 0)
        overlay_layout.addStretch(1)

        card = QFrame()
        self.card = card
        card.setObjectName("SuccessOverlayCard")
        card.setMinimumWidth(560)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 24, 28, 24)
        card_layout.setSpacing(16)

        header_row = QHBoxLayout()
        header_row.setSpacing(14)

        icon_badge = QLabel("OK")
        icon_badge.setObjectName("SuccessOverlayIcon")
        icon_badge.setAlignment(Qt.AlignCenter)
        icon_badge.setFixedSize(56, 56)

        title_wrap = QVBoxLayout()
        title_wrap.setSpacing(4)

        title_label = QLabel(title)
        title_label.setObjectName("SuccessOverlayTitle")
        title_label.setWordWrap(True)

        summary_label = QLabel(summary)
        summary_label.setObjectName("SuccessOverlaySummary")
        summary_label.setWordWrap(True)
        summary_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        title_wrap.addWidget(title_label)
        title_wrap.addWidget(summary_label)
        header_row.addWidget(icon_badge, 0, Qt.AlignTop)
        header_row.addLayout(title_wrap, 1)

        detail_label = QLabel("Ringkasan hasil")
        detail_label.setObjectName("SuccessOverlaySectionTitle")
        detail_label.setVisible(bool(self._details_text))

        self.details_box = QPlainTextEdit()
        self.details_box.setObjectName("SuccessOverlayDetails")
        self.details_box.setReadOnly(True)
        self.details_box.setPlainText(self._details_text)
        self.details_box.setMinimumHeight(140)
        self.details_box.setVisible(bool(self._details_text))

        button_row = QHBoxLayout()
        button_row.setSpacing(10)
        button_row.addStretch(1)

        close_button = QPushButton("Tutup")
        close_button.setObjectName("SuccessOverlayPrimaryButton")
        close_button.clicked.connect(lambda: self.done(QDialog.Accepted))
        close_button.setAutoDefault(True)
        close_button.setDefault(True)

        button_row.addWidget(close_button)

        card_layout.addLayout(header_row)
        card_layout.addWidget(detail_label)
        card_layout.addWidget(self.details_box)
        card_layout.addLayout(button_row)

        overlay_layout.addWidget(card, 0, Qt.AlignCenter)
        overlay_layout.addStretch(1)
        outer.addWidget(overlay)

        self.setStyleSheet(
            f"""
            QDialog#SuccessOverlayDialog {{
                background: transparent;
            }}
            QFrame#SuccessOverlayBackdrop {{
                background: rgba(11, 18, 32, 168);
                border-radius: 0px;
            }}
            QFrame#SuccessOverlayCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #f5fff9
                );
                border: 1px solid #cce9db;
                border-radius: 24px;
            }}
            QLabel#SuccessOverlayIcon {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #16a56f,
                    stop: 1 #27c281
                );
                color: white;
                border-radius: 28px;
                font: 800 11pt "Segoe UI";
            }}
            QLabel#SuccessOverlayTitle {{
                color: {TEXT};
                font: 700 15pt "Segoe UI";
            }}
            QLabel#SuccessOverlaySummary {{
                color: {MUTED};
                font: 10pt "Segoe UI";
                line-height: 1.4em;
            }}
            QLabel#SuccessOverlaySectionTitle {{
                color: {TEXT};
                font: 700 10pt "Segoe UI";
            }}
            QPlainTextEdit#SuccessOverlayDetails {{
                background: #f6fff9;
                color: #25324a;
                border: 1px solid #cce9db;
                border-radius: 16px;
                padding: 12px;
                font: 9.5pt "Consolas";
                selection-background-color: #cfeedd;
            }}
            QPushButton#SuccessOverlayPrimaryButton {{
                min-width: 120px;
                padding: 10px 18px;
                border: none;
                border-radius: 14px;
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #169766,
                    stop: 1 #27c281
                );
                color: white;
                font: 700 10pt "Segoe UI";
            }}
            QPushButton#SuccessOverlayPrimaryButton:hover {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #118557,
                    stop: 1 #1fb172
                );
            }}
            """
        )

    def showEvent(self, event) -> None:
        parent = self.parentWidget()
        if parent is not None:
            self.resize(parent.size())
            self.move(parent.mapToGlobal(QPoint(0, 0)))
        if self._blur_target is not None and self._blur_target.graphicsEffect() is None:
            blur = QGraphicsBlurEffect(self._blur_target)
            blur.setBlurRadius(8.0)
            self._blur_target.setGraphicsEffect(blur)
            self._owns_blur_effect = True
            self._blur_effect = blur
        super().showEvent(event)
        QTimer.singleShot(0, self._start_open_animation)

    def done(self, result: int) -> None:
        if self._is_finalizing:
            self._clear_blur()
            super().done(result)
            return
        self._close_result = result
        self._start_close_animation()

    def hideEvent(self, event) -> None:
        self._clear_blur()
        super().hideEvent(event)

    def closeEvent(self, event) -> None:
        if not self._is_finalizing:
            event.ignore()
            self.done(QDialog.Accepted)
            return
        self._clear_blur()
        super().closeEvent(event)

    def _clear_blur(self) -> None:
        if self._owns_blur_effect and self._blur_target is not None:
            self._blur_target.setGraphicsEffect(None)
            self._owns_blur_effect = False
        self._blur_effect = None

    @staticmethod
    def _scaled_rect(rect: QRect, scale: float) -> QRect:
        width = max(1, int(rect.width() * scale))
        height = max(1, int(rect.height() * scale))
        center = rect.center()
        scaled = QRect(0, 0, width, height)
        scaled.moveCenter(center)
        return scaled

    def _start_open_animation(self) -> None:
        if self._open_animation_started:
            return
        end_rect = self.card.geometry()
        if not end_rect.isValid():
            return

        self._open_animation_started = True
        start_rect = self._scaled_rect(end_rect, 0.985)
        self.card.setGeometry(start_rect)
        self.setWindowOpacity(0.0)

        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(170)
        opacity_animation.setStartValue(0.0)
        opacity_animation.setEndValue(1.0)
        opacity_animation.setEasingCurve(QEasingCurve.OutCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(180)
        geometry_animation.setStartValue(start_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.OutCubic)

        animation.start()
        self._open_animation = animation

    def _start_close_animation(self) -> None:
        if self._is_finalizing:
            return

        current_rect = self.card.geometry()
        if not current_rect.isValid():
            self._finalize_close()
            return

        end_rect = self._scaled_rect(current_rect, 0.99)
        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(140)
        opacity_animation.setStartValue(self.windowOpacity())
        opacity_animation.setEndValue(0.0)
        opacity_animation.setEasingCurve(QEasingCurve.InCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(140)
        geometry_animation.setStartValue(current_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.InCubic)

        animation.finished.connect(self._finalize_close)
        animation.start()
        self._close_animation = animation

    def _finalize_close(self) -> None:
        if self._is_finalizing:
            return
        self._is_finalizing = True
        try:
            self._clear_blur()
            super().done(self._close_result)
        finally:
            self._is_finalizing = False


class FileDetailOverlayDialog(QDialog):
    def __init__(
        self,
        parent: Optional[QWidget],
        result: MatchResult,
        missing_labels: List[str],
        suggestion_text: str,
        *,
        show_compare_actions: bool = False,
    ) -> None:
        super().__init__(parent)
        self._blur_target = parent.centralWidget() if isinstance(parent, QMainWindow) else parent
        self._owns_blur_effect = False
        self._is_finalizing = False
        self._close_result = QDialog.Accepted
        self._open_animation_started = False
        self._blur_effect: Optional[QGraphicsBlurEffect] = None
        self._result = result
        self._show_compare_actions = show_compare_actions

        self.setModal(True)
        self.setObjectName("FileDetailOverlayDialog")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setMinimumSize(860, 620)
        self.setWindowOpacity(0.0)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        overlay = QFrame()
        self.overlay = overlay
        overlay.setObjectName("FileDetailOverlayBackdrop")
        overlay_layout = QVBoxLayout(overlay)
        overlay_layout.setContentsMargins(0, 0, 0, 0)
        overlay_layout.addStretch(1)

        card = QFrame()
        self.card = card
        card.setObjectName("FileDetailOverlayCard")
        card.setMinimumWidth(760)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 24, 28, 24)
        card_layout.setSpacing(16)

        header_row = QHBoxLayout()
        header_row.setSpacing(14)

        icon_badge = QLabel("i")
        icon_badge.setObjectName("FileDetailOverlayIcon")
        icon_badge.setAlignment(Qt.AlignCenter)
        icon_badge.setFixedSize(52, 52)

        title_wrap = QVBoxLayout()
        title_wrap.setSpacing(4)

        title_label = QLabel("Detail File Terpilih")
        title_label.setObjectName("FileDetailOverlayTitle")
        title_label.setWordWrap(True)

        summary_label = QLabel(f"{result.status_text} | {result.target_relative_path}")
        summary_label.setObjectName("FileDetailOverlaySummary")
        summary_label.setWordWrap(True)
        summary_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        title_wrap.addWidget(title_label)
        title_wrap.addWidget(summary_label)
        header_row.addWidget(icon_badge, 0, Qt.AlignTop)
        header_row.addLayout(title_wrap, 1)

        scroll = QScrollArea()
        scroll.setObjectName("FileDetailOverlayScroll")
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        body = QWidget()
        body.setObjectName("FileDetailOverlayBody")
        body_layout = QVBoxLayout(body)
        body_layout.setContentsMargins(0, 0, 0, 0)
        body_layout.setSpacing(14)

        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(12)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)

        info_cards = [
            ("Status", result.status_text),
            ("Path Target", str(result.target_path)),
            ("Relative Path", result.target_relative_path),
            ("Ukuran", f"{MatchResultTableModel._format_size(result.size)} ({result.size} bytes)"),
            ("Ditemukan di", result.exact_folder_labels),
            ("Tidak Ada di", ", ".join(missing_labels) if missing_labels else "-"),
            ("Mode", result.match_type),
        ]

        for index, (label_text, value_text) in enumerate(info_cards):
            grid.addWidget(self._create_info_card(label_text, value_text), index // 2, index % 2)

        body_layout.addLayout(grid)

        exact_card = self._create_text_card("Path Cocok", result.exact_paths_text)
        diff_card = self._create_text_card("Path Beda", result.diff_paths_text)
        body_layout.addWidget(exact_card)
        body_layout.addWidget(diff_card)

        if suggestion_text:
            suggestion_card = QFrame()
            suggestion_card.setObjectName("FileDetailSuggestionCard")
            suggestion_layout = QVBoxLayout(suggestion_card)
            suggestion_layout.setContentsMargins(14, 12, 14, 12)
            suggestion_layout.setSpacing(10)

            suggestion_title = QLabel("Saran Tindakan")
            suggestion_title.setObjectName("FileDetailSuggestionTitle")

            suggestion_label = QLabel(suggestion_text)
            suggestion_label.setObjectName("FileDetailSuggestionText")
            suggestion_label.setWordWrap(True)

            suggestion_layout.addWidget(suggestion_title)
            suggestion_layout.addWidget(suggestion_label)

            if show_compare_actions:
                suggestion_actions = QHBoxLayout()
                suggestion_actions.setSpacing(10)
                suggestion_actions.addStretch(1)

                copy_compare_button = QPushButton("Salin ke Folder Pembanding")
                copy_compare_button.setObjectName("FileDetailSecondaryButton")
                copy_compare_button.clicked.connect(lambda: self._trigger_compare_action("copy"))

                move_compare_button = QPushButton("Pindah ke Folder Pembanding")
                move_compare_button.setObjectName("FileDetailGhostButton")
                move_compare_button.clicked.connect(lambda: self._trigger_compare_action("move"))

                suggestion_actions.addWidget(copy_compare_button)
                suggestion_actions.addWidget(move_compare_button)
                suggestion_layout.addLayout(suggestion_actions)

            body_layout.addWidget(suggestion_card)

        scroll.setWidget(body)

        button_row = QHBoxLayout()
        button_row.setSpacing(10)
        button_row.addStretch(1)

        reveal_button = QPushButton("Tampilkan di Explorer")
        reveal_button.setObjectName("FileDetailSecondaryButton")
        reveal_button.setIcon(QIcon("assets/open_folder.svg"))
        reveal_button.clicked.connect(self._reveal_in_explorer)

        copy_button = QPushButton("Copy Path")
        copy_button.setObjectName("FileDetailSecondaryButton")
        copy_button.clicked.connect(self._copy_selected_path)

        close_button = QPushButton("Tutup")
        close_button.setObjectName("FileDetailPrimaryButton")
        close_button.clicked.connect(lambda: self.done(QDialog.Accepted))
        close_button.setAutoDefault(True)
        close_button.setDefault(True)

        button_row.addWidget(reveal_button)
        button_row.addWidget(copy_button)
        button_row.addWidget(close_button)

        card_layout.addLayout(header_row)
        card_layout.addWidget(scroll, 1)
        card_layout.addLayout(button_row)

        overlay_layout.addWidget(card, 0, Qt.AlignCenter)
        overlay_layout.addStretch(1)
        outer.addWidget(overlay)

        self.setStyleSheet(
            f"""
            QDialog#FileDetailOverlayDialog {{
                background: transparent;
            }}
            QFrame#FileDetailOverlayBackdrop {{
                background: rgba(11, 18, 32, 168);
                border-radius: 0px;
            }}
            QFrame#FileDetailOverlayCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #f5f9ff
                );
                border: 1px solid #d6e0f4;
                border-radius: 24px;
            }}
            QLabel#FileDetailOverlayIcon {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #2c67f2,
                    stop: 1 #5a8dff
                );
                color: white;
                border-radius: 26px;
                font: 700 18pt "Segoe UI";
            }}
            QLabel#FileDetailOverlayTitle {{
                color: {TEXT};
                font: 700 15pt "Segoe UI";
            }}
            QLabel#FileDetailOverlaySummary {{
                color: {MUTED};
                font: 10pt "Segoe UI";
            }}
            QScrollArea#FileDetailOverlayScroll {{
                background: transparent;
                border: none;
            }}
            QWidget#FileDetailOverlayBody {{
                background: transparent;
            }}
            QFrame#FileDetailInfoCard {{
                background: #f8fbff;
                border: 1px solid #dbe5f7;
                border-radius: 16px;
            }}
            QLabel#FileDetailInfoTitle {{
                color: {MUTED};
                font: 700 9pt "Segoe UI";
            }}
            QLabel#FileDetailInfoValue {{
                color: {TEXT};
                font: 9.5pt "Segoe UI";
            }}
            QFrame#FileDetailTextCard {{
                background: #f8fbff;
                border: 1px solid #dbe5f7;
                border-radius: 16px;
            }}
            QLabel#FileDetailTextTitle {{
                color: {TEXT};
                font: 700 10pt "Segoe UI";
            }}
            QPlainTextEdit#FileDetailTextBox {{
                background: #fcfdff;
                color: #25324a;
                border: 1px solid #dbe5f7;
                border-radius: 12px;
                padding: 10px;
                font: 9.3pt "Consolas";
                selection-background-color: #cfe0ff;
            }}
            QFrame#FileDetailSuggestionCard {{
                background: #f2f8ff;
                border: 1px solid #d2e1f7;
                border-radius: 16px;
            }}
            QLabel#FileDetailSuggestionTitle {{
                color: {TEXT};
                font: 700 10pt "Segoe UI";
            }}
            QLabel#FileDetailSuggestionText {{
                color: #35527a;
                font: 9.5pt "Segoe UI";
            }}
            QPushButton#FileDetailPrimaryButton {{
                min-width: 120px;
                padding: 10px 18px;
                border: none;
                border-radius: 14px;
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 {PRIMARY},
                    stop: 1 #4d84ff
                );
                color: white;
                font: 700 10pt "Segoe UI";
            }}
            QPushButton#FileDetailPrimaryButton:hover {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 {PRIMARY_DARK},
                    stop: 1 #386cf0
                );
            }}
            QPushButton#FileDetailSecondaryButton {{
                min-width: 120px;
                padding: 10px 18px;
                border-radius: 14px;
                border: 1px solid #d3def2;
                background: white;
                color: {TEXT};
                font: 600 10pt "Segoe UI";
            }}
            QPushButton#FileDetailSecondaryButton:hover {{
                background: #f4f8ff;
            }}
            QPushButton#FileDetailGhostButton {{
                min-width: 120px;
                padding: 10px 18px;
                border: none;
                border-radius: 14px;
                background: #edf3ff;
                color: {TEXT};
                font: 600 10pt "Segoe UI";
            }}
            QPushButton#FileDetailGhostButton:hover {{
                background: #e2ecff;
            }}
            """
        )

    def _create_info_card(self, title: str, value: str) -> QWidget:
        card = QFrame()
        card.setObjectName("FileDetailInfoCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setSpacing(6)

        title_label = QLabel(title)
        title_label.setObjectName("FileDetailInfoTitle")

        value_label = QLabel(value or "-")
        value_label.setObjectName("FileDetailInfoValue")
        value_label.setWordWrap(True)
        value_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        layout.addWidget(title_label)
        layout.addWidget(value_label)
        return card

    def _create_text_card(self, title: str, value: str) -> QWidget:
        card = QFrame()
        card.setObjectName("FileDetailTextCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 12, 14, 14)
        layout.setSpacing(8)

        title_label = QLabel(title)
        title_label.setObjectName("FileDetailTextTitle")

        text_box = QPlainTextEdit()
        text_box.setObjectName("FileDetailTextBox")
        text_box.setReadOnly(True)
        text_box.setPlainText(value or "-")
        text_box.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        text_box.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        text_box.setMinimumHeight(64)

        metrics = text_box.fontMetrics()
        line_count = max(1, len((value or "-").splitlines()))
        content_height = (metrics.lineSpacing() * min(line_count, 6)) + 28
        text_box.setFixedHeight(max(64, min(content_height, 176)))

        layout.addWidget(title_label)
        layout.addWidget(text_box)
        return card

    def showEvent(self, event) -> None:
        parent = self.parentWidget()
        if parent is not None:
            self.resize(parent.size())
            self.move(parent.mapToGlobal(QPoint(0, 0)))
        if self._blur_target is not None and self._blur_target.graphicsEffect() is None:
            blur = QGraphicsBlurEffect(self._blur_target)
            blur.setBlurRadius(8.0)
            self._blur_target.setGraphicsEffect(blur)
            self._owns_blur_effect = True
            self._blur_effect = blur
        super().showEvent(event)
        QTimer.singleShot(0, self._start_open_animation)

    def done(self, result: int) -> None:
        if self._is_finalizing:
            self._clear_blur()
            super().done(result)
            return
        self._close_result = result
        self._start_close_animation()

    def hideEvent(self, event) -> None:
        self._clear_blur()
        super().hideEvent(event)

    def closeEvent(self, event) -> None:
        if not self._is_finalizing:
            event.ignore()
            self.done(QDialog.Accepted)
            return
        self._clear_blur()
        super().closeEvent(event)

    def _clear_blur(self) -> None:
        if self._owns_blur_effect and self._blur_target is not None:
            self._blur_target.setGraphicsEffect(None)
            self._owns_blur_effect = False
        self._blur_effect = None

    @staticmethod
    def _scaled_rect(rect: QRect, scale: float) -> QRect:
        width = max(1, int(rect.width() * scale))
        height = max(1, int(rect.height() * scale))
        center = rect.center()
        scaled = QRect(0, 0, width, height)
        scaled.moveCenter(center)
        return scaled

    def _start_open_animation(self) -> None:
        if self._open_animation_started:
            return
        end_rect = self.card.geometry()
        if not end_rect.isValid():
            return

        self._open_animation_started = True
        start_rect = self._scaled_rect(end_rect, 0.985)
        self.card.setGeometry(start_rect)
        self.setWindowOpacity(0.0)

        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(170)
        opacity_animation.setStartValue(0.0)
        opacity_animation.setEndValue(1.0)
        opacity_animation.setEasingCurve(QEasingCurve.OutCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(180)
        geometry_animation.setStartValue(start_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.OutCubic)

        animation.start()
        self._open_animation = animation

    def _start_close_animation(self) -> None:
        if self._is_finalizing:
            return

        current_rect = self.card.geometry()
        if not current_rect.isValid():
            self._finalize_close()
            return

        end_rect = self._scaled_rect(current_rect, 0.99)
        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(140)
        opacity_animation.setStartValue(self.windowOpacity())
        opacity_animation.setEndValue(0.0)
        opacity_animation.setEasingCurve(QEasingCurve.InCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(140)
        geometry_animation.setStartValue(current_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.InCubic)

        animation.finished.connect(self._finalize_close)
        animation.start()
        self._close_animation = animation

    def _finalize_close(self) -> None:
        if self._is_finalizing:
            return
        self._is_finalizing = True
        try:
            self._clear_blur()
            super().done(self._close_result)
        finally:
            self._is_finalizing = False

    def _copy_selected_path(self) -> None:
        parent = self.parentWidget()
        if parent is not None and hasattr(parent, "copy_selected_path"):
            parent.copy_selected_path()

    def _reveal_in_explorer(self) -> None:
        try:
            path = str(self._result.target_path)
            if sys.platform == "win32":
                subprocess.Popen(f'explorer /select,"{os.path.normpath(path)}"')
            elif sys.platform == "darwin":  # macOS
                subprocess.Popen(['open', '-R', path])
            else:  # linux
                subprocess.Popen(['xdg-open', os.path.dirname(path)])
        except Exception as e:
            QMessageBox.warning(self, "Gagal Membuka File", f"Tidak dapat membuka file explorer:\n\n{str(e)}")

    def _trigger_compare_action(self, operation: str) -> None:
        parent = self.parentWidget()
        if parent is None:
            return
        # Pass self._result explicitly instead of relying on current_selected_result
        # which can be cleared by the focus-release handler when this dialog closes.
        sync_fn = getattr(parent, "_sync_selected_result_to_compare_folders", None)
        if sync_fn is None:
            return
        self.done(QDialog.Accepted)
        QTimer.singleShot(180, lambda op=operation, r=self._result: sync_fn(op, explicit_result=r))


class UpdateStatusDialog(QDialog):
    def __init__(self, parent, latest_version: str, current_version: str, changelog: str, has_update: bool):
        super().__init__(parent)
        self._blur_target = parent.centralWidget() if isinstance(parent, QMainWindow) else parent
        self._owns_blur_effect = False
        self._is_finalizing = False
        self._close_result = QDialog.Rejected
        self._open_animation_started = False
        self._blur_effect: Optional[QGraphicsBlurEffect] = None
        self._processing = False

        self.setModal(True)
        self.setObjectName("UpdateStatusDialog")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setMinimumSize(680, 480)
        self.setWindowOpacity(0.0)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        overlay = QFrame()
        self.overlay = overlay
        overlay.setObjectName("OverlayBackdrop")
        overlay_layout = QVBoxLayout(overlay)
        overlay_layout.setContentsMargins(0, 0, 0, 0)
        overlay_layout.addStretch(1)

        card = QFrame()
        self.card = card
        card.setObjectName("OverlayCard")
        card.setMinimumWidth(560)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 24, 28, 24)
        card_layout.setSpacing(16)

        header_row = QHBoxLayout()
        header_row.setSpacing(14)
        
        self.icon_badge = QLabel("🎉" if has_update else "✨")
        self.icon_badge.setObjectName("OverlayIcon")
        self.icon_badge.setAlignment(Qt.AlignCenter)
        self.icon_badge.setFixedSize(56, 56)

        title_wrap = QVBoxLayout()
        title_wrap.setSpacing(4)
        
        self.title_label = QLabel(f"Pembaruan Tersedia: v{latest_version}" if has_update else f"Aplikasi Terkini (v{current_version})")
        self.title_label.setObjectName("OverlayTitle")
        self.title_label.setWordWrap(True)

        self.summary_label = QLabel(f"Versi saat ini v{current_version}. Pembaruan baru tersedia untuk diunduh." if has_update else "Anda sudah menggunakan aplikasi versi terbaru.")
        self.summary_label.setObjectName("OverlaySummary")
        self.summary_label.setWordWrap(True)

        title_wrap.addWidget(self.title_label)
        title_wrap.addWidget(self.summary_label)
        header_row.addWidget(self.icon_badge, 0, Qt.AlignTop)
        header_row.addLayout(title_wrap, 1)

        card_layout.addLayout(header_row)

        if changelog:
            detail_label = QLabel("Catatan Rilis (Changelog)")
            detail_label.setObjectName("OverlaySectionTitle")
            card_layout.addWidget(detail_label)

            from PySide6.QtWidgets import QTextBrowser
            from PySide6.QtGui import QDesktopServices
            from PySide6.QtCore import QUrl
            import re

            self.changelog_view = QTextBrowser()
            self.changelog_view.setObjectName("OverlayDetails")
            self.changelog_view.setMinimumHeight(280)

            # Ekstraksi tag <details> untuk mensimulasikan expand/collapse
            self.is_details_expanded = False
            self.changelog_main = changelog
            self.changelog_summary = "Riwayat Versi"
            self.changelog_details = ""
            
            detail_match = re.search(r'(.*?)<details>\s*<summary>(.*?)</summary>(.*?)</details>(.*)', changelog, re.DOTALL | re.IGNORECASE)
            
            if detail_match:
                self.changelog_main = detail_match.group(1).strip()
                summary_raw = detail_match.group(2).strip()
                self.changelog_summary = re.sub(r'</?strong>', '', summary_raw, flags=re.IGNORECASE).strip()
                self.changelog_details = detail_match.group(3).strip()
                # Bersihkan tag br
                self.changelog_details = re.sub(r'<br\s*/?>', '\n', self.changelog_details, flags=re.IGNORECASE)
                
                if detail_match.group(4):
                    self.changelog_details += "\n\n" + detail_match.group(4).strip()

            def render_changelog():
                if detail_match:
                    if self.is_details_expanded:
                        md = f"{self.changelog_main}\n\n&nbsp;\n\n<a href='toggle_details' style='color:#1f5eff; text-decoration:none;'><b>▼ Sembunyikan {self.changelog_summary}</b></a>\n\n&nbsp;\n\n---\n\n{self.changelog_details}"
                    else:
                        md = f"{self.changelog_main}\n\n&nbsp;\n\n<a href='toggle_details' style='color:#1f5eff; text-decoration:none;'><b>▶ Tampilkan {self.changelog_summary}</b></a>\n\n&nbsp;"
                else:
                    md = self.changelog_main
                
                self.changelog_view.setMarkdown(md)

            def on_anchor_clicked(url):
                if url.toString() == "toggle_details":
                    self.is_details_expanded = not self.is_details_expanded
                    render_changelog()
                else:
                    QDesktopServices.openUrl(url)

            self.changelog_view.setOpenLinks(False)
            self.changelog_view.anchorClicked.connect(on_anchor_clicked)
            
            render_changelog()
            card_layout.addWidget(self.changelog_view, 1)
        else:
            card_layout.addStretch(1)

        button_row = QHBoxLayout()
        button_row.setSpacing(10)
        button_row.addStretch(1)

        if has_update:
            self.btn_nanti = QPushButton("Nanti")
            self.btn_nanti.setObjectName("OverlaySecondaryButton")
            self.btn_nanti.clicked.connect(lambda: self.done(QDialog.Rejected))
            
            self.btn_unduh = QPushButton("Unduh Sekarang")
            self.btn_unduh.setObjectName("OverlayPrimaryButton")
            self.btn_unduh.clicked.connect(lambda: self.done(QDialog.Accepted))
            
            button_row.addWidget(self.btn_nanti)
            button_row.addWidget(self.btn_unduh)
        else:
            self.btn_tutup = QPushButton("Tutup")
            self.btn_tutup.setObjectName("OverlayPrimaryButton")
            self.btn_tutup.clicked.connect(lambda: self.done(QDialog.Accepted))
            button_row.addWidget(self.btn_tutup)

        card_layout.addLayout(button_row)
        overlay_layout.addWidget(card, 0, Qt.AlignCenter)
        overlay_layout.addStretch(1)
        outer.addWidget(overlay)

        self._apply_stylesheet()

    def _apply_stylesheet(self) -> None:
        self.setStyleSheet(
            f"""
            QDialog#UpdateStatusDialog {{
                background: transparent;
            }}
            QFrame#OverlayBackdrop {{
                background: rgba(11, 18, 32, 168);
                border-radius: 0px;
            }}
            QFrame#OverlayCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #f2f7ff
                );
                border: 1px solid #c7d8f9;
                border-radius: 24px;
            }}
            QFrame#OverlayCard[successMode="true"] {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #f4fdf8
                );
                border: 1px solid #c3edd3;
            }}
            QLabel#OverlayIcon {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #2b61df,
                    stop: 1 #3f86ff
                );
                color: white;
                border-radius: 28px;
                font: 800 16pt "Segoe UI";
            }}
            QLabel#OverlayIcon[successMode="true"] {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #1b8a4f,
                    stop: 1 #22c366
                );
            }}
            QLabel#OverlayTitle {{
                color: {TEXT};
                font: 700 15pt "Segoe UI";
            }}
            QLabel#OverlaySummary {{
                color: {MUTED};
                font: 10pt "Segoe UI";
                line-height: 1.4em;
            }}
            QLabel#OverlaySectionTitle {{
                color: {TEXT};
                font: 700 10pt "Segoe UI";
            }}
            QTextBrowser#OverlayDetails {{
                background: #f8fbff;
                color: #25324a;
                border: 1px solid #dce5f6;
                border-radius: 16px;
                padding: 12px;
                font: 9.5pt "Segoe UI";
                selection-background-color: #dce8ff;
            }}
            QTextBrowser#OverlayDetails a {{
                color: #1f5eff;
                text-decoration: none;
                font-weight: bold;
            }}
            QTextBrowser#OverlayDetails[successMode="true"] {{
                background: #f4fff8;
                border: 1px solid #b7ecd0;
                selection-background-color: #c8f5dc;
            }}
            QPushButton#OverlayPrimaryButton {{
                min-width: 130px;
                padding: 10px 18px;
                border: none;
                border-radius: 14px;
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #1f5eff,
                    stop: 1 #3f86ff
                );
                color: white;
                font: 700 10pt "Segoe UI";
            }}
            QPushButton#OverlayPrimaryButton[successMode="true"] {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #169352,
                    stop: 1 #25d26f
                );
            }}
            QPushButton#OverlayPrimaryButton:hover {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #174ce6,
                    stop: 1 #2b61df
                );
            }}
            QPushButton#OverlayPrimaryButton[successMode="true"]:hover {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #147f47,
                    stop: 1 #1fb861
                );
            }}
            QPushButton#OverlaySecondaryButton {{
                min-width: 100px;
                padding: 10px 18px;
                border-radius: 14px;
                border: 1px solid #c7d8f9;
                background: white;
                color: {TEXT};
                font: 600 10pt "Segoe UI";
            }}
            QPushButton#OverlaySecondaryButton[successMode="true"] {{
                border: 1px solid #bceccf;
                color: #0f6c3a;
            }}
            QPushButton#OverlaySecondaryButton:hover {{
                background: #f0f4fc;
            }}
            QPushButton#OverlaySecondaryButton[successMode="true"]:hover {{
                background: #eaffee;
            }}
            """
        )

    def showEvent(self, event) -> None:
        parent = self.parentWidget()
        if parent is not None:
            self.resize(parent.size())
            self.move(parent.mapToGlobal(QPoint(0, 0)))
        if getattr(self, "_blur_target", None) is not None and self._blur_target.graphicsEffect() is None:
            from PySide6.QtWidgets import QGraphicsBlurEffect
            blur = QGraphicsBlurEffect(self._blur_target)
            blur.setBlurRadius(8.0)
            self._blur_target.setGraphicsEffect(blur)
            self._owns_blur_effect = True
            self._blur_effect = blur
        super().showEvent(event)
        QTimer.singleShot(0, self._start_open_animation)

    def done(self, result: int) -> None:
        if self._processing:
            return
        if self._is_finalizing:
            self._clear_blur()
            super().done(result)
            return
        self._close_result = result
        self._start_close_animation()

    def hideEvent(self, event) -> None:
        self._clear_blur()
        super().hideEvent(event)

    def closeEvent(self, event) -> None:
        if self._processing:
            event.ignore()
            return
        if not self._is_finalizing:
            event.ignore()
            self.done(QDialog.Rejected)
            return
        self._clear_blur()
        super().closeEvent(event)

    def _clear_blur(self) -> None:
        if self._owns_blur_effect and self._blur_target is not None:
            self._blur_target.setGraphicsEffect(None)
            self._owns_blur_effect = False
        self._blur_effect = None

    @staticmethod
    def _scaled_rect(rect: QRect, scale: float) -> QRect:
        width = max(1, int(rect.width() * scale))
        height = max(1, int(rect.height() * scale))
        center = rect.center()
        scaled = QRect(0, 0, width, height)
        scaled.moveCenter(center)
        return scaled

    def _start_open_animation(self) -> None:
        if self._open_animation_started:
            return
        end_rect = self.card.geometry()
        if not end_rect.isValid():
            return
        self._open_animation_started = True
        start_rect = self._scaled_rect(end_rect, 0.985)
        self.card.setGeometry(start_rect)
        self.setWindowOpacity(0.0)
        animation = QParallelAnimationGroup(self)
        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(170)
        opacity_animation.setStartValue(0.0)
        opacity_animation.setEndValue(1.0)
        opacity_animation.setEasingCurve(QEasingCurve.OutCubic)
        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(180)
        geometry_animation.setStartValue(start_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.OutCubic)
        animation.start()
        self._open_animation = animation

    def _start_close_animation(self) -> None:
        if self._is_finalizing:
            return
        current_rect = self.card.geometry()
        if not current_rect.isValid():
            self._finalize_close()
            return
        end_rect = self._scaled_rect(current_rect, 0.99)
        animation = QParallelAnimationGroup(self)
        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(140)
        opacity_animation.setStartValue(self.windowOpacity())
        opacity_animation.setEndValue(0.0)
        opacity_animation.setEasingCurve(QEasingCurve.InCubic)
        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(140)
        geometry_animation.setStartValue(current_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.InCubic)
        animation.finished.connect(self._finalize_close)
        animation.start()
        self._close_animation = animation

    def _finalize_close(self) -> None:
        if self._is_finalizing:
            return
        self._is_finalizing = True
        try:
            self._clear_blur()
            super().done(self._close_result)
        finally:
            self._is_finalizing = False


class UpdateDownloadDialog(QDialog):
    progress_updated = Signal(int, str)
    download_error = Signal(str)
    download_finished = Signal(bool)

    def __init__(self, parent, target_version: str, download_url: str):
        super().__init__(parent)
        self._blur_target = parent.centralWidget() if isinstance(parent, QMainWindow) else parent
        self._owns_blur_effect = False
        self._is_finalizing = False
        self._close_result = QDialog.Rejected
        self._open_animation_started = False
        self._blur_effect: Optional[QGraphicsBlurEffect] = None
        self._processing = True

        self.setModal(True)
        self.setObjectName("UpdateDownloadDialog")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setMinimumSize(680, 480)
        self.setWindowOpacity(0.0)
        
        self.target_version = target_version
        self.download_url = download_url
        self._is_cancelled = False
        self.worker_thread = None

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        overlay = QFrame()
        self.overlay = overlay
        overlay.setObjectName("OverlayBackdrop")
        overlay_layout = QVBoxLayout(overlay)
        overlay_layout.setContentsMargins(0, 0, 0, 0)
        overlay_layout.addStretch(1)

        card = QFrame()
        self.card = card
        card.setObjectName("OverlayCard")
        card.setMinimumWidth(560)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 24, 28, 24)
        card_layout.setSpacing(20)

        header_row = QHBoxLayout()
        header_row.setSpacing(14)
        
        icon_badge = QLabel("⬇️")
        icon_badge.setObjectName("OverlayIcon")
        icon_badge.setAlignment(Qt.AlignCenter)
        icon_badge.setFixedSize(56, 56)

        title_wrap = QVBoxLayout()
        title_wrap.setSpacing(4)
        self.title_label = QLabel(f"Mengunduh pembaruan v{self.target_version}...")
        self.title_label.setObjectName("OverlayTitle")
        
        self.status_label = QLabel("Mempersiapkan unduhan...")
        self.status_label.setObjectName("OverlaySummary")
        self.status_label.setWordWrap(True)

        title_wrap.addWidget(self.title_label)
        title_wrap.addWidget(self.status_label)
        header_row.addWidget(icon_badge, 0, Qt.AlignTop)
        header_row.addLayout(title_wrap, 1)
        card_layout.addLayout(header_row)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setObjectName("OverlayProgress")
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setMinimumHeight(14)
        card_layout.addWidget(self.progress_bar)

        self.progress_text = QLabel("0%")
        self.progress_text.setObjectName("OverlaySectionTitle")
        self.progress_text.setAlignment(Qt.AlignCenter)
        card_layout.addWidget(self.progress_text)
        
        button_row = QHBoxLayout()
        button_row.setSpacing(10)
        button_row.addStretch(1)

        self.cancel_button = QPushButton("Batalkan Unduhan")
        self.cancel_button.setObjectName("OverlaySecondaryButton")
        self.cancel_button.clicked.connect(self._cancel_download)
        button_row.addWidget(self.cancel_button)

        card_layout.addLayout(button_row)
        overlay_layout.addWidget(card, 0, Qt.AlignCenter)
        overlay_layout.addStretch(1)
        outer.addWidget(overlay)

        self._apply_stylesheet()
        
        self.progress_updated.connect(self._update_progress)
        self.download_finished.connect(self._finish_download)
        self.download_error.connect(self._handle_error)

    def _apply_stylesheet(self) -> None:
        self.setStyleSheet(
            f"""
            QDialog#UpdateDownloadDialog {{
                background: transparent;
            }}
            QFrame#OverlayBackdrop {{
                background: rgba(11, 18, 32, 168);
                border-radius: 0px;
            }}
            QFrame#OverlayCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #f2f7ff
                );
                border: 1px solid #c7d8f9;
                border-radius: 24px;
            }}
            QLabel#OverlayIcon {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #2b61df,
                    stop: 1 #3f86ff
                );
                color: white;
                border-radius: 28px;
                font: 800 18pt "Segoe UI";
            }}
            QLabel#OverlayTitle {{
                color: {TEXT};
                font: 700 15pt "Segoe UI";
            }}
            QLabel#OverlaySummary {{
                color: {MUTED};
                font: 10pt "Segoe UI";
                line-height: 1.4em;
            }}
            QLabel#OverlaySectionTitle {{
                color: {PRIMARY};
                font: 800 11pt "Segoe UI";
            }}
            QProgressBar#OverlayProgress {{
                background: #dce5f6;
                border: none;
                border-radius: 7px;
            }}
            QProgressBar#OverlayProgress::chunk {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 0,
                    stop: 0 #2b61df, stop: 1 #3f86ff
                );
                border-radius: 7px;
            }}
            QPushButton#OverlaySecondaryButton {{
                min-width: 140px;
                padding: 10px 18px;
                border-radius: 14px;
                border: 1px solid #f4c2c2;
                background: #fffafa;
                color: #b52a42;
                font: 600 10pt "Segoe UI";
            }}
            QPushButton#OverlaySecondaryButton:hover {{
                background: #ffeeee;
                border: 1px solid #eca1a1;
            }}
            QPushButton#OverlaySecondaryButton:disabled {{
                background: #f0f4fc;
                color: #9ab0d5;
                border: 1px solid #dfe7f4;
            }}
            """
        )

    def showEvent(self, event) -> None:
        parent = self.parentWidget()
        if parent is not None:
            self.resize(parent.size())
            self.move(parent.mapToGlobal(QPoint(0, 0)))
        if getattr(self, "_blur_target", None) is not None and self._blur_target.graphicsEffect() is None:
            from PySide6.QtWidgets import QGraphicsBlurEffect
            blur = QGraphicsBlurEffect(self._blur_target)
            blur.setBlurRadius(8.0)
            self._blur_target.setGraphicsEffect(blur)
            self._owns_blur_effect = True
            self._blur_effect = blur
        super().showEvent(event)
        QTimer.singleShot(0, self._start_open_animation)

    def done(self, result: int) -> None:
        if getattr(self, "_is_finalizing", False):
            self._clear_blur()
            super().done(result)
            return
        self._close_result = result
        self._start_close_animation()

    def hideEvent(self, event) -> None:
        self._clear_blur()
        super().hideEvent(event)

    def closeEvent(self, event) -> None:
        if self._processing and not self._is_cancelled:
            event.ignore()
            return
        if not getattr(self, "_is_finalizing", False):
            event.ignore()
            self.done(QDialog.Rejected)
            return
        self._clear_blur()
        super().closeEvent(event)

    def _clear_blur(self) -> None:
        if getattr(self, "_owns_blur_effect", False) and getattr(self, "_blur_target", None) is not None:
            self._blur_target.setGraphicsEffect(None)
            self._owns_blur_effect = False
        self._blur_effect = None

    @staticmethod
    def _scaled_rect(rect: QRect, scale: float) -> QRect:
        width = max(1, int(rect.width() * scale))
        height = max(1, int(rect.height() * scale))
        center = rect.center()
        scaled = QRect(0, 0, width, height)
        scaled.moveCenter(center)
        return scaled

    def _start_open_animation(self) -> None:
        if getattr(self, "_open_animation_started", False):
            return
        end_rect = self.card.geometry()
        if not end_rect.isValid():
            return
        self._open_animation_started = True
        start_rect = self._scaled_rect(end_rect, 0.985)
        self.card.setGeometry(start_rect)
        self.setWindowOpacity(0.0)
        animation = QParallelAnimationGroup(self)
        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(170)
        opacity_animation.setStartValue(0.0)
        opacity_animation.setEndValue(1.0)
        opacity_animation.setEasingCurve(QEasingCurve.OutCubic)
        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(180)
        geometry_animation.setStartValue(start_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.OutCubic)
        animation.start()
        self._open_animation = animation

    def _start_close_animation(self) -> None:
        if getattr(self, "_is_finalizing", False):
            return
        current_rect = self.card.geometry()
        if not current_rect.isValid():
            self._finalize_close()
            return
        end_rect = self._scaled_rect(current_rect, 0.99)
        animation = QParallelAnimationGroup(self)
        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(140)
        opacity_animation.setStartValue(self.windowOpacity())
        opacity_animation.setEndValue(0.0)
        opacity_animation.setEasingCurve(QEasingCurve.InCubic)
        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(140)
        geometry_animation.setStartValue(current_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.InCubic)
        animation.finished.connect(self._finalize_close)
        animation.start()
        self._close_animation = animation

    def _finalize_close(self) -> None:
        if getattr(self, "_is_finalizing", False):
            return
        self._is_finalizing = True
        try:
            self._clear_blur()
            super().done(self._close_result)
        finally:
            self._is_finalizing = False

    def _cancel_download(self):
        self._is_cancelled = True
        self.status_label.setText("Membatalkan...")
        self.cancel_button.setEnabled(False)
        self.cancel_button.setText("Membatalkan...")
        
    def start_download(self):
        import tempfile
        import sys
        import os
        import urllib.parse
        from uuid import uuid4
        
        parsed_url = urllib.parse.urlparse(self.download_url)
        url_path = urllib.parse.unquote(parsed_url.path)
        original_filename = os.path.basename(url_path)
        
        if not original_filename:
            is_frozen = getattr(sys, 'frozen', False)
            ext = ".exe" if (is_frozen and sys.platform == "win32") else ".py"
            original_filename = f"folder_compare{ext}"

        name, ext = os.path.splitext(original_filename)
        self.dest_path = os.path.join(tempfile.gettempdir(), f"{name}_v{self.target_version}_{uuid4().hex[:8]}{ext}")
            
        def worker_func():
            import urllib.request
            try:
                req = urllib.request.Request(self.download_url, headers={"User-Agent": "FolderCompareApp"})
                with urllib.request.urlopen(req, timeout=15) as response:
                    total_size = int(response.getheader("Content-Length", 0) or 0)
                    downloaded = 0
                    with open(self.dest_path, "wb") as f:
                        while not self._is_cancelled:
                            chunk = response.read(8192)
                            if not chunk:
                                break
                            f.write(chunk)
                            downloaded += len(chunk)
                            pct = int((downloaded / total_size) * 100) if total_size > 0 else 0
                            size_mb = f"{downloaded / 1024 / 1024:.1f} MB"
                            total_mb = f"{total_size / 1024 / 1024:.1f} MB" if total_size > 0 else ""
                            self.progress_updated.emit(pct, f"Terunduh: {size_mb} / {total_mb}")
                            
                if not self._is_cancelled:
                    self.download_finished.emit(True)
                else:
                    self.download_finished.emit(False)
            except Exception as e:
                self.download_error.emit(str(e))

        self.worker_thread = threading.Thread(target=worker_func, daemon=True)
        self.worker_thread.start()

    def _update_progress(self, percentage: int, text: str):
        self.progress_bar.setValue(percentage)
        self.progress_text.setText(f"{percentage}%")
        self.status_label.setText(text)

    def _handle_error(self, err: str):
        self._processing = False
        QMessageBox.warning(self, "Gagal", f"Gagal mengunduh pembaruan:\n\n{err}")
        self.reject()

    def _finish_download(self, success: bool):
        self._processing = False
        if not success:
            if os.path.exists(self.dest_path):
                try: os.remove(self.dest_path)
                except: pass
            self.reject()
            return
            
        self.status_label.setText("Unduhan selesai. Mempersiapkan instalasi...")
        self.progress_bar.setValue(100)
        self.progress_text.setText("100%")
        self.accept()

    def get_downloaded_path(self) -> str:
        return self.dest_path


class ProcessingOverlayDialog(QDialog):
    def __init__(self, parent: Optional[QWidget], title: str, summary: str) -> None:
        super().__init__(parent)
        self._blur_target = parent.centralWidget() if isinstance(parent, QMainWindow) else parent
        self._owns_blur_effect = False

        self.setModal(True)
        self.setObjectName("ProcessingOverlayDialog")
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.setMinimumSize(620, 320)
        self.setWindowOpacity(0.0)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        overlay = QFrame()
        self.overlay = overlay
        overlay.setObjectName("ProcessingOverlayBackdrop")
        overlay_layout = QVBoxLayout(overlay)
        overlay_layout.setContentsMargins(0, 0, 0, 0)
        overlay_layout.addStretch(1)

        card = QFrame()
        self.card = card
        card.setObjectName("ProcessingOverlayCard")
        card.setMinimumWidth(500)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 24, 28, 24)
        card_layout.setSpacing(16)

        title_label = QLabel(title)
        title_label.setObjectName("ProcessingOverlayTitle")
        title_label.setAlignment(Qt.AlignCenter)

        summary_label = QLabel(summary)
        summary_label.setObjectName("ProcessingOverlaySummary")
        summary_label.setAlignment(Qt.AlignCenter)
        summary_label.setWordWrap(True)

        progress = QProgressBar()
        progress.setObjectName("ProcessingOverlayProgress")
        progress.setRange(0, 0)
        progress.setTextVisible(False)

        footnote = QLabel("Mohon tunggu, proses ini bisa memakan waktu tergantung jumlah file.")
        footnote.setObjectName("ProcessingOverlayFootnote")
        footnote.setAlignment(Qt.AlignCenter)
        footnote.setWordWrap(True)

        card_layout.addWidget(title_label)
        card_layout.addWidget(summary_label)
        card_layout.addWidget(progress)
        card_layout.addWidget(footnote)

        overlay_layout.addWidget(card, 0, Qt.AlignCenter)
        overlay_layout.addStretch(1)
        outer.addWidget(overlay)

        self.setStyleSheet(
            f"""
            QDialog#ProcessingOverlayDialog {{
                background: transparent;
            }}
            QFrame#ProcessingOverlayBackdrop {{
                background: rgba(11, 18, 32, 168);
                border-radius: 0px;
            }}
            QFrame#ProcessingOverlayCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #ffffff,
                    stop: 1 #f7fbff
                );
                border: 1px solid #d7e1f4;
                border-radius: 24px;
            }}
            QLabel#ProcessingOverlayTitle {{
                color: {TEXT};
                font: 700 15pt "Segoe UI";
            }}
            QLabel#ProcessingOverlaySummary {{
                color: {MUTED};
                font: 10pt "Segoe UI";
            }}
            QLabel#ProcessingOverlayFootnote {{
                color: {MUTED};
                font: 9.2pt "Segoe UI";
            }}
            QProgressBar#ProcessingOverlayProgress {{
                background: #edf2fb;
                border: none;
                border-radius: 8px;
                min-height: 12px;
                max-height: 12px;
            }}
            QProgressBar#ProcessingOverlayProgress::chunk {{
                background: {PRIMARY};
                border-radius: 8px;
            }}
            """
        )

    def showEvent(self, event) -> None:
        parent = self.parentWidget()
        if parent is not None:
            self.resize(parent.size())
            self.move(parent.mapToGlobal(QPoint(0, 0)))
        if self._blur_target is not None and self._blur_target.graphicsEffect() is None:
            blur = QGraphicsBlurEffect(self._blur_target)
            blur.setBlurRadius(8.0)
            self._blur_target.setGraphicsEffect(blur)
            self._owns_blur_effect = True
        super().showEvent(event)
        QTimer.singleShot(0, self._start_open_animation)

    def closeEvent(self, event) -> None:
        self._clear_blur()
        super().closeEvent(event)

    def hideEvent(self, event) -> None:
        self._clear_blur()
        super().hideEvent(event)

    def _clear_blur(self) -> None:
        if self._owns_blur_effect and self._blur_target is not None:
            self._blur_target.setGraphicsEffect(None)
            self._owns_blur_effect = False

    @staticmethod
    def _scaled_rect(rect: QRect, scale: float) -> QRect:
        width = max(1, int(rect.width() * scale))
        height = max(1, int(rect.height() * scale))
        center = rect.center()
        scaled = QRect(0, 0, width, height)
        scaled.moveCenter(center)
        return scaled

    def _start_open_animation(self) -> None:
        end_rect = self.card.geometry()
        if not end_rect.isValid():
            return

        start_rect = self._scaled_rect(end_rect, 0.985)
        self.card.setGeometry(start_rect)
        self.setWindowOpacity(0.0)

        animation = QParallelAnimationGroup(self)

        opacity_animation = QPropertyAnimation(self, b"windowOpacity", animation)
        opacity_animation.setDuration(140)
        opacity_animation.setStartValue(0.0)
        opacity_animation.setEndValue(1.0)
        opacity_animation.setEasingCurve(QEasingCurve.OutCubic)

        geometry_animation = QPropertyAnimation(self.card, b"geometry", animation)
        geometry_animation.setDuration(150)
        geometry_animation.setStartValue(start_rect)
        geometry_animation.setEndValue(end_rect)
        geometry_animation.setEasingCurve(QEasingCurve.OutCubic)

        animation.start()
        self._open_animation = animation


class SafeApplication(QApplication):
    errorRaised = Signal(str, str, str)

    def __init__(self, argv: List[str]) -> None:
        super().__init__(argv)
        self._error_presenter = None
        self._presenting_error = False
        self.errorRaised.connect(self._dispatch_error)

    def set_error_presenter(self, presenter) -> None:
        self._error_presenter = presenter

    def notify(self, receiver, event) -> bool:
        try:
            return super().notify(receiver, event)
        except Exception as exc:
            sys.stderr.write(
                "".join(traceback.format_exception(type(exc), exc, exc.__traceback__)) + os.linesep
            )
            self.report_error(
                "Terjadi kesalahan pada antarmuka aplikasi",
                str(exc) or exc.__class__.__name__,
                "".join(traceback.format_exception(type(exc), exc, exc.__traceback__)),
            )
            return False

    def report_error(self, title: str, summary: str, details: str) -> None:
        if self._presenting_error:
            sys.stderr.write(f"{title}\n{summary}\n{details}\n")
            return
        self.errorRaised.emit(title, summary, details)

    def _dispatch_error(self, title: str, summary: str, details: str) -> None:
        if self._presenting_error:
            return

        self._presenting_error = True
        try:
            if callable(self._error_presenter):
                self._error_presenter(title, summary, details)
            else:
                QMessageBox.critical(None, title, summary)
        finally:
            self._presenting_error = False


class FolderCompareDeleteApp(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        
        icon_path = Path(__file__).parent / "assets" / "app_icon.png"
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))

        self.resize(1560, 940)
        self.setMinimumSize(1280, 780)

        self.result_rows: List[MatchResult] = []
        self.compare_folder_rows: List[Dict[str, Any]] = []
        self.scan_thread: Optional[threading.Thread] = None
        self.delete_thread: Optional[threading.Thread] = None
        self.ui_queue: queue.Queue[Tuple[str, object]] = queue.Queue()
        self.openpyxl_available = False
        self.current_selected_result: Optional[MatchResult] = None
        self._last_queued_progress = -1.0
        self._last_queued_progress_text = ""
        self._pending_progress: Optional[Tuple[float, str]] = None
        self._pending_scan_results: Optional[List[MatchResult]] = None
        self._awaiting_scan_finalize = False
        self._pending_delete_result: Optional[Dict[str, Any]] = None
        self._pending_transfer_result: Optional[Dict[str, Any]] = None
        self._pending_bulk_sync_result: Optional[Dict[str, Any]] = None
        self.delete_confirm_dialog: Optional[ConfirmOverlayDialog] = None
        self.transfer_confirm_dialog: Optional[ConfirmOverlayDialog] = None
        self.delete_processing_dialog: Optional[ProcessingOverlayDialog] = None
        self.transfer_thread: Optional[threading.Thread] = None
        self.undo_thread: Optional[threading.Thread] = None
        self.undo_processing_dialog: Optional[ProcessingOverlayDialog] = None
        self._pending_undo_result: Optional[Dict[str, Any]] = None
        self._progress_lock = threading.Lock()
        self.history_entries: List[HistoryEntry] = []
        self.history_limit = 200
        self.undo_stack: List[UndoAction] = []
        self.undo_limit = 20
        
        self.app_data_dir = Path.home() / ".folder_compare_app"
        self.app_data_dir.mkdir(parents=True, exist_ok=True)
        self.undo_root = self.app_data_dir / "folder_compare_delete_app_undo"
        self.undo_root.mkdir(parents=True, exist_ok=True)
        self.trash_db_path = self.app_data_dir / "trash_db.json"
        self.trash_entries: List[TrashEntry] = []
        self.sidebar_icon_cache_dir = self.app_data_dir / "folder_compare_delete_sidebar_icons"
        self.file_detail_dialog: Optional[FileDetailOverlayDialog] = None
        
        # Load local database for internal app trash
        self._load_trash_db()

        self.stat_labels: Dict[str, QLabel] = {}
        self.detail_labels: Dict[str, QLabel] = {}

        self._build_ui()
        self._apply_styles()
        self._update_trash_sidebar_badge()
        self.progress_animation = QPropertyAnimation(self.progress_bar, b"value", self)
        self.progress_animation.setEasingCurve(QEasingCurve.OutCubic)
        self.progress_animation.finished.connect(self._on_progress_animation_finished)
        app_instance = QApplication.instance()
        if isinstance(app_instance, SafeApplication):
            app_instance.set_error_presenter(self.show_error_dialog)
        self._check_excel_support()
        self._reset_detail_panel()
        self._refresh_stats()
        self._update_table_empty_state()
        self._update_history_empty_state()
        self._set_progress(0, "Siap untuk scan folder.")

        self.filter_timer = QTimer(self)
        self.filter_timer.setSingleShot(True)
        self.filter_timer.setInterval(180)
        self.filter_timer.timeout.connect(self._apply_debounced_filter)

        self.queue_timer = QTimer(self)
        self.queue_timer.timeout.connect(self._poll_queue)
        self.queue_timer.start(33)

        self.show_only_matches_checkbox.toggled.connect(self._on_filter_changed)
        self.allow_delete_orange_checkbox.toggled.connect(self._on_delete_scope_changed)
        self.allow_delete_red_checkbox.toggled.connect(self._on_delete_scope_changed)
        self._update_delete_action_controls()
        self.add_compare_folder_row()
        self.add_compare_folder_row()

        QTimer.singleShot(500, self._check_update_success)

    def _check_update_success(self) -> None:
        try:
            update_status_file = self.app_data_dir / "update_status.json"
            if update_status_file.exists():
                import json
                with open(update_status_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                update_status_file.unlink()
                
                target_version = data.get("target_version")
                changelog = data.get("changelog")
                if target_version:
                    def parse_version(v):
                        return [int(x) if x.isdigit() else x for x in v.split('.')]
                    
                    if parse_version(APP_VERSION) >= parse_version(target_version):
                        # Panggil modal dialog status yang sudah dimodifikasi atau baru dengan pesan sukses
                        success_dialog = UpdateStatusDialog(self, APP_VERSION, data.get("old_version", ""), changelog, False)
                        
                        # Terapkan styling properti success
                        success_dialog.card.setProperty("successMode", "true")
                        success_dialog.icon_badge.setProperty("successMode", "true")
                        if hasattr(success_dialog, "changelog_view"):
                            success_dialog.changelog_view.setProperty("successMode", "true")
                        if hasattr(success_dialog, "btn_tutup"):
                            success_dialog.btn_tutup.setProperty("successMode", "true")
                            
                        # Muat ulang style QFrame untuk memicu successMode
                        success_dialog.card.style().unpolish(success_dialog.card)
                        success_dialog.card.style().polish(success_dialog.card)
                        
                        success_dialog.icon_badge.setText("✓")
                        success_dialog.title_label.setText(f"Pembaruan Selesai!")
                        success_dialog.summary_label.setText(f"Aplikasi Anda berhasil ditingkatkan ke versi {APP_VERSION}.")
                        success_dialog.exec()
        except Exception as e:
            print(f"Failed to check update success: {e}")

    def closeEvent(self, event) -> None:
        # Trash internal dibuat persisten antar sesi, jadi saat aplikasi ditutup
        # kita hanya menyimpan indeksnya dan tidak membersihkan file-file trash.
        self._save_trash_db()
        super().closeEvent(event)

    def _build_sidebar(self) -> QWidget:
        dock = QWidget()
        dock.setObjectName("SidebarDock")
        dock_layout = QVBoxLayout(dock)
        dock_layout.setContentsMargins(0, 8, 0, 8)
        dock_layout.setSpacing(0)
        dock_layout.addStretch(1)

        sidebar = QFrame()
        sidebar.setObjectName("Sidebar")
        sidebar.setFixedWidth(72)
        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(10, 14, 10, 14)
        layout.setSpacing(12)

        logo = QLabel("FC")
        logo.setObjectName("SidebarLogo")
        logo.setAlignment(Qt.AlignCenter)
        logo.setFixedSize(44, 44)
        layout.addWidget(logo, 0, Qt.AlignHCenter)
        layout.addSpacing(6)

        self.btn_nav_dashboard = QPushButton()
        self.btn_nav_dashboard.setToolTip("Dashboard")
        self.btn_nav_dashboard.setObjectName("SidebarButtonActive")
        self.btn_nav_dashboard.setCursor(Qt.PointingHandCursor)
        self.btn_nav_dashboard.clicked.connect(lambda: self._switch_page(0))
        self.btn_nav_dashboard.setFixedSize(44, 44)

        self.btn_nav_history = QPushButton()
        self.btn_nav_history.setToolTip("Riwayat & Undo")
        self.btn_nav_history.setObjectName("SidebarButton")
        self.btn_nav_history.setCursor(Qt.PointingHandCursor)
        self.btn_nav_history.clicked.connect(lambda: self._switch_page(1))
        self.btn_nav_history.setFixedSize(44, 44)

        self.btn_nav_trash = QPushButton()
        self.btn_nav_trash.setToolTip("Trash Internal")
        self.btn_nav_trash.setObjectName("SidebarButton")
        self.btn_nav_trash.setCursor(Qt.PointingHandCursor)
        self.btn_nav_trash.clicked.connect(lambda: self._switch_page(2))
        self.btn_nav_trash.setFixedSize(44, 44)
        self.trash_nav_host = QWidget()
        self.trash_nav_host.setObjectName("SidebarBadgeHost")
        self.trash_nav_host.setFixedSize(44, 44)
        self.btn_nav_trash.setParent(self.trash_nav_host)
        self.btn_nav_trash.move(0, 0)
        self.trash_nav_badge = QLabel("0", self.trash_nav_host)
        self.trash_nav_badge.setObjectName("SidebarBadge")
        self.trash_nav_badge.setAlignment(Qt.AlignCenter)
        self.trash_nav_badge.setFixedHeight(22)
        self.trash_nav_badge.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        self.trash_nav_badge.hide()
        self.trash_nav_badge.raise_()
        
        self.btn_nav_check_update = QPushButton()
        self.btn_nav_check_update.setToolTip("Periksa Pembaruan")
        self.btn_nav_check_update.setObjectName("SidebarButton")
        self.btn_nav_check_update.setCursor(Qt.PointingHandCursor)
        self.btn_nav_check_update.clicked.connect(self.check_for_updates)
        self.btn_nav_check_update.setFixedSize(44, 44)

        layout.addWidget(self.btn_nav_dashboard, 0, Qt.AlignHCenter)
        layout.addWidget(self.btn_nav_history, 0, Qt.AlignHCenter)
        layout.addWidget(self.trash_nav_host, 0, Qt.AlignHCenter)
        layout.addStretch(1)
        layout.addWidget(self.btn_nav_check_update, 0, Qt.AlignHCenter)

        dock_layout.addWidget(sidebar, 0, Qt.AlignHCenter)
        dock_layout.addStretch(1)

        self._apply_sidebar_icons(active_index=0)
        self._update_trash_sidebar_badge()
        return dock

    def _build_ui(self) -> None:
        central = QWidget()
        self.setCentralWidget(central)
        central.setObjectName("AppShell")

        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(20, 18, 20, 18)
        main_layout.setSpacing(18)

        main_layout.addWidget(self._build_sidebar())

        content_widget = QWidget()
        content_widget.setObjectName("AppContent")
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(16)

        content_layout.addWidget(self._build_header())

        self.main_stack = QStackedWidget()

        self.page_dashboard = QWidget()
        page_dash_layout = QVBoxLayout(self.page_dashboard)
        page_dash_layout.setContentsMargins(0, 0, 0, 0)
        page_dash_layout.setSpacing(16)
        page_dash_layout.addLayout(self._build_stat_cards())

        self.main_splitter = QSplitter(Qt.Horizontal)
        self.main_splitter.setObjectName("MainSplitter")
        self.main_splitter.setChildrenCollapsible(False)
        self.main_splitter.setHandleWidth(10)
        self.main_splitter.addWidget(self._build_left_panel())
        self.main_splitter.addWidget(self._build_right_panel())
        self.main_splitter.setSizes([420, 1080])
        page_dash_layout.addWidget(self.main_splitter, 1)

        self.main_stack.addWidget(self.page_dashboard)

        self.page_history = QWidget()
        page_hist_layout = QVBoxLayout(self.page_history)
        page_hist_layout.setContentsMargins(0, 0, 0, 0)
        page_hist_layout.addWidget(self._build_history_panel(), 1)
        self.main_stack.addWidget(self.page_history)

        self.page_trash = QWidget()
        page_trash_layout = QVBoxLayout(self.page_trash)
        page_trash_layout.setContentsMargins(0, 0, 0, 0)
        
        page_trash_layout.addWidget(self._build_trash_page(), 1)
        self.main_stack.addWidget(self.page_trash)

        content_layout.addWidget(self.main_stack, 1)

        footer = QFrame()
        footer.setObjectName("FooterBar")
        footer_layout = QHBoxLayout(footer)
        footer_layout.setContentsMargins(16, 10, 16, 10)
        footer_layout.setSpacing(12)

        self.status_label = QLabel("Siap untuk scan folder.")
        self.status_label.setObjectName("FooterLabel")
        self.status_label.setWordWrap(False)
        footer_layout.addWidget(self.status_label)

        footer_layout.addStretch(1)

        self.footer_meta_label = QLabel(f"Developed by {APP_DEVELOPER}  |  v{APP_VERSION}")
        self.footer_meta_label.setObjectName("FooterMetaLabel")
        footer_layout.addWidget(self.footer_meta_label, 0, Qt.AlignRight | Qt.AlignVCenter)

        content_layout.addWidget(footer)
        main_layout.addWidget(content_widget, 1)

    def _switch_page(self, index: int) -> None:
        self.main_stack.setCurrentIndex(index)
        
        self.btn_nav_dashboard.setObjectName("SidebarButton")
        self.btn_nav_history.setObjectName("SidebarButton")
        self.btn_nav_trash.setObjectName("SidebarButton")

        if index == 0:
            self.btn_nav_dashboard.setObjectName("SidebarButtonActive")
        elif index == 1:
            self.btn_nav_history.setObjectName("SidebarButtonActive")
        elif index == 2:
            self.btn_nav_trash.setObjectName("SidebarButtonActive")
            
        self._refresh_widget_style(self.btn_nav_dashboard)
        self._refresh_widget_style(self.btn_nav_history)
        self._refresh_widget_style(self.btn_nav_trash)
        self._apply_sidebar_icons(active_index=index)

    def check_for_updates(self) -> None:
        self.status_label.setText("Memeriksa pembaruan...")
        self.btn_nav_check_update.setEnabled(False)

        if not hasattr(self, "update_spinner_timer"):
            self.update_spinner_timer = QTimer(self)
            self.update_spinner_timer.timeout.connect(self._animate_update_icon)
        self.update_angle = 0
        from PySide6.QtGui import QIcon, QPixmap
        icon_path = self._sidebar_icon_variant("sync.svg", "#ffffff")
        self._update_original_pixmap = QIcon(str(icon_path)).pixmap(22, 22)
        self.update_spinner_timer.start(50)

        def worker():
            import urllib.request
            import json
            try:
                url = "https://api.github.com/repos/AlvinPradanaAntony/folder-compare-delete-vibecode/releases/latest"
                req = urllib.request.Request(url, headers={"User-Agent": "FolderCompareApp"})
                with urllib.request.urlopen(req, timeout=10) as response:
                    data = json.loads(response.read().decode("utf-8"))
                    latest_version = data.get("tag_name", "").lstrip("v")
                    release_url = data.get("html_url", "")
                    changelog = data.get("body", "Tidak ada catatan rilis.")
                    assets = data.get("assets", [])
                    download_url = None
                    import platform
                    import sys
                    import os
                    system = platform.system().lower()
                    is_frozen = getattr(sys, 'frozen', False)
                    exe_path = sys.executable if is_frozen else ""
                    
                    is_installed = False
                    if system == "windows" and is_frozen:
                        lower_exe = exe_path.lower()
                        if "program files" in lower_exe or "appdata\\local\\programs" in lower_exe:
                            is_installed = True
                    elif system == "linux" and is_frozen:
                        if exe_path.startswith("/usr/") or exe_path.startswith("/opt/"):
                            is_installed = True
                    elif system == "darwin" and is_frozen:
                        if "/Applications/" in exe_path:
                            is_installed = True

                    best_score = -1
                    for asset in assets:
                        name = asset.get("name", "").lower()
                        score = -1
                        
                        if system == "windows":
                            if is_installed and ("setup" in name or "install" in name or name.endswith(".msi")):
                                score = 100
                            elif not is_installed and ("portable" in name and name.endswith(".exe")):
                                score = 100
                            elif name.endswith(".exe"):
                                score = 50
                        elif system == "darwin":
                            if not is_installed and ("portable" in name or name.endswith(".zip") or name.endswith(".tar.gz")):
                                score = 100
                            elif is_installed and (name.endswith(".dmg") or name.endswith(".pkg")):
                                score = 100
                            elif name.endswith(".dmg"):
                                score = 50
                        elif system == "linux":
                            if not is_installed and ("appimage" in name or "portable" in name):
                                score = 100
                            elif is_installed and (name.endswith(".deb") or name.endswith(".rpm")):
                                score = 100
                            elif name.endswith(".appimage") or name.endswith(".tar.gz"):
                                score = 50
                        elif name.endswith(".py"):
                            score = 10

                        if score > best_score:
                            best_score = score
                            download_url = asset.get("browser_download_url")
                    
                self.ui_queue.put(("update_check_done", {
                    "latest_version": latest_version,
                    "release_url": release_url,
                    "download_url": download_url,
                    "changelog": changelog,
                    "current_version": APP_VERSION
                }))
            except Exception as e:
                self.ui_queue.put(("update_check_error", str(e)))

        threading.Thread(target=worker, daemon=True).start()

    def _animate_update_icon(self) -> None:
        if not hasattr(self, "_update_original_pixmap"):
            return
        self.update_angle = (self.update_angle + 15) % 360
        from PySide6.QtGui import QIcon, QPainter, QPixmap
        from PySide6.QtCore import Qt
        
        size = self._update_original_pixmap.size()
        rotated_pixmap = QPixmap(size)
        rotated_pixmap.fill(Qt.transparent)
        
        painter = QPainter(rotated_pixmap)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setRenderHint(QPainter.SmoothPixmapTransform)
        
        painter.translate(size.width() / 2, size.height() / 2)
        painter.rotate(self.update_angle)
        painter.translate(-size.width() / 2, -size.height() / 2)
        
        painter.drawPixmap(0, 0, self._update_original_pixmap)
        painter.end()
        
        self.btn_nav_check_update.setIcon(QIcon(rotated_pixmap))

    def _build_header(self) -> QWidget:
        header = QFrame()
        header.setObjectName("HeaderCard")
        layout = QHBoxLayout(header)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(24)

        left_wrap = QHBoxLayout()
        left_wrap.setSpacing(18)

        logo_card = QFrame()
        logo_card.setObjectName("HeaderLogoCard")
        logo_card.setFixedSize(84, 84)
        logo_layout = QVBoxLayout(logo_card)
        logo_layout.setContentsMargins(0, 0, 0, 0)
        logo_layout.setSpacing(0)

        logo_mark = QLabel("FC")
        logo_mark.setObjectName("HeaderLogoText")
        logo_mark.setAlignment(Qt.AlignCenter)
        logo_layout.addWidget(logo_mark)

        text_group = QVBoxLayout()
        text_group.setSpacing(6)

        eyebrow = QLabel("FILE OPERATIONS SUITE")
        eyebrow.setObjectName("HeaderEyebrow")

        title = QLabel("Folder Compare Delete")
        title.setObjectName("HeaderTitle")

        header_meta = QLabel(f"Version {APP_VERSION}  |  Developed by {APP_DEVELOPER}")
        header_meta.setObjectName("HeaderMeta")

        subtitle = QLabel(
            "Bandingkan isi beberapa folder, identifikasi duplikat dan perbedaan secara visual, lalu proses file dari Folder A dengan lebih aman."
        )
        subtitle.setObjectName("HeaderSubtitle")
        subtitle.setWordWrap(True)

        meta_row = QHBoxLayout()
        meta_row.setSpacing(8)
        meta_row.addWidget(self._create_header_chip("Multi-folder compare"))
        meta_row.addWidget(self._create_header_chip("Visual diff review"))
        meta_row.addWidget(self._create_header_chip("Safe delete flow"))
        meta_row.addStretch(1)

        text_group.addWidget(eyebrow)
        text_group.addWidget(title)
        text_group.addWidget(header_meta)
        text_group.addWidget(subtitle)
        text_group.addLayout(meta_row)

        left_wrap.addWidget(logo_card, 0, Qt.AlignTop)
        left_wrap.addLayout(text_group, 1)
        layout.addLayout(left_wrap, 1)

        hero = QFrame()
        hero.setObjectName("HeroInfo")
        hero_layout = QVBoxLayout(hero)
        hero_layout.setContentsMargins(18, 16, 18, 16)
        hero_layout.setSpacing(4)

        tip_title = QLabel("Workflow")
        tip_title.setObjectName("HeroInfoTitle")
        tip_body = QLabel(
            "1. Pilih Folder A dan folder pembanding.\n"
            "2. Scan via nama+ukuran atau hash+ukuran.\n"
            "3. Tinjau hasil hijau sebelum hapus."
        )
        tip_body.setObjectName("HeroInfoBody")
        tip_body.setWordWrap(True)

        hero_layout.addWidget(tip_title)
        hero_layout.addWidget(tip_body)
        layout.addWidget(hero, 0)

        return header

    def _create_header_chip(self, text: str) -> QWidget:
        chip = QFrame()
        chip.setObjectName("HeaderChip")
        chip_layout = QHBoxLayout(chip)
        chip_layout.setContentsMargins(10, 6, 10, 6)
        chip_layout.setSpacing(6)

        dot = QFrame()
        dot.setObjectName("HeaderChipDot")
        dot.setFixedSize(8, 8)

        label = QLabel(text)
        label.setObjectName("HeaderChipLabel")

        chip_layout.addWidget(dot, 0, Qt.AlignVCenter)
        chip_layout.addWidget(label, 0, Qt.AlignVCenter)
        return chip

    def _build_stat_cards(self) -> QHBoxLayout:
        layout = QHBoxLayout()
        layout.setSpacing(12)

        cards = [
            ("total", "Total Hasil", CYAN, "Semua file target yang dianalisis"),
            ("exact", "Duplikat Hijau", GREEN, "Aman dipertimbangkan dihapus"),
            ("diff", "Perbedaan Merah", RED, "Nama sama tetapi isi berbeda"),
            ("only", "Hanya di A", ORANGE, "Tidak ditemukan di folder lain"),
        ]
        for key, title, accent, subtitle in cards:
            layout.addWidget(self._create_stat_card(title, accent, subtitle, key))

        return layout

    def _create_stat_card(self, title: str, accent: str, subtitle: str, key: str) -> QWidget:
        card = QFrame()
        card.setObjectName("StatCard")
        background_map = {
            CYAN: (("#1d63f2", "#3b82ff"), "#5d9bff"),
            GREEN: (("#0d9b6b", "#24c78b"), "#5ee1af"),
            RED: (("#d94b6a", "#ff6f8b"), "#ff96ab"),
            ORANGE: (("#d88916", "#ffb23d"), "#ffc96f"),
        }
        (gradient_start, gradient_end), card_border = background_map.get(accent, ((SURFACE, SURFACE_ALT), BORDER))
        card.setStyleSheet(
            f"""
            QFrame#StatCard {{
                background: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 {gradient_start},
                    stop: 1 {gradient_end}
                );
                border: 1px solid {card_border};
                border-radius: 20px;
            }}
            QLabel#StatTitle {{
                color: rgba(255, 255, 255, 0.88);
                font: 700 10pt "Segoe UI";
                background: transparent;
            }}
            QLabel#StatValue {{
                color: #ffffff;
                font: 700 25px "Segoe UI";
                background: transparent;
            }}
            QLabel#StatSubtitle {{
                color: rgba(255, 255, 255, 0.9);
                font: 9pt "Segoe UI";
                background: transparent;
            }}
            """
        )
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(18, 16, 18, 16)
        card_layout.setSpacing(6)

        body = QWidget()
        body_layout = QVBoxLayout(body)
        body_layout.setContentsMargins(0, 0, 0, 0)
        body_layout.setSpacing(6)

        title_label = QLabel(title)
        title_label.setObjectName("StatTitle")
        value_label = QLabel("0")
        value_label.setObjectName("StatValue")
        subtitle_label = QLabel(subtitle)
        subtitle_label.setObjectName("StatSubtitle")
        subtitle_label.setWordWrap(True)

        self.stat_labels[key] = value_label

        body_layout.addWidget(title_label)
        body_layout.addWidget(value_label)
        body_layout.addWidget(subtitle_label)

        card_layout.addWidget(body)
        return card

    def _build_left_panel(self) -> QWidget:
        container = QFrame()
        container.setObjectName("SurfaceCard")
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setObjectName("PanelScroll")
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.viewport().setObjectName("PanelViewport")

        body = QWidget()
        body.setObjectName("PanelScrollBody")
        layout = QVBoxLayout(body)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(14)

        heading = QLabel("Konfigurasi Scan")
        heading.setObjectName("SectionTitle")
        description = QLabel("Atur folder target, folder pembanding, mode pencocokan, dan aksi hasil scan.")
        description.setObjectName("SectionSubtitle")
        description.setWordWrap(True)
        layout.addWidget(heading)
        layout.addWidget(description)

        layout.addWidget(self._build_folder_input("Folder A (target hapus)", is_target=True))
        layout.addWidget(self._build_compare_group())
        layout.addWidget(self._build_option_group())
        layout.addWidget(self._build_action_group())
        layout.addStretch(1)

        scroll.setWidget(body)
        container_layout.addWidget(scroll)
        return container

    def _build_folder_input(self, label_text: str, is_target: bool = False) -> QWidget:
        card = QFrame()
        card.setObjectName("SubCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        label = QLabel(label_text)
        label.setObjectName("FieldLabel")

        row = QHBoxLayout()
        row.setSpacing(10)

        line_edit = FolderPathLineEdit("Pilih atau drop folder...")
        line_edit.setClearButtonEnabled(True)
        browse_button = QPushButton("Pilih Folder")
        browse_button.setObjectName("GhostButton")
        browse_button.clicked.connect(lambda: self.pick_folder(line_edit))
        line_edit.folderDropped.connect(lambda path: self.status_label.setText(f"Folder didrop: {path}"))

        status_label = self._create_path_status_label()
        self._bind_path_field(line_edit, status_label)

        row.addWidget(line_edit, 1)
        row.addWidget(browse_button, 0)

        layout.addWidget(label)
        layout.addLayout(row)
        layout.addWidget(status_label)

        if is_target:
            self.target_folder_edit = line_edit
            self.target_folder_status_label = status_label

        return card

    def _build_compare_group(self) -> QWidget:
        card = QFrame()
        card.setObjectName("SubCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        title = QLabel("Folder Pembanding")
        title.setObjectName("FieldLabel")
        description = QLabel("Tambahkan satu atau lebih folder untuk dibandingkan terhadap Folder A.")
        description.setObjectName("MutedText")
        description.setWordWrap(True)

        self.compare_list_widget = QWidget()
        self.compare_list_layout = QVBoxLayout(self.compare_list_widget)
        self.compare_list_layout.setContentsMargins(0, 0, 0, 0)
        self.compare_list_layout.setSpacing(8)

        buttons = QHBoxLayout()
        buttons.setSpacing(8)

        add_button = QPushButton("+ Tambah Folder")
        add_button.setObjectName("GhostButton")
        add_button.clicked.connect(self.add_compare_folder_row)

        remove_button = QPushButton("- Hapus Terakhir")
        remove_button.setObjectName("OutlineButton")
        remove_button.clicked.connect(self.remove_compare_folder_row)

        buttons.addWidget(add_button)
        buttons.addWidget(remove_button)
        buttons.addStretch(1)

        layout.addWidget(title)
        layout.addWidget(description)
        layout.addWidget(self.compare_list_widget)
        layout.addLayout(buttons)
        return card

    def _build_option_group(self) -> QWidget:
        wrapper = QWidget()
        layout = QVBoxLayout(wrapper)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        method_card = QFrame()
        method_card.setObjectName("SubCard")
        method_layout = QVBoxLayout(method_card)
        method_layout.setContentsMargins(14, 14, 14, 14)
        method_layout.setSpacing(8)

        method_title = QLabel("Metode Pencocokan")
        method_title.setObjectName("FieldLabel")

        self.compare_mode_group = QButtonGroup(self)
        self.compare_mode_name_size = QRadioButton("Nama file + ukuran (cepat)")
        self.compare_mode_hash = QRadioButton("Hash SHA-256 + ukuran (akurat)")
        self.compare_mode_name_size.setChecked(True)
        self.compare_mode_group.addButton(self.compare_mode_name_size)
        self.compare_mode_group.addButton(self.compare_mode_hash)

        method_layout.addWidget(method_title)
        method_layout.addWidget(self.compare_mode_name_size)
        method_layout.addWidget(self.compare_mode_hash)

        extra_card = QFrame()
        extra_card.setObjectName("SubCard")
        extra_layout = QVBoxLayout(extra_card)
        extra_layout.setContentsMargins(14, 14, 14, 14)
        extra_layout.setSpacing(8)

        extra_title = QLabel("Opsi Tampilan dan Hapus")
        extra_title.setObjectName("FieldLabel")

        self.include_subfolders_checkbox = QCheckBox("Sertakan subfolder")
        self.include_subfolders_checkbox.setChecked(True)
        self.show_only_matches_checkbox = QCheckBox("Tampilkan hanya file dengan kecocokan / perbedaan")
        self.delete_mode_internal_trash = QRadioButton("Hapus ke Trash Internal")
        self.delete_mode_permanent = QRadioButton("Hapus permanen")
        self.delete_mode_internal_trash.setChecked(True)
        self.allow_delete_orange_checkbox = QCheckBox("Izinkan hapus hasil oranye (hanya di Folder A)")
        self.allow_delete_red_checkbox = QCheckBox("Izinkan hapus hasil merah (nama sama, isi berbeda)")
        self.delete_scope_hint_label = QLabel("Default aman: hanya hasil hijau yang dapat dihapus.")
        self.delete_scope_hint_label.setObjectName("MutedText")
        self.delete_scope_hint_label.setWordWrap(True)

        self.delete_mode_group = QButtonGroup(self)
        self.delete_mode_group.addButton(self.delete_mode_internal_trash)
        self.delete_mode_group.addButton(self.delete_mode_permanent)

        extra_layout.addWidget(extra_title)
        extra_layout.addWidget(self.include_subfolders_checkbox)
        extra_layout.addWidget(self.show_only_matches_checkbox)
        extra_layout.addSpacing(6)
        extra_layout.addWidget(self.delete_mode_internal_trash)
        extra_layout.addWidget(self.delete_mode_permanent)
        extra_layout.addSpacing(6)
        extra_layout.addWidget(self.allow_delete_orange_checkbox)
        extra_layout.addWidget(self.allow_delete_red_checkbox)
        extra_layout.addWidget(self.delete_scope_hint_label)

        layout.addWidget(method_card)
        layout.addWidget(extra_card)
        return wrapper

    def _build_action_group(self) -> QWidget:
        card = QFrame()
        card.setObjectName("SubCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        self.scan_button = QPushButton("Scan dan Bandingkan")
        self.scan_button.setObjectName("PrimaryButton")
        self.scan_button.clicked.connect(self.start_scan)

        first_row = QHBoxLayout()
        first_row.setSpacing(8)
        self.export_csv_button = QPushButton("CSV")
        self.export_csv_button.setObjectName("GhostButton")
        self.export_csv_button.clicked.connect(self.export_csv)

        self.export_excel_button = QPushButton("Excel")
        self.export_excel_button.setObjectName("GhostButton")
        self.export_excel_button.clicked.connect(self.export_excel)

        self.clear_button = QPushButton("Reset")
        self.clear_button.setObjectName("OutlineButton")
        self.clear_button.clicked.connect(self.clear_results)

        first_row.addWidget(self.export_csv_button)
        first_row.addWidget(self.export_excel_button)
        first_row.addWidget(self.clear_button)

        second_row = QHBoxLayout()
        second_row.setSpacing(8)

        self.delete_button = QPushButton("Hapus Terpilih")
        self.delete_button.setObjectName("DangerButton")
        self.delete_button.clicked.connect(self.delete_selected)

        self.delete_all_button = QPushButton("Hapus Semua Hijau")
        self.delete_all_button.setObjectName("DangerButton")
        self.delete_all_button.clicked.connect(self.delete_all_results)

        second_row.addWidget(self.delete_button)
        second_row.addWidget(self.delete_all_button)

        third_row = QHBoxLayout()
        third_row.setSpacing(8)

        self.copy_button = QPushButton("Salin Terpilih")
        self.copy_button.setObjectName("GhostButton")
        self.copy_button.clicked.connect(lambda: self.transfer_selected_files("copy"))

        self.move_button = QPushButton("Pindah Terpilih")
        self.move_button.setObjectName("OutlineButton")
        self.move_button.clicked.connect(lambda: self.transfer_selected_files("move"))

        third_row.addWidget(self.copy_button)
        third_row.addWidget(self.move_button)

        layout.addWidget(self.scan_button)
        layout.addLayout(first_row)
        layout.addLayout(second_row)
        layout.addLayout(third_row)

        self.export_csv_button.setEnabled(False)
        self.export_excel_button.setEnabled(False)
        self.delete_button.setEnabled(False)
        self.delete_all_button.setEnabled(False)
        self.copy_button.setEnabled(False)
        self.move_button.setEnabled(False)
        return card

    def _build_right_panel(self) -> QWidget:
        container = QFrame()
        container.setObjectName("SurfaceCard")
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setObjectName("PanelScroll")
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.viewport().setObjectName("PanelViewport")

        body = QWidget()
        body.setObjectName("PanelScrollBody")
        body_layout = QVBoxLayout(body)
        body_layout.setContentsMargins(18, 18, 18, 18)
        body_layout.setSpacing(12)

        heading = QLabel("Visual Diff Result")
        heading.setObjectName("SectionTitle")

        body_layout.addWidget(heading)
        body_layout.addWidget(self._build_navigation_card())
        self.progress_card = self._build_progress_card()
        self.progress_card.setVisible(False)
        body_layout.addWidget(self.progress_card)
        body_layout.addWidget(self._build_results_table())
        self.detail_panel = self._build_detail_panel()
        self.detail_panel.setVisible(False)
        body_layout.addStretch(1)

        scroll.setWidget(body)
        layout.addWidget(scroll)
        return container

    def _build_progress_card(self) -> QWidget:
        card = QFrame()
        card.setObjectName("SubCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        top = QHBoxLayout()
        top.setSpacing(8)

        label = QLabel("Progress Scan")
        label.setObjectName("FieldLabel")
        self.progress_badge = QLabel("0%")
        self.progress_badge.setObjectName("ProgressBadge")
        self.progress_badge.setAlignment(Qt.AlignCenter)

        top.addWidget(label)
        top.addStretch(1)
        top.addWidget(self.progress_badge)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 1000)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(False)

        layout.addLayout(top)
        layout.addWidget(self.progress_bar)
        return card

    def _build_navigation_card(self) -> QWidget:
        card = QFrame()
        card.setObjectName("SubCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        top_row = QHBoxLayout()
        top_row.setSpacing(10)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Cari status, path, relative path, atau folder...")
        self.search_input.setClearButtonEnabled(True)
        self.search_input.textChanged.connect(self._schedule_search_filter)

        self.reset_filter_button = QPushButton("Reset Filter")
        self.reset_filter_button.setObjectName("OutlineButton")
        self.reset_filter_button.clicked.connect(self._reset_quick_filters)

        top_row.addWidget(self.search_input, 1)
        top_row.addWidget(self.reset_filter_button)

        chip_row = QHBoxLayout()
        chip_row.setSpacing(8)

        self.quick_filter_group = QButtonGroup(self)
        self.quick_filter_group.setExclusive(True)
        self.quick_filter_buttons: Dict[str, QPushButton] = {}

        chip_specs = [
            ("all", "Semua", CYAN),
            ("exact_match", "Duplikat", GREEN),
            ("different_content", "Berbeda", RED),
            ("only_target", "Hanya di A", ORANGE),
        ]

        for key, label, color in chip_specs:
            button = QPushButton(label)
            button.setCheckable(True)
            button.setObjectName("FilterChip")
            button.setProperty("chipKey", key)
            button.setProperty("chipColor", color)
            button.clicked.connect(self._on_filter_changed)
            self.quick_filter_group.addButton(button)
            self.quick_filter_buttons[key] = button
            chip_row.addWidget(button)

        self.quick_filter_buttons["all"].setChecked(True)
        chip_row.addStretch(1)

        self.sync_button = QPushButton("Sync")
        self.sync_button.setObjectName("PrimaryButton")
        sync_icon_path = self._sidebar_icon_variant("sync.svg", "#ffffff")
        if sync_icon_path.exists():
            self.sync_button.setIcon(QIcon(str(sync_icon_path)))
        self.sync_button.clicked.connect(self.sync_selected_green)
        self.sync_button.setEnabled(False)
        chip_row.addWidget(self.sync_button)

        layout.addLayout(top_row)
        layout.addLayout(chip_row)
        return card

    def _build_results_table(self) -> QWidget:
        card = QFrame()
        card.setObjectName("TableCard")
        card.setMinimumHeight(360)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(12, 12, 12, 12)

        self.table_stack_host = QWidget()
        self.table_stack = QStackedLayout(self.table_stack_host)
        self.table_stack.setContentsMargins(0, 0, 0, 0)
        self.table_stack.setStackingMode(QStackedLayout.StackOne)

        self.results_table = ResponsiveTableWidget()
        self.table_model = MatchResultTableModel()
        self.table_proxy = MatchResultFilterProxyModel()
        self.table_proxy.setSourceModel(self.table_model)
        self.results_table.setModel(self.table_proxy)
        self.results_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.results_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.results_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.results_table.setAlternatingRowColors(False)
        self.results_table.setWordWrap(False)
        self.results_table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.results_table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.results_table.setTextElideMode(Qt.ElideMiddle)
        self.results_table.setShowGrid(True)
        self.results_table.verticalHeader().setVisible(False)
        self.results_table.verticalHeader().setDefaultSectionSize(34)
        self.results_table.setSortingEnabled(True)
        self.results_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.results_table.customContextMenuRequested.connect(self._show_results_table_context_menu)

        header = self.results_table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setMinimumSectionSize(90)
        header.setSectionsMovable(False)
        header.setSortIndicatorShown(True)
        for column in range(self.table_model.columnCount()):
            header.setSectionResizeMode(column, QHeaderView.Interactive)

        self._apply_default_table_widths()
        self.results_table.sortByColumn(0, Qt.AscendingOrder)

        selection_model = self.results_table.selectionModel()
        selection_model.selectionChanged.connect(self._on_table_selection_changed)
        selection_model.currentRowChanged.connect(self._on_current_row_changed)
        self.results_table.doubleClicked.connect(self._open_detail_dialog_from_index)
        self.results_table.focusReleased.connect(self._on_table_focus_released)

        self.table_empty_state = self._build_table_empty_state()

        self.table_stack.addWidget(self.results_table)
        self.table_stack.addWidget(self.table_empty_state)
        layout.addWidget(self.table_stack_host)
        return card

    def _build_table_empty_state(self) -> QWidget:
        empty = QFrame()
        empty.setObjectName("TableEmptyState")
        layout = QVBoxLayout(empty)
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(10)
        layout.setAlignment(Qt.AlignCenter)

        icon = QLabel("[]")
        icon.setObjectName("EmptyStateIcon")
        icon.setAlignment(Qt.AlignCenter)

        self.empty_state_title = QLabel("Belum ada hasil scan")
        self.empty_state_title.setObjectName("EmptyStateTitle")
        self.empty_state_title.setAlignment(Qt.AlignCenter)

        self.empty_state_description = QLabel(
            "Pilih folder target dan folder pembanding, lalu jalankan Scan dan Bandingkan untuk melihat hasil."
        )
        self.empty_state_description.setObjectName("EmptyStateDescription")
        self.empty_state_description.setAlignment(Qt.AlignCenter)
        self.empty_state_description.setWordWrap(True)

        layout.addWidget(icon)
        layout.addWidget(self.empty_state_title)
        layout.addWidget(self.empty_state_description)
        return empty

    def _build_detail_panel(self) -> QWidget:
        card = QFrame()
        card.setObjectName("SubCard")
        card.setMinimumHeight(320)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(12)

        top = QHBoxLayout()
        top.setSpacing(8)
        title = QLabel("Detail File Terpilih")
        title.setObjectName("FieldLabel")
        copy_button = QPushButton("Copy Path")
        copy_button.setObjectName("GhostButton")
        copy_button.clicked.connect(self.copy_selected_path)

        top.addWidget(title)
        top.addStretch(1)
        top.addWidget(copy_button)

        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(12)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)
        grid.setRowStretch(4, 1)

        fields = [
            ("status", "Status"),
            ("target", "Path Target"),
            ("relative", "Relative Path"),
            ("size", "Ukuran"),
            ("found", "Ditemukan di"),
            ("missing", "Tidak Ada di"),
            ("mode", "Mode"),
            ("exact", "Path Cocok"),
            ("diff", "Path Beda"),
        ]
        positions = {
            "status": (0, 0),
            "target": (1, 0),
            "relative": (2, 0),
            "size": (3, 0),
            "found": (4, 0),
            "missing": (0, 1),
            "mode": (1, 1),
            "exact": (2, 1),
            "diff": (3, 1),
        }

        for key, title_text in fields:
            if key not in positions:
                continue
            row, column = positions[key]
            grid.addWidget(self._create_detail_card(key, title_text), row, column)

        self.missing_compare_suggestion = QFrame()
        self.missing_compare_suggestion.setObjectName("SuggestionCard")
        self.missing_compare_suggestion.setVisible(False)
        suggestion_layout = QHBoxLayout(self.missing_compare_suggestion)
        suggestion_layout.setContentsMargins(14, 12, 14, 12)
        suggestion_layout.setSpacing(12)

        self.missing_compare_suggestion_label = QLabel("")
        self.missing_compare_suggestion_label.setObjectName("SuggestionText")
        self.missing_compare_suggestion_label.setWordWrap(True)

        self.missing_compare_copy_button = QPushButton("Salin ke Folder Pembanding")
        self.missing_compare_copy_button.setObjectName("GhostButton")
        self.missing_compare_copy_button.clicked.connect(self.copy_to_compare_folders)

        self.missing_compare_move_button = QPushButton("Pindah ke Folder Pembanding")
        self.missing_compare_move_button.setObjectName("OutlineButton")
        self.missing_compare_move_button.clicked.connect(self.move_to_compare_folders)

        suggestion_layout.addWidget(self.missing_compare_suggestion_label, 1)
        suggestion_layout.addWidget(self.missing_compare_copy_button, 0)
        suggestion_layout.addWidget(self.missing_compare_move_button, 0)

        layout.addLayout(top)
        layout.addLayout(grid)
        layout.addWidget(self.missing_compare_suggestion)
        return card

    def _build_history_panel(self) -> QWidget:
        card = QFrame()
        card.setObjectName("SurfaceCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(18)

        hero = QFrame()
        hero.setObjectName("HistoryHeroCard")
        hero_layout = QVBoxLayout(hero)
        hero_layout.setContentsMargins(20, 18, 20, 18)
        hero_layout.setSpacing(14)

        hero_top = QHBoxLayout()
        hero_top.setSpacing(16)

        hero_mark = QLabel("HS")
        hero_mark.setObjectName("HistoryHeroMark")
        hero_mark.setAlignment(Qt.AlignCenter)
        hero_mark.setFixedSize(58, 58)

        hero_text = QVBoxLayout()
        hero_text.setSpacing(5)

        title = QLabel("Riwayat Aksi")
        title.setObjectName("SectionTitle")

        description = QLabel(
            "Semua aksi penting aplikasi dicatat di sini, termasuk scan, hapus, salin, pindah, export, restore, dan status hasilnya."
        )
        description.setObjectName("SectionSubtitle")
        description.setWordWrap(True)

        self.history_summary_label = QLabel("Belum ada aktivitas yang tercatat.")
        self.history_summary_label.setObjectName("HistoryHeroSummary")
        self.history_summary_label.setWordWrap(True)

        hero_text.addWidget(title)
        hero_text.addWidget(description)
        hero_text.addWidget(self.history_summary_label)

        hero_top.addWidget(hero_mark, 0, Qt.AlignTop)
        hero_top.addLayout(hero_text, 1)
        hero_layout.addLayout(hero_top)

        metrics_row = QHBoxLayout()
        metrics_row.setSpacing(12)

        self.history_total_value = QLabel("0")
        self.history_success_value = QLabel("0")
        self.history_issue_value = QLabel("0")
        self.history_undo_value = QLabel("0 siap")
        metrics_row.addWidget(self._create_history_metric_card("Total Aksi", self.history_total_value, "Semua event yang tersimpan"))
        metrics_row.addWidget(self._create_history_metric_card("Aksi Sukses", self.history_success_value, "Operasi yang selesai baik"))
        metrics_row.addWidget(self._create_history_metric_card("Perlu Perhatian", self.history_issue_value, "Warning, sebagian gagal, atau error"))
        metrics_row.addWidget(self._create_history_metric_card("Undo", self.history_undo_value, "Aksi terakhir yang masih bisa dibatalkan"))

        hero_layout.addLayout(metrics_row)
        layout.addWidget(hero)

        table_card = QFrame()
        table_card.setObjectName("SubCard")
        table_layout = QVBoxLayout(table_card)
        table_layout.setContentsMargins(16, 16, 16, 16)
        table_layout.setSpacing(14)

        top = QHBoxLayout()
        top.setSpacing(10)

        title_wrap = QVBoxLayout()
        title_wrap.setSpacing(4)

        section_title = QLabel("Timeline Aktivitas")
        section_title.setObjectName("FieldLabel")
        section_subtitle = QLabel("Tinjau aksi terbaru, statusnya, dan detail proses yang telah dijalankan aplikasi.")
        section_subtitle.setObjectName("MutedText")
        section_subtitle.setWordWrap(True)

        self.history_count_badge = QLabel("0 aksi")
        self.history_count_badge.setObjectName("HistoryCountBadge")
        self.history_count_badge.setAlignment(Qt.AlignCenter)

        self.undo_button = QPushButton("Undo Terakhir")
        self.undo_button.setObjectName("GhostButton")
        self.undo_button.clicked.connect(self.undo_last_action)
        self.undo_button.setEnabled(False)

        self.clear_history_button = QPushButton("Bersihkan")
        self.clear_history_button.setObjectName("OutlineButton")
        self.clear_history_button.clicked.connect(self.clear_history)
        self.clear_history_button.setEnabled(False)

        title_wrap.addWidget(section_title)
        title_wrap.addWidget(section_subtitle)

        top.addLayout(title_wrap, 1)
        top.addStretch(1)
        top.addWidget(self.history_count_badge)
        top.addWidget(self.undo_button)
        top.addWidget(self.clear_history_button)

        self.history_stack_host = QWidget()
        self.history_stack = QStackedLayout(self.history_stack_host)
        self.history_stack.setContentsMargins(0, 0, 0, 0)
        self.history_stack.setStackingMode(QStackedLayout.StackOne)

        self.history_table = QTableView()
        self.history_table.setObjectName("HistoryTable")
        self.history_model = HistoryTableModel()
        self.history_table.setModel(self.history_model)
        self.history_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.history_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.history_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.history_table.setAlternatingRowColors(False)
        self.history_table.setWordWrap(True)
        self.history_table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.history_table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.history_table.setTextElideMode(Qt.ElideRight)
        self.history_table.setShowGrid(True)
        self.history_table.setSortingEnabled(False)
        self.history_table.verticalHeader().setVisible(False)
        self.history_table.verticalHeader().setDefaultSectionSize(34)
        self.history_table.setMinimumHeight(220)

        self._configure_history_table_columns()

        self.history_empty_state = self._build_history_empty_state()
        self.history_stack.addWidget(self.history_table)
        self.history_stack.addWidget(self.history_empty_state)

        table_layout.addLayout(top)
        table_layout.addWidget(self.history_stack_host, 1)
        layout.addWidget(table_card, 1)
        return card

    def _create_history_metric_card(self, title: str, value_label: QLabel, subtitle: str) -> QWidget:
        card = QFrame()
        card.setObjectName("HistoryMetricCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setSpacing(4)

        title_label = QLabel(title)
        title_label.setObjectName("HistoryMetricTitle")
        value_label.setObjectName("HistoryMetricValue")
        subtitle_label = QLabel(subtitle)
        subtitle_label.setObjectName("HistoryMetricSubtitle")
        subtitle_label.setWordWrap(True)

        layout.addWidget(title_label)
        layout.addWidget(value_label)
        layout.addWidget(subtitle_label)
        return card

    def _configure_history_table_columns(self) -> None:
        if not hasattr(self, "history_table"):
            return

        history_header = self.history_table.horizontalHeader()
        history_header.setStretchLastSection(False)
        history_header.setMinimumSectionSize(90)
        history_header.setSectionsMovable(False)

        for column in range(self.history_model.columnCount()):
            history_header.setSectionResizeMode(column, QHeaderView.Interactive)

        # Kolom detail dibuat fleksibel agar tabel selalu mengisi lebar parent.
        history_header.setSectionResizeMode(4, QHeaderView.Stretch)

        self.history_table.setColumnWidth(0, 145)
        self.history_table.setColumnWidth(1, 130)
        self.history_table.setColumnWidth(2, 200)
        self.history_table.setColumnWidth(3, 90)

    def _build_trash_page(self) -> QWidget:
        trash_card = QFrame()
        trash_card.setObjectName("SurfaceCard")
        trash_layout = QVBoxLayout(trash_card)
        trash_layout.setContentsMargins(24, 24, 24, 24)
        trash_layout.setSpacing(18)

        hero = QFrame()
        hero.setObjectName("TrashHeroCard")
        hero_layout = QVBoxLayout(hero)
        hero_layout.setContentsMargins(20, 18, 20, 18)
        hero_layout.setSpacing(14)

        hero_top = QHBoxLayout()
        hero_top.setSpacing(16)

        hero_mark = QLabel("TR")
        hero_mark.setObjectName("TrashHeroMark")
        hero_mark.setAlignment(Qt.AlignCenter)
        hero_mark.setFixedSize(58, 58)

        hero_text = QVBoxLayout()
        hero_text.setSpacing(5)

        title = QLabel("Trash Internal Aplikasi")
        title.setObjectName("SectionTitle")
        description = QLabel(
            "File yang dihapus dalam mode Trash Internal dipindahkan ke penyimpanan aman aplikasi. Anda dapat memulihkan file kapan saja atau menghapusnya permanen jika sudah tidak diperlukan."
        )
        description.setObjectName("SectionSubtitle")
        description.setWordWrap(True)

        self.trash_summary_label = QLabel("Trash masih kosong dan siap dipakai.")
        self.trash_summary_label.setObjectName("TrashHeroSummary")
        self.trash_summary_label.setWordWrap(True)

        hero_text.addWidget(title)
        hero_text.addWidget(description)
        hero_text.addWidget(self.trash_summary_label)

        hero_top.addWidget(hero_mark, 0, Qt.AlignTop)
        hero_top.addLayout(hero_text, 1)
        hero_layout.addLayout(hero_top)

        metrics_row = QHBoxLayout()
        metrics_row.setSpacing(12)

        self.trash_count_value = QLabel("0")
        self.trash_size_value = QLabel("0 B")
        self.trash_selection_value = QLabel("0 dipilih")
        metrics_row.addWidget(self._create_trash_metric_card("Item Trash", self.trash_count_value, "Jumlah file yang tersimpan"))
        metrics_row.addWidget(self._create_trash_metric_card("Total Ukuran", self.trash_size_value, "Estimasi ruang yang dipakai"))
        metrics_row.addWidget(self._create_trash_metric_card("Pilihan Aktif", self.trash_selection_value, "File yang akan diproses"))

        hero_layout.addLayout(metrics_row)
        trash_layout.addWidget(hero)

        table_card = QFrame()
        table_card.setObjectName("SubCard")
        table_layout = QVBoxLayout(table_card)
        table_layout.setContentsMargins(16, 16, 16, 16)
        table_layout.setSpacing(14)

        top_row = QHBoxLayout()
        top_row.setSpacing(10)

        top_title_wrap = QVBoxLayout()
        top_title_wrap.setSpacing(4)
        table_title = QLabel("Daftar File Trash")
        table_title.setObjectName("FieldLabel")
        table_subtitle = QLabel("Pilih file yang ingin dipulihkan kembali atau dihapus permanen dari Trash Internal.")
        table_subtitle.setObjectName("MutedText")
        table_subtitle.setWordWrap(True)
        top_title_wrap.addWidget(table_title)
        top_title_wrap.addWidget(table_subtitle)

        actions_wrap = QHBoxLayout()
        actions_wrap.setSpacing(10)

        self.trash_restore_selected_button = QPushButton("Pulihkan Terpilih")
        self.trash_restore_selected_button.setObjectName("GhostButton")
        self.trash_restore_selected_button.clicked.connect(self.restore_selected_trash_entries)

        self.trash_delete_selected_button = QPushButton("Hapus Permanen Terpilih")
        self.trash_delete_selected_button.setObjectName("DangerButton")
        self.trash_delete_selected_button.clicked.connect(self.delete_selected_trash_entries_permanently)

        self.trash_delete_all_button = QPushButton("Kosongkan Trash")
        self.trash_delete_all_button.setObjectName("DangerButton")
        self.trash_delete_all_button.clicked.connect(self.delete_all_trash_entries_permanently)

        actions_wrap.addWidget(self.trash_restore_selected_button)
        actions_wrap.addWidget(self.trash_delete_selected_button)
        actions_wrap.addWidget(self.trash_delete_all_button)

        top_row.addLayout(top_title_wrap, 1)
        top_row.addLayout(actions_wrap)

        self.trash_table = QTableWidget(0, 6)
        self.trash_table.setObjectName("TrashTable")
        self.trash_table.setHorizontalHeaderLabels(["Pilih", "Nama File", "Path Asal", "Waktu Hapus", "Ukuran", "Undo"])
        self.trash_table.setSelectionMode(QAbstractItemView.NoSelection)
        self.trash_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.trash_table.setAlternatingRowColors(False)
        self.trash_table.setWordWrap(False)
        self.trash_table.setTextElideMode(Qt.ElideMiddle)
        self.trash_table.setShowGrid(True)
        self.trash_table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.trash_table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.trash_table.verticalHeader().setVisible(False)
        self.trash_table.verticalHeader().setDefaultSectionSize(52)
        self.trash_table.setMinimumHeight(300)
        self._configure_trash_table_columns(apply_default_widths=True)
        self.trash_table.cellClicked.connect(self._toggle_trash_row_check)
        self.trash_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.trash_table.customContextMenuRequested.connect(self._show_trash_table_context_menu)

        self.trash_stack_host = QWidget()
        self.trash_stack = QStackedLayout(self.trash_stack_host)
        self.trash_stack.setContentsMargins(0, 0, 0, 0)
        self.trash_stack.setStackingMode(QStackedLayout.StackOne)

        self.trash_empty_state = self._build_trash_empty_state()
        self.trash_stack.addWidget(self.trash_table)
        self.trash_stack.addWidget(self.trash_empty_state)

        table_layout.addLayout(top_row)
        table_layout.addWidget(self.trash_stack_host, 1)
        trash_layout.addWidget(table_card, 1)

        self._refresh_trash_page()
        return trash_card

    def _configure_trash_table_columns(self, apply_default_widths: bool = False) -> None:
        if not hasattr(self, "trash_table"):
            return

        trash_header = self.trash_table.horizontalHeader()
        trash_header.setStretchLastSection(False)
        trash_header.setMinimumSectionSize(72)
        trash_header.setSectionsMovable(False)

        for column in range(self.trash_table.columnCount()):
            trash_header.setSectionResizeMode(column, QHeaderView.Interactive)

        # Path asal menyerap sisa ruang agar tabel responsif terhadap parent.
        trash_header.setSectionResizeMode(2, QHeaderView.Stretch)

        if apply_default_widths:
            default_widths = {
                0: 56,
                1: 220,
                3: 170,
                4: 100,
                5: 116,
            }
            for column, width in default_widths.items():
                self.trash_table.setColumnWidth(column, width)

    def _create_trash_undo_button(self, entry_id: str) -> QWidget:
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(8, 6, 8, 6)
        layout.setSpacing(0)

        restore_button = QPushButton("Undo")
        restore_button.setObjectName("TableGhostButton")
        restore_button.setCursor(Qt.PointingHandCursor)
        restore_button.setMinimumWidth(84)
        restore_button.setMinimumHeight(30)
        restore_button.clicked.connect(lambda _checked=False, target_entry_id=entry_id: self.restore_trash_entries([target_entry_id]))

        layout.addWidget(restore_button, 0, Qt.AlignCenter)
        return container

    def _create_trash_checkbox(self, entry_id: str, checked: bool = False) -> QWidget:
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        checkbox = QCheckBox()
        checkbox.setCursor(Qt.PointingHandCursor)
        checkbox.setChecked(checked)
        checkbox.setProperty("trashEntryId", entry_id)
        checkbox.toggled.connect(lambda _checked=False: self._update_trash_selection_state())

        layout.addWidget(checkbox, 0, Qt.AlignCenter)
        return container

    def _create_trash_metric_card(self, title: str, value_label: QLabel, subtitle: str) -> QWidget:
        card = QFrame()
        card.setObjectName("TrashMetricCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setSpacing(4)

        title_label = QLabel(title)
        title_label.setObjectName("TrashMetricTitle")
        value_label.setObjectName("TrashMetricValue")
        subtitle_label = QLabel(subtitle)
        subtitle_label.setObjectName("TrashMetricSubtitle")
        subtitle_label.setWordWrap(True)

        layout.addWidget(title_label)
        layout.addWidget(value_label)
        layout.addWidget(subtitle_label)
        return card

    def _build_trash_empty_state(self) -> QWidget:
        empty = QFrame()
        empty.setObjectName("TrashEmptyState")
        layout = QVBoxLayout(empty)
        layout.setContentsMargins(24, 30, 24, 30)
        layout.setSpacing(10)
        layout.setAlignment(Qt.AlignCenter)

        icon = QLabel("[]")
        icon.setObjectName("TrashEmptyIcon")
        icon.setAlignment(Qt.AlignCenter)

        title = QLabel("Trash Internal Masih Kosong")
        title.setObjectName("TrashEmptyTitle")
        title.setAlignment(Qt.AlignCenter)

        description = QLabel(
            "Saat Anda menghapus file dengan mode Trash Internal, file akan muncul di halaman ini untuk dipulihkan atau dihapus permanen."
        )
        description.setObjectName("TrashEmptyDescription")
        description.setAlignment(Qt.AlignCenter)
        description.setWordWrap(True)

        hint = QLabel("Tip: gunakan mode Trash Internal jika Anda ingin aman untuk undo dan restore antar sesi.")
        hint.setObjectName("TrashEmptyHint")
        hint.setAlignment(Qt.AlignCenter)
        hint.setWordWrap(True)

        layout.addWidget(icon)
        layout.addWidget(title)
        layout.addWidget(description)
        layout.addWidget(hint)
        return empty

    def _build_history_empty_state(self) -> QWidget:
        empty = QFrame()
        empty.setObjectName("HistoryEmptyState")
        layout = QVBoxLayout(empty)
        layout.setContentsMargins(24, 30, 24, 30)
        layout.setSpacing(10)
        layout.setAlignment(Qt.AlignCenter)

        icon = QLabel("::")
        icon.setObjectName("HistoryEmptyIcon")
        icon.setAlignment(Qt.AlignCenter)

        title = QLabel("Belum ada riwayat aksi")
        title.setObjectName("HistoryEmptyTitle")
        title.setAlignment(Qt.AlignCenter)

        description = QLabel(
            "Riwayat scan, hapus, salin, pindah, export, dan aksi penting lainnya akan muncul di sini."
        )
        description.setObjectName("HistoryEmptyDescription")
        description.setAlignment(Qt.AlignCenter)
        description.setWordWrap(True)

        hint = QLabel("Mulai dari scan folder atau aksi file lain untuk membangun timeline aktivitas di halaman ini.")
        hint.setObjectName("HistoryEmptyHint")
        hint.setAlignment(Qt.AlignCenter)
        hint.setWordWrap(True)

        layout.addWidget(icon)
        layout.addWidget(title)
        layout.addWidget(description)
        layout.addWidget(hint)
        return empty

    def _create_detail_card(self, key: str, title: str) -> QWidget:
        card = QFrame()
        card.setObjectName("DetailCard")
        card.setMinimumHeight(78)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(6)

        title_label = QLabel(title)
        title_label.setObjectName("DetailTitle")
        value_label = QLabel("-")
        value_label.setObjectName("DetailValue")
        value_label.setWordWrap(True)
        value_label.setMinimumHeight(22)
        value_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        layout.addWidget(title_label)
        layout.addWidget(value_label)

        self.detail_labels[key] = value_label
        return card

    def _create_path_status_label(self) -> QLabel:
        label = QLabel("Belum ada folder dipilih.")
        label.setObjectName("PathStatus")
        label.setWordWrap(True)
        return label

    def _bind_path_field(self, line_edit: QLineEdit, status_label: QLabel) -> None:
        line_edit.textChanged.connect(
            lambda _text, edit=line_edit, status=status_label: self._update_path_field_state(edit, status)
        )
        self._update_path_field_state(line_edit, status_label)

    def _normalize_folder_path(self, raw_path: str) -> str:
        cleaned = raw_path.strip().strip('"')
        if not cleaned:
            return ""
        return os.path.normpath(str(Path(cleaned).expanduser()))

    def _update_path_field_state(self, line_edit: QLineEdit, status_label: QLabel) -> None:
        raw_value = line_edit.text()
        normalized = self._normalize_folder_path(raw_value)

        if not normalized:
            line_edit.setProperty("pathState", "empty")
            status_label.setProperty("pathState", "empty")
            status_label.setText("Belum ada folder dipilih.")
        elif Path(normalized).is_dir():
            line_edit.setProperty("pathState", "valid")
            status_label.setProperty("pathState", "valid")
            status_label.setText(f"Folder valid: {normalized}")
        else:
            line_edit.setProperty("pathState", "invalid")
            status_label.setProperty("pathState", "invalid")
            status_label.setText(f"Folder belum valid / tidak ditemukan: {normalized}")

        line_edit.setToolTip(normalized or "Belum ada folder dipilih")
        status_label.setToolTip(status_label.text())
        self._refresh_widget_style(line_edit)
        self._refresh_widget_style(status_label)

    def _refresh_widget_style(self, widget: QWidget) -> None:
        style = widget.style()
        style.unpolish(widget)
        style.polish(widget)
        widget.update()

    def _stylesheet_url(self, path: Path) -> str:
        return path.resolve().as_posix()

    def _asset_path(self, name: str) -> Path:
        return Path(__file__).resolve().parent / "assets" / name

    def _sidebar_icon_variant(self, source_name: str, color: str) -> Path:
        source_path = self._asset_path(source_name)
        self.sidebar_icon_cache_dir.mkdir(parents=True, exist_ok=True)
        variant_path = self.sidebar_icon_cache_dir / f"{source_path.stem}_{color.strip('#')}.svg"
        try:
            svg_text = source_path.read_text(encoding="utf-8")
            tinted_text = svg_text.replace("currentColor", color)
            if not variant_path.exists() or variant_path.read_text(encoding="utf-8") != tinted_text:
                variant_path.write_text(tinted_text, encoding="utf-8")
        except Exception:
            return source_path
        return variant_path

    def _apply_sidebar_icons(self, active_index: int) -> None:
        icon_specs = [
            (self.btn_nav_dashboard, "dashboard.svg", active_index == 0),
            (self.btn_nav_history, "history.svg", active_index == 1),
            (self.btn_nav_trash, "trash-can.svg", active_index == 2),
            (self.btn_nav_check_update, "sync.svg", False),
        ]
        for button, source_name, is_active in icon_specs:
            color = "#ffffff" if is_active else "#8ea0c3"
            icon_path = self._sidebar_icon_variant(source_name, color)
            button.setIcon(QIcon(str(icon_path)))
            button.setIconSize(QSize(18, 18))

    def _apply_styles(self) -> None:
        checkbox_check_url = self._stylesheet_url(self._asset_path("checkbox_check.svg"))
        sort_up_url = self._stylesheet_url(self._asset_path("sort_up.svg"))
        sort_down_url = self._stylesheet_url(self._asset_path("sort_down.svg"))
        self.setStyleSheet(
            f"""
            QMainWindow {{
                background: {BG_COLOR};
                color: {TEXT};
            }}
            QWidget#AppShell {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #f4f7fb, stop:1 #eef4ff);
            }}
            QWidget#AppContent {{
                background: transparent;
            }}
            QLabel {{
                color: {TEXT};
            }}
            QFrame#HeaderCard {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #173fbb, stop:0.55 #2b61df, stop:1 #3f86ff);
                border-radius: 24px;
            }}
            QFrame#HeaderLogoCard {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 rgba(255,255,255,0.18), stop:1 rgba(255,255,255,0.08));
                border: 1px solid rgba(255, 255, 255, 0.18);
                border-radius: 22px;
            }}
            QLabel#HeaderLogoText {{
                color: white;
                font: 800 24pt "Segoe UI";
                letter-spacing: 1px;
            }}
            QLabel#HeaderEyebrow {{
                color: rgba(236, 243, 255, 0.76);
                font: 700 8.5pt "Segoe UI";
                letter-spacing: 1.6px;
            }}
            QLabel#HeaderTitle {{
                color: white;
                font: 700 24pt "Segoe UI Semibold";
                letter-spacing: 0.3px;
            }}
            QLabel#HeaderMeta {{
                color: rgba(236, 243, 255, 0.88);
                font: 600 9.5pt "Segoe UI";
                letter-spacing: 0.2px;
            }}
            QLabel#HeaderSubtitle {{
                color: rgba(242, 246, 255, 0.92);
                font: 10.5pt "Segoe UI";
                line-height: 1.35;
            }}
            QFrame#HeroInfo {{
                background: rgba(255, 255, 255, 0.12);
                border: 1px solid rgba(255, 255, 255, 0.18);
                border-radius: 18px;
                min-width: 290px;
            }}
            QLabel#HeroInfoTitle {{
                color: white;
                font: 700 11.5pt "Segoe UI Semibold";
            }}
            QLabel#HeroInfoBody {{
                color: rgba(241, 246, 255, 0.95);
                font: 9.5pt "Segoe UI";
            }}
            QFrame#HeaderChip {{
                background: rgba(255, 255, 255, 0.12);
                border: 1px solid rgba(255, 255, 255, 0.16);
                border-radius: 12px;
            }}
            QFrame#HeaderChipDot {{
                background: #b9d4ff;
                border-radius: 4px;
            }}
            QLabel#HeaderChipLabel {{
                color: rgba(248, 250, 255, 0.92);
                font: 700 8.8pt "Segoe UI";
            }}
            QFrame#SurfaceCard, QFrame#StatCard, QFrame#SubCard, QFrame#TableCard {{
                background: {SURFACE};
                border: 1px solid {BORDER};
                border-radius: 20px;
            }}
            QWidget#PanelScrollBody {{
                background: #f8fbff;
                border-radius: 20px;
            }}
            QWidget#PanelViewport {{
                background: #f8fbff;
                border-radius: 20px;
            }}
            QScrollArea#PanelScroll {{
                background: #f8fbff;
                border: none;
                border-radius: 20px;
            }}
            QSplitter#MainSplitter::handle {{
                background: transparent;
            }}
            QSplitter#MainSplitter::handle:hover {{
                background: #dde6f6;
                border-radius: 4px;
            }}
            QLabel#SectionTitle {{
                font: 700 14pt "Segoe UI";
                color: {TEXT};
            }}
            QLabel#SectionSubtitle {{
                color: {MUTED};
                font: 9.5pt "Segoe UI";
            }}
            QFrame#TrashHeroCard {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #f8fbff, stop:1 #eef5ff);
                border: 1px solid #d7e5fb;
                border-radius: 22px;
            }}
            QFrame#HistoryHeroCard {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #f8fbff, stop:1 #eef4ff);
                border: 1px solid #d7e3fa;
                border-radius: 22px;
            }}
            QLabel#HistoryHeroMark {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #173a78, stop:1 #2b64e8);
                color: white;
                border: 1px solid rgba(255, 255, 255, 0.22);
                border-radius: 18px;
                font: 700 16pt "Segoe UI";
            }}
            QLabel#HistoryHeroSummary {{
                color: #35527a;
                font: 600 9.5pt "Segoe UI";
            }}
            QLabel#TrashHeroMark {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #2f66f1, stop:1 #5d8cff);
                color: white;
                border: 1px solid rgba(255, 255, 255, 0.28);
                border-radius: 18px;
                font: 700 18pt "Segoe UI";
            }}
            QLabel#TrashHeroSummary {{
                color: #31547f;
                font: 600 9.5pt "Segoe UI";
            }}
            QFrame#TrashMetricCard {{
                background: rgba(255, 255, 255, 0.82);
                border: 1px solid #dce7f8;
                border-radius: 16px;
            }}
            QFrame#HistoryMetricCard {{
                background: rgba(255, 255, 255, 0.84);
                border: 1px solid #dce5f6;
                border-radius: 16px;
            }}
            QLabel#HistoryMetricTitle {{
                color: {MUTED};
                font: 700 8.8pt "Segoe UI";
                text-transform: uppercase;
            }}
            QLabel#HistoryMetricValue {{
                color: {TEXT};
                font: 700 17pt "Segoe UI Semibold";
            }}
            QLabel#HistoryMetricSubtitle {{
                color: #6e7f9a;
                font: 8.8pt "Segoe UI";
            }}
            QLabel#TrashMetricTitle {{
                color: {MUTED};
                font: 700 8.8pt "Segoe UI";
                text-transform: uppercase;
            }}
            QLabel#TrashMetricValue {{
                color: {TEXT};
                font: 700 17pt "Segoe UI Semibold";
            }}
            QLabel#TrashMetricSubtitle {{
                color: #6e7f9a;
                font: 8.8pt "Segoe UI";
            }}
            QLabel#FieldLabel {{
                font: 700 10pt "Segoe UI";
                color: {TEXT};
            }}
            QLabel#MutedText {{
                font: 9pt "Segoe UI";
                color: {MUTED};
            }}
            QLabel#StatTitle {{
                font: 700 10pt "Segoe UI";
                color: {MUTED};
            }}
            QLabel#StatValue {{
                font: 700 25px "Segoe UI";
                color: {TEXT};
            }}
            QLabel#StatSubtitle {{
                font: 9pt "Segoe UI";
                color: {MUTED};
            }}
            QLineEdit {{
                background: white;
                color: {TEXT};
                border: 1px solid {BORDER};
                border-radius: 12px;
                padding: 10px 12px;
                selection-background-color: #cfe0ff;
                min-height: 20px;
            }}
            QLineEdit[pathState="valid"] {{
                border: 1px solid #a8e2c4;
                background: #fbfffd;
            }}
            QLineEdit[pathState="invalid"] {{
                border: 1px solid #ffc3ce;
                background: #fffafb;
            }}
            QLineEdit:focus {{
                border: 1px solid {PRIMARY};
            }}
            QLineEdit::placeholder {{
                color: #8d98ab;
            }}
            QPushButton {{
                border-radius: 12px;
                padding: 10px 14px;
                font: 600 9.5pt "Segoe UI";
            }}
            QPushButton#PrimaryButton {{
                background: {PRIMARY};
                color: white;
                border: none;
            }}
            QPushButton#PrimaryButton:hover {{
                background: {PRIMARY_DARK};
            }}
            QPushButton#GhostButton {{
                background: #edf3ff;
                color: {TEXT};
                border: 1px solid #dce6fb;
            }}
            QPushButton#GhostButton:hover {{
                background: #e2ecff;
            }}
            QPushButton#TableGhostButton {{
                background: #edf3ff;
                color: {TEXT};
                border: 1px solid #dce6fb;
                border-radius: 10px;
                padding: 6px 12px;
                font: 600 9pt "Segoe UI";
            }}
            QPushButton#TableGhostButton:hover {{
                background: #e2ecff;
            }}
            QPushButton#OutlineButton {{
                background: white;
                color: {TEXT};
                border: 1px solid {BORDER};
            }}
            QPushButton#OutlineButton:hover {{
                background: {SURFACE_ALT};
            }}
            QPushButton#FilterChip {{
                background: #f6f9ff;
                color: {TEXT};
                border: 1px solid #d8e2f4;
                border-radius: 16px;
                padding: 8px 14px;
                font: 700 9pt "Segoe UI";
            }}
            QPushButton#FilterChip:hover {{
                background: #eef4ff;
                border: 1px solid #c6d6f4;
            }}
            QPushButton#FilterChip:checked {{
                background: #e9f1ff;
                color: {PRIMARY};
                border: 1px solid #aecaef;
            }}
            QPushButton#FilterChip[chipKey="all"] {{
                background: #f6f9ff;
                color: #315cc9;
                border: 1px solid #cfdcf6;
            }}
            QPushButton#FilterChip[chipKey="all"]:hover {{
                background: #edf3ff;
                border: 1px solid #bdd0f4;
            }}
            QPushButton#FilterChip[chipKey="all"]:checked {{
                background: #e7efff;
                color: #1f5eff;
                border: 1px solid #9fbdf0;
            }}
            QPushButton#FilterChip[chipKey="exact_match"] {{
                background: #f4fcf8;
                color: #1f7a58;
                border: 1px solid #cdebdc;
            }}
            QPushButton#FilterChip[chipKey="exact_match"]:hover {{
                background: #ebfaf2;
                border: 1px solid #b4e2cc;
            }}
            QPushButton#FilterChip[chipKey="exact_match"]:checked {{
                background: #dff7ea;
                color: #156746;
                border: 1px solid #8fd0b0;
            }}
            QPushButton#FilterChip[chipKey="different_content"] {{
                background: #fff5f7;
                color: #b23a50;
                border: 1px solid #f4ccd5;
            }}
            QPushButton#FilterChip[chipKey="different_content"]:hover {{
                background: #ffedf1;
                border: 1px solid #efb8c4;
            }}
            QPushButton#FilterChip[chipKey="different_content"]:checked {{
                background: #ffe2e9;
                color: #982c42;
                border: 1px solid #e79aaa;
            }}
            QPushButton#FilterChip[chipKey="only_target"] {{
                background: #fff8ef;
                color: #b56b18;
                border: 1px solid #f4dcc0;
            }}
            QPushButton#FilterChip[chipKey="only_target"]:hover {{
                background: #fff1df;
                border: 1px solid #efcba1;
            }}
            QPushButton#FilterChip[chipKey="only_target"]:checked {{
                background: #ffe8c8;
                color: #99540c;
                border: 1px solid #e5b06d;
            }}
            QPushButton#DangerButton {{
                background: #fff0f3;
                color: #9d263a;
                border: 1px solid #ffd2da;
            }}
            QPushButton#DangerButton:hover {{
                background: #ffe5eb;
            }}
            QFrame#SuggestionCard {{
                background: #f4fbf7;
                border: 1px solid #d6ecdf;
                border-radius: 16px;
            }}
            QLabel#SuggestionText {{
                color: #23583f;
                font: 9.5pt "Segoe UI";
                line-height: 1.35;
            }}
            QPushButton:disabled {{
                background: #f0f3f8;
                color: #99a4b8;
                border: 1px solid #e2e7f0;
            }}
            QRadioButton, QCheckBox {{
                spacing: 8px;
                color: {TEXT};
                font: 9.5pt "Segoe UI";
            }}
            QRadioButton::indicator, QCheckBox::indicator {{
                width: 18px;
                height: 18px;
                background: #f7faff;
                border: 1px solid #c7d5ec;
            }}
            QCheckBox::indicator {{
                border-radius: 5px;
            }}
            QRadioButton::indicator {{
                border-radius: 9px;
            }}
            QCheckBox::indicator:hover, QRadioButton::indicator:hover {{
                border: 1px solid #9fb6dc;
                background: #eef4ff;
            }}
            QCheckBox::indicator:checked, QRadioButton::indicator:checked {{
                background: #2d64f1;
                border: 1px solid #2d64f1;
            }}
            QCheckBox::indicator:unchecked, QRadioButton::indicator:unchecked {{
                background: #f9fbff;
                border: 1px solid #c7d5ec;
            }}
            QCheckBox::indicator:disabled, QRadioButton::indicator:disabled {{
                background: #eef2f8;
                border: 1px solid #d9e1ee;
            }}
            QCheckBox::indicator:checked {{
                image: url({checkbox_check_url});
            }}
            QRadioButton::indicator:checked {{
                background: qradialgradient(
                    cx: 0.5, cy: 0.5, radius: 0.7,
                    fx: 0.5, fy: 0.5,
                    stop: 0 #2d64f1,
                    stop: 0.45 #2d64f1,
                    stop: 0.46 #f7faff,
                    stop: 1 #f7faff
                );
                border: 1px solid #2d64f1;
            }}
            QWidget#SidebarDock {{
                background: transparent;
            }}
            QFrame#Sidebar {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #15233f, stop:1 #0e1930);
                border: 1px solid rgba(130, 155, 208, 0.18);
                border-radius: 24px;
            }}
            QLabel#SidebarLogo {{
                color: #ffffff;
                font: 900 13pt "Segoe UI";
                background: qlineargradient(x1: 0, y1: 0, x2: 1, y2: 1, stop: 0 #3f86ff, stop: 1 #1f5eff);
                border: 1px solid rgba(255, 255, 255, 0.16);
                border-radius: 14px;
                padding: 4px;
            }}
            QLabel#SidebarHint {{
                color: rgba(214, 226, 248, 0.58);
                font: 700 7pt "Segoe UI";
                letter-spacing: 1.4px;
            }}
            QWidget#SidebarBadgeHost {{
                background: transparent;
            }}
            QPushButton#SidebarButton {{
                background: rgba(255, 255, 255, 0.04);
                border: 1px solid transparent;
                border-radius: 14px;
                color: #cbd5e1;
                padding: 0;
            }}
            QPushButton#SidebarButton:hover {{
                background: rgba(255, 255, 255, 0.09);
                border: 1px solid rgba(173, 194, 232, 0.14);
            }}
            QPushButton#SidebarButtonActive {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #2aa8d8, stop:1 #1f5eff);
                border: 1px solid rgba(255, 255, 255, 0.2);
                border-radius: 14px;
                color: #ffffff;
                padding: 0;
            }}
            QLabel#SidebarBadge {{
                background: #f51212;
                color: #ffffff;
                border: 2px solid #13213b;
                border-radius: 11px;
                font: 800 8pt "Segoe UI";
                padding: 0 2px;
                padding-bottom: 2px;
            }}
            QProgressBar {{
                background: #edf2fb;
                border: none;
                border-radius: 6px;
                min-height: 8px;
                max-height: 8px;
            }}
            QProgressBar::chunk {{
                background: {PRIMARY};
                border-radius: 6px;
            }}
            QLabel#ProgressBadge {{
                background: #edf3ff;
                color: {PRIMARY};
                border-radius: 12px;
                font: 700 9pt "Segoe UI";
                padding: 4px 10px;
                min-width: 48px;
            }}
            QFrame#LegendChip {{
                background: #f7f9fc;
                border: 1px solid {BORDER};
                border-radius: 12px;
            }}
            QLabel#LegendLabel {{
                font: 700 9pt "Segoe UI";
                color: {TEXT};
            }}
            QTableView {{
                background: white;
                border: 1px solid {BORDER};
                border-radius: 16px;
                gridline-color: #edf1f7;
                color: {TEXT};
                selection-background-color: transparent;
                selection-color: {TEXT};
                font: 9pt "Segoe UI";
                outline: 0;
            }}
            QTableView::item {{
                padding: 4px 6px;
            }}
            QTableView::item:selected:active {{
                background: #dce8ff;
                color: {TEXT};
            }}
            QTableView::item:selected:!active {{
                background: transparent;
                color: {TEXT};
            }}
            QHeaderView::section {{
                background: #f0f4fb;
                color: {TEXT};
                padding: 10px 18px 10px 8px;
                border: none;
                border-right: 1px solid #e5ebf5;
                border-bottom: 1px solid #e5ebf5;
                font: 700 9pt "Segoe UI";
            }}
            QHeaderView::up-arrow {{
                image: url({sort_up_url});
                width: 12px;
                height: 12px;
            }}
            QHeaderView::down-arrow {{
                image: url({sort_down_url});
                width: 12px;
                height: 12px;
            }}
            QFrame#DetailCard {{
                background: #f8fafc;
                border: 1px solid {BORDER};
                border-radius: 14px;
            }}
            QFrame#TableEmptyState {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #fbfdff, stop:1 #f4f8ff);
                border: 1px dashed #cdd9ee;
                border-radius: 16px;
            }}
            QFrame#HistoryEmptyState {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #fbfdff, stop:1 #f7faff);
                border: 1px dashed #d4def1;
                border-radius: 16px;
            }}
            QLabel#HistoryEmptyIcon {{
                color: #8ca7d7;
                font: 700 22pt "Consolas";
                letter-spacing: 3px;
            }}
            QLabel#EmptyStateIcon {{
                color: #8eabdd;
                font: 700 22pt "Consolas";
                letter-spacing: 2px;
            }}
            QLabel#EmptyStateTitle {{
                color: {TEXT};
                font: 700 13pt "Segoe UI Semibold";
            }}
            QLabel#EmptyStateDescription {{
                color: {MUTED};
                font: 9.5pt "Segoe UI";
            }}
            QLabel#HistoryEmptyTitle {{
                color: {TEXT};
                font: 700 12pt "Segoe UI Semibold";
            }}
            QLabel#HistoryEmptyDescription {{
                color: {MUTED};
                font: 9.2pt "Segoe UI";
            }}
            QLabel#HistoryEmptyHint {{
                color: #5c77ab;
                font: 600 9pt "Segoe UI";
            }}
            QFrame#TrashEmptyState {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #fbfdff, stop:1 #f2f7ff);
                border: 1px dashed #d5e1f4;
                border-radius: 18px;
            }}
            QLabel#TrashEmptyIcon {{
                color: #8aa8dc;
                font: 700 22pt "Consolas";
                letter-spacing: 3px;
            }}
            QLabel#TrashEmptyTitle {{
                color: {TEXT};
                font: 700 13pt "Segoe UI Semibold";
            }}
            QLabel#TrashEmptyDescription {{
                color: {MUTED};
                font: 9.5pt "Segoe UI";
            }}
            QLabel#TrashEmptyHint {{
                color: #5b78b0;
                font: 600 9pt "Segoe UI";
            }}
            QLabel#DetailTitle {{
                color: {MUTED};
                font: 700 9pt "Segoe UI";
            }}
            QLabel#DetailValue {{
                color: {TEXT};
                font: 9pt "Segoe UI";
            }}
            QLabel#HistoryCountBadge {{
                background: #edf3ff;
                color: {PRIMARY};
                border-radius: 12px;
                font: 700 9pt "Segoe UI";
                padding: 4px 10px;
                min-width: 58px;
            }}
            QTableView#HistoryTable {{
                border-radius: 14px;
            }}
            QTableWidget#TrashTable {{
                background: #fbfdff;
                color: {TEXT};
                border: 1px solid #dbe4f4;
                border-radius: 16px;
                gridline-color: #e8eef8;
                font: 9pt "Segoe UI";
            }}
            QTableWidget#TrashTable::item {{
                padding: 6px 8px;
            }}
            QTableWidget#TrashTable QTableCornerButton::section {{
                background: #f2f6fd;
                border: none;
                border-right: 1px solid #e3eaf7;
                border-bottom: 1px solid #e3eaf7;
            }}
            QTableWidget#TrashTable QHeaderView::section {{
                background: #f2f6fd;
                color: {TEXT};
                padding: 11px 16px 11px 8px;
                border: none;
                border-right: 1px solid #e3eaf7;
                border-bottom: 1px solid #e3eaf7;
                font: 700 9pt "Segoe UI";
            }}
            QLabel#PathStatus {{
                font: 8.8pt "Segoe UI";
                color: {MUTED};
            }}
            QLabel#PathStatus[pathState="valid"] {{
                color: #1c8f5b;
            }}
            QLabel#PathStatus[pathState="invalid"] {{
                color: #b24154;
            }}
            QFrame#FooterBar {{
                background: #182235;
                border-radius: 16px;
            }}
            QLabel#FooterLabel {{
                color: #edf2ff;
                font: 9.5pt "Segoe UI";
            }}
            QLabel#FooterMetaLabel {{
                color: rgba(237, 242, 255, 0.72);
                font: 600 9pt "Segoe UI";
            }}
            QScrollArea {{
                background: transparent;
                border: none;
            }}
            QScrollArea#PanelScroll {{
                background: #f8fbff;
                border: none;
            }}
            QAbstractScrollArea {{
                background: #f8fbff;
                border: none;
            }}
            QScrollBar:vertical {{
                background: transparent;
                width: 12px;
                margin: 4px 0 4px 0;
            }}
            QScrollBar::handle:vertical {{
                background: #c5d2e8;
                min-height: 36px;
                border-radius: 6px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: #aebdd9;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
                background: transparent;
            }}
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                background: transparent;
            }}
            QScrollBar:horizontal {{
                background: transparent;
                height: 12px;
                margin: 0 4px 0 4px;
            }}
            QScrollBar::handle:horizontal {{
                background: #c5d2e8;
                min-width: 36px;
                border-radius: 6px;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: #aebdd9;
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
                background: transparent;
            }}
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{
                background: transparent;
            }}
            QMessageBox {{
                background: #ffffff;
                padding: 12px 14px 10px 14px;
            }}
            QMessageBox QWidget {{
                background: #ffffff;
                color: {TEXT};
            }}
            QMessageBox QLabel {{
                background: transparent;
                color: {TEXT};
                font: 9.5pt "Segoe UI";
                padding: 0px;
            }}
            QMessageBox QPushButton {{
                min-width: 92px;
                padding: 8px 16px;
                border-radius: 12px;
                border: 1px solid #d8e2f4;
                background: #edf3ff;
                color: {TEXT};
                font: 600 9.5pt "Segoe UI";
            }}
            QMessageBox QPushButton:hover {{
                background: #e2ecff;
                border: 1px solid #c6d6f4;
            }}
            """
        )
        self.status_label.setObjectName("FooterLabel")

    def _check_excel_support(self) -> None:
        try:
            import openpyxl  # noqa: F401
            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False
        self.export_excel_button.setEnabled(False)

    def add_compare_folder_row(self) -> None:
        row_index = len(self.compare_folder_rows) + 1

        row_widget = QFrame()
        row_layout = QVBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 0, 0, 0)
        row_layout.setSpacing(6)

        top_row = QHBoxLayout()
        top_row.setContentsMargins(0, 0, 0, 0)
        top_row.setSpacing(8)

        label = QLabel(f"Folder {row_index}")
        label.setObjectName("FieldLabel")
        label.setMinimumWidth(72)

        line_edit = FolderPathLineEdit("Pilih atau drop folder pembanding...")
        line_edit.setClearButtonEnabled(True)

        browse_button = QPushButton("Browse")
        browse_button.setObjectName("OutlineButton")
        browse_button.clicked.connect(lambda: self.pick_folder(line_edit))
        line_edit.folderDropped.connect(lambda path: self.status_label.setText(f"Folder didrop: {path}"))

        status_label = self._create_path_status_label()
        self._bind_path_field(line_edit, status_label)

        top_row.addWidget(label)
        top_row.addWidget(line_edit, 1)
        top_row.addWidget(browse_button)

        row_layout.addLayout(top_row)
        row_layout.addWidget(status_label)

        self.compare_folder_rows.append(
            {"widget": row_widget, "edit": line_edit, "label": label, "status": status_label}
        )
        self.compare_list_layout.addWidget(row_widget)

    def remove_compare_folder_row(self) -> None:
        if len(self.compare_folder_rows) <= 1:
            QMessageBox.information(self, APP_TITLE, "Minimal harus ada satu folder pembanding.")
            return

        last = self.compare_folder_rows.pop()
        widget = last["widget"]
        widget.setParent(None)
        widget.deleteLater()

    def pick_folder(self, line_edit: QLineEdit) -> None:
        start_dir = line_edit.text().strip() or str(Path.home())
        folder = QFileDialog.getExistingDirectory(self, "Pilih Folder", start_dir)
        if folder:
            normalized = self._normalize_folder_path(folder)
            line_edit.setText(normalized)
            line_edit.setCursorPosition(0)
            line_edit.setFocus()
            self.status_label.setText(f"Folder dipilih: {normalized}")

    def start_scan(self) -> None:
        if self.scan_thread and self.scan_thread.is_alive():
            QMessageBox.information(self, APP_TITLE, "Proses scan masih berjalan.")
            return

        target_folder = self._normalize_folder_path(self.target_folder_edit.text())
        compare_folders = [
            self._normalize_folder_path(row["edit"].text())
            for row in self.compare_folder_rows
            if self._normalize_folder_path(row["edit"].text())
        ]

        self.target_folder_edit.setText(target_folder)
        for row in self.compare_folder_rows:
            normalized = self._normalize_folder_path(row["edit"].text())
            if normalized != row["edit"].text():
                row["edit"].setText(normalized)

        if not target_folder:
            QMessageBox.warning(self, APP_TITLE, "Folder A wajib diisi.")
            return
        if not compare_folders:
            QMessageBox.warning(self, APP_TITLE, "Minimal isi satu folder pembanding.")
            return
        if any(Path(target_folder) == Path(folder) for folder in compare_folders):
            QMessageBox.warning(self, APP_TITLE, "Folder A tidak boleh sama dengan folder pembanding.")
            return

        self.clear_results(reset_status=False)
        self._set_action_state(scanning=True)
        self.progress_card.setVisible(True)
        self._last_queued_progress = -1.0
        self._last_queued_progress_text = ""
        with self._progress_lock:
            self._pending_progress = None
        self._set_progress(0, "Memulai scan folder...")
        self._record_history(
            "Scan dan bandingkan",
            "Diproses",
            (
                f"Folder A: {target_folder}\n"
                f"Folder pembanding: {len(compare_folders)} folder\n"
                f"Mode: {'hash+ukuran' if self._current_compare_mode() == 'hash' else 'nama+ukuran'}\n"
                f"Subfolder: {'ya' if self.include_subfolders_checkbox.isChecked() else 'tidak'}"
            ),
            "info",
            file_name="[Semua File]"
        )

        self.scan_thread = threading.Thread(
            target=self._scan_worker,
            args=(
                target_folder,
                compare_folders,
                self._current_compare_mode(),
                self.include_subfolders_checkbox.isChecked(),
            ),
            daemon=True,
        )
        self.scan_thread.start()

    def _current_compare_mode(self) -> str:
        return "hash" if self.compare_mode_hash.isChecked() else "name_size"

    def _current_delete_mode(self) -> str:
        return "permanent" if self.delete_mode_permanent.isChecked() else "internal_trash"

    def _is_result_deletable(self, result: MatchResult) -> bool:
        if result.exact_matches:
            return True
        if result.same_name_different_content:
            return self.allow_delete_red_checkbox.isChecked()
        if result.only_in_target:
            return self.allow_delete_orange_checkbox.isChecked()
        return False

    def _deletable_results(self) -> List[MatchResult]:
        return [row for row in self.result_rows if self._is_result_deletable(row)]

    def _deletable_result_count(self) -> int:
        return len(self._deletable_results())

    def _delete_scope_counts(self, rows: List[MatchResult]) -> Dict[str, int]:
        counts = {"exact_match": 0, "different_content": 0, "only_target": 0}
        for row in rows:
            counts[row.tree_tag] = counts.get(row.tree_tag, 0) + 1
        return counts

    def _delete_scope_text(self) -> str:
        labels = ["hijau"]
        if self.allow_delete_orange_checkbox.isChecked():
            labels.append("oranye")
        if self.allow_delete_red_checkbox.isChecked():
            labels.append("merah")
        return ", ".join(labels)

    def _update_delete_action_controls(self, deleting: bool = False) -> None:
        advanced_scope = self.allow_delete_orange_checkbox.isChecked() or self.allow_delete_red_checkbox.isChecked()
        self.delete_all_button.setText("Hapus Semua Sesuai Opsi" if advanced_scope else "Hapus Semua Hijau")

        if advanced_scope:
            self.delete_scope_hint_label.setText(
                "Mode hapus lanjutan aktif. Periksa ulang hasil merah/oranye sebelum menghapus."
            )
        else:
            self.delete_scope_hint_label.setText("Default aman: hanya hasil hijau yang dapat dihapus.")

        has_deletable = self._deletable_result_count() > 0
        self.delete_button.setEnabled(not deleting and has_deletable)
        self.delete_all_button.setEnabled(not deleting and has_deletable)

    def _update_transfer_action_controls(self, operating: bool = False) -> None:
        has_results = bool(self.result_rows)
        self.copy_button.setEnabled(not operating and has_results)
        self.move_button.setEnabled(not operating and has_results)
        self.sync_button.setEnabled(not operating and has_results)

    def _on_delete_scope_changed(self) -> None:
        if self.scan_thread and self.scan_thread.is_alive():
            return
        self._update_delete_action_controls(deleting=bool(self.delete_thread and self.delete_thread.is_alive()))
        if self.result_rows:
            self.status_label.setText(f"Opsi hapus diperbarui. Kategori aktif: {self._delete_scope_text()}.")

    def _set_action_state(self, scanning: bool) -> None:
        self.scan_button.setEnabled(not scanning)
        self.export_csv_button.setEnabled(False)
        self.export_excel_button.setEnabled(False)
        self.delete_button.setEnabled(False)
        self.delete_all_button.setEnabled(False)
        self.copy_button.setEnabled(False)
        self.move_button.setEnabled(False)
        self.sync_button.setEnabled(False)

    def _set_delete_processing_state(self, deleting: bool) -> None:
        self.scan_button.setEnabled(not deleting)
        self.export_csv_button.setEnabled(not deleting and bool(self.result_rows))
        self.export_excel_button.setEnabled(not deleting and bool(self.result_rows) and self.openpyxl_available)
        self._update_delete_action_controls(deleting=deleting)
        self._update_transfer_action_controls(operating=deleting or bool(self.transfer_thread and self.transfer_thread.is_alive()))

    def _set_transfer_processing_state(self, transferring: bool) -> None:
        self.scan_button.setEnabled(not transferring and not (self.delete_thread and self.delete_thread.is_alive()))
        self.export_csv_button.setEnabled(not transferring and bool(self.result_rows))
        self.export_excel_button.setEnabled(not transferring and bool(self.result_rows) and self.openpyxl_available)
        self._update_delete_action_controls(
            deleting=transferring or bool(self.delete_thread and self.delete_thread.is_alive())
        )
        self._update_transfer_action_controls(operating=transferring)

    def _set_undo_processing_state(self, undoing: bool) -> None:
        has_results = bool(self.result_rows)
        self.scan_button.setEnabled(not undoing)
        self.export_csv_button.setEnabled(not undoing and has_results)
        self.export_excel_button.setEnabled(not undoing and has_results and self.openpyxl_available)
        self._update_delete_action_controls(
            deleting=undoing or bool(self.delete_thread and self.delete_thread.is_alive())
        )
        self._update_transfer_action_controls(
            operating=undoing or bool(self.transfer_thread and self.transfer_thread.is_alive())
        )

    def _scan_worker(self, target_folder: str, compare_folders: List[str], mode: str, include_subfolders: bool) -> None:
        try:
            self._queue_progress(2.0, "Menyiapkan proses scan...")

            compare_paths = [Path(folder) for folder in compare_folders]
            folder_plan = [("A", Path(target_folder), "Folder A")] + [
                (f"F{idx}", folder_path, f"folder pembanding F{idx}")
                for idx, folder_path in enumerate(compare_paths, start=1)
            ]
            file_counts: Dict[str, int] = {}
            total_folders = len(folder_plan)

            for idx, (label, folder_path, folder_name) in enumerate(folder_plan, start=1):
                self._queue_progress(
                    self._progress_in_range(2.0, 10.0, idx - 1, total_folders),
                    f"Menghitung file di {folder_name}...",
                )
                file_counts[label] = self._count_files(folder_path, include_subfolders)
                self._queue_progress(
                    self._progress_in_range(2.0, 10.0, idx, total_folders),
                    f"Perkiraan file {folder_name}: {file_counts[label]} item.",
                )

            total_collect_files = file_counts.get("A", 0) + sum(file_counts.get(f"F{idx}", 0) for idx in range(1, len(compare_folders) + 1))
            collected_so_far = 0
            collect_stride = self._progress_stride(total_collect_files, 220)

            def emit_collect_progress(processed: int, folder_name: str) -> None:
                if total_collect_files <= 0:
                    return
                if processed not in {1, total_collect_files} and processed % collect_stride != 0:
                    return
                self._queue_progress(
                    self._progress_in_range(10.0, 45.0, processed, total_collect_files),
                    f"Mengambil metadata file dari {folder_name} ({processed}/{total_collect_files})...",
                )

            target_files = self._collect_files(
                Path(target_folder),
                include_subfolders,
                base_label="A",
                progress_callback=lambda folder_done: emit_collect_progress(folder_done, "Folder A"),
            )
            collected_so_far += len(target_files)

            compare_groups: List[Tuple[str, Path, List[FileRecord]]] = []
            for idx, folder_path in enumerate(compare_paths, start=1):
                label = f"F{idx}"
                folder_name = f"folder pembanding {label}"
                base_processed = collected_so_far
                compare_records = self._collect_files(
                    folder_path,
                    include_subfolders,
                    base_label=label,
                    progress_callback=lambda folder_done, base=base_processed, name=folder_name: emit_collect_progress(
                        base + folder_done,
                        name,
                    ),
                )
                collected_so_far += len(compare_records)
                compare_groups.append((label, folder_path, compare_records))

            self._queue_progress(45.0, "Menyusun indeks data pembanding...")
            results = self._build_comparison_results(target_files, compare_groups, mode)
            self.ui_queue.put(("scan_done", results))
        except Exception as exc:
            self.ui_queue.put(
                (
                    "scan_error",
                    {
                        "title": "Terjadi kesalahan saat scan dan bandingkan",
                        "summary": str(exc) or exc.__class__.__name__,
                        "details": "".join(traceback.format_exception(type(exc), exc, exc.__traceback__)),
                    },
                )
            )

    def _build_comparison_results(
        self,
        target_files: List[FileRecord],
        compare_groups: List[Tuple[str, Path, List[FileRecord]]],
        mode: str,
    ) -> List[MatchResult]:
        by_name_size: Dict[Tuple[str, int], List[FileRecord]] = {}
        by_relative_size: Dict[Tuple[str, int], List[FileRecord]] = {}
        by_relative_name: Dict[str, List[FileRecord]] = {}
        by_hash: Dict[Tuple[int, str], List[FileRecord]] = {}
        all_compare_labels = [label for label, _, _ in compare_groups]

        flat_compare_records: List[FileRecord] = []
        total_compare_records = sum(len(records) for _, _, records in compare_groups)
        indexed_records = 0
        index_stride = self._progress_stride(total_compare_records, 160)
        for _, _, records in compare_groups:
            for record in records:
                flat_compare_records.append(record)
                by_name_size.setdefault((record.path.name.lower(), record.size), []).append(record)
                by_relative_size.setdefault((record.relative_path.lower(), record.size), []).append(record)
                by_relative_name.setdefault(record.relative_path.lower(), []).append(record)
                indexed_records += 1
                if total_compare_records > 0 and (
                    indexed_records in {1, total_compare_records} or indexed_records % index_stride == 0
                ):
                    self._queue_progress(
                        self._progress_in_range(45.0, 55.0, indexed_records, total_compare_records),
                        f"Menyusun indeks data pembanding ({indexed_records}/{total_compare_records})...",
                    )
        if total_compare_records == 0:
            self._queue_progress(55.0, "Tidak ada file pada folder pembanding yang perlu diindeks.")

        if mode == "hash":
            total_hash_items = len(flat_compare_records) + len(target_files)
            processed = 0
            hash_stride = self._progress_stride(total_hash_items, 180)
            for record in flat_compare_records:
                record.sha256 = self._hash_file(record.path)
                by_hash.setdefault((record.size, record.sha256), []).append(record)
                processed += 1
                if processed in {1, total_hash_items} or processed % hash_stride == 0:
                    self._queue_progress(
                        self._progress_in_range(55.0, 82.0, processed, total_hash_items),
                        f"Hash file pembanding {processed}/{total_hash_items}...",
                    )

        results: List[MatchResult] = []
        total_targets = len(target_files)
        compare_progress_start = 82.0 if mode == "hash" else 55.0
        compare_progress_end = 98.5
        compare_stride = self._progress_stride(total_targets, 220)

        for idx, target in enumerate(target_files, start=1):
            exact_matches: List[FileRecord] = []
            same_name_different_content: List[FileRecord] = []

            if mode == "name_size":
                exact_matches.extend(by_name_size.get((target.path.name.lower(), target.size), []))
                exact_matches.extend(
                    item
                    for item in by_relative_size.get((target.relative_path.lower(), target.size), [])
                    if item not in exact_matches
                )
                same_name_different_content = [
                    item for item in by_relative_name.get(target.relative_path.lower(), []) if item.size != target.size
                ]
                match_type = "nama+ukuran"
            else:
                target.sha256 = self._hash_file(target.path)
                exact_matches = by_hash.get((target.size, target.sha256), [])
                same_name_different_content = [
                    item
                    for item in by_relative_name.get(target.relative_path.lower(), [])
                    if item.sha256 != target.sha256 or item.size != target.size
                ]
                match_type = "hash+ukuran"
                total_hash_steps = len(flat_compare_records) + len(target_files)
                processed_hash_steps = len(flat_compare_records) + idx
                if processed_hash_steps in {1, total_hash_steps} or processed_hash_steps % max(1, self._progress_stride(total_hash_steps, 180)) == 0:
                    self._queue_progress(
                        self._progress_in_range(55.0, 82.0, processed_hash_steps, total_hash_steps),
                        f"Hash file target {idx}/{total_targets}...",
                    )

            exact_labels = {item.base_label for item in exact_matches}
            diff_labels = {item.base_label for item in same_name_different_content}
            involved_labels = exact_labels.union(diff_labels)
            missing = [label for label in all_compare_labels if label not in involved_labels]

            results.append(
                MatchResult(
                    target_path=target.path,
                    target_relative_path=target.relative_path,
                    size=target.size,
                    match_type=match_type,
                    exact_matches=sorted(exact_matches, key=lambda x: (x.base_label, str(x.path))),
                    same_name_different_content=sorted(
                        same_name_different_content,
                        key=lambda x: (x.base_label, str(x.path)),
                    ),
                    missing_from_folders=missing,
                    only_in_target=not exact_matches and not same_name_different_content,
                )
            )
            if idx in {1, total_targets} or idx % compare_stride == 0:
                self._queue_progress(
                    self._progress_in_range(compare_progress_start, compare_progress_end, idx, total_targets),
                    f"Membandingkan file {idx}/{total_targets}...",
                )

        if total_targets == 0:
            self._queue_progress(compare_progress_end, "Tidak ada file di Folder A yang perlu dibandingkan.")

        self._queue_progress(99.5, "Menyiapkan tampilan visual...")
        return results

    def _count_files(self, folder: Path, include_subfolders: bool) -> int:
        if not folder.exists() or not folder.is_dir():
            raise FileNotFoundError(f"Folder tidak valid: {folder}")

        total = 0
        iterator: Iterable[Path] = folder.rglob("*") if include_subfolders else folder.glob("*")
        for path in iterator:
            if path.is_file():
                total += 1
        return total

    @staticmethod
    def _progress_in_range(start: float, end: float, completed: int, total: int) -> float:
        if total <= 0:
            return end
        fraction = max(0.0, min(1.0, completed / total))
        return start + ((end - start) * fraction)

    @staticmethod
    def _progress_stride(total: int, target_updates: int = 180) -> int:
        if total <= 0:
            return 1
        return max(1, total // target_updates)

    def _collect_files(
        self,
        folder: Path,
        include_subfolders: bool,
        base_label: str,
        progress_callback: Optional[Callable[[int], None]] = None,
    ) -> List[FileRecord]:
        if not folder.exists() or not folder.is_dir():
            raise FileNotFoundError(f"Folder tidak valid: {folder}")

        records: List[FileRecord] = []
        iterator: Iterable[Path] = folder.rglob("*") if include_subfolders else folder.glob("*")
        for path in iterator:
            if path.is_file():
                try:
                    size = path.stat().st_size
                except OSError:
                    continue
                records.append(
                    FileRecord(
                        path=path,
                        base_folder=folder,
                        base_label=base_label,
                        relative_path=str(path.relative_to(folder)),
                        size=size,
                    )
                )
                if progress_callback is not None:
                    progress_callback(len(records))
        return records

    def _hash_file(self, path: Path) -> str:
        sha = hashlib.sha256()
        with path.open("rb") as file_handle:
            while True:
                chunk = file_handle.read(CHUNK_SIZE)
                if not chunk:
                    break
                sha.update(chunk)
        return sha.hexdigest()

    def _format_size(self, size: int) -> str:
        value = float(size)
        units = ["B", "KB", "MB", "GB", "TB"]
        for unit in units:
            if value < 1024 or unit == units[-1]:
                return f"{value:.2f} {unit}" if unit != "B" else f"{int(value)} B"
            value /= 1024
        return f"{size} B"

    @staticmethod
    def _format_progress_value(value: float) -> str:
        rounded = round(value, 1)
        if abs(rounded - round(rounded)) < 0.05:
            return f"{int(round(rounded))}%"
        return f"{rounded:.1f}%"

    def _set_progress(self, value: float, text: str) -> None:
        safe_value = max(0.0, min(100.0, value))
        target_value = int(round(safe_value * 10))
        current_value = self.progress_bar.value()

        self.progress_badge.setText(self._format_progress_value(safe_value))
        self.status_label.setText(text)

        if target_value == current_value:
            return

        if target_value <= current_value or target_value == 0:
            self.progress_animation.stop()
            self.progress_bar.setValue(target_value)
            return

        duration = max(140, min(520, 120 + abs(target_value - current_value)))
        self.progress_animation.stop()
        self.progress_animation.setDuration(duration)
        self.progress_animation.setStartValue(current_value)
        self.progress_animation.setEndValue(target_value)
        self.progress_animation.start()

    def _queue_progress(self, value: float, text: str) -> None:
        safe_value = max(0.0, min(100.0, value))
        rounded_value = round(safe_value, 1)

        if rounded_value not in {0.0, 100.0} and rounded_value < self._last_queued_progress:
            rounded_value = self._last_queued_progress
        if rounded_value == self._last_queued_progress and text == self._last_queued_progress_text:
            return

        self._last_queued_progress = rounded_value
        self._last_queued_progress_text = text
        with self._progress_lock:
            self._pending_progress = (rounded_value, text)

    def _poll_queue(self) -> None:
        pending_progress: Optional[Tuple[float, str]] = None
        with self._progress_lock:
            if self._pending_progress is not None:
                pending_progress = self._pending_progress
                self._pending_progress = None

        if pending_progress is not None:
            value, text = pending_progress
            self._set_progress(float(value), str(text))

        processed_messages = 0
        try:
            while processed_messages < 12:
                message = self.ui_queue.get_nowait()
                self._handle_queue_message(message)
                processed_messages += 1
        except queue.Empty:
            return

    def _handle_queue_message(self, message: Tuple[str, object]) -> None:
        kind, payload = message
        if kind == "progress":
            value, text = payload  # type: ignore[misc]
            self._set_progress(float(value), str(text))
            return

        if kind == "scan_done":
            self.scan_button.setEnabled(True)
            with self._progress_lock:
                self._pending_progress = None
            self._pending_scan_results = payload  # type: ignore[assignment]
            self._awaiting_scan_finalize = True
            self._set_progress(100, "Menyiapkan tampilan hasil scan...")
            self._maybe_finalize_scan_after_progress()
            return

        if kind == "scan_error":
            self.scan_button.setEnabled(True)
            self._pending_scan_results = None
            self._awaiting_scan_finalize = False
            with self._progress_lock:
                self._pending_progress = None
            self.progress_animation.stop()
            self._set_progress(0, "Terjadi kesalahan saat scan.")
            self.progress_card.setVisible(False)
            if isinstance(payload, dict):
                self._record_history(
                    "Scan dan bandingkan",
                    "Gagal",
                    str(payload.get("summary", "Terjadi kesalahan saat scan.")),
                    "error",
                    file_name="[Semua File]"
                )
                self.show_error_dialog(
                    str(payload.get("title", APP_TITLE)),
                    str(payload.get("summary", "Terjadi kesalahan saat scan.")),
                    str(payload.get("details", payload.get("summary", ""))),
                )
            else:
                self._record_history(
                    "Scan dan bandingkan",
                    "Gagal",
                    str(payload),
                    "error",
                    file_name="[Semua File]"
                )
                self.show_error_dialog(
                    "Terjadi kesalahan saat scan dan bandingkan",
                    str(payload),
                    str(payload),
                )
            return

        if kind == "delete_done":
            self._pending_delete_result = payload if isinstance(payload, dict) else {}
            QTimer.singleShot(0, self._finalize_delete_results)
            return

        if kind == "transfer_done":
            self._pending_transfer_result = payload if isinstance(payload, dict) else {}
            QTimer.singleShot(0, self._finalize_transfer_results)
            return

        if kind == "bulk_compare_sync_done":
            self._pending_bulk_sync_result = payload if isinstance(payload, dict) else {}
            QTimer.singleShot(0, self._finalize_bulk_sync_green_wrapper)
            return

        if kind == "undo_done":
            self._pending_undo_result = payload if isinstance(payload, dict) else {}
            QTimer.singleShot(0, self._finalize_undo_results)
            return

        if kind == "update_check_error":
            if hasattr(self, "update_spinner_timer"):
                self.update_spinner_timer.stop()
                self._apply_sidebar_icons(active_index=self.main_stack.currentIndex())
            self.btn_nav_check_update.setEnabled(True)
            self.status_label.setText("Gagal memeriksa pembaruan.")
            QMessageBox.warning(self, "Pembaruan", f"Gagal mengecek pembaruan:\n\n{payload}")
            return
            
        if kind == "update_check_done":
            if hasattr(self, "update_spinner_timer"):
                self.update_spinner_timer.stop()
                self._apply_sidebar_icons(active_index=self.main_stack.currentIndex())
            self.btn_nav_check_update.setEnabled(True)
            data = payload
            latest = data.get("latest_version")
            current = data.get("current_version")
            release_url = data.get("release_url")
            download_url = data.get("download_url")
            changelog = data.get("changelog")
            
            def parse_version(v):
                return [int(x) if x.isdigit() else x for x in v.split('.')]
                
            try:
                has_update = parse_version(latest) > parse_version(current)
                if has_update:
                    self.status_label.setText(f"Pembaruan tersedia: v{latest}")
                else:
                    self.status_label.setText("Aplikasi sudah versi terbaru.")

                status_dialog = UpdateStatusDialog(self, latest, current, changelog, has_update)
                if status_dialog.exec() == QDialog.DialogCode.Accepted and has_update:
                    if download_url:
                        dialog = UpdateDownloadDialog(self, latest, download_url)
                        dialog.start_download()
                        if dialog.exec() == QDialog.DialogCode.Accepted:
                            downloaded_path = dialog.get_downloaded_path()
                            import subprocess
                            import sys
                            import os
                            import tempfile
                            import platform
                            
                            # Simpan status niat pembaruan untuk ditampilkan modal sukses di sesi berikutnya
                            update_status_file = self.app_data_dir / "update_status.json"
                            try:
                                import json
                                with open(update_status_file, "w", encoding="utf-8") as f:
                                    json.dump({
                                        "target_version": latest,
                                        "old_version": current,
                                        "changelog": changelog
                                    }, f)
                            except Exception as e:
                                print(f"Failed to write update status: {e}")

                            try:
                                system = platform.system().lower()
                                lower_path = downloaded_path.lower()
                                
                                is_archive = lower_path.endswith(('.zip', '.tar.gz', '.tgz', '.rar'))
                                if is_archive:
                                    QMessageBox.information(
                                        self, 
                                        "Pembaruan Manual Diperlukan", 
                                        f"File unduhan merupakan arsip ({os.path.basename(downloaded_path)}).\n\nSilakan ekstrak file ini dan ganti file aplikasi lama Anda secara manual."
                                    )
                                    # Buka folder tempat file diunduh
                                    folder_dest = os.path.dirname(downloaded_path)
                                    if system == "windows":
                                        os.startfile(folder_dest)
                                    elif system == "darwin":
                                        subprocess.Popen(["open", folder_dest])
                                    else:
                                        subprocess.Popen(["xdg-open", folder_dest])
                                    sys.exit(0)

                                is_installer = False
                                if system == "windows":
                                    is_installer = any(x in lower_path for x in ['.msi', 'setup', 'install'])
                                elif system == "darwin":
                                    is_installer = any(x in lower_path for x in ['.dmg', '.pkg'])
                                elif system == "linux":
                                    is_installer = any(x in lower_path for x in ['.deb', '.rpm'])

                                if is_installer:
                                    if system == "darwin":
                                        subprocess.Popen(["open", downloaded_path])
                                    elif system == "linux":
                                        # Gunakan software-center sebagai fallback jika xdg-open tidak tahu cara menangani deb
                                        try:
                                            subprocess.Popen(["xdg-open", downloaded_path])
                                        except Exception:
                                            subprocess.Popen(["gnome-software", "--local-filename", downloaded_path])
                                    else:
                                        if hasattr(os, 'startfile'):
                                            os.startfile(downloaded_path)
                                        else:
                                            subprocess.Popen([downloaded_path], shell=True)
                                    sys.exit(0)
                                else:
                                    is_frozen = getattr(sys, 'frozen', False)
                                    
                                    # Fallback jika ternyata yang berjalan adalah AppImage native Linux yg dilarikan ke /tmp/ virtual disk
                                    appimage_env = os.environ.get("APPIMAGE")
                                    if system == "linux" and is_frozen and appimage_env and os.path.exists(appimage_env):
                                        target_path = appimage_env
                                    else:
                                        target_path = sys.executable if is_frozen else os.path.abspath(sys.argv[0])
                                    
                                    # Hapus environment variable yg berhubungan dgn PyInstaller agar app baru tidak error DLL
                                    env = os.environ.copy()
                                    env.pop("_MEIPASS", None)
                                    env.pop("_MEIPASS2", None)
                                    env.pop("_PYIBoot_SPLASH", None)

                                    if system == "windows":
                                        bat_path = os.path.join(tempfile.gettempdir(), "updater.bat")
                                        with open(bat_path, "w") as f:
                                            f.write('@echo off\n')
                                            f.write('set _MEIPASS=\n')
                                            f.write('set _MEIPASS2=\n')
                                            f.write('timeout /t 2 /nobreak > NUL\n')
                                            f.write(f'move /Y "{downloaded_path}" "{target_path}"\n')
                                            if is_frozen:
                                                f.write(f'start "" "{target_path}"\n')
                                            else:
                                                f.write(f'start "" "{sys.executable}" "{target_path}"\n')
                                            f.write('del "%~f0"\n')
                                        subprocess.Popen([bat_path], shell=True, env=env)
                                    else:
                                        sh_path = os.path.join(tempfile.gettempdir(), "updater.sh")
                                        with open(sh_path, "w") as f:
                                            f.write('#!/bin/bash\n')
                                            f.write('unset _MEIPASS _MEIPASS2 _PYIBoot_SPLASH\n')
                                            f.write('sleep 2\n')
                                            
                                            if system == "darwin" and target_path.endswith("/MacOS/FolderCompare"):
                                                # Overwrite seluruh .app bundle di parent folder jika Mac OS App Bundle
                                                bundle_path = os.path.abspath(os.path.join(target_path, "../../.."))
                                                f.write(f'mv -f "{downloaded_path}" "{bundle_path}"\n')
                                                f.write(f'chmod -R +x "{bundle_path}"\n')
                                                f.write(f'nohup open "{bundle_path}" >/dev/null 2>&1 &\n')
                                            else:
                                                f.write(f'mv -f "{downloaded_path}" "{target_path}"\n')
                                                f.write(f'chmod +x "{target_path}"\n')
                                                if is_frozen:
                                                    f.write(f'nohup "{target_path}" >/dev/null 2>&1 &\n')
                                                else:
                                                    f.write(f'nohup "{sys.executable}" "{target_path}" >/dev/null 2>&1 &\n')
                                            f.write('rm -f "$0"\n')
                                        os.chmod(sh_path, 0o755)
                                        subprocess.Popen([sh_path], shell=False, start_new_session=True, env=env)
                                    sys.exit(0)
                            except Exception as e:
                                QMessageBox.critical(self, "Gagal", f"Gagal menjalankan pembaruan:\n\n{e}")
                    else:
                        import webbrowser
                        webbrowser.open(release_url)
                        self.status_label.setText("Membuka browser untuk mengunduh rilis terbaru...")
            except Exception as e:
                self.status_label.setText("Gagal memeriksa versi pembaruan.")
            return

    def _on_progress_animation_finished(self) -> None:
        self._maybe_finalize_scan_after_progress()

    def _maybe_finalize_scan_after_progress(self) -> None:
        if not self._awaiting_scan_finalize or self._pending_scan_results is None:
            return
        if self.progress_animation.state() == QAbstractAnimation.Running:
            return
        if self.progress_bar.value() < self.progress_bar.maximum():
            return

        self._awaiting_scan_finalize = False
        QTimer.singleShot(0, self._finalize_scan_results)

    def _finalize_scan_results(self) -> None:
        if self._pending_scan_results is None:
            return

        self.result_rows = self._pending_scan_results
        self._pending_scan_results = None

        self._apply_default_table_widths()
        self._populate_table(recompute_widths=False)
        self._refresh_stats()
        self.progress_animation.stop()
        self.progress_bar.setValue(self.progress_bar.maximum())
        self.progress_badge.setText("100%")
        self.status_label.setText("Scan selesai. Data siap dianalisis.")
        self._record_history(
            "Scan dan bandingkan",
            "Sukses",
            (
                f"Total hasil: {len(self.result_rows)}\n"
                f"Duplikat: {sum(1 for row in self.result_rows if row.exact_matches)}\n"
                f"Berbeda: {sum(1 for row in self.result_rows if row.same_name_different_content)}\n"
                f"Hanya di A: {sum(1 for row in self.result_rows if row.only_in_target)}"
            ),
            "success",
            file_name="[Semua File]"
        )
        self.export_csv_button.setEnabled(True)
        if self.openpyxl_available:
            self.export_excel_button.setEnabled(True)
        self._update_transfer_action_controls()

        deletable_count = self._deletable_result_count()
        if deletable_count > 0:
            self._update_delete_action_controls()
        else:
            self._update_delete_action_controls()
            QMessageBox.information(
                self,
                APP_TITLE,
                "Tidak ada hasil yang sesuai dengan izin hapus saat ini. Aktifkan opsi merah/oranye bila memang ingin menghapus kategori tersebut.",
            )

        QTimer.singleShot(40, self._finalize_table_layout_after_scan)
        self.progress_card.setVisible(False)

    def _finalize_table_layout_after_scan(self) -> None:
        if self.result_rows and self.table_proxy.rowCount() > 0:
            self._apply_responsive_column_widths()

    def _finalize_delete_results(self) -> None:
        payload_map = self._pending_delete_result or {}
        self._pending_delete_result = None

        deleted_paths = [Path(path) for path in payload_map.get("deleted_paths", [])]
        deleted_count = int(payload_map.get("deleted_count", len(deleted_paths)))
        errors = [str(item) for item in payload_map.get("errors", [])]
        undo_payload = payload_map.get("undo_action") if isinstance(payload_map.get("undo_action"), dict) else None
        trash_entries_payload = payload_map.get("trash_entries", []) if isinstance(payload_map.get("trash_entries", []), list) else []
        active_delete_dialog = self.delete_confirm_dialog

        if active_delete_dialog is not None and not errors:
            active_delete_dialog.set_processing(
                True,
                f"Penghapusan selesai. Menyegarkan hasil visual untuk {deleted_count} file...",
            )
            active_delete_dialog.flush_visual_state()

        path_set = set(deleted_paths)
        self.result_rows = [row for row in self.result_rows if row.target_path not in path_set]
        self._apply_default_table_widths()
        self._populate_table(recompute_widths=False)
        self._refresh_stats()
        self.current_selected_result = None
        self._reset_detail_panel()

        remaining = self.table_proxy.rowCount()
        self.status_label.setText(f"Selesai menghapus {deleted_count} file. Sisa hasil terlihat: {remaining}.")
        self._set_delete_processing_state(False)
        self._add_trash_entries([item for item in trash_entries_payload if isinstance(item, dict)])

        if self.result_rows and self.table_proxy.rowCount() > 0:
            QTimer.singleShot(40, self._finalize_table_layout_after_delete)

        op_list = undo_payload.get("operations", []) if undo_payload else []
        if len(op_list) == 1 and "destination" in op_list[0]:
            target_path = op_list[0]["destination"]
            success_details = f"File: {Path(target_path).name}\nPath: {target_path}\nSisa hasil terlihat: {remaining}"
            warn_details = f"File: {Path(target_path).name}\nBerhasil: {deleted_count} | Gagal: {len(errors)}"
        else:
            success_details = f"Jumlah file diproses: {deleted_count}\nSisa hasil terlihat: {remaining}"
            warn_details = f"Berhasil: {deleted_count} | Gagal: {len(errors)}"

        if errors:
            self._record_history(
                "Penghapusan file",
                "Sebagian gagal" if deleted_count > 0 else "Gagal",
                warn_details,
                "warning" if deleted_count > 0 else "error",
            )
            self._push_undo_action(undo_payload)
            if active_delete_dialog is not None:
                active_delete_dialog.force_close(QDialog.Rejected)
                self.delete_confirm_dialog = None
            error_preview = os.linesep.join(errors[:10])
            QTimer.singleShot(
                180,
                lambda preview=error_preview, total_errors=len(errors): self.show_error_dialog(
                    "Sebagian file gagal diproses",
                    f"{total_errors} file gagal diproses saat penghapusan.",
                    preview,
                ),
            )
        else:
            success_summary = f"Berhasil memproses {deleted_count} file dari Folder A."
            self._record_history(
                "Penghapusan file",
                "Sukses",
                success_details,
                "success",
            )
            self._push_undo_action(undo_payload)
            if active_delete_dialog is not None:
                active_delete_dialog.show_success_state(success_summary, success_details)
                active_delete_dialog.flush_visual_state()
            else:
                self.show_success_dialog(
                    "Penghapusan Berhasil",
                    success_summary,
                    success_details,
                )

    def _finalize_table_layout_after_delete(self) -> None:
        if self.result_rows and self.table_proxy.rowCount() > 0:
            self._apply_responsive_column_widths()

    def _finalize_transfer_results(self) -> None:
        payload_map = self._pending_transfer_result or {}
        self._pending_transfer_result = None

        operation = str(payload_map.get("operation", "copy"))
        processed_count = int(payload_map.get("processed_count", 0))
        error_count = int(payload_map.get("error_count", 0))
        destination_root = str(payload_map.get("destination_root", ""))
        errors = [str(item) for item in payload_map.get("errors", [])]
        undo_payload = payload_map.get("undo_action") if isinstance(payload_map.get("undo_action"), dict) else None
        active_transfer_dialog = self.transfer_confirm_dialog
        self.transfer_confirm_dialog = None

        self.transfer_thread = None
        self._set_transfer_processing_state(False)
        self._refresh_missing_compare_suggestion(self.current_selected_result)

        if operation in {"copy_compare_sync", "move_compare_sync"}:
            operation_label = "dipindahkan" if operation == "move_compare_sync" else "disalin"
            source_path = str(payload_map.get("source_path", ""))
            created_labels = payload_map.get("created_labels", [])
            created_paths = payload_map.get("created_paths", [])
            replaced_labels = payload_map.get("replaced_labels", [])
            
            if processed_count > 0:
                if operation == "move_compare_sync":
                    self.result_rows = [row for row in self.result_rows if str(row.target_path) != source_path]
                else:
                    for row in self.result_rows:
                        if str(row.target_path) == source_path:
                            row.temp_synced_labels.extend(created_labels + replaced_labels)
                            row.temp_synced_paths.extend(created_paths)
                            break
                
                self.status_label.setText(
                    f"File berhasil {operation_label}. Perubahan telah diterapkan pada tabel (tanpa perlu scan ulang)."
                )
                self._apply_default_table_widths()
                self._populate_table(recompute_widths=False)
                self._refresh_stats()
                self._reset_detail_panel()

            if errors:
                self._record_history(
                    "Sinkronisasi folder pembanding",
                    "Sebagian gagal" if processed_count > 0 else "Gagal",
                    f"Berhasil: {processed_count} | Target: {destination_root or '-'} | Operasi: {operation_label}",
                    "warning" if processed_count > 0 else "error",
                    file_name=Path(source_path).name
                )
                self._push_undo_action(undo_payload)
                if active_transfer_dialog is not None:
                    active_transfer_dialog.force_close(QDialog.Rejected)
                
                error_details = (
                    f"Path relatif: {payload_map.get('relative_path', '-')}{os.linesep}"
                    f"Folder pembanding target: {destination_root or '-'}{os.linesep}{os.linesep}"
                    f"{os.linesep.join(errors[:10])}"
                )
                QTimer.singleShot(
                    180,
                    lambda p_count=processed_count, e_count=error_count, op_lbl=operation_label, det=error_details: self.show_error_dialog(
                        "Sinkronisasi Folder Pembanding Belum Selesai",
                        f"{p_count} file berhasil {op_lbl}, {e_count} file gagal diproses.",
                        det,
                    )
                )
                return

            success_summary = f"{processed_count} file berhasil {operation_label} ke folder pembanding."
            success_details = (
                f"Path relatif: {payload_map.get('relative_path', '-')}{os.linesep}"
                f"Folder pembanding target: {destination_root or '-'}{os.linesep}"
                "Catatan: hasil scan dibersihkan. Silakan scan ulang untuk melihat kondisi terbaru."
            )
            self._record_history(
                "Sinkronisasi folder pembanding",
                "Sukses",
                f"Berhasil: {processed_count} | Target: {destination_root or '-'} | Operasi: {operation_label}",
                "success",
                file_name=Path(source_path).name
            )
            self._push_undo_action(undo_payload)
            if active_transfer_dialog is not None:
                active_transfer_dialog.show_success_state(success_summary, success_details)
                active_transfer_dialog.flush_visual_state()
            else:
                self.show_success_dialog(
                    "Sinkronisasi Folder Pembanding Berhasil",
                    success_summary,
                    success_details,
                )
            return

        operation_label = "dipindahkan" if operation == "move" else "disalin"
        title = "Pindah File Selesai" if operation == "move" else "Salin File Selesai"

        if operation == "move" and processed_count > 0:
            self.clear_results(reset_status=False)
            self.status_label.setText(
                "File berhasil dipindahkan. Hasil scan dibersihkan karena struktur sumber sudah berubah."
            )
        elif processed_count > 0:
            self.status_label.setText(f"{processed_count} file berhasil {operation_label}.")

        detail_lines = [
            f"Jumlah file berhasil: {processed_count}",
            f"Folder tujuan: {destination_root or '-'}",
        ]
        if operation == "move" and processed_count > 0:
            detail_lines.append("Catatan: hasil scan dibersihkan. Silakan scan ulang untuk melihat kondisi terbaru.")

        if errors:
            self._record_history(
                "Pindah file terpilih" if operation == "move" else "Salin file terpilih",
                "Sebagian gagal" if processed_count > 0 else "Gagal",
                f"Berhasil: {processed_count} | Gagal: {error_count} | Tujuan: {destination_root or '-'}",
                "warning" if processed_count > 0 else "error",
            )
            self._push_undo_action(undo_payload)
            error_preview = os.linesep.join(errors[:10])
            if active_transfer_dialog is not None:
                active_transfer_dialog.force_close(QDialog.Rejected)
            error_details = f"{os.linesep.join(detail_lines)}{os.linesep}{os.linesep}{error_preview}"
            QTimer.singleShot(
                180,
                lambda ttl=title, p_count=processed_count, e_count=error_count, op_lbl=operation_label, det=error_details: self.show_error_dialog(
                    ttl,
                    f"{p_count} file berhasil {op_lbl}, {e_count} file gagal diproses.",
                    det
                )
            )
            return

        self._record_history(
            "Pindah file terpilih" if operation == "move" else "Salin file terpilih",
            "Sukses",
            f"Jumlah file: {processed_count} | Tujuan: {destination_root or '-'}",
            "success",
        )
        self._push_undo_action(undo_payload)
        if active_transfer_dialog is not None:
            active_transfer_dialog.show_success_state(
                f"{processed_count} file berhasil {operation_label}.",
                os.linesep.join(detail_lines),
            )
            active_transfer_dialog.flush_visual_state()
        else:
            self.show_success_dialog(title, f"{processed_count} file berhasil {operation_label}.", os.linesep.join(detail_lines))

    def _finalize_bulk_sync_green_wrapper(self) -> None:
        payload_map = self._pending_bulk_sync_result or {}
        self._pending_bulk_sync_result = None
        self._finalize_bulk_sync_green(payload_map)

    def _finalize_bulk_sync_green(self, payload_map: Dict[str, Any]) -> None:
        processed_count = int(payload_map.get("processed_count", 0))
        error_count = int(payload_map.get("error_count", 0))
        errors = [str(item) for item in payload_map.get("errors", [])]
        undo_payload = payload_map.get("undo_action") if isinstance(payload_map.get("undo_action"), dict) else None
        bulk_updates = payload_map.get("bulk_updates", [])
        active_transfer_dialog = self.transfer_confirm_dialog
        self.transfer_confirm_dialog = None

        self.transfer_thread = None
        self._set_transfer_processing_state(False)

        if bulk_updates:
            for update in bulk_updates:
                target_path = update["target_path"]
                for row in self.result_rows:
                    if str(row.target_path) == target_path:
                        row.temp_synced_labels.extend(update["created_labels"])
                        row.temp_synced_paths.extend(update["created_paths"])
                        break
            
            self._apply_default_table_widths()
            self._populate_table(recompute_widths=False)
            self._refresh_stats()
            self._reset_detail_panel()
            self.status_label.setText(f"{len(bulk_updates)} baris berhasil disinkronkan tanpa perlu scan ulang.")

        if errors:
            self._record_history(
                "Sync Hijau Massal",
                "Sebagian gagal" if processed_count > 0 else "Gagal",
                f"Berhasil: {processed_count} | Gagal: {error_count}",
                "warning" if processed_count > 0 else "error",
            )
            self._push_undo_action(undo_payload)
            if active_transfer_dialog is not None:
                active_transfer_dialog.force_close(QDialog.Rejected)
            error_details = f"{os.linesep.join(errors[:15])}"
            QTimer.singleShot(
                180,
                lambda e_count=error_count, det=error_details: self.show_error_dialog(
                    "Sinkronisasi Massal Belum Selesai",
                    f"Beberapa file ({e_count}) gagal disinkronkan ke folder pembanding.",
                    det
                )
            )
            return

        success_summary = f"{processed_count} operasi penyalinan sukses ke seluruh folder pembanding."
        success_details = "Status sinkronisasi diperbarui instan pada tabel utama."
        
        self._record_history(
            "Sync Hijau Massal",
            "Sukses",
            f"Diaplikasikan pada {len(bulk_updates)} sumber duplikat",
            "success",
        )
        self._push_undo_action(undo_payload)
        
        if active_transfer_dialog is not None:
            active_transfer_dialog.show_success_state(success_summary, success_details)
            active_transfer_dialog.flush_visual_state()
        else:
            self.show_success_dialog(
                "Sync Hijau Massal Berhasil",
                success_summary,
                success_details,
            )

    def _finalize_undo_results(self) -> None:
        payload_map = self._pending_undo_result or {}
        self._pending_undo_result = None

        label = str(payload_map.get("label", "Aksi terakhir"))
        detail = str(payload_map.get("detail", ""))
        restored_count = int(payload_map.get("restored_count", 0))
        error_count = int(payload_map.get("error_count", 0))
        errors = [str(item) for item in payload_map.get("errors", [])]
        action_dir = str(payload_map.get("action_dir", ""))
        restored_trash_entry_ids = [str(item) for item in payload_map.get("restored_trash_entry_ids", [])]

        self.undo_thread = None
        self._set_undo_processing_state(False)
        if self.undo_processing_dialog is not None:
            self.undo_processing_dialog.close()
            self.undo_processing_dialog.deleteLater()
            self.undo_processing_dialog = None
        self._refresh_undo_button()
        
        action_file_hint = str(payload_map.get("file_name_hint", "-"))

        if errors:
            error_action = "Restore dari Trash" if label == "Pulihkan dari Trash Internal" else "Undo"
            error_title = "Pemulihan Trash Belum Selesai" if label == "Pulihkan dari Trash Internal" else "Undo Belum Selesai"
            error_summary = (
                f"Pemulihan dari trash berhasil memulihkan {restored_count} operasi dan gagal pada {error_count} operasi."
                if label == "Pulihkan dari Trash Internal"
                else f"Undo untuk '{label}' berhasil memulihkan {restored_count} operasi dan gagal pada {error_count} operasi."
            )
            self._record_history(
                error_action,
                "Sebagian gagal" if restored_count > 0 else "Gagal",
                f"{label} | Berhasil: {restored_count} | Gagal: {error_count}",
                "warning" if restored_count > 0 else "error",
            )
            self.status_label.setText(
                "Pemulihan file dari trash belum sepenuhnya berhasil."
                if label == "Pulihkan dari Trash Internal"
                else f"Undo untuk '{label}' belum sepenuhnya berhasil."
            )
            error_details = f"{detail}{os.linesep}{os.linesep}{os.linesep.join(errors[:10])}"
            QTimer.singleShot(
                180,
                lambda t=error_title, s=error_summary, d=error_details: self.show_error_dialog(t, s, d)
            )
            return

        self._remove_trash_entries(restored_trash_entry_ids)
        self._remove_trash_entries_from_undo_stack(restored_trash_entry_ids)
        self._cleanup_undo_action_dir(action_dir)
        success_action = "Undo"
        self._record_history(success_action, "Sukses", f"Aksi: {label}\nMode: {detail}\nOperasi dipulihkan: {restored_count}", "success", file_name=action_file_hint)
        if label in {"Penghapusan file", "Sinkronisasi folder pembanding", "Pindah file terpilih", "Pulihkan dari Trash Internal"}:
            self.clear_results(reset_status=False)
            if label == "Pulihkan dari Trash Internal":
                self.status_label.setText(
                    "File dari trash berhasil dipulihkan. Hasil scan dibersihkan, silakan scan ulang untuk melihat kondisi terbaru."
                )
            else:
                self.status_label.setText(
                    f"Undo berhasil untuk aksi: {label}. Hasil scan dibersihkan, silakan scan ulang untuk melihat kondisi terbaru."
                )
        else:
            if label == "Pulihkan dari Trash Internal":
                self.status_label.setText("File dari trash berhasil dipulihkan.")
            else:
                self.status_label.setText(f"Undo berhasil untuk aksi: {label}.")
        success_details = f"{detail}{os.linesep}Operasi dipulihkan: {restored_count}"
        self.show_success_dialog(
            "Pemulihan Trash Berhasil" if label == "Pulihkan dari Trash Internal" else "Undo Berhasil",
            "File yang dipilih berhasil dikembalikan dari trash internal."
            if label == "Pulihkan dari Trash Internal"
            else f"Aksi '{label}' berhasil dibatalkan.",
            success_details,
        )

    def show_error_dialog(self, title: str, summary: str, details: str) -> None:
        self.status_label.setText(summary or "Terjadi kesalahan pada aplikasi.")
        try:
            dialog = ErrorOverlayDialog(self, title, summary, details)
            dialog.exec()
        except Exception:
            QMessageBox.critical(self, title or APP_TITLE, summary or "Terjadi kesalahan pada aplikasi.")

    def show_success_dialog(self, title: str, summary: str, details: str = "") -> None:
        self.status_label.setText(summary or "Aksi berhasil diproses.")
        try:
            dialog = SuccessOverlayDialog(self, title, summary, details)
            dialog.exec()
        except Exception:
            QMessageBox.information(self, title or APP_TITLE, summary or "Aksi berhasil diproses.")

    def _history_detail_text(self, detail: str) -> str:
        normalized_lines = [line.strip() for line in str(detail).splitlines() if line.strip()]
        if not normalized_lines:
            return "-"
        return "\n".join(normalized_lines)

    def _record_history(self, action: str, status: str, detail: str, tone: str = "info", file_name: str = "-") -> None:
        if file_name == "-":
            for line in str(detail).splitlines():
                line = line.strip()
                for keyword in ("Sumber: ", "File: ", "Path asli: ", "Tujuan: ", "Ke: ", "Path: "):
                    if line.startswith(keyword):
                        parts = line.split(": ", 1)
                        if len(parts) == 2:
                            val = parts[1].strip()
                            extracted_name = Path(val).name
                            if extracted_name:
                                file_name = extracted_name
                                break
                if file_name != "-":
                    break
                    
        entry = HistoryEntry(
            timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            action=action,
            file_name=file_name,
            status=status,
            detail=self._history_detail_text(detail),
            tone=tone,
        )
        self.history_entries.insert(0, entry)
        if len(self.history_entries) > self.history_limit:
            self.history_entries = self.history_entries[: self.history_limit]
        self.history_model.set_rows(self.history_entries)
        if hasattr(self, "history_table"):
            self.history_table.resizeRowsToContents()
        self._update_history_empty_state()
        self._refresh_history_summary()

    def _refresh_history_summary(self) -> None:
        count = len(self.history_entries)
        self.history_count_badge.setText(f"{count} aksi")
        success_count = sum(1 for entry in self.history_entries if entry.tone == "success")
        issue_count = sum(1 for entry in self.history_entries if entry.tone in {"warning", "error"})
        if hasattr(self, "history_total_value"):
            self.history_total_value.setText(str(count))
        if hasattr(self, "history_success_value"):
            self.history_success_value.setText(str(success_count))
        if hasattr(self, "history_issue_value"):
            self.history_issue_value.setText(str(issue_count))
        if hasattr(self, "history_undo_value"):
            self.history_undo_value.setText(f"{len(self.undo_stack)} siap")
        if hasattr(self, "history_summary_label"):
            if count:
                latest_entry = self.history_entries[0]
                self.history_summary_label.setText(
                    f"Aksi terbaru: {latest_entry.action} ({latest_entry.status}) pada {latest_entry.timestamp}."
                )
            else:
                self.history_summary_label.setText("Belum ada aktivitas yang tercatat.")
        self.clear_history_button.setEnabled(count > 0)
        self._refresh_undo_button()

    def _update_history_empty_state(self) -> None:
        if not hasattr(self, "history_stack"):
            return
        if self.history_entries:
            self.history_stack.setCurrentWidget(self.history_table)
        else:
            self.history_stack.setCurrentWidget(self.history_empty_state)
        self._refresh_history_summary()

    def clear_history(self) -> None:
        self.history_entries = []
        self.history_model.set_rows([])
        self._update_history_empty_state()
        self.status_label.setText("Riwayat aksi dibersihkan.")

    def _refresh_undo_button(self) -> None:
        if not hasattr(self, "undo_button"):
            return
        has_undo = bool(self.undo_stack) and not (self.undo_thread and self.undo_thread.is_alive())
        self.undo_button.setEnabled(has_undo)
        if self.undo_stack:
            last_action = self.undo_stack[-1]
            self.undo_button.setToolTip(f"Batalkan aksi terakhir: {last_action.label}")
        else:
            self.undo_button.setToolTip("Belum ada aksi yang bisa di-undo.")

    def _refresh_trash_page(self) -> None:
        if not hasattr(self, "trash_table"):
            return

        entries = list(self.trash_entries)
        selected_ids = set(self._selected_trash_entry_ids())
        self.trash_table.setRowCount(len(entries))

        total_size = sum(entry.size for entry in entries)
        if hasattr(self, "trash_summary_label"):
            self.trash_summary_label.setText(
                f"{len(entries)} file tersimpan dengan total ukuran {self._format_size(total_size)}."
                if entries
                else "Trash masih kosong. File yang dihapus dengan mode ini akan tersimpan aman di sini."
            )
        if hasattr(self, "trash_count_value"):
            self.trash_count_value.setText(str(len(entries)))
        if hasattr(self, "trash_size_value"):
            self.trash_size_value.setText(self._format_size(total_size))

        for row_index, entry in enumerate(entries):
            self.trash_table.setCellWidget(row_index, 0, self._create_trash_checkbox(entry.entry_id, entry.entry_id in selected_ids))

            select_item = QTableWidgetItem()
            select_item.setFlags(Qt.ItemIsEnabled)
            select_item.setData(Qt.UserRole, entry.entry_id)
            self.trash_table.setItem(row_index, 0, select_item)

            file_name_item = QTableWidgetItem(Path(entry.original_path).name)
            file_name_item.setToolTip(entry.original_path)
            self.trash_table.setItem(row_index, 1, file_name_item)

            original_path_item = QTableWidgetItem(entry.original_path)
            original_path_item.setToolTip(entry.original_path)
            self.trash_table.setItem(row_index, 2, original_path_item)

            deleted_at_item = QTableWidgetItem(entry.deleted_at)
            self.trash_table.setItem(row_index, 3, deleted_at_item)

            size_item = QTableWidgetItem(self._format_size(entry.size))
            size_item.setTextAlignment(Qt.AlignCenter)
            self.trash_table.setItem(row_index, 4, size_item)

            self.trash_table.setCellWidget(row_index, 5, self._create_trash_undo_button(entry.entry_id))

        self.trash_stack.setCurrentWidget(self.trash_table if entries else self.trash_empty_state)
        self.trash_delete_all_button.setEnabled(bool(entries))
        self._update_trash_selection_state()
        self._update_trash_sidebar_badge()
        
        # Simpan data terbaru ke json database setiap kali ada refresh
        self._save_trash_db()

    def _toggle_trash_row_check(self, row_index: int, column_index: int) -> None:
        if column_index in {0, 5} or not hasattr(self, "trash_table"):
            return

        checkbox = self._trash_row_checkbox(row_index)
        if checkbox is None:
            return

        checkbox.setChecked(not checkbox.isChecked())

    def _update_trash_selection_state(self) -> None:
        selected_count = len(self._selected_trash_entry_ids()) if hasattr(self, "trash_table") else 0
        has_entries = bool(self.trash_entries)
        if hasattr(self, "trash_selection_value"):
            self.trash_selection_value.setText(f"{selected_count} dipilih")
        if hasattr(self, "trash_restore_selected_button"):
            self.trash_restore_selected_button.setEnabled(selected_count > 0)
        if hasattr(self, "trash_delete_selected_button"):
            self.trash_delete_selected_button.setEnabled(selected_count > 0)
        if hasattr(self, "trash_delete_all_button"):
            self.trash_delete_all_button.setEnabled(has_entries)

    def _update_trash_sidebar_badge(self) -> None:
        if not hasattr(self, "trash_nav_badge") or not hasattr(self, "btn_nav_trash"):
            return

        count = len(self.trash_entries)
        if count <= 0:
            self.trash_nav_badge.hide()
            self.btn_nav_trash.setToolTip("Trash Internal")
            return

        badge_text = "99+" if count > 99 else str(count)
        badge_width = max(22, self.trash_nav_badge.fontMetrics().horizontalAdvance(badge_text) + 14)
        self.trash_nav_badge.setText(badge_text)
        self.trash_nav_badge.setFixedSize(badge_width, 22)

        button_rect = self.btn_nav_trash.geometry()
        badge_x = min(
            self.trash_nav_host.width() - badge_width,
            button_rect.x() + button_rect.width() - max(10, badge_width // 2),
        )
        badge_y = max(0, button_rect.y() - 4)
        self.trash_nav_badge.move(badge_x, badge_y)
        self.trash_nav_badge.show()
        self.trash_nav_badge.raise_()
        self.btn_nav_trash.setToolTip(f"Trash Internal ({count} file)")

    def _selected_trash_entry_ids(self) -> List[str]:
        selected_ids: List[str] = []
        if not hasattr(self, "trash_table"):
            return selected_ids
        for row_index in range(self.trash_table.rowCount()):
            checkbox = self._trash_row_checkbox(row_index)
            if checkbox is None or not checkbox.isChecked():
                continue
            entry_id = checkbox.property("trashEntryId")
            if entry_id:
                selected_ids.append(str(entry_id))
        return selected_ids

    def _trash_row_checkbox(self, row_index: int) -> Optional[QCheckBox]:
        if not hasattr(self, "trash_table"):
            return None

        checkbox_host = self.trash_table.cellWidget(row_index, 0)
        if checkbox_host is None:
            return None
        return checkbox_host.findChild(QCheckBox)

    def _find_trash_entry(self, entry_id: str) -> Optional[TrashEntry]:
        for entry in self.trash_entries:
            if entry.entry_id == entry_id:
                return entry
        return None

    def _trash_storage_dir(self) -> Path:
        trash_dir = self.undo_root / "trash_items"
        trash_dir.mkdir(parents=True, exist_ok=True)
        return trash_dir

    def _normalize_trash_entries(self) -> bool:
        normalized_entries: List[TrashEntry] = []
        seen_entry_ids: set[str] = set()
        seen_trash_paths: set[str] = set()
        changed = False

        for entry in self.trash_entries:
            entry_id = str(entry.entry_id).strip()
            trash_path = str(Path(entry.trash_path))
            if not entry_id or not trash_path:
                changed = True
                continue
            if entry_id in seen_entry_ids or trash_path in seen_trash_paths:
                changed = True
                continue
            if not Path(trash_path).exists():
                changed = True
                continue

            seen_entry_ids.add(entry_id)
            seen_trash_paths.add(trash_path)
            normalized_entries.append(
                TrashEntry(
                    entry_id=entry_id,
                    original_path=str(entry.original_path),
                    trash_path=trash_path,
                    deleted_at=str(entry.deleted_at),
                    size=int(entry.size),
                )
            )

        if changed or len(normalized_entries) != len(self.trash_entries):
            self.trash_entries = normalized_entries
            return True
        return False

    def _load_trash_db(self) -> None:
        if not hasattr(self, "trash_db_path") or not self.trash_db_path.exists():
            return
        
        try:
            with open(self.trash_db_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    for item in data:
                        self.trash_entries.append(
                            TrashEntry(
                                entry_id=item.get("entry_id", ""),
                                original_path=item.get("original_path", ""),
                                trash_path=str(Path(item.get("trash_path", ""))),
                                deleted_at=item.get("deleted_at", ""),
                                size=item.get("size", 0),
                            )
                        )
            if self._normalize_trash_entries():
                self._save_trash_db()
        except Exception as e:
            print(f"Error loading trash DB: {e}")

    def _save_trash_db(self) -> None:
        if not hasattr(self, "trash_db_path"):
            return
            
        try:
            self._normalize_trash_entries()
            # Pastikan folder app data ada
            self.trash_db_path.parent.mkdir(parents=True, exist_ok=True)
            
            data = [
                {
                    "entry_id": entry.entry_id,
                    "original_path": entry.original_path,
                    "trash_path": entry.trash_path,
                    "deleted_at": entry.deleted_at,
                    "size": entry.size,
                }
                for entry in self.trash_entries
            ]
            with open(self.trash_db_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving trash DB: {e}")

    def _remove_trash_entries(self, entry_ids: List[str]) -> None:
        if not entry_ids:
            return
        wanted = set(entry_ids)
        self.trash_entries = [entry for entry in self.trash_entries if entry.entry_id not in wanted]
        self._refresh_trash_page()

    def _remove_trash_entries_from_undo_stack(self, entry_ids: List[str]) -> None:
        if not entry_ids:
            return
        wanted = set(entry_ids)
        updated_stack: List[UndoAction] = []
        for action in self.undo_stack:
            filtered_operations = [operation for operation in action.operations if operation.get("trash_entry_id") not in wanted]
            if filtered_operations:
                action.operations = filtered_operations
                updated_stack.append(action)
            else:
                self._cleanup_undo_action_dir(action.action_dir)
        self.undo_stack = updated_stack
        self._refresh_undo_button()

    def _add_trash_entries(self, payload_items: List[Dict[str, Any]]) -> None:
        for item in payload_items:
            entry_id = str(item.get("entry_id", "")).strip()
            if not entry_id:
                continue
            self.trash_entries.append(
                TrashEntry(
                    entry_id=entry_id,
                    original_path=str(item.get("original_path", "")),
                    trash_path=str(item.get("trash_path", "")),
                    deleted_at=str(item.get("deleted_at", "")),
                    size=int(item.get("size", 0)),
                )
            )
        self._normalize_trash_entries()
        self._refresh_trash_page()

    def restore_trash_entries(self, entry_ids: List[str]) -> None:
        selected_entries = [entry for entry_id in entry_ids if (entry := self._find_trash_entry(entry_id)) is not None]
        if not selected_entries:
            self.show_error_dialog(
                "Peringatan", 
                "Pilih minimal satu file dari trash.", 
                "Anda harus mencentang kotak di sebelah file yang ingin dipulihkan."
            )
            return
        if self.undo_thread and self.undo_thread.is_alive():
            self.show_error_dialog(
                "Perhatian", 
                "Proses Masih Berjalan", 
                "Mohon tunggu hingga aksi sebelumnya selesai sebelum memulihkan file lain."
            )
            return

        if len(selected_entries) == 1:
            action_detail = f"File: {Path(selected_entries[0].trash_path).name}\nKe: {selected_entries[0].original_path}"
        else:
            action_detail = f"Jumlah file: {len(selected_entries)}"

        undo_action = UndoAction(
            label="Pulihkan dari Trash Internal",
            detail=action_detail,
            operations=[
                self._serialize_undo_operation(
                    "move_path",
                    source=entry.trash_path,
                    destination=entry.original_path,
                    trash_entry_id=entry.entry_id,
                )
                for entry in selected_entries
            ],
            file_name_hint=selected_entries[0].original_path if len(selected_entries)==1 else f"{len(selected_entries)} file"
        )

        self.status_label.setText(f"Memulihkan {len(selected_entries)} file dari trash internal...")
        self._record_history("Restore dari Trash", "Diproses", action_detail, "info")
        self._set_undo_processing_state(True)
        self.undo_button.setEnabled(False)
        dialog = ProcessingOverlayDialog(
            self,
            "Memulihkan File dari Trash",
            f"Sedang mengembalikan {len(selected_entries)} file ke lokasi asal.",
        )
        self.undo_processing_dialog = dialog
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()

        self.undo_thread = threading.Thread(target=self._undo_worker, args=(undo_action,), daemon=True)
        self.undo_thread.start()

    def restore_selected_trash_entries(self) -> None:
        self.restore_trash_entries(self._selected_trash_entry_ids())

    def delete_selected_trash_entries_permanently(self, force_entry_ids: Optional[List[str]] = None) -> None:
        if isinstance(force_entry_ids, list):
            entry_ids = force_entry_ids
        else:
            entry_ids = self._selected_trash_entry_ids()
            
        if not entry_ids:
            self.show_error_dialog(
                "Peringatan",
                "Pilih minimal satu file dari Trash Internal.",
                "Centang kotak di sisi kiri tabel di samping nama file untuk memilih item."
            )
            return

        entries = [self._find_trash_entry(eid) for eid in entry_ids]
        entries = [e for e in entries if e is not None]
        preview = "\n".join(f"[{Path(e.original_path).name}] {e.original_path}" for e in entries[:10])
        extra_details = "" if len(entries) <= 10 else f"\n... dan {len(entries) - 10} file lainnya"
        
        is_all_entries = len(entry_ids) == len(self.trash_entries) and len(self.trash_entries) > 0
        dialog_title = "Kosongkan Trash Internal" if is_all_entries else "Hapus Permanen dari Trash"

        confirm_dialog = ConfirmOverlayDialog(
            self,
            dialog_title,
            f"{len(entry_ids)} file di trash internal akan dihapus secara permanen.",
            f"{preview}{extra_details}",
            detail_title="Daftar File",
            confirm_button_text="Hapus Permanen",
            confirm_footnote="Tindakan ini tidak dapat dibatalkan atau dikembalikan.",
        )
        
        confirm_dialog.confirmRequested.connect(confirm_dialog.accept)
        
        if confirm_dialog.exec() != QDialog.Accepted:
            return

        deleted_count = 0
        errors: List[str] = []
        for entry_id in entry_ids:
            entry = self._find_trash_entry(entry_id)
            if entry is None:
                continue
            try:
                trash_path = Path(entry.trash_path)
                if trash_path.exists():
                    trash_path.unlink()
                deleted_count += 1
            except Exception as exc:
                errors.append(f"{entry.original_path}: {exc}")

        self._remove_trash_entries(entry_ids)
        self._remove_trash_entries_from_undo_stack(entry_ids)
        
        entry_list = [self._find_trash_entry(eid) for eid in entry_ids]
        entry_list = [e for e in entry_list if e is not None]
        if len(entry_list) == 1:
            single_val = f"File: {Path(entry_list[0].trash_path).name}\nPath asli: {entry_list[0].original_path}\n"
        else:
            single_val = ""
            
        self._record_history(
            "Hapus permanen dari Trash",
            "Sebagian gagal" if errors and deleted_count > 0 else "Sukses" if not errors else "Gagal",
            f"{single_val}Berhasil: {deleted_count} | Gagal: {len(errors)}",
            "warning" if errors else "success",
        )
        if errors:
            self.show_error_dialog(
                "Hapus Permanen dari Trash Belum Selesai",
                f"{deleted_count} file berhasil dihapus permanen, {len(errors)} file gagal diproses.",
                os.linesep.join(errors[:10]),
            )
        else:
            self.show_success_dialog(
                "Trash Diperbarui",
                f"{deleted_count} file berhasil dihapus permanen dari trash.",
                "File yang sudah dihapus permanen tidak dapat dipulihkan kembali.",
            )

    def delete_all_trash_entries_permanently(self) -> None:
        if not self.trash_entries:
            self.show_error_dialog(
                "Info",
                "Trash internal sudah kosong.",
                "Tidak ada file yang bisa dihapus."
            )
            return

        all_entry_ids = [entry.entry_id for entry in self.trash_entries]
        self.delete_selected_trash_entries_permanently(all_entry_ids)

    def _create_undo_action_dir(self, label: str) -> Path:
        safe_label = "".join(char if char.isalnum() else "_" for char in label.lower()).strip("_") or "aksi"
        action_dir = self.undo_root / f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{safe_label}_{uuid4().hex[:8]}"
        action_dir.mkdir(parents=True, exist_ok=True)
        return action_dir

    def _cleanup_undo_action_dir(self, action_dir: str) -> None:
        if not action_dir:
            return
        try:
            shutil.rmtree(action_dir, ignore_errors=True)
        except Exception:
            pass

    def _push_undo_action(self, payload: Optional[Dict[str, Any]]) -> None:
        if not payload:
            return

        operations = payload.get("operations", [])
        if not isinstance(operations, list) or not operations:
            self._cleanup_undo_action_dir(str(payload.get("action_dir", "")))
            return

        action = UndoAction(
            label=str(payload.get("label", "Aksi terakhir")),
            detail=str(payload.get("detail", "")),
            operations=[
                {str(key): str(value) for key, value in item.items()}
                for item in operations
                if isinstance(item, dict)
            ],
            action_dir=str(payload.get("action_dir", "")),
            file_name_hint=str(payload.get("file_name", "-"))
        )
        if not action.operations:
            self._cleanup_undo_action_dir(action.action_dir)
            return

        self.undo_stack.append(action)
        while len(self.undo_stack) > self.undo_limit:
            discarded = self.undo_stack.pop(0)
            self._cleanup_undo_action_dir(discarded.action_dir)
        self._refresh_history_summary()

    @staticmethod
    def _serialize_undo_operation(kind: str, **kwargs: str) -> Dict[str, str]:
        operation = {"kind": kind}
        operation.update({key: str(value) for key, value in kwargs.items()})
        return operation

    def _backup_file_for_undo(self, source: Path, action_dir: Path, prefix: str) -> Path:
        backup_dir = action_dir / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f"{prefix}_{uuid4().hex}{source.suffix}"
        shutil.copy2(source, backup_path)
        return backup_path

    def undo_last_action(self) -> None:
        if self.scan_thread and self.scan_thread.is_alive():
            QMessageBox.information(self, APP_TITLE, "Tunggu sampai proses scan selesai sebelum menjalankan undo.")
            return
        if self.delete_thread and self.delete_thread.is_alive():
            QMessageBox.information(self, APP_TITLE, "Tunggu sampai proses penghapusan selesai sebelum menjalankan undo.")
            return
        if self.transfer_thread and self.transfer_thread.is_alive():
            QMessageBox.information(self, APP_TITLE, "Tunggu sampai proses salin/pindah selesai sebelum menjalankan undo.")
            return
        if self.undo_thread and self.undo_thread.is_alive():
            QMessageBox.information(self, APP_TITLE, "Undo aksi sebelumnya masih berjalan.")
            return
        if not self.undo_stack:
            QMessageBox.information(self, APP_TITLE, "Belum ada aksi yang bisa di-undo.")
            return

        action = self.undo_stack.pop()
        self._refresh_history_summary()
        self.status_label.setText(f"Menjalankan undo untuk aksi: {action.label}...")
        self._record_history("Undo", "Diproses", f"Aksi: {action.label}\nFile: {action.file_name_hint}", "info")
        self._set_undo_processing_state(True)
        self.undo_button.setEnabled(False)

        dialog = ProcessingOverlayDialog(self, "Undo Sedang Diproses", f"Mengembalikan aksi: {action.label}")
        self.undo_processing_dialog = dialog
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()

        self.undo_thread = threading.Thread(target=self._undo_worker, args=(action,), daemon=True)
        self.undo_thread.start()

    def _undo_worker(self, action: UndoAction) -> None:
        restored_count = 0
        errors: List[str] = []
        restored_trash_entry_ids: List[str] = []

        for operation in action.operations:
            kind = operation.get("kind", "")
            try:
                if kind == "delete_path":
                    path = Path(operation["path"])
                    if path.exists():
                        path.unlink()
                    restored_count += 1
                elif kind == "move_path":
                    source = Path(operation["source"])
                    destination = Path(operation["destination"])
                    if not source.exists():
                        raise FileNotFoundError(f"Path sumber undo tidak ditemukan: {source}")
                    if destination.exists():
                        raise FileExistsError(f"Path tujuan undo sudah ada: {destination}")
                    destination.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(source), str(destination))
                    restored_count += 1
                    if operation.get("trash_entry_id"):
                        restored_trash_entry_ids.append(str(operation["trash_entry_id"]))
                elif kind == "restore_copy":
                    backup = Path(operation["backup"])
                    destination = Path(operation["destination"])
                    if not backup.exists():
                        raise FileNotFoundError(f"Backup undo tidak ditemukan: {backup}")
                    destination.parent.mkdir(parents=True, exist_ok=True)
                    if destination.exists():
                        destination.unlink()
                    shutil.copy2(backup, destination)
                    restored_count += 1
                else:
                    raise ValueError(f"Operasi undo tidak dikenal: {kind}")
            except Exception as exc:
                errors.append(f"{kind}: {exc}")

        self.ui_queue.put(
            (
                "undo_done",
                {
                    "label": action.label,
                    "detail": action.detail,
                    "restored_count": restored_count,
                    "error_count": len(errors),
                    "errors": errors,
                    "action_dir": action.action_dir,
                    "restored_trash_entry_ids": restored_trash_entry_ids,
                    "file_name_hint": action.file_name_hint,
                },
            )
        )

    def _refresh_stats(self) -> None:
        total = len(self.result_rows)
        exact = sum(1 for row in self.result_rows if row.exact_matches)
        diff = sum(1 for row in self.result_rows if row.same_name_different_content)
        only = sum(1 for row in self.result_rows if row.only_in_target)

        self.stat_labels["total"].setText(str(total))
        self.stat_labels["exact"].setText(str(exact))
        self.stat_labels["diff"].setText(str(diff))
        self.stat_labels["only"].setText(str(only))

    def _apply_default_table_widths(self) -> None:
        default_widths = [170, 240, 180, 110, 120, 260, 260, 140, 110]
        for column, width in enumerate(default_widths):
            self.results_table.setColumnWidth(column, width)

    def _apply_responsive_column_widths(self) -> None:
        min_widths = [140, 200, 170, 110, 120, 260, 260, 140, 110]
        max_widths = [240, 420, 320, 140, 180, 520, 520, 220, 150]
        flexible_columns = [1, 2, 5, 6]

        self.results_table.resizeColumnsToContents()

        computed_widths: List[int] = []
        for column in range(self.table_model.columnCount()):
            content_width = self.results_table.columnWidth(column) + 18
            target_width = max(min_widths[column], min(content_width, max_widths[column]))
            computed_widths.append(target_width)

        viewport_width = self.results_table.viewport().width()
        used_width = sum(computed_widths)

        if viewport_width > used_width and flexible_columns:
            extra = viewport_width - used_width
            per_column = max(extra // len(flexible_columns), 0)
            for column in flexible_columns:
                computed_widths[column] = min(computed_widths[column] + per_column, max_widths[column])

        for column, width in enumerate(computed_widths):
            self.results_table.setColumnWidth(column, width)

    def _update_table_empty_state(self) -> None:
        if not self.result_rows:
            self.empty_state_title.setText("Belum ada hasil scan")
            self.empty_state_description.setText(
                "Pilih folder target dan folder pembanding, lalu jalankan Scan dan Bandingkan untuk melihat hasil."
            )
            self.table_stack.setCurrentWidget(self.table_empty_state)
            return

        if self.table_proxy.rowCount() == 0:
            self.empty_state_title.setText("Tidak ada hasil yang cocok")
            self.empty_state_description.setText(
                "Coba ubah kata kunci pencarian, quick filter status, atau opsi filter tampilan untuk menampilkan data."
            )
            self.table_stack.setCurrentWidget(self.table_empty_state)
            return

        self.table_stack.setCurrentWidget(self.results_table)

    def _populate_table(self, recompute_widths: bool = True, refresh_source: bool = True) -> None:
        previously_selected_target = self.current_selected_result.target_path if self.current_selected_result else None

        self.results_table.setUpdatesEnabled(False)
        if refresh_source:
            self.table_model.set_rows(self.result_rows)
        self.table_proxy.set_matches_only(self.show_only_matches_checkbox.isChecked())
        self.table_proxy.set_status_filter(self._current_quick_filter())
        self.table_proxy.set_search_text(self.search_input.text() if hasattr(self, "search_input") else "")

        if self.results_table.selectionModel() is not None:
            self.results_table.clearSelection()
            self.results_table.selectionModel().setCurrentIndex(QModelIndex(), QItemSelectionModel.SelectionFlag.NoUpdate)

        if recompute_widths and self.table_proxy.rowCount() > 0:
            self._apply_responsive_column_widths()
        elif recompute_widths:
            self._apply_default_table_widths()

        self._update_table_empty_state()
        self.results_table.setUpdatesEnabled(True)

        selected_row_index = self._find_table_row_by_target_path(previously_selected_target) if previously_selected_target else None
        if selected_row_index is not None:
            proxy_index = self.table_proxy.index(selected_row_index, 0)
            self.results_table.selectRow(selected_row_index)
            self.results_table.setCurrentIndex(proxy_index)
            self.results_table.scrollTo(proxy_index)
            self._update_detail_from_row(selected_row_index)
        else:
            self.current_selected_result = None
            self._reset_detail_panel()

    def _schedule_search_filter(self) -> None:
        self.filter_timer.start()

    def _apply_debounced_filter(self) -> None:
        self._populate_table(recompute_widths=False, refresh_source=False)
        if self.result_rows:
            self.status_label.setText("Hasil pencarian diperbarui.")

    def _on_filter_changed(self, *args) -> None:
        self.filter_timer.stop()
        self._populate_table(recompute_widths=False, refresh_source=False)
        if self.result_rows:
            self.status_label.setText("Filter tampilan diperbarui.")

    def _current_quick_filter(self) -> str:
        for key, button in getattr(self, "quick_filter_buttons", {}).items():
            if button.isChecked():
                return key
        return "all"

    def _reset_quick_filters(self) -> None:
        self.filter_timer.stop()
        if hasattr(self, "search_input"):
            self.search_input.blockSignals(True)
            self.search_input.clear()
            self.search_input.blockSignals(False)
        if hasattr(self, "quick_filter_buttons"):
            self.quick_filter_buttons["all"].setChecked(True)
        self._on_filter_changed()

    def _on_table_selection_changed(self, *args) -> None:
        selected_rows = self.results_table.selectionModel().selectedRows()
        if not selected_rows:
            self.current_selected_result = None
            self._reset_detail_panel()
            return
        self._update_detail_from_row(selected_rows[0].row())

    def _on_current_row_changed(self, current: QModelIndex, previous: QModelIndex) -> None:
        if current.isValid():
            self._update_detail_from_row(current.row())

    def _on_table_focus_released(self) -> None:
        QTimer.singleShot(0, self._handle_table_focus_release)

    def _handle_table_focus_release(self) -> None:
        focus_widget = QApplication.focusWidget()
        if focus_widget is not None and hasattr(self, "detail_panel"):
            if focus_widget is self.detail_panel or self.detail_panel.isAncestorOf(focus_widget):
                return
        if focus_widget is not None and self.file_detail_dialog is not None and self.file_detail_dialog.isVisible():
            if focus_widget is self.file_detail_dialog or self.file_detail_dialog.isAncestorOf(focus_widget):
                return
        self.current_selected_result = None
        self._reset_detail_panel()

    def _result_for_table_row(self, row_index: int) -> Optional[MatchResult]:
        if row_index < 0:
            return None
        proxy_index = self.table_proxy.index(row_index, 0)
        if not proxy_index.isValid():
            return None
        result = proxy_index.data(Qt.UserRole)
        return result if isinstance(result, MatchResult) else None

    def _find_table_row_by_target_path(self, target_path: Path) -> Optional[int]:
        for row_index in range(self.table_proxy.rowCount()):
            result = self._result_for_table_row(row_index)
            if result is not None and result.target_path == target_path:
                return row_index
        return None

    def _update_detail_from_row(self, row_index: int) -> None:
        result = self._result_for_table_row(row_index)
        if result is None:
            self.current_selected_result = None
            self._reset_detail_panel()
            return

        self.current_selected_result = result
        self.detail_labels["status"].setText(result.status_text)
        self.detail_labels["target"].setText(str(result.target_path))
        self.detail_labels["relative"].setText(result.target_relative_path)
        self.detail_labels["size"].setText(f"{self._format_size(result.size)} ({result.size} bytes)")
        self.detail_labels["found"].setText(result.exact_folder_labels)
        self.detail_labels["missing"].setText(result.missing_display_text)
        self.detail_labels["mode"].setText(result.match_type)
        self.detail_labels["exact"].setText(result.exact_paths_text)
        self.detail_labels["diff"].setText(result.diff_paths_text)

    def _compare_sync_suggestion_text(self, result: Optional[MatchResult]) -> str:
        if result is None:
            return ""

        plan = self._compare_sync_plan(result)
        create_labels = [label for label in plan["create_labels"] if label]
        replace_labels = [label for label in plan["replace_labels"] if label]
        if not plan["source"] or (not create_labels and not replace_labels):
            return ""

        if create_labels and replace_labels:
            return (
                f"File ini bisa disalin atau dipindahkan ke folder pembanding. "
                f"Tujuan baru: {', '.join(create_labels)}. "
                f"Tujuan yang akan diganti: {', '.join(replace_labels)}."
            )
        if replace_labels:
            return (
                "Versi file pada folder pembanding berbeda. "
                f"Anda dapat menyalin atau memindahkan versi dari Folder A untuk menggantikan file di {', '.join(replace_labels)}."
            )
        return (
            "File ini belum ada di beberapa folder pembanding. "
            f"Anda dapat menyalin atau memindahkannya ke {', '.join(create_labels)}."
        )

    def _show_results_table_context_menu(self, pos: QPoint) -> None:
        index = self.results_table.indexAt(pos)
        if not index.isValid():
            return

        result = self.table_proxy.data(index, Qt.UserRole)
        if not isinstance(result, MatchResult):
            return

        menu = QMenu(self)
        reveal_action = QAction("Tampilkan di File Explorer", self)
        
        menu.addAction(reveal_action)

        action = menu.exec(self.results_table.viewport().mapToGlobal(pos))
        if action == reveal_action:
            self._reveal_in_explorer(str(result.target_path))

    def _show_trash_table_context_menu(self, pos: QPoint) -> None:
        index = self.trash_table.indexAt(pos)
        if not index.isValid():
            return

        # Ambil data entry_id dari kolom 0
        entry_id_item = self.trash_table.item(index.row(), 0)
        if not entry_id_item:
            return

        entry_id = entry_id_item.data(Qt.UserRole)
        if not entry_id:
            return

        entry = self._find_trash_entry(str(entry_id))
        if not entry:
            return

        menu = QMenu(self)
        reveal_trash_action = QAction("Tampilkan File di Trash", self)
        
        menu.addAction(reveal_trash_action)

        action = menu.exec(self.trash_table.viewport().mapToGlobal(pos))
        if action == reveal_trash_action:
            self._reveal_in_explorer(str(entry.trash_path))

    def _reveal_in_explorer(self, path: str) -> None:
        try:
            if sys.platform == "win32":
                subprocess.Popen(f'explorer /select,"{os.path.normpath(path)}"')
            elif sys.platform == "darwin":  # macOS
                subprocess.Popen(['open', '-R', path])
            else:  # linux
                subprocess.Popen(['xdg-open', os.path.dirname(path)])
        except Exception as e:
            QMessageBox.warning(self, "Gagal Membuka File", f"Tidak dapat membuka file explorer:\n\n{str(e)}")

    def _open_detail_dialog_from_index(self, index: QModelIndex) -> None:
        if not index.isValid():
            return

        self._update_detail_from_row(index.row())
        result = self.current_selected_result
        if result is None:
            return

        missing_labels = self._actual_missing_compare_labels(result)
        suggestion_text = self._compare_sync_suggestion_text(result)
        plan = self._compare_sync_plan(result)
        show_compare_actions = bool(plan["source"] and (plan["create_labels"] or plan["replace_labels"]))

        dialog = FileDetailOverlayDialog(
            self,
            result,
            missing_labels,
            suggestion_text,
            show_compare_actions=show_compare_actions,
        )
        self.file_detail_dialog = dialog
        try:
            dialog.exec()
        finally:
            self.file_detail_dialog = None

    def _reset_detail_panel(self) -> None:
        defaults = {
            "status": "Pilih satu item untuk melihat detail.",
            "target": "-",
            "relative": "-",
            "size": "-",
            "found": "-",
            "missing": "-",
            "mode": "-",
            "exact": "-",
            "diff": "-",
        }
        for key, value in defaults.items():
            self.detail_labels[key].setText(value)
        if hasattr(self, "detail_panel"):
            self.detail_panel.setVisible(False)
        if hasattr(self, "missing_compare_suggestion"):
            self.missing_compare_suggestion.setVisible(False)

    def copy_selected_path(self) -> None:
        if self.current_selected_result is None:
            QMessageBox.information(self, APP_TITLE, "Pilih satu item terlebih dahulu.")
            return

        QApplication.clipboard().setText(str(self.current_selected_result.target_path))
        self.status_label.setText("Path file target berhasil disalin ke clipboard.")
        self._record_history("Salin path", "Sukses", str(self.current_selected_result.target_path), "info")

    def _selected_results(self) -> List[MatchResult]:
        selection_model = self.results_table.selectionModel()
        if selection_model is None:
            return []

        selected_rows = sorted({index.row() for index in selection_model.selectedRows()})
        results = [result for row_index in selected_rows if (result := self._result_for_table_row(row_index)) is not None]
        if results:
            return results
        if self.current_selected_result is not None:
            return [self.current_selected_result]
        return []

    def _sources_for_result(self, result: MatchResult) -> List[Tuple[str, str, Path]]:
        sources: List[Tuple[str, str, Path]] = [("A", result.target_relative_path, result.target_path)]
        for record in result.exact_matches:
            sources.append((record.base_label, record.relative_path, record.path))
        for record in result.same_name_different_content:
            sources.append((record.base_label, record.relative_path, record.path))

        deduplicated: List[Tuple[str, str, Path]] = []
        seen_paths: set[str] = set()
        for base_label, relative_path, path in sources:
            normalized = os.path.normcase(str(path))
            if normalized in seen_paths:
                continue
            seen_paths.add(normalized)
            deduplicated.append((base_label, relative_path, path))
        return deduplicated

    def _compare_folder_path_for_label(self, label: str) -> Optional[Path]:
        if not label.startswith("F"):
            return None
        try:
            index = int(label[1:]) - 1
        except ValueError:
            return None
        # Labels F1, F2, … are assigned during scan by enumerating only the
        # *non-empty* compare-folder rows (same as in start_scan). We must
        # apply the same filter here so that an empty row in the middle does
        # not cause the label to resolve to the wrong (or missing) folder.
        filled_paths = [
            self._normalize_folder_path(row["edit"].text())
            for row in self.compare_folder_rows
            if self._normalize_folder_path(row["edit"].text())
        ]
        if not (0 <= index < len(filled_paths)):
            return None
        raw_path = filled_paths[index]
        path = Path(raw_path)
        return path if path.exists() and path.is_dir() else None

    def _actual_missing_compare_labels(self, result: MatchResult) -> List[str]:
        labels: List[str] = []
        for missing_label in result.missing_from_folders:
            compare_folder = self._compare_folder_path_for_label(missing_label)
            if compare_folder is None:
                continue
            destination = compare_folder / result.target_relative_path
            if not destination.exists():
                labels.append(missing_label)
        return labels

    def _compare_sync_plan(self, result: MatchResult) -> Dict[str, Any]:
        source = result.target_path if result.target_path.exists() and result.target_path.is_file() else None
        if source is None:
            return {"source": None, "create": [], "replace": [], "create_labels": [], "replace_labels": []}

        create_paths: List[str] = []
        create_labels: List[str] = []
        for missing_label in self._actual_missing_compare_labels(result):
            compare_folder = self._compare_folder_path_for_label(missing_label)
            if compare_folder is None:
                continue
            create_paths.append(str(compare_folder / result.target_relative_path))
            create_labels.append(missing_label)

        replace_map: Dict[str, str] = {}
        for record in result.same_name_different_content:
            if not record.path:
                continue
            replace_map[record.base_label] = str(record.path)

        return {
            "source": str(source),
            "create": create_paths,
            "replace": list(replace_map.values()),
            "create_labels": create_labels,
            "replace_labels": list(replace_map.keys()),
        }

    def _refresh_missing_compare_suggestion(self, result: Optional[MatchResult]) -> None:
        if not hasattr(self, "missing_compare_suggestion"):
            return
        if result is None:
            self.missing_compare_suggestion.setVisible(False)
            return

        suggestion_text = self._compare_sync_suggestion_text(result)
        if not suggestion_text:
            self.missing_compare_suggestion.setVisible(False)
            return

        self.missing_compare_suggestion_label.setText(suggestion_text)
        self.missing_compare_suggestion.setVisible(True)

    def copy_to_compare_folders(self) -> None:
        self._sync_selected_result_to_compare_folders("copy")

    def move_to_compare_folders(self) -> None:
        self._sync_selected_result_to_compare_folders("move")

    def _sync_selected_result_to_compare_folders(self, operation: str, explicit_result: Optional["MatchResult"] = None) -> None:
        result = explicit_result if explicit_result is not None else self.current_selected_result
        if result is None:
            QMessageBox.information(self, APP_TITLE, "Pilih satu item terlebih dahulu.")
            return
        if self.transfer_thread and self.transfer_thread.is_alive():
            QMessageBox.information(self, APP_TITLE, "Proses salin/pindah file lain masih berjalan.")
            return

        plan = self._compare_sync_plan(result)
        create_paths = [str(path) for path in plan["create"]]
        replace_paths = [str(path) for path in plan["replace"]]
        create_labels = [str(label) for label in plan["create_labels"]]
        replace_labels = [str(label) for label in plan["replace_labels"]]
        source = str(plan["source"] or "")

        if not source or (not create_paths and not replace_paths):
            QMessageBox.information(
                self,
                APP_TITLE,
                "Tidak ada folder pembanding yang relevan untuk aksi ini.",
            )
            self._refresh_missing_compare_suggestion(result)
            return

        operation_text = "dipindahkan" if operation == "move" else "disalin"
        summary_lines = [f"File dari Folder A akan {operation_text} ke folder pembanding."]
        if create_labels:
            summary_lines.append(f"Tujuan baru: {', '.join(create_labels)}.")
        if replace_labels:
            summary_lines.append(f"Akan mengganti file yang sudah ada di: {', '.join(replace_labels)}.")
        summary_lines.append("Struktur perbandingan akan berubah dan hasil scan perlu diperbarui.")

        details_lines = [f"Sumber: {source}", f"Path relatif: {result.target_relative_path}"]
        if create_labels:
            details_lines.append(f"Folder pembanding yang akan ditambah: {', '.join(create_labels)}")
        if replace_labels:
            details_lines.append(f"Folder pembanding yang akan diganti: {', '.join(replace_labels)}")

        confirm_dialog = ConfirmOverlayDialog(
            self,
            "Konfirmasi Sinkronisasi Folder Pembanding",
            "\n".join(summary_lines),
            "\n".join(details_lines),
            detail_title="Rencana sinkronisasi",
            confirm_button_text="Ya, Sinkronkan",
            confirm_footnote="Periksa kembali folder pembanding tujuan sebelum melanjutkan sinkronisasi file.",
            processing_footnote="Sinkronisasi file sedang berjalan. Mohon tunggu sampai proses selesai.",
            processing_button_text="Menyinkronkan",
            success_title="Sinkronisasi Folder Pembanding Berhasil",
            success_detail_title="Ringkasan sinkronisasi",
            success_footnote="Sinkronisasi selesai. Anda dapat menutup dialog ini.",
        )
        self.transfer_confirm_dialog = confirm_dialog
        confirm_dialog.confirmRequested.connect(
            lambda src=source, creates=list(create_paths), replaces=list(replace_paths), create_lbls=list(create_labels), replace_lbls=list(replace_labels), rel=result.target_relative_path, op=operation: self._start_compare_sync_from_dialog(
                src, creates, replaces, create_lbls, replace_lbls, rel, op
            )
        )

        dialog_result = confirm_dialog.exec()
        if self.transfer_confirm_dialog is confirm_dialog:
            self.transfer_confirm_dialog = None
        confirm_dialog.deleteLater()

        if dialog_result != QDialog.Accepted:
            return

    def _start_compare_sync_from_dialog(
        self,
        source: str,
        create_paths: List[str],
        replace_paths: List[str],
        create_labels: List[str],
        replace_labels: List[str],
        relative_path: str,
        operation: str,
    ) -> None:
        if self.transfer_thread and self.transfer_thread.is_alive():
            return

        dialog = self.transfer_confirm_dialog
        if dialog is None:
            return

        operation_text = "pemindahan" if operation == "move" else "penyalinan"
        self.status_label.setText(
            "Memindahkan file ke folder pembanding..." if operation == "move" else "Menyalin file ke folder pembanding..."
        )
        dialog.set_processing(
            True,
            f"Sedang memproses {operation_text} file ke folder pembanding. Mohon tunggu...",
        )
        self._set_transfer_processing_state(True)
        dialog.flush_visual_state()
        self._record_history(
            "Sinkronisasi folder pembanding",
            "Diproses",
            (
                f"Operasi: {'pindah' if operation == 'move' else 'salin'}\n"
                f"Sumber: {source}\n"
                f"Target: {', '.join(create_labels + replace_labels) if (create_labels or replace_labels) else '-'}"
            ),
            "info",
            file_name=Path(source).name
        )

        QTimer.singleShot(
            0,
            lambda src=source, creates=list(create_paths), replaces=list(replace_paths), create_lbls=list(create_labels), replace_lbls=list(replace_labels), rel=relative_path, op=operation: self._launch_compare_sync_worker(
                src, creates, replaces, create_lbls, replace_lbls, rel, op
            ),
        )

    def _launch_compare_sync_worker(
        self,
        source: str,
        create_paths: List[str],
        replace_paths: List[str],
        create_labels: List[str],
        replace_labels: List[str],
        relative_path: str,
        operation: str,
    ) -> None:
        if self.transfer_thread and self.transfer_thread.is_alive():
            return
        self.transfer_thread = threading.Thread(
            target=self._compare_sync_worker,
            args=(source, create_paths, replace_paths, create_labels, replace_labels, relative_path, operation),
            daemon=True,
        )
        self.transfer_thread.start()

    def _compare_sync_worker(
        self,
        source_path: str,
        create_paths: List[str],
        replace_paths: List[str],
        create_labels: List[str],
        replace_labels: List[str],
        relative_path: str,
        operation: str,
    ) -> None:
        processed_count = 0
        errors: List[str] = []
        source = Path(source_path)
        all_targets = [(path, False) for path in create_paths] + [(path, True) for path in replace_paths]
        undo_operations: List[Dict[str, str]] = []
        action_dir = self._create_undo_action_dir("compare_sync")
        source_backup_path: Optional[Path] = None

        for destination_path, should_replace in all_targets:
            destination = Path(destination_path)
            try:
                if not source.exists() or not source.is_file():
                    raise FileNotFoundError(f"File sumber tidak ditemukan: {source}")

                destination.parent.mkdir(parents=True, exist_ok=True)
                if should_replace and destination.exists():
                    backup_path = self._backup_file_for_undo(destination, action_dir, "replace")
                    destination.unlink()
                    undo_operations.append(
                        self._serialize_undo_operation(
                            "restore_copy",
                            backup=str(backup_path),
                            destination=str(destination),
                        )
                    )
                elif not should_replace and destination.exists():
                    raise FileExistsError(f"File tujuan sudah ada: {destination}")

                shutil.copy2(source, destination)
                undo_operations.insert(0, self._serialize_undo_operation("delete_path", path=str(destination)))
                processed_count += 1
            except Exception as exc:
                errors.append(f"{destination}: {exc}")

        if operation == "move" and processed_count == len(all_targets) and source.exists():
            try:
                source_backup_path = self._backup_file_for_undo(source, action_dir, "source")
                source.unlink()
                undo_operations.append(
                    self._serialize_undo_operation(
                        "restore_copy",
                        backup=str(source_backup_path),
                        destination=str(source),
                    )
                )
            except Exception as exc:
                errors.append(f"{source}: {exc}")

        undo_payload = {
            "label": "Sinkronisasi folder pembanding",
            "detail": f"Path relatif: {relative_path}\nTarget: {', '.join(create_labels + replace_labels)}",
            "operations": undo_operations,
            "action_dir": str(action_dir),
            "file_name": Path(source_path).name,
        } if undo_operations else None
        if undo_payload is None:
            self._cleanup_undo_action_dir(str(action_dir))

        self.ui_queue.put(
            (
                "transfer_done",
                {
                    "operation": "move_compare_sync" if operation == "move" else "copy_compare_sync",
                    "processed_count": processed_count,
                    "error_count": len(errors),
                    "destination_root": ", ".join(create_labels + replace_labels),
                    "relative_path": relative_path,
                    "source_path": source_path,
                    "created_labels": create_labels,
                    "created_paths": create_paths,
                    "replaced_labels": replace_labels,
                    "errors": errors,
                    "undo_action": undo_payload,
                },
            )
        )

    @staticmethod
    def _unique_destination_path(destination: Path) -> Path:
        if not destination.exists():
            return destination

        counter = 2
        while True:
            candidate = destination.with_name(f"{destination.stem} ({counter}){destination.suffix}")
            if not candidate.exists():
                return candidate
            counter += 1

    def _show_transfer_notice(self, title: str, summary: str, details: str = "") -> None:
        self.show_error_dialog(title, summary, details)

    def transfer_selected_files(self, operation: str) -> None:
        if self.scan_thread and self.scan_thread.is_alive():
            self._show_transfer_notice(
                "Aksi Belum Bisa Diproses",
                "Tunggu sampai proses scan selesai sebelum menyalin atau memindahkan file.",
            )
            return
        if self.delete_thread and self.delete_thread.is_alive():
            self._show_transfer_notice(
                "Aksi Belum Bisa Diproses",
                "Tunggu sampai proses penghapusan selesai terlebih dahulu.",
            )
            return
        if self.transfer_thread and self.transfer_thread.is_alive():
            self._show_transfer_notice(
                "Aksi Belum Bisa Diproses",
                "Proses salin/pindah file masih berjalan.",
            )
            return

        selected_results = self._selected_results()
        if not selected_results:
            self._show_transfer_notice(
                "Belum Ada Pilihan",
                "Pilih minimal satu item hasil scan untuk disalin atau dipindahkan.",
            )
            return

        source_entries: List[Tuple[str, str, str]] = []
        seen_paths: set[str] = set()
        for result in selected_results:
            for base_label, relative_path, path in self._sources_for_result(result):
                normalized = os.path.normcase(str(path))
                if normalized in seen_paths:
                    continue
                seen_paths.add(normalized)
                source_entries.append((base_label, relative_path, str(path)))

        if not source_entries:
            self._show_transfer_notice(
                "Tidak Ada File Sumber",
                "Tidak ada file sumber yang dapat diproses dari pilihan saat ini.",
            )
            return

        destination_root = QFileDialog.getExistingDirectory(
            self,
            "Pilih folder tujuan untuk menyalin file" if operation == "copy" else "Pilih folder tujuan untuk memindahkan file",
            str(Path.home()),
        )
        if not destination_root:
            return

        operation_title = "Pindah File" if operation == "move" else "Salin File"
        operation_label = "dipindahkan" if operation == "move" else "disalin"
        details_lines = [
            f"Jumlah file sumber: {len(source_entries)}",
            f"Folder tujuan: {destination_root}",
            "Struktur sumber akan dipertahankan berdasarkan label folder (A, F1, F2, dst).",
        ]

        confirm_dialog = ConfirmOverlayDialog(
            self,
            operation_title,
            f"{len(source_entries)} file terkait akan {operation_label} ke folder tujuan terpilih.",
            "\n".join(details_lines),
            detail_title="Rencana transfer",
            confirm_button_text="Ya, Proses Transfer",
            confirm_footnote="Pastikan folder tujuan sudah benar sebelum melanjutkan proses transfer file.",
            processing_footnote="Transfer file sedang berjalan. Mohon tunggu sampai proses selesai.",
            processing_button_text="Mentransfer",
            success_title="Transfer File Berhasil",
            success_detail_title="Ringkasan transfer",
            success_footnote="Transfer selesai. Anda dapat menutup dialog ini.",
        )
        self.transfer_confirm_dialog = confirm_dialog
        confirm_dialog.confirmRequested.connect(
            lambda entries=list(source_entries), destination=str(destination_root), op=operation: self._start_selected_transfer_from_dialog(
                entries,
                destination,
                op,
            )
        )

        dialog_result = confirm_dialog.exec()
        if self.transfer_confirm_dialog is confirm_dialog:
            self.transfer_confirm_dialog = None
        confirm_dialog.deleteLater()

        if dialog_result != QDialog.Accepted:
            return

    def _start_selected_transfer_from_dialog(
        self,
        source_entries: List[Tuple[str, str, str]],
        destination_root: str,
        operation: str,
    ) -> None:
        if self.transfer_thread and self.transfer_thread.is_alive():
            return

        dialog = self.transfer_confirm_dialog
        if dialog is None:
            return

        operation_text = "pemindahan" if operation == "move" else "penyalinan"
        self.status_label.setText(
            "Memproses pemindahan file terkait..." if operation == "move" else "Memproses penyalinan file terkait..."
        )
        dialog.set_processing(
            True,
            f"Sedang memproses {operation_text} file ke folder tujuan. Mohon tunggu...",
        )
        self._set_transfer_processing_state(True)
        dialog.flush_visual_state()
        self._record_history(
            "Pindah file terpilih" if operation == "move" else "Salin file terpilih",
            "Diproses",
            f"Jumlah file sumber: {len(source_entries)}\nFolder tujuan: {destination_root}",
            "info",
        )

        QTimer.singleShot(
            0,
            lambda entries=list(source_entries), destination=str(destination_root), op=operation: self._launch_selected_transfer_worker(
                entries,
                destination,
                op,
            ),
        )

    def _launch_selected_transfer_worker(
        self,
        source_entries: List[Tuple[str, str, str]],
        destination_root: str,
        operation: str,
    ) -> None:
        if self.transfer_thread and self.transfer_thread.is_alive():
            return
        self.transfer_thread = threading.Thread(
            target=self._transfer_worker,
            args=(list(source_entries), destination_root, operation),
            daemon=True,
        )
        self.transfer_thread.start()

    def _transfer_worker(self, source_entries: List[Tuple[str, str, str]], destination_root: str, operation: str) -> None:
        processed_count = 0
        errors: List[str] = []
        destination_root_path = Path(destination_root)
        undo_operations: List[Dict[str, str]] = []

        for base_label, relative_path, source_path in source_entries:
            source = Path(source_path)
            try:
                if not source.exists() or not source.is_file():
                    raise FileNotFoundError(f"File tidak ditemukan: {source}")

                target = self._unique_destination_path(destination_root_path / base_label / Path(relative_path))
                target.parent.mkdir(parents=True, exist_ok=True)

                if operation == "move":
                    shutil.move(str(source), str(target))
                    undo_operations.append(
                        self._serialize_undo_operation(
                            "move_path",
                            source=str(target),
                            destination=str(source),
                        )
                    )
                else:
                    shutil.copy2(source, target)
                    undo_operations.append(
                        self._serialize_undo_operation("delete_path", path=str(target))
                    )
                processed_count += 1
            except Exception as exc:
                errors.append(f"{source}: {exc}")

        undo_payload = {
            "label": "Pindah file terpilih" if operation == "move" else "Salin file terpilih",
            "detail": f"Folder tujuan: {destination_root}",
            "operations": undo_operations,
            "file_name": Path(source_entries[0][2]).name if len(source_entries) == 1 else f"{len(source_entries)} file",
        } if undo_operations else None

        self.ui_queue.put(
            (
                "transfer_done",
                {
                    "operation": operation,
                    "processed_count": processed_count,
                    "error_count": len(errors),
                    "destination_root": destination_root,
                    "errors": errors,
                    "undo_action": undo_payload,
                },
            )
        )

    def sync_selected_green(self) -> None:
        if self.scan_thread and self.scan_thread.is_alive():
            self._show_transfer_notice(
                "Aksi Belum Bisa Diproses",
                "Tunggu sampai proses scan selesai sebelum menyinkronkan file.",
            )
            return
        if self.delete_thread and self.delete_thread.is_alive():
            self._show_transfer_notice(
                "Aksi Belum Bisa Diproses",
                "Tunggu sampai proses penghapusan selesai terlebih dahulu.",
            )
            return
        if self.transfer_thread and self.transfer_thread.is_alive():
            self._show_transfer_notice(
                "Aksi Belum Bisa Diproses",
                "Proses transfer file masih berjalan.",
            )
            return

        selection_model = self.results_table.selectionModel()
        selected_rows = sorted({index.row() for index in selection_model.selectedRows()}) if selection_model else []
        
        results_to_sync: List[MatchResult] = []
        if selected_rows:
            for row_index in selected_rows:
                result = self._result_for_table_row(row_index)
                if result and result.tree_tag == "exact_match" and self._actual_missing_compare_labels(result):
                    results_to_sync.append(result)
            if not results_to_sync:
                QMessageBox.information(
                    self, 
                    APP_TITLE, 
                    "Pilihan Anda tidak memiliki baris data duplikat (hijau) yang mising folder pembanding."
                )
                return
        else:
            for result in self.result_rows:
                if result.tree_tag == "exact_match" and self._actual_missing_compare_labels(result):
                    results_to_sync.append(result)
            
            if not results_to_sync:
                QMessageBox.information(
                    self, 
                    APP_TITLE, 
                    "Tidak ada data hijau (duplikat) yang memiliki file mising untuk disinkronkan."
                )
                return
            
            if QMessageBox.question(
                self, 
                "Sync Semua Hijau Missing", 
                f"Anda tidak sedang memilih baris. Sinkronkan semua {len(results_to_sync)} data duplikat yang mising tersebut ke folder pembanding masing-masing?",
            ) != QMessageBox.StandardButton.Yes:
                return

        total_files = 0
        destinations_set = set()
        
        for result in results_to_sync:
            plan = self._compare_sync_plan(result)
            total_files += len(plan["create"])
            destinations_set.update(plan["create_labels"])

        if total_files == 0:
            QMessageBox.information(self, APP_TITLE, "Folder pembanding yang tertuju kemungkinan belum disetel.")
            return

        dest_str = ", ".join(sorted(destinations_set))
        summary = f"Akan menyinkronkan (salin massal) {len(results_to_sync)} file dari Folder A."
        details_lines = [
            f"Total penyalinan file individu: {total_files}",
            f"Ke folder target: {dest_str}",
            "Catatan: Hanya akan menambahkan file yang saat ini missing di kolom 'Tidak ada di'."
        ]

        confirm_dialog = ConfirmOverlayDialog(
            self,
            "Konfirmasi Sinkronisasi Folder Pembanding",
            summary,
            "\n".join(details_lines),
            detail_title="Detail Sync",
            confirm_button_text="Proses Sinkronisasi",
            success_title="Sinkronisasi Hijau Berhasil",
            success_detail_title="Status Eksekusi",
            processing_button_text="Menyinkronkan",
            confirm_footnote="Otomatis menyalin data Folder A ke masing-masing folder pembanding.",
            success_footnote="Seluruh operasi berhasil dan segera diperbarui pada antarmuka."
        )
        self.transfer_confirm_dialog = confirm_dialog
        confirm_dialog.confirmRequested.connect(
            lambda: self._execute_bulk_sync_green_from_dialog(results_to_sync)
        )
        dialog_result = confirm_dialog.exec()
        if self.transfer_confirm_dialog is confirm_dialog:
            self.transfer_confirm_dialog = None
        confirm_dialog.deleteLater()

    def _execute_bulk_sync_green_from_dialog(self, results: List[MatchResult]) -> None:
        if self.transfer_thread and self.transfer_thread.is_alive():
            return

        dialog = self.transfer_confirm_dialog
        if dialog is None:
            return

        self.status_label.setText("Memulai sinkronisasi file missing...")
        dialog.set_processing(
            True,
            "Sedang menyalin file ke folder pembanding. Mohon tunggu...",
        )
        self._set_transfer_processing_state(True)
        dialog.flush_visual_state()

        self._record_history(
            "Sync Massal Pembanding",
            "Diproses",
            f"Sinkronisasi atas {len(results)} file hijau (duplikat).",
            "info"
        )
        QTimer.singleShot(0, lambda: self._launch_bulk_sync_green_worker(results))

    def _launch_bulk_sync_green_worker(self, results: List[MatchResult]) -> None:
        if self.transfer_thread and self.transfer_thread.is_alive():
            return
            
        tasks: List[Dict[str, Any]] = []
        for result in results:
            plan = self._compare_sync_plan(result)
            if plan["source"] and plan["create"]:
                tasks.append({
                    "source_path": plan["source"],
                    "target_path": str(result.target_path),
                    "create_paths": plan["create"],
                    "create_labels": plan["create_labels"],
                })
            
        self.transfer_thread = threading.Thread(
            target=self._bulk_sync_green_worker,
            args=(tasks,),
            daemon=True
        )
        self.transfer_thread.start()

    def _bulk_sync_green_worker(self, tasks: List[Dict[str, Any]]) -> None:
        processed_count = 0
        errors: List[str] = []
        undo_operations: List[Dict[str, str]] = []
        action_dir = self._create_undo_action_dir("bulk_sync")
        
        all_created_labels: List[str] = []
        bulk_updates: List[Dict[str, Any]] = []

        for task in tasks:
            source = Path(task["source_path"])
            local_processed = 0
            success_creates = []
            success_labels = []
            if not source.exists() or not source.is_file():
                errors.append(f"File sumber tidak ditemukan: {source}")
                continue

            for dest_idx, dest_path_str in enumerate(task["create_paths"]):
                dest_path = Path(dest_path_str)
                dest_label = task["create_labels"][dest_idx]
                try:
                    dest_path.parent.mkdir(parents=True, exist_ok=True)
                    if dest_path.exists():
                        raise FileExistsError(f"File tujuan sudah ada: {dest_path}")
                    shutil.copy2(source, dest_path)
                    undo_operations.insert(0, self._serialize_undo_operation("delete_path", path=str(dest_path)))
                    success_creates.append(str(dest_path))
                    success_labels.append(dest_label)
                    local_processed += 1
                except Exception as e:
                    errors.append(f"{dest_path}: {e}")

            if local_processed > 0:
                bulk_updates.append({
                    "target_path": task["target_path"],
                    "created_paths": success_creates,
                    "created_labels": success_labels,
                })
                processed_count += local_processed
                all_created_labels.extend(success_labels)

        undo_payload = {
            "label": "Sinkronisasi Massal (Hijau)",
            "detail": f"Sinkronisasi ke label folder: {', '.join(set(all_created_labels))}",
            "operations": undo_operations,
            "action_dir": str(action_dir),
            "file_name": f"{len(tasks)} file massal",
        } if undo_operations else None
        if not undo_payload:
            self._cleanup_undo_action_dir(str(action_dir))

        self.ui_queue.put((
            "bulk_compare_sync_done",
            {
                "bulk_updates": bulk_updates,
                "processed_count": processed_count,
                "error_count": len(errors),
                "errors": errors,
                "undo_action": undo_payload,
            }
        ))

    def clear_results(self, reset_status: bool = True) -> None:
        had_results = bool(self.result_rows)
        if self.delete_confirm_dialog is not None:
            self.delete_confirm_dialog.force_close(QDialog.Rejected)
            self.delete_confirm_dialog = None
        if self.transfer_confirm_dialog is not None:
            self.transfer_confirm_dialog.force_close(QDialog.Rejected)
            self.transfer_confirm_dialog = None
        if self.delete_processing_dialog is not None:
            self.delete_processing_dialog.close()
            self.delete_processing_dialog = None
        if self.undo_processing_dialog is not None:
            self.undo_processing_dialog.close()
            self.undo_processing_dialog = None
        self.result_rows = []
        self._pending_scan_results = None
        self._awaiting_scan_finalize = False
        self.current_selected_result = None
        self.table_model.set_rows([])
        if self.results_table.selectionModel() is not None:
            self.results_table.clearSelection()
            self.results_table.selectionModel().setCurrentIndex(QModelIndex(), QItemSelectionModel.SelectionFlag.NoUpdate)

        self.export_csv_button.setEnabled(False)
        self.export_excel_button.setEnabled(False)
        self.delete_button.setEnabled(False)
        self.delete_all_button.setEnabled(False)
        self.copy_button.setEnabled(False)
        self.move_button.setEnabled(False)
        self.sync_button.setEnabled(False)

        for label in self.stat_labels.values():
            label.setText("0")

        self.progress_animation.stop()
        self.progress_bar.setValue(0)
        self.progress_badge.setText("0%")
        self.progress_card.setVisible(False)
        self._last_queued_progress = -1.0
        self._last_queued_progress_text = ""
        with self._progress_lock:
            self._pending_progress = None
        self._update_table_empty_state()
        self._reset_detail_panel()

        if reset_status:
            self.status_label.setText("Hasil dibersihkan. Silakan pilih folder baru.")
            if had_results:
                self._record_history(
                    "Reset hasil",
                    "Sukses",
                    "Hasil scan dibersihkan dari tampilan aplikasi.",
                    "info",
                )

    def delete_selected(self) -> None:
        selected_rows = sorted({index.row() for index in self.results_table.selectionModel().selectedRows()})
        if not selected_rows:
            QMessageBox.information(self, APP_TITLE, "Pilih minimal satu file dari hasil scan.")
            return

        deletable_results: List[MatchResult] = []
        skipped_counts = {"different_content": 0, "only_target": 0}
        for row_index in selected_rows:
            result = self._result_for_table_row(row_index)
            if result is None:
                continue
            if self._is_result_deletable(result):
                deletable_results.append(result)
            else:
                skipped_counts[result.tree_tag] = skipped_counts.get(result.tree_tag, 0) + 1

        if not deletable_results:
            QMessageBox.information(
                self,
                APP_TITLE,
                "Pilihan Anda belum termasuk kategori yang diizinkan untuk dihapus. Aktifkan opsi merah/oranye jika memang diperlukan.",
            )
            return
        skipped_messages: List[str] = []
        if skipped_counts.get("different_content", 0):
            skipped_messages.append(
                f"{skipped_counts['different_content']} item merah diabaikan karena izin hapus merah belum aktif."
            )
        if skipped_counts.get("only_target", 0):
            skipped_messages.append(
                f"{skipped_counts['only_target']} item oranye diabaikan karena izin hapus oranye belum aktif."
            )
        if skipped_messages:
            QMessageBox.information(self, APP_TITLE, " ".join(skipped_messages))

        self._confirm_and_delete(deletable_results)

    def delete_all_results(self) -> None:
        deletable_results = self._deletable_results()
        if not deletable_results:
            QMessageBox.information(
                self,
                APP_TITLE,
                "Belum ada hasil yang sesuai dengan izin hapus saat ini.",
            )
            return
        self._confirm_and_delete(deletable_results)

    def _confirm_and_delete(self, results: List[MatchResult]) -> None:
        if self.delete_thread and self.delete_thread.is_alive():
            QMessageBox.information(self, APP_TITLE, "Proses penghapusan masih berjalan.")
            return

        delete_paths = [result.target_path for result in results]
        scope_counts = self._delete_scope_counts(results)
        preview = "\n".join(f"[{result.status_text}] {result.target_path}" for result in results[:10])
        extra = "" if len(delete_paths) <= 10 else f"\n... dan {len(delete_paths) - 10} file lainnya"
        mode_text = "Trash Internal" if self._current_delete_mode() == "internal_trash" else "hapus permanen"
        summary_lines = [
            f"Anda akan memproses {len(delete_paths)} file dari Folder A.",
            (
                f"Kategori: hijau {scope_counts['exact_match']} | "
                f"merah {scope_counts['different_content']} | "
                f"oranye {scope_counts['only_target']}."
            ),
            f"Mode penghapusan: {mode_text}.",
        ]
        if scope_counts["different_content"] or scope_counts["only_target"]:
            summary_lines.append("Perhatian: pilihan ini mencakup hasil merah/oranye yang berisiko lebih tinggi.")
        summary = "\n".join(summary_lines)
        details = f"{preview}{extra}"

        confirm_dialog = ConfirmOverlayDialog(
            self,
            "Konfirmasi Penghapusan File",
            summary,
            details,
        )
        self.delete_confirm_dialog = confirm_dialog
        confirm_dialog.confirmRequested.connect(
            lambda selected_paths=list(delete_paths): self._start_delete_from_confirm_dialog(selected_paths)
        )

        result = confirm_dialog.exec()
        if self.delete_confirm_dialog is confirm_dialog:
            self.delete_confirm_dialog = None
        confirm_dialog.deleteLater()

        if result != QDialog.Accepted:
            return

    def _start_delete_from_confirm_dialog(self, paths: List[Path]) -> None:
        if self.delete_thread and self.delete_thread.is_alive():
            return

        dialog = self.delete_confirm_dialog
        if dialog is None:
            return

        self.status_label.setText("Memproses penghapusan file terpilih...")
        dialog.set_processing(True, f"Sedang memproses {len(paths)} file dari Folder A. Mohon tunggu...")
        self._set_delete_processing_state(True)
        dialog.flush_visual_state()

        use_internal_trash = self._current_delete_mode() == "internal_trash"
        trash_text = "Trash Internal" if use_internal_trash else "Permanen"
        if len(paths) == 1:
            detail_msg = f"File: {Path(paths[0]).name}\nPath: {paths[0]}\nMode: {trash_text}"
        else:
            detail_msg = f"Jumlah file: {len(paths)}\nMode: {trash_text}"

        self._record_history(
            "Penghapusan file",
            "Diproses",
            detail_msg,
            "warning",
        )

        QTimer.singleShot(
            0,
            lambda delete_paths=list(paths), trash_mode=use_internal_trash: self._launch_delete_worker(
                delete_paths, trash_mode
            ),
        )

    def _launch_delete_worker(self, paths: List[Path], use_internal_trash: bool) -> None:
        if self.delete_thread and self.delete_thread.is_alive():
            return

        self.delete_thread = threading.Thread(
            target=self._delete_worker,
            args=(list(paths), use_internal_trash),
            daemon=True,
        )
        self.delete_thread.start()

    def _delete_worker(self, paths: List[Path], use_internal_trash: bool) -> None:
        deleted_count = 0
        deleted_paths: List[Path] = []
        errors: List[str] = []
        undo_payload: Optional[Dict[str, Any]] = None
        action_dir = self._create_undo_action_dir("delete")
        undo_operations: List[Dict[str, str]] = []
        trash_entries_payload: List[Dict[str, Any]] = []
        trash_storage_dir = self._trash_storage_dir()

        for path in paths:
            try:
                if not path.exists() or not path.is_file():
                    raise FileNotFoundError(f"File tidak ditemukan: {path}")
                original_size = path.stat().st_size
                if use_internal_trash:
                    trash_target = trash_storage_dir / f"{uuid4().hex}{path.suffix}"
                    shutil.move(str(path), str(trash_target))
                    entry_id = uuid4().hex
                    trash_entries_payload.append(
                        {
                            "entry_id": entry_id,
                            "original_path": str(path),
                            "trash_path": str(trash_target),
                            "deleted_at": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            "size": original_size,
                        }
                    )
                    undo_operations.append(
                        self._serialize_undo_operation(
                            "move_path",
                            source=str(trash_target),
                            destination=str(path),
                            trash_entry_id=entry_id,
                        )
                    )
                else:
                    backup_path = self._backup_file_for_undo(path, action_dir, "delete")
                    path.unlink()
                    undo_operations.append(
                        self._serialize_undo_operation(
                            "restore_copy",
                            backup=str(backup_path),
                            destination=str(path),
                        )
                    )
                deleted_count += 1
                deleted_paths.append(path)
            except Exception as exc:
                errors.append(f"{path}: {exc}")

        if undo_operations:
            undo_payload = {
                "label": "Penghapusan file",
                "detail": (
                    f"Jumlah file: {deleted_count}\n"
                    f"Mode: {'Trash Internal' if use_internal_trash else 'Permanen'}"
                ),
                "operations": undo_operations,
                "action_dir": str(action_dir),
                "file_name": Path(paths[0]).name if len(paths) == 1 else f"{deleted_count} file",
            }
        else:
            self._cleanup_undo_action_dir(str(action_dir))

        self.ui_queue.put(
            (
                "delete_done",
                {
                    "deleted_count": deleted_count,
                    "deleted_paths": [str(path) for path in deleted_paths],
                    "errors": errors,
                    "undo_action": undo_payload,
                    "trash_entries": trash_entries_payload,
                },
            )
        )

    def export_csv(self) -> None:
        if not self.result_rows:
            QMessageBox.information(self, APP_TITLE, "Belum ada hasil scan untuk disimpan.")
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Simpan hasil ke CSV",
            str(Path.home() / "hasil_scan.csv"),
            "CSV files (*.csv)",
        )
        if not save_path:
            return

        undo_payload: Optional[Dict[str, Any]]
        existing_csv = Path(save_path)
        if existing_csv.exists() and existing_csv.is_file():
            action_dir = self._create_undo_action_dir("export_csv")
            backup_path = self._backup_file_for_undo(existing_csv, action_dir, "export_csv")
            undo_payload = {
                "label": "Export CSV",
                "detail": save_path,
                "operations": [
                    self._serialize_undo_operation(
                        "restore_copy",
                        backup=str(backup_path),
                        destination=save_path,
                    )
                ],
                "action_dir": str(action_dir),
            }
        else:
            undo_payload = {
                "label": "Export CSV",
                "detail": save_path,
                "operations": [self._serialize_undo_operation("delete_path", path=save_path)],
            }

        with open(save_path, "w", newline="", encoding="utf-8-sig") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(
                [
                    "status",
                    "target_path",
                    "target_relative_path",
                    "size_bytes",
                    "size_display",
                    "found_in_folders",
                    "exact_match_paths",
                    "different_content_paths",
                    "missing_in_folders",
                    "match_type",
                ]
            )
            for row in self.result_rows:
                writer.writerow(
                    [
                        row.status_text,
                        str(row.target_path),
                        row.target_relative_path,
                        row.size,
                        self._format_size(row.size),
                        row.exact_folder_labels,
                        row.exact_paths_text,
                        row.diff_paths_text,
                        ", ".join(row.missing_from_folders) if row.missing_from_folders else "-",
                        row.match_type,
                    ]
                )

        self._push_undo_action(undo_payload)
        self._record_history("Export CSV", "Sukses", save_path, "success")
        QMessageBox.information(self, APP_TITLE, f"Berhasil menyimpan CSV:\n{save_path}")

    def export_excel(self) -> None:
        if not self.result_rows:
            QMessageBox.information(self, APP_TITLE, "Belum ada hasil scan untuk disimpan.")
            return
        if not self.openpyxl_available:
            QMessageBox.warning(self, APP_TITLE, "Modul openpyxl belum terpasang. Jalankan: pip install openpyxl")
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Simpan hasil ke Excel",
            str(Path.home() / "hasil_scan.xlsx"),
            "Excel files (*.xlsx)",
        )
        if not save_path:
            return

        undo_payload: Optional[Dict[str, Any]]
        existing_excel = Path(save_path)
        if existing_excel.exists() and existing_excel.is_file():
            action_dir = self._create_undo_action_dir("export_excel")
            backup_path = self._backup_file_for_undo(existing_excel, action_dir, "export_excel")
            undo_payload = {
                "label": "Export Excel",
                "detail": save_path,
                "operations": [
                    self._serialize_undo_operation(
                        "restore_copy",
                        backup=str(backup_path),
                        destination=save_path,
                    )
                ],
                "action_dir": str(action_dir),
            }
        else:
            undo_payload = {
                "label": "Export Excel",
                "detail": save_path,
                "operations": [self._serialize_undo_operation("delete_path", path=save_path)],
            }

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Hasil Scan"
        headers = [
            "Status",
            "Path File di Folder A",
            "Relative Path",
            "Ukuran (bytes)",
            "Ukuran",
            "Ditemukan di Folder",
            "Path File yang Cocok",
            "Path Nama Sama Isi Beda",
            "Tidak Ada di Folder",
            "Pencocokan",
        ]
        worksheet.append(headers)

        fill_green = PatternFill("solid", fgColor="E8FFF4")
        fill_red = PatternFill("solid", fgColor="FFF0F2")
        fill_orange = PatternFill("solid", fgColor="FFF6E8")

        for row in self.result_rows:
            worksheet.append(
                [
                    row.status_text,
                    str(row.target_path),
                    row.target_relative_path,
                    row.size,
                    self._format_size(row.size),
                    row.exact_folder_labels,
                    row.exact_paths_text,
                    row.diff_paths_text,
                    ", ".join(row.missing_from_folders) if row.missing_from_folders else "-",
                    row.match_type,
                ]
            )
            excel_row = worksheet.max_row
            fill = fill_green if row.tree_tag == "exact_match" else fill_red if row.tree_tag == "different_content" else fill_orange
            for column in range(1, len(headers) + 1):
                worksheet.cell(row=excel_row, column=column).fill = fill

        for column_letter, width in {"A": 22, "B": 50, "C": 30, "D": 16, "E": 14, "F": 18, "G": 70, "H": 70, "I": 22, "J": 14}.items():
            worksheet.column_dimensions[column_letter].width = width

        workbook.save(save_path)
        self._push_undo_action(undo_payload)
        self._record_history("Export Excel", "Sukses", save_path, "success")
        QMessageBox.information(self, APP_TITLE, f"Berhasil menyimpan Excel:\n{save_path}")

def _forward_exception_to_app(title: str, summary: str, details: str) -> None:
    sys.stderr.write(f"{title}\n{summary}\n{details}\n")
    app = QApplication.instance()
    if isinstance(app, SafeApplication):
        app.report_error(title, summary, details)
    else:
        QMessageBox.critical(None, title or APP_TITLE, summary or "Terjadi kesalahan pada aplikasi.")


def _handle_uncaught_exception(exc_type, exc_value, exc_traceback) -> None:
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return

    _forward_exception_to_app(
        "Terjadi kesalahan tak terduga",
        str(exc_value) or exc_type.__name__,
        "".join(traceback.format_exception(exc_type, exc_value, exc_traceback)),
    )


def _handle_thread_exception(args: threading.ExceptHookArgs) -> None:
    _forward_exception_to_app(
        "Terjadi kesalahan pada proses latar belakang",
        str(args.exc_value) or args.exc_type.__name__,
        "".join(traceback.format_exception(args.exc_type, args.exc_value, args.exc_traceback)),
    )


def main() -> int:
    if sys.platform == "win32":
        try:
            import ctypes
            myappid = f'tonzdev.foldercomparedeleteapp.{APP_VERSION}'
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
        except Exception:
            pass

    sys.excepthook = _handle_uncaught_exception
    threading.excepthook = _handle_thread_exception

    app = SafeApplication(sys.argv)
    app.setStyle("Fusion")

    window = FolderCompareDeleteApp()
    window.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
